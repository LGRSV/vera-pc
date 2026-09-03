"""
Os gráficos da planilha de taxa de falha que o gestor mandou.

Pedido (25/08): «faça a partir dessa planilha por favor os gráficos». Então a fonte
é o arquivo dele, não os dados desta base — as abas originais ficam intactas e os
gráficos entram em abas novas, lendo o que está escrito lá.

A planilha tem dois níveis que não batem entre si: os quadros de resumo e o ROL
itemizado das ocorrências. Em vez de escolher um em silêncio, cada nível ganha a
própria aba de gráficos e a divergência fica escrita numa terceira.

Rodar: python3 scripts/graficos_taxa_falha_upload.py <planilha.xlsx> [saida.xlsx]
"""

import json
import os
import sys
from collections import Counter

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PADRAO = os.path.join(RAIZ, "dist", "TAXA_DE_FALHA_COM_GRAFICOS.xlsx")

TITULO = Font(bold=True, color="FFFFFF", size=10)
FUNDO = PatternFill("solid", fgColor="1F3864")
BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)
MESES = ["jan", "fev", "mar", "abr", "mai", "jun",
         "jul", "ago", "set", "out", "nov", "dez"]


def texto(v):
    return "" if v is None else str(v).strip()


def ler(caminho):
    """Lê os dois níveis da planilha do gestor: os quadros e o rol itemizado.

    São duas aberturas do mesmo arquivo: `wb` é o que vai ser gravado de volta, com
    as fórmulas originais preservadas; `vals` é a leitura só dos valores calculados.
    Quem lê o número tem de ser o segundo — a taxa e parte do parque são fórmula na
    planilha do gestor, e copiar o texto delas para uma aba nova aponta para células
    erradas.
    """
    wb = openpyxl.load_workbook(caminho)
    vals = openpyxl.load_workbook(caminho, data_only=True)
    ws = vals["Taxa de falha"]
    taxa, familia = {}, None
    for linha in ws.iter_rows(values_only=True):
        primeiro = texto(linha[0]).upper()
        if primeiro in ("RELIGADORES", "REGULADORES"):
            familia = "Religador" if primeiro == "RELIGADORES" else "Regulador"
        elif familia and primeiro.startswith("20"):
            taxa.setdefault(familia, []).append(
                {"ano": texto(linha[0]), "parque": linha[1], "ocorrencias": linha[2],
                 "falharam": linha[3], "taxa": linha[4]})

    ws = vals["Falhados"]
    quadros, cab, atual = {}, None, None
    rol, derrubados, modo = [], [], None
    for linha in ws.iter_rows(values_only=True):
        primeiro = texto(linha[0])
        alto = primeiro.upper()
        if alto.startswith("RELIGADORES —") or alto.startswith("REGULADORES —"):
            atual = "Religador" if alto.startswith("RELIGADORES") else "Regulador"
            cab, modo = None, "quadro"
            continue
        if alto.startswith("ROL DAS"):
            modo, atual = "rol", None
            continue
        if alto.startswith("O QUE A REVISÃO DERRUBOU"):
            modo = "derrubado"
            continue
        if not primeiro:
            continue
        if modo == "quadro" and atual:
            if primeiro == "Ano":
                cab = [texto(c) for c in linha[1:] if texto(c)]
            elif cab and primeiro.startswith("20"):
                quadros.setdefault(atual, {})[primeiro] = dict(zip(cab, linha[1:]))
        elif modo == "rol" and primeiro != "Ativo":
            rol.append({"ativo": primeiro, "familia": texto(linha[1]),
                        "ano": texto(linha[2]), "peca": texto(linha[3]),
                        "ss": texto(linha[4]), "data": texto(linha[5]),
                        "troca": texto(linha[6])})
        elif modo == "derrubado" and primeiro != "Ativo":
            derrubados.append({"ativo": primeiro, "familia": texto(linha[1]),
                               "ano": texto(linha[2]), "peca": texto(linha[3])})
    return wb, taxa, quadros, rol, derrubados


def grafico(ws, tipo, titulo, colunas, faixa, ancora, altura=7.5, largura=14,
            pct=False, legenda=True):
    """Gráfico nativo, com as correções que o Excel exige: categoria como texto,
    eixos declarados visíveis e o de categoria embaixo."""
    ch = LineChart() if tipo == "linha" else BarChart()
    ch.title = titulo
    ch.height, ch.width, ch.style = altura, largura, 2
    if tipo != "linha":
        ch.type, ch.gapWidth = "col", 60
    for col in colunas:
        ch.add_data(Reference(ws, min_col=col, min_row=faixa[0] - 1, max_row=faixa[1]),
                    titles_from_data=True)
    ref = f"'{ws.title}'!$A${faixa[0]}:$A${faixa[1]}"
    for s in ch.series:
        s.cat = AxDataSource(strRef=StrRef(f=ref))
        if tipo == "linha":
            s.smooth = False
            s.marker.symbol, s.marker.size = "circle", 6
    ch.x_axis.delete = ch.y_axis.delete = False
    ch.x_axis.axPos, ch.y_axis.axPos = "b", "l"
    if pct:
        ch.y_axis.numFmt = "0.00%"
    if not legenda:
        ch.legend = None
    ws.add_chart(ch, ancora)


def quadro(ws, titulo, colunas, linhas, pct=False):
    """Escreve um bloco rótulo/valores e devolve (primeira, última) linha."""
    ws.append([])
    ws.append([titulo] + colunas)
    for c in range(1, len(colunas) + 2):
        cel = ws.cell(row=ws.max_row, column=c)
        cel.font, cel.fill = TITULO, FUNDO
        cel.alignment = Alignment(horizontal="center", wrap_text=True)
    primeira = ws.max_row + 1
    for rotulo, valores in linhas:
        ws.append([rotulo] + list(valores))
    for linha in ws.iter_rows(min_row=primeira, max_row=ws.max_row):
        for cel in linha:
            cel.border = BORDA
            if cel.column > 1:
                cel.alignment = Alignment(horizontal="center")
                if pct:
                    cel.number_format = "0.00%"
    return primeira, ws.max_row


def aba_quadros(wb, taxa, quadros):
    ws = wb.create_sheet("Gráficos · quadros", 1)
    ws.column_dimensions["A"].width = 26
    for c in "BCDEF":
        ws.column_dimensions[c].width = 14
    ws.append(["GRÁFICOS DOS QUADROS DE RESUMO DA PLANILHA"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.append(["Lidos das abas «Taxa de falha» e «Falhados», sem recalcular nada."])

    anos = [r["ano"] for r in taxa["Religador"]]
    a = quadro(ws, "Taxa de falha", anos,
               [(f, [r["taxa"] for r in taxa[f]]) for f in ("Religador", "Regulador")],
               pct=True)
    b = quadro(ws, "Equipamentos que falharam", anos,
               [(f, [r["falharam"] for r in taxa[f]]) for f in ("Religador", "Regulador")])
    c = quadro(ws, "Parque do ano", anos,
               [(f, [r["parque"] for r in taxa[f]]) for f in ("Religador", "Regulador")])
    d = quadro(ws, "Ocorrências", anos,
               [(f, [r["ocorrencias"] for r in taxa[f]]) for f in ("Religador", "Regulador")])

    blocos = {}
    for fam in ("Religador", "Regulador"):
        q = quadros.get(fam, {})
        anos_q = sorted(q)
        pecas = [p for p in next(iter(q.values()), {}) if p.lower() != "total"]
        blocos[fam] = quadro(ws, f"{fam} — por peça", anos_q,
                             [(p, [q[an].get(p) for an in anos_q]) for p in pecas])

    n = len(anos) + 1
    grafico(ws, "barra", "Taxa de falha por tipo e ano", list(range(2, n + 1)), a,
            "H3", pct=True)
    grafico(ws, "barra", "Equipamentos que falharam", list(range(2, n + 1)), b, "H19")
    grafico(ws, "barra", "Parque do ano", list(range(2, n + 1)), c, "H35")
    grafico(ws, "barra", "Ocorrências", list(range(2, n + 1)), d, "R3")
    for i, fam in enumerate(("Religador", "Regulador")):
        largura = len(quadros.get(fam, {})) + 1
        grafico(ws, "barra", f"{fam} — por peça", list(range(2, largura + 1)),
                blocos[fam], f"R{19 + i * 16}")
    return ws


def aba_rol(wb, rol, derrubados):
    ws = wb.create_sheet("Gráficos · rol", 2)
    ws.column_dimensions["A"].width = 26
    for c in "BCDEF":
        ws.column_dimensions[c].width = 14
    ws.append(["GRÁFICOS DO ROL ITEMIZADO"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.append([f"Contados linha a linha nas {len(rol)} ocorrências confirmadas do rol. "
               "O rol tem 2024, que os quadros de resumo não trazem."])

    anos = sorted({r["ano"] for r in rol})
    fam_ano = Counter((r["familia"], r["ano"]) for r in rol)
    a = quadro(ws, "Ocorrências no rol", anos,
               [(f.capitalize(), [fam_ano[(f, an)] for an in anos])
                for f in ("religador", "regulador")])

    ativos = {}
    for f in ("religador", "regulador"):
        ativos[f] = [len({r["ativo"] for r in rol if r["familia"] == f and r["ano"] == an})
                     for an in anos]
    b = quadro(ws, "Equipamentos distintos", anos,
               [(f.capitalize(), ativos[f]) for f in ("religador", "regulador")])

    blocos = {}
    for f in ("religador", "regulador"):
        pecas = sorted({r["peca"] for r in rol if r["familia"] == f})
        cont = Counter((r["peca"], r["ano"]) for r in rol if r["familia"] == f)
        blocos[f] = quadro(ws, f"{f.capitalize()} — por peça", anos,
                           [(p.capitalize(), [cont[(p, an)] for an in anos])
                            for p in pecas])

    troca = Counter((r["troca"], r["ano"]) for r in rol)
    c = quadro(ws, "Troca já executada", anos,
               [(t.capitalize(), [troca[(t, an)] for an in anos]) for t in ("sim", "não")])

    mes = Counter((r["data"][3:5], r["ano"]) for r in rol if len(r["data"]) >= 10)
    d = quadro(ws, "Ocorrências por mês", anos,
               [(m, [mes[(f"{i:02d}", an)] for an in anos]) for i, m in enumerate(MESES, 1)])

    der = Counter((r["familia"], r["ano"]) for r in derrubados)
    e = quadro(ws, f"A revisão derrubou ({len(derrubados)})", anos,
               [(f.capitalize(), [der[(f, an)] for an in anos])
                for f in ("religador", "regulador")])

    n = len(anos) + 1
    cols = list(range(2, n + 1))
    grafico(ws, "barra", "Ocorrências no rol, por ano", cols, a, "H3")
    grafico(ws, "barra", "Equipamentos distintos, por ano", cols, b, "H19")
    grafico(ws, "barra", "Religador — peça por ano", cols, blocos["religador"], "H35")
    grafico(ws, "barra", "Regulador — peça por ano", cols, blocos["regulador"], "R3")
    grafico(ws, "barra", "Troca já executada", cols, c, "R19")
    grafico(ws, "linha", "Ocorrências por mês", cols, d, "R35", largura=17)
    grafico(ws, "barra", "O que a revisão derrubou", cols, e, "H51")
    return ws


def aba_divergencias(wb, taxa, quadros, rol):
    ws = wb.create_sheet("Onde a planilha diverge", 3)
    ws.column_dimensions["A"].width = 112
    linhas = ["ONDE A PLANILHA DIVERGE DE SI MESMA", ""]

    for fam, chave in (("Religador", "religador"), ("Regulador", "regulador")):
        for r in taxa[fam]:
            ano = r["ano"][:4]
            no_rol = sum(1 for x in rol if x["familia"] == chave and x["ano"] == ano)
            ativos = len({x["ativo"] for x in rol
                          if x["familia"] == chave and x["ano"] == ano})
            linhas.append(
                f"{fam} {ano}: a aba «Taxa de falha» diz {r['ocorrencias']} ocorrências e "
                f"{r['falharam']} equipamentos; o rol tem {no_rol} linhas em {ativos} "
                f"equipamentos. É a diferença que faz a taxa.")
    linhas += [
        "",
        "O rol traz 2024, que nenhum quadro de resumo mostra. Por isso a aba de gráficos "
        "do rol tem três anos e a dos quadros tem dois.",
        "",
    ]
    for fam, chave in (("Religador", "religador"), ("Regulador", "regulador")):
        for ano, dados in sorted(quadros.get(fam, {}).items()):
            soma = sum(v for k, v in dados.items()
                       if k.lower() != "total" and isinstance(v, (int, float)))
            total = dados.get("Total")
            if total is not None and soma != total:
                linhas.append(f"{fam} {ano}, quadro por peça: as colunas somam {soma}, "
                              f"mas a coluna Total diz {total}.")
            no_rol = Counter(x["peca"] for x in rol
                             if x["familia"] == chave and x["ano"] == ano)
            difs = [f"{p} {dados.get(p.capitalize(), dados.get(p, 0))}→{n}"
                    for p, n in sorted(no_rol.items())
                    if (dados.get(p.capitalize(), dados.get(p, 0)) or 0) != n]
            if difs:
                linhas.append(f"{fam} {ano}, peça a peça, quadro → rol: {', '.join(difs)}.")
    linhas += [
        "",
        "A nota da aba «Taxa de falha» diz que 2026 foi ajustado pela fração decorrida "
        "(61%), mas a taxa gravada é a divisão direta: religador 31 ÷ 1.197 = 2,59% e "
        "regulador 7 ÷ 202 = 3,47%. Não há anualização no número.",
        "",
        "O parque também não é o mesmo dos dois lados: esta planilha usa 1.197 "
        "religadores e 202 reguladores em 2026; a aba «Acumulado 2026» usa a base de "
        "janeiro do gestor mais a expansão realizada, que fecha agosto em 1.294 e 190. "
        "Com 1.294 no lugar de 1.197, a taxa do religador cai de 2,59% para 2,40%.",
        "",
        "Nada aqui foi corrigido — os gráficos saem do que está escrito em cada nível, e "
        "as abas originais ficaram como vieram.",
    ]
    for t in linhas:
        ws.append([t])
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True,
                                                               vertical="top")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    return ws


ACUMULADO = [
    ("Parque", "parque"), ("Entrantes — com o acervo herdado", "com_acervo"),
    ("Indisponibilidade acumulada — só 2026", "so_2026"),
    ("Resolvidos — todas as cadeias", "resolvidos"),
    ("Resolvidos pelo COEP", "coep"),
]
TIPOS = {"RL": "Religador", "RT": "Regulador"}


def serie_acumulada(p, tipo):
    """As mesmas colunas da figura da página, mês a mês."""
    dados, cum = p["series"][tipo], p["cumulativo"][tipo]
    acervo = cum["acervo_em_janeiro"]
    linhas = []
    for d, m in zip(dados, cum["meses"]):
        linhas.append((d["rotulo"], {
            "parque": d["parque"],
            "com_acervo": m["entraram_acumulado"],
            "so_2026": m["entraram_acumulado"] - acervo,
            "resolvidos": m["resolvidos_acumulado"],
            "coep": m["resolvidos_coep_acumulado"],
        }))
    return linhas, acervo, cum


def aba_acumulado(wb):
    """A figura do acumulado da página, agora como gráfico nativo do Excel.

    São dois gráficos e não um: parque na casa do milhar e contagem nas centenas
    não convivem num eixo só. Na página as duas faixas dividem o mesmo eixo do
    tempo; aqui ficam lado a lado, que é o que o Excel sabe fazer sem distorcer.
    """
    with open(os.path.join(RAIZ, "data", "missao", "parque_2026.json"),
              encoding="utf-8") as fh:
        p = json.load(fh)

    ws = wb.create_sheet("Acumulado 2026", 1)
    ws.column_dimensions["A"].width = 34
    for c in "BCDEFGHI":
        ws.column_dimensions[c].width = 11
    ws.append(["ACUMULADO DE 2026 — A FIGURA DA PÁGINA, MÊS A MÊS"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.append(["Parque, entrantes com o acervo herdado, indisponibilidade só do ano, "
               "resolvidos por qualquer posto e resolvidos pelo COEP."])

    for tipo, nome in TIPOS.items():
        linhas, acervo, cum = serie_acumulada(p, tipo)
        meses = [r[0] for r in linhas]
        faixa = quadro(ws, f"{nome} — 2026", meses,
                       [(rot, [v[chave] for _, v in linhas]) for rot, chave in ACUMULADO])
        primeira, ultima = faixa
        col_parque, col_series = primeira, list(range(primeira + 1, ultima + 1))

        # o parque tem escala própria: começar do zero achata a curva inteira
        vals = [v["parque"] for _, v in linhas]
        piso = (min(vals) // 10) * 10
        teto = -(-max(vals) // 10) * 10

        linha_meses = primeira - 1
        grafico_linha(ws, f"{nome} · parque, mês a mês", [col_parque], meses,
                      linha_meses, "H3" if tipo == "RL" else "H21",
                      piso=piso, teto=teto)
        grafico_linha(ws, f"{nome} · acumulado no ano", col_series, meses,
                      linha_meses, "R3" if tipo == "RL" else "R21")

        ws.append([])
        ws.append([f"Acervo em 1º de janeiro: {acervo}. A conta do ano fecha: "
                   f"{acervo} + {cum['entraram_no_ano']} − {cum['resolvidos_no_ano']} = "
                   f"{cum['meses'][-1]['fila']}, as cadeias abertas hoje. A distância "
                   f"entre «com o acervo» e «só 2026» é exatamente o acervo."])
    return ws


def grafico_linha(ws, titulo, linhas_dados, meses, linha_meses, ancora,
                  piso=None, teto=None):
    """Uma linha por série da faixa; as categorias são os meses, na horizontal.

    A linha dos meses vem por parâmetro, não deduzida da primeira linha de dados —
    deduzir só acerta no gráfico que começa logo abaixo do cabeçalho, e no outro
    aponta para a série de cima.
    """
    ch = LineChart()
    ch.title = titulo
    ch.height, ch.width, ch.style = 8, 17, 2
    for linha in linhas_dados:
        ch.add_data(Reference(ws, min_col=1, min_row=linha,
                              max_col=len(meses) + 1, max_row=linha),
                    titles_from_data=True, from_rows=True)
    ref = (f"'{ws.title}'!$B${linha_meses}:"
           f"${get_column_letter(len(meses) + 1)}${linha_meses}")
    for s in ch.series:
        s.cat = AxDataSource(strRef=StrRef(f=ref))
        s.smooth = False
        s.marker.symbol, s.marker.size = "circle", 6
    ch.x_axis.delete = ch.y_axis.delete = False
    ch.x_axis.axPos, ch.y_axis.axPos = "b", "l"
    if piso is not None:
        ch.y_axis.scaling.min, ch.y_axis.scaling.max = piso, teto
    if len(linhas_dados) == 1:
        ch.legend = None
    ws.add_chart(ch, ancora)
    return ch


def montar(entrada, saida):
    wb, taxa, quadros, rol, derrubados = ler(entrada)
    aba_acumulado(wb)
    aba_quadros(wb, taxa, quadros)
    aba_rol(wb, rol, derrubados)
    aba_divergencias(wb, taxa, quadros, rol)
    os.makedirs(os.path.dirname(saida), exist_ok=True)
    wb.save(saida)
    return wb, rol, derrubados


if __name__ == "__main__":
    entrada = sys.argv[1]
    saida = sys.argv[2] if len(sys.argv) > 2 else PADRAO
    wb, rol, derrubados = montar(entrada, saida)
    print(f"gravado: {saida}")
    for ws in wb.worksheets:
        print(f"  {ws.title}: {len(ws._charts)} gráficos")
    print(f"  rol: {len(rol)} ocorrências · derrubados: {len(derrubados)}")
