"""
Quem passou pelo posto do COEP em 2026 — a conta de verdade.

O que atrapalha essa conta: SS repassada não tem data de conclusão. Sai vazia na
base. Quem contar «SS do COEP ainda sem conclusão» como se estivesse no posto hoje
puxa para 2026 uma SS de 2020 que saiu do COEP no mesmo ano. Foi o que deu 442 SS
herdadas numa primeira contagem — número falso.

A saída certa vem da cadeia de repasse: quando a SS foi repassada, ela saiu do posto
no dia em que a SS seguinte foi aberta. Então, para cada SS do COEP:

    saída  =  data de conclusão, se houver
              senão, abertura da SS seguinte (SS_APOS_REPASSE)
              senão, ainda está no posto

E «passou pelo COEP em 2026» é a SS cujo intervalo [chegada, saída] cruza o ano.

A conta principal é de EQUIPAMENTO, não de SS: o mesmo religador pode ter três SS no
posto no mesmo ano e continua sendo um equipamento.

Grava data/missao/coep_2026.json e dist/COEP_2026.xlsx.

Rodar: python3 scripts/coep_2026.py
"""

import csv
import datetime
import json
import os
import re
from collections import Counter, defaultdict

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ARQ_SS = os.path.join(RAIZ, "data", "missao", "ss_ocorrencia.json")
ARQ_CARTEIRA = os.path.join(RAIZ, "data", "raw", "equipamentos_especiais.csv")
SAIDA_JSON = os.path.join(RAIZ, "data", "missao", "coep_2026.json")
SAIDA_XLSX = os.path.join(RAIZ, "dist", "COEP_2026.xlsx")

POSTO = "ETO-COEP"
INICIO = datetime.datetime(2026, 1, 1)
FIM = datetime.datetime(2026, 8, 18, 23, 59)
RE_SS = re.compile(r"([A-Z][A-Z-]*)\s+0*(\d+)/(\d{4})")


def norm(numero):
    m = RE_SS.match((numero or "").strip().upper())
    return f"{m.group(1)} {int(m.group(2))}/{m.group(3)}" if m else (numero or "").strip().upper()


def dia(texto):
    try:
        return datetime.datetime.strptime((texto or "")[:19], "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def _ler(nome, chave):
    caminho = os.path.join(RAIZ, "data", "missao", nome)
    if not os.path.exists(caminho):
        return {}
    with open(caminho, encoding="utf-8") as fh:
        return json.load(fh).get(chave) or {}


def localidades():
    """Cidade de cada equipamento — a carteira só cobre os 129; o resto vem da base."""
    caminho = os.path.join(RAIZ, "data", "missao", "ssos_min.json")
    saida = {}
    if not os.path.exists(caminho):
        return saida
    with open(caminho, encoding="utf-8") as fh:
        for linha in json.load(fh):
            cod = (linha.get("NUM_TRAFO") or "").strip()
            nome = (linha.get("LOCALIDADE") or "").strip()
            if cod and nome:
                saida.setdefault(cod, nome)
    return saida


def leitura_das_canceladas():
    """SS e ativos em que a leitura do texto confirmou volta à operação.

    O SGM não exporta o motivo do cancelamento — foi lido no texto, em duas frentes:
    as 131 canceladas dos 129 da carteira e a varredura das 585 de todos os postos.
    """
    ss, ativos = set(), set()
    caminho = os.path.join(RAIZ, "data", "missao", "m5_canceladas.json")
    if os.path.exists(caminho):
        with open(caminho, encoding="utf-8") as fh:
            for linha in json.load(fh)["ss"]:
                if linha.get("categoria") == "cancelada_em_operacao":
                    ss.add(norm(linha["numero_ss"]))
    caminho = os.path.join(RAIZ, "data", "missao", "m6_canceladas_global.json")
    if os.path.exists(caminho):
        with open(caminho, encoding="utf-8") as fh:
            m6 = json.load(fh)
        for linha in m6.get("ss", []):
            n = linha.get("numero_ss") or linha.get("ss")
            if n:
                ss.add(norm(n))
        ativos |= {k for k, v in (m6.get("ativos") or {}).items() if v.get("em_operacao")}
    return ss, ativos


def indexar():
    with open(ARQ_SS, encoding="utf-8") as fh:
        base = json.load(fh)
    idx = {}
    for x in base:
        x["_id"] = norm(x["SS_ORIGINAL"])
        x["_abriu"] = dia(x.get("DTA_ABERTURA"))
        x["_ocorreu"] = dia(x.get("DTA_OCORRENCIA"))
        x["_concluiu"] = dia(x.get("DTA_CONCLUSAO"))
        antes = idx.get(x["_id"])
        if antes is None or (x["_abriu"] and antes["_abriu"] and x["_abriu"] > antes["_abriu"]):
            idx[x["_id"]] = x
    seguinte = {x["_id"]: norm(x["SS_APOS_REPASSE"]) for x in idx.values()
                if x.get("SS_APOS_REPASSE")}
    return idx, seguinte


def carteira():
    """A carteira consolidada do gestor — 129 ativos, com o marcador de concluída."""
    with open(ARQ_CARTEIRA, encoding="utf-8") as fh:
        linhas = [x for x in csv.DictReader(fh, delimiter=";")
                  if (x.get("Ativo") or "").strip().isdigit()]
    todos, resolvidos = {}, set()
    for x in linhas:
        cod = x["Ativo"].strip()
        todos[cod] = {
            "tipo": "religador" if (x.get("Tipo") or "").strip() == "79" else "regulador",
            "localidade": (x.get("Localidade") or "").strip(),
            "ss_da_carteira": (x.get("SS aberta") or "").strip(),
            "parecer_coep": (x.get("Parecer COEP") or "").strip(),
            "check": (x.get("Check de concluídas") or "").strip(),
            "criticidade": (x.get("Criticidade") or "").strip(),
        }
        if (x.get("SS aberta") or "").strip().upper() == "CONCLUÍDA":
            resolvidos.add(cod)
    return todos, resolvidos


def montar():
    idx, seguinte = indexar()
    antecessor = {v: k for k, v in seguinte.items()}
    cart, resolvidos = carteira()

    def saida(x):
        if x["_concluiu"]:
            return x["_concluiu"], "conclusão da SS", ""
        prox = seguinte.get(x["_id"])
        if prox and prox in idx and idx[prox]["_abriu"]:
            return idx[prox]["_abriu"], "repasse — abertura da SS seguinte", idx[prox]["POSTO_SGM"]
        return None, "ainda no posto", ""

    no_posto, por_ativo = [], defaultdict(list)
    por_ativo_ss = defaultdict(list)   # os registros crus das SS do COEP, por equipamento
    for x in idx.values():
        if x["POSTO_SGM"] != POSTO or not x["_abriu"]:
            continue
        quando, como, destino = saida(x)
        limite = quando or FIM
        if not (x["_abriu"] <= FIM and limite >= INICIO):
            continue
        item = {
            "ss": x["SS_ORIGINAL"], "ativo": x["EQUIPAMENTO"],
            "tipo": x["TIPO_ATIVO"].lower(), "status": x["STATUS"],
            "chegou": x["_abriu"].strftime("%d/%m/%Y"),
            "saiu": quando.strftime("%d/%m/%Y") if quando else "",
            "como_apurou_a_saida": como, "foi_para": destino,
            "dias_no_posto": (limite - x["_abriu"]).days,
            "chegou_em_2026": x["_abriu"] >= INICIO,
            "saiu_em_2026": bool(quando and INICIO <= quando <= FIM),
            "segue_no_posto": quando is None or quando > FIM,
            "pendencia": x.get("PENDENCIA_DO_ATIVO", ""),
            "ano_da_ss": x["ANO_SS"],
        }
        no_posto.append(item)
        por_ativo[x["EQUIPAMENTO"]].append(item)
        por_ativo_ss[x["EQUIPAMENTO"]].append(x)

    cidade = localidades()
    ativos = []
    for cod, itens in sorted(por_ativo.items()):
        c = cart.get(cod)
        ativos.append({
            "ativo": cod, "tipo": itens[0]["tipo"], "ss_no_coep_em_2026": len(itens),
            "ss": " | ".join(i["ss"] for i in itens),
            # ordenar dd/mm/aaaa como texto compara o dia primeiro — ordena por data
            "primeira_chegada": min((i["chegou"] for i in itens),
                                    key=lambda t: (t[6:], t[3:5], t[:2])),
            "dias_no_posto": max(i["dias_no_posto"] for i in itens),
            "chegou_em_2026": any(i["chegou_em_2026"] for i in itens),
            "ja_estava_de_antes": any(not i["chegou_em_2026"] for i in itens),
            "saiu_em_2026": any(i["saiu_em_2026"] for i in itens),
            "segue_no_posto": any(i["segue_no_posto"] for i in itens),
            "na_carteira": bool(c),
            "resolvido_na_carteira": cod in resolvidos,
            "parecer_coep": (c or {}).get("parecer_coep", ""),
            "criticidade": (c or {}).get("criticidade", ""),
            "localidade": (c or {}).get("localidade") or cidade.get(cod, ""),
            "ss_da_carteira": (c or {}).get("ss_da_carteira", ""),
        })

    codigos = {a["ativo"] for a in ativos}
    # já passou pelo COEP alguma vez, em qualquer ano?
    coep_de_sempre = defaultdict(set)
    for x in idx.values():
        if x["POSTO_SGM"] == POSTO and x["_abriu"]:
            coep_de_sempre[x["EQUIPAMENTO"]].add(x["_abriu"].year)
    resolvidos_fora = []
    for cod in sorted(resolvidos - codigos):
        c = cart[cod]
        ultima = [x for x in idx.values() if x["EQUIPAMENTO"] == cod]
        ultima.sort(key=lambda x: x["_abriu"] or INICIO, reverse=True)
        u = ultima[0] if ultima else None
        anos = sorted(coep_de_sempre.get(cod, ()))
        if anos:
            motivo = ("passou pelo COEP em " + ", ".join(str(a) for a in anos) +
                      " — saiu do posto antes de 2026 e só o fechamento veio depois")
        elif u:
            motivo = (f"nunca teve SS no COEP; a demanda ficou no {u['POSTO_SGM']} "
                      "e foi resolvida por lá")
        else:
            motivo = "sem SS de religador ou regulador nesta base"
        resolvidos_fora.append({
            "ativo": cod, "tipo": c["tipo"], "localidade": c["localidade"],
            "parecer_coep": c["parecer_coep"],
            "passou_pelo_coep_em": ", ".join(str(a) for a in anos) or "nunca",
            "ultima_ss_conhecida": u["SS_ORIGINAL"] if u else "",
            "posto_da_ultima_ss": u["POSTO_SGM"] if u else "",
            "abertura_da_ultima_ss": (u["DTA_ABERTURA"] or "")[:10] if u else "",
            "motivo": motivo,
        })

    # Visão 2: quem o COEP resolveu em 2026 — pela cadeia da demanda, não pela carteira.
    # A carteira é a foto do que ainda está pendente; ela não guarda o que fechou e saiu.
    # Por isso ela perde justamente o que o gestor lembra de ter resolvido: demanda velha,
    # de 2024 e 2025, fechada agora.
    ss_confirmadas, ativos_em_operacao = leitura_das_canceladas()
    cidade = localidades()
    ss_do_ativo = defaultdict(list)
    for y in idx.values():
        ss_do_ativo[y["EQUIPAMENTO"]].append(y)

    def raiz(x):
        cur, visto = x, set()
        while cur["_id"] in antecessor and antecessor[cur["_id"]] in idx \
                and antecessor[cur["_id"]] not in visto:
            visto.add(cur["_id"])
            cur = idx[antecessor[cur["_id"]]]
        return cur

    def ponta(x):
        cur, visto = x, set()
        while cur["_id"] in seguinte and seguinte[cur["_id"]] in idx \
                and seguinte[cur["_id"]] not in visto:
            visto.add(cur["_id"])
            cur = idx[seguinte[cur["_id"]]]
        return cur

    LOTE = {datetime.date(2026, 6, 29), datetime.date(2026, 6, 30)}
    resolvidos_do_coep, candidatos = [], 0
    for cod, itens in sorted(por_ativo_ss.items()):
        escolhida = None
        for x in itens:
            fim = ponta(x)
            if fim["STATUS"] in ("SS ATENDIDA", "SS CANCELADA") and fim["_concluiu"] \
                    and INICIO <= fim["_concluiu"] <= FIM:
                inicio = raiz(x)
                # o ano da demanda é o da OCORRÊNCIA, régua do gestor: a abertura da SS
                # vem depois do fato, e o número da SS às vezes carrega outro ano ainda
                marco = inicio["_ocorreu"] or inicio["_abriu"]
                cand = {"x": x, "raiz": inicio, "ponta": fim,
                        "ano": marco.year if marco else None, "marco": marco}
                if escolhida is None or (cand["ano"] or 9999) < (escolhida["ano"] or 9999):
                    escolhida = cand
        if not escolhida:
            continue
        candidatos += 1
        fim = escolhida["ponta"]
        c = cart.get(cod, {})
        primeiro_ataque = "primeiro ataque" in (c.get("parecer_coep") or "").lower()
        # Régua do gestor (22/08): o que derruba o cancelamento é nota nova PARA ESSE ATIVO
        # NO POSTO DO COEP. Nota aberta em outro posto é outra frente de trabalho, não a
        # demanda voltando para a mesa do COEP.
        voltou = [y for y in ss_do_ativo[cod]
                  if y["_abriu"] and y["_abriu"] > fim["_concluiu"] and y["POSTO_SGM"] == POSTO]
        nota_de_volta = min(voltou, key=lambda y: y["_abriu"]) if voltou else None
        # informativo: nota nova em qualquer posto, e nota que segue pendente hoje
        fora_do_coep = [y for y in ss_do_ativo[cod]
                        if y["_abriu"] and y["_abriu"] > fim["_concluiu"] and y["POSTO_SGM"] != POSTO]
        pendente_hoje = [y for y in ss_do_ativo[cod]
                         if y["_abriu"] and y["_abriu"] > fim["_concluiu"]
                         and y["STATUS"] == "SS PENDENTE"]
        confirmada = (fim["_id"] in ss_confirmadas or cod in ativos_em_operacao)
        no_lote = fim["STATUS"] == "SS CANCELADA" and fim["_concluiu"].date() in LOTE
        if nota_de_volta:
            entra = False
            porque = (f"voltou para o COEP: {nota_de_volta['SS_ORIGINAL']} aberta em "
                      f"{nota_de_volta['_abriu'].strftime('%d/%m/%Y')}, depois do fechamento")
        elif primeiro_ataque:
            entra, porque = False, "resolvido no primeiro ataque do DMSL — não é trabalho do posto"
        else:
            entra, porque = True, ""
        if entra:
            if fim["STATUS"] == "SS ATENDIDA":
                prova = "SS atendida — serviço executado"
            elif confirmada:
                prova = "cancelada, com leitura que confirma volta à operação"
            else:
                prova = "resolvido por cancelamento — nenhuma nota nova no COEP depois"
        else:
            prova = ""
        resolvidos_do_coep.append({
            "ativo": cod, "tipo": escolhida["x"]["TIPO_ATIVO"].lower(),
            "ano_da_demanda": escolhida["ano"],
            "ss_que_abriu_a_demanda": escolhida["raiz"]["SS_ORIGINAL"],
            "posto_que_abriu": escolhida["raiz"]["POSTO_SGM"],
            "ocorrencia_da_demanda": (escolhida["marco"].strftime("%d/%m/%Y")
                                      if escolhida["marco"] else ""),
            "ss_no_coep": escolhida["x"]["SS_ORIGINAL"],
            "ss_que_fechou": fim["SS_ORIGINAL"], "posto_que_fechou": fim["POSTO_SGM"],
            "como_terminou": fim["STATUS"],
            "data_do_fechamento": fim["_concluiu"].strftime("%d/%m/%Y"),
            "dias_da_demanda": ((fim["_concluiu"] - escolhida["marco"]).days
                                if escolhida["marco"] else None),
            "conta_como_resolvido_pelo_coep": entra, "porque_nao": porque, "prova": prova,
            "nota_nova_no_coep": nota_de_volta["SS_ORIGINAL"] if nota_de_volta else "",
            "nota_nova_em_outro_posto": ", ".join(
                f"{y['SS_ORIGINAL']} ({y['POSTO_SGM']})" for y in sorted(
                    fora_do_coep, key=lambda z: z["_abriu"])[:3]),
            "tem_nota_pendente_hoje": bool(pendente_hoje),
            "cancelada_no_lote_de_junho": no_lote,
            "esta_na_carteira": bool(c), "parecer_coep": c.get("parecer_coep", ""),
            "localidade": c.get("localidade") or cidade.get(cod, ""),
        })

    contam = [r for r in resolvidos_do_coep if r["conta_como_resolvido_pelo_coep"]]
    por_ano = Counter(r["ano_da_demanda"] for r in contam)
    por_prova = Counter(r["prova"].split(" —")[0] for r in contam)

    conta = {
        "equipamentos_que_passaram": len(ativos),
        "candidatos_a_resolvido": candidatos,
        "resolvidos_pelo_coep": len(contam),
        "resolvidos_por_ano_da_demanda": {str(k): v for k, v in sorted(por_ano.items())},
        "resolvidos_por_prova": dict(por_prova),
        "tirados_por_volta_ao_coep": sum(1 for r in resolvidos_do_coep
                                         if r["porque_nao"].startswith("voltou")),
        "resolvidos_com_nota_pendente_em_outro_posto": sum(
            1 for r in contam if r["tem_nota_pendente_hoje"]),
        "resolvidos_no_lote_de_junho": sum(1 for r in contam if r["cancelada_no_lote_de_junho"]),
        "tirados_por_primeiro_ataque_dmsl": sum(1 for r in resolvidos_do_coep
                                                if "primeiro ataque" in r["porque_nao"]),
        "por_tipo": dict(Counter(a["tipo"] for a in ativos)),
        "ss_no_posto": len(no_posto),
        "chegaram_em_2026": sum(1 for a in ativos if a["chegou_em_2026"]),
        "ja_estavam_de_antes": sum(1 for a in ativos if a["ja_estava_de_antes"]),
        "sairam_do_posto_em_2026": sum(1 for a in ativos if a["saiu_em_2026"]),
        "seguem_no_posto_em_18_08": sum(1 for a in ativos if a["segue_no_posto"]),
        "na_carteira_consolidada": sum(1 for a in ativos if a["na_carteira"]),
        "fora_da_carteira": sum(1 for a in ativos if not a["na_carteira"]),
        "resolvidos_na_carteira": sum(1 for a in ativos if a["resolvido_na_carteira"]),
        "resolvidos_na_carteira_total": len(resolvidos),
        "resolvidos_sem_passagem_pelo_coep_em_2026": len(resolvidos_fora),
    }

    MES_PT = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
    def _mes_br(br):
        return f"{br[6:10]}-{br[3:5]}"
    primeira = {}
    for i in no_posto:
        if not i["chegou_em_2026"]:
            continue
        m = _mes_br(i["chegou"])
        if i["ativo"] not in primeira or m < primeira[i["ativo"]]:
            primeira[i["ativo"]] = m
    cheg = Counter(primeira.values())
    fech = Counter(_mes_br(r["data_do_fechamento"]) for r in resolvidos_do_coep
                   if r["conta_como_resolvido_pelo_coep"])
    def _fila_em(dia):
        """Equipamentos com SS no posto naquele dia — a fila de verdade."""
        no_posto = set()
        for i in no_posto_ss_cache:
            cheg = i["_cheg"]
            sai = i["_sai"]
            if cheg and cheg <= dia and (sai is None or sai > dia):
                no_posto.add(i["ativo"])
        return len(no_posto)
    def _d(br):
        return datetime.date(int(br[6:10]), int(br[3:5]), int(br[:2])) if br else None
    no_posto_ss_cache = [{"ativo": i["ativo"], "_cheg": _d(i["chegou"]),
                          "_sai": _d(i["saiu"]) if i["saiu"] else None} for i in no_posto]
    fins = {m: (datetime.date(2026, m + 1, 1) - datetime.timedelta(days=1)) if m < 8
            else datetime.date(2026, 8, 18) for m in range(1, 9)}
    curva_mensal = [{"mes": f"2026-{m:02d}", "rotulo": f"{MES_PT[m - 1]}/2026",
                     "chegaram": cheg.get(f"2026-{m:02d}", 0),
                     "resolvidos": fech.get(f"2026-{m:02d}", 0),
                     "no_posto": _fila_em(fins[m])}
                    for m in range(1, 9)]

    pacote = {
        "gerado_em": "2026-08-22", "posicao": "18/08/2026",
        "curva_mensal": curva_mensal,
        "herdados": sum(1 for a in ativos if a["ja_estava_de_antes"]),
        "fonte": "EQP_SS_OCORRENCIA_11082026 (10.386 SS de religador e regulador) e a carteira "
                 "consolidada do gestor (Relação dos Equipamentos Indisponíveis ETO, "
                 "ATUALIZADA 16 — 129 ativos)",
        "premissas": PREMISSAS,
        "conta": conta,
        "ativos": ativos,
        "resolvidos_do_coep": resolvidos_do_coep,
        "ss": no_posto,
        "resolvidos_sem_passagem": resolvidos_fora,
    }
    with open(SAIDA_JSON, "w", encoding="utf-8") as fh:
        json.dump(pacote, fh, ensure_ascii=False, indent=1)
    return pacote


PREMISSAS = [
    "Passou pelo COEP em 2026 é a SS que esteve no posto em algum momento do ano — não só a "
    "que chegou em 2026. SS que chegou em 2025 e só saiu em março de 2026 passou pelo posto.",
    "A data de saída não está na base para SS repassada: o campo de conclusão vem vazio. A saída "
    "é a abertura da SS seguinte, aquela gravada em SS_APOS_REPASSE. Sem isso a conta erra feio "
    "— SS de 2020 e 2021 entram como se ainda estivessem no posto.",
    "Ordem de apuração da saída: conclusão da SS, se houver; senão a abertura da SS seguinte; "
    "senão a SS ainda está no posto. Das 694 SS do COEP na base: 153 pela conclusão, 486 pelo "
    "repasse, 55 ainda no posto.",
    "A conta principal é de EQUIPAMENTO, não de SS. O mesmo religador com três SS no posto no "
    "mesmo ano é um equipamento. O número de SS vai ao lado.",
    "O ano vai até 18/08/2026, a posição do relatório. Quem ainda está no posto nessa data conta "
    "como tendo passado.",
    "Chegou em 2026 e já estava de antes somam mais que o total porque o mesmo equipamento pode "
    "ter uma SS herdada e outra nova no mesmo ano.",
    "A carteira consolidada é a Relação dos Equipamentos Indisponíveis ETO, versão ATUALIZADA 16, "
    "com 129 ativos. Resolvido é o ativo cuja coluna «SS aberta» está marcada CONCLUÍDA — 52.",
    "Resolvido na carteira sem SS no COEP dentro de 2026 não é erro: ou o equipamento passou pelo "
    "posto em ano anterior e o fechamento veio depois, ou quem resolveu foi outro posto.",
    "Só religador e regulador. A base de ocorrência traz 8.835 SS de religador e 1.551 de "
    "regulador, com data de ocorrência em 100% das linhas.",
    "VISÃO 2 — resolvido pelo COEP em 2026: a demanda passou pelo posto dentro de 2026 e a "
    "cadeia dela fechou dentro de 2026, com SS atendida ou cancelada.",
    "A conta NÃO sai da carteira. A carteira é a foto do que ainda está pendente; o que fechou "
    "e saiu não fica registrado nela. Foi por isso que a primeira contagem, feita pela carteira, "
    "achou só 29 e perdeu justamente o que o gestor lembrava: demanda velha, de 2024 e 2025, "
    "fechada agora.",
    "A demanda é a cadeia inteira de SS, do primeiro posto ao último — e o ano dela é o da DATA "
    "DE OCORRÊNCIA da primeira SS, não o da SS do COEP nem o número da SS. É isso que mostra que "
    "o posto fechou caso de 2024 e 2025.",
    "O número da SS não serve para datar: ETO-COEP 00149/2025 foi aberta em 29/06/2026, com "
    "ocorrência em 11/07/2025. Numerar não é abrir, e abrir não é o fato acontecer.",
    "O posto que fecha não precisa ser o COEP (gestor, 22/08). Fechamento no ETO-RD-PS, no "
    "ETO-PROT, no ETO-RD-AR e nos demais conta igual: o COEP diagnosticou e despachou, quem "
    "executou foi a ponta. O ETO-TELE também conta, desde que haja parecer do COEP ou que o "
    "equipamento tenha estado no posto antes — e os 15 fechados lá têm SS no COEP, 11 deles "
    "com parecer registrado na carteira.",
    "Onde os 71 fecharam: 40 no ETO-COEP, 15 no ETO-TELE, 5 no ETO-RD-PS, 5 no ETO-RD-AR, "
    "4 no ETO-PROT, 1 no ETO-RD-PA e 1 no ETO-RD-PO.",
    "Régua do gestor para o cancelamento (22/08): cancelado é resolvido, DESDE QUE não tenham "
    "aberto outra nota para esse ativo no posto do COEP depois. Se abriram, a demanda voltou "
    "para a mesa do posto e continua pendente — não conta. Isso derruba 19 dos 90 candidatos.",
    "O que derruba é nota nova NO COEP. Nota aberta em outro posto depois do fechamento é outra "
    "frente de trabalho, não a demanda voltando — vai anotada na planilha, em coluna própria, "
    "mas não tira o equipamento da conta. Seis dos resolvidos têm nota pendente em outro posto.",
    "Não conta o resolvido no primeiro ataque do DMSL — a demanda morreu na mão da DMSL, o posto "
    "não trabalhou nela.",
    "A planilha diz como cada um foi resolvido: 26 por SS atendida, com serviço executado; 9 "
    "por cancelamento com leitura do texto confirmando volta à operação; 36 por cancelamento "
    "sem nota nova no COEP depois, que é a régua do gestor.",
    "O SGM não exporta o motivo do cancelamento — é a lacuna que obrigou a régua acima. Onde a "
    "leitura do texto confirmou volta à operação, isso vai escrito na coluna da prova.",
    "Vinte dos resolvidos foram cancelados no lote de 29 e 30 de junho. Vai marcado em coluna "
    "própria: pela régua do gestor eles contam, mas quem quiser conferir o lote sabe quais são.",
]


def planilha(pacote):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    tit = Font(bold=True, color="FFFFFF", size=10)
    fundo = PatternFill("solid", fgColor="1F3864")
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    cinza = PatternFill("solid", fgColor="EFEFEF")

    def cabecalho(ws, colunas, larguras):
        ws.append(colunas)
        for c, (col, larg) in enumerate(zip(colunas, larguras), 1):
            cel = ws.cell(row=1, column=c)
            cel.font, cel.fill = tit, fundo
            cel.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[cel.column_letter].width = larg
        ws.freeze_panes = "A2"

    def fechar(ws):
        for linha in ws.iter_rows(min_row=2):
            for cel in linha:
                cel.border = borda
                cel.alignment = Alignment(vertical="top", wrap_text=True)

    sn = lambda v: "sim" if v else "não"
    c = pacote["conta"]

    # VISÃO 1 — quem passou pelo posto em 2026
    ws = wb.active
    ws.title = f"1 · Passaram pelo COEP ({c['equipamentos_que_passaram']})"
    cabecalho(ws, ["Ativo", "Tipo", "Localidade", "SS no COEP em 2026", "Quantas SS",
                   "Primeira chegada", "Dias no posto", "Ainda no posto em 18/08",
                   "Está na carteira", "Parecer COEP", "Criticidade"],
              [14, 12, 20, 40, 10, 14, 12, 14, 12, 26, 12])
    for a in sorted(pacote["ativos"], key=lambda x: -x["dias_no_posto"]):
        ws.append([a["ativo"], a["tipo"], a["localidade"], a["ss"], a["ss_no_coep_em_2026"],
                   a["primeira_chegada"], a["dias_no_posto"], sn(a["segue_no_posto"]),
                   sn(a["na_carteira"]), a["parecer_coep"], a["criticidade"]])
    fechar(ws)

    # VISÃO 2 — quem o COEP resolveu em 2026
    ws = wb.create_sheet(f"2 · Resolvidos pelo COEP ({c['resolvidos_pelo_coep']})")
    cabecalho(ws, ["Ativo", "Tipo", "Ano em que a demanda nasceu", "Ocorrência", "Conta",
                   "Prova", "Por que não conta", "Nota nova no COEP",
                   "Nota nova em outro posto", "Tem nota pendente hoje",
                   "Cancelada no lote de 29-30/06", "SS que abriu a demanda",
                   "Posto que abriu", "SS no COEP", "SS que fechou", "Posto que fechou",
                   "Como terminou", "Fechou em", "Dias da demanda", "Está na carteira",
                   "Parecer COEP", "Localidade"],
              [14, 12, 12, 12, 9, 46, 46, 20, 34, 12, 13, 22, 14, 22, 22, 14, 15, 12, 12,
               12, 24, 20])
    ordem = sorted(pacote["resolvidos_do_coep"],
                   key=lambda r: (not r["conta_como_resolvido_pelo_coep"],
                                  r["ano_da_demanda"] or 9999, r["ativo"]))
    for r in ordem:
        ws.append([r["ativo"], r["tipo"], r["ano_da_demanda"], r["ocorrencia_da_demanda"],
                   sn(r["conta_como_resolvido_pelo_coep"]), r["prova"], r["porque_nao"],
                   r["nota_nova_no_coep"], r["nota_nova_em_outro_posto"],
                   sn(r["tem_nota_pendente_hoje"]), sn(r["cancelada_no_lote_de_junho"]),
                   r["ss_que_abriu_a_demanda"], r["posto_que_abriu"], r["ss_no_coep"],
                   r["ss_que_fechou"], r["posto_que_fechou"], r["como_terminou"],
                   r["data_do_fechamento"], r["dias_da_demanda"],
                   sn(r["esta_na_carteira"]), r["parecer_coep"], r["localidade"]])
    fechar(ws)
    for n, r in enumerate(ordem, 2):
        if not r["conta_como_resolvido_pelo_coep"]:
            for cel in ws[n]:
                cel.fill = cinza

    # o método, para as duas
    ws = wb.create_sheet("Como foi feito")
    cabecalho(ws, ["Passo", "O que foi feito"], [8, 130])
    for n, texto in enumerate(pacote["premissas"], 1):
        ws.append([n, texto])
    fechar(ws)

    os.makedirs(os.path.dirname(SAIDA_XLSX), exist_ok=True)
    wb.save(SAIDA_XLSX)


def main():
    pacote = montar()
    c = pacote["conta"]
    print(f"VISÃO 1 — passaram pelo COEP em 2026: {c['equipamentos_que_passaram']} equipamentos "
          f"({c['por_tipo']['religador']} RL + {c['por_tipo']['regulador']} RT) em "
          f"{c['ss_no_posto']} SS; {c['seguem_no_posto_em_18_08']} ainda no posto em 18/08")
    print(f"VISÃO 2 — resolvidos PELO COEP em 2026: {c['resolvidos_pelo_coep']} "
          f"(de {c['candidatos_a_resolvido']} candidatos)")
    print(f"  por ano em que a demanda nasceu.: {c['resolvidos_por_ano_da_demanda']}")
    print(f"  por prova.......................: {c['resolvidos_por_prova']}")
    print(f"  tirados por voltar ao COEP......: {c['tirados_por_volta_ao_coep']}")
    print(f"  com nota pendente em outro posto: "
          f"{c['resolvidos_com_nota_pendente_em_outro_posto']}")
    print(f"  cancelados no lote de 29-30/06..: {c['resolvidos_no_lote_de_junho']}")
    print(f"  tirados por primeiro ataque DMSL: {c['tirados_por_primeiro_ataque_dmsl']}")
    print(f"gravado: {SAIDA_JSON}")
    planilha(pacote)
    print(f"gravado: {SAIDA_XLSX}")


if __name__ == "__main__":
    main()
