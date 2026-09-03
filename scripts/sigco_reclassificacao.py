"""
O caso SIGCO consolidado — obras de RL/RT fora do par 8495/8481.

Junta as cinco frentes que encontraram obra no projeto errado, sem duplicar:

  auditoria      M2: veredito sigco_errado com destino no par, ou cruzada dentro dele
  trafo_texto    obra cujo texto diz «trafo auxiliar» de religador (régua do gestor:
                 trafo auxiliar pertence ao projeto do equipamento)
  vinculo_ss     obra declarada em SS de RL/RT (campo NUM_OBRA, com o zero à esquerda
                 devolvido) cuja SS declarante é serviço do próprio equipamento
  resolvidos     obra ligada aos ativos que o gestor resolveu em 2026
  regra_trafo    código de trafo auxiliar no texto da obra — prefixo 51 (padrão) ou 57,
                 com os 8 dígitos finais iguais aos do equipamento pai. Validada por
                 código, pelo texto das obras e pelas coordenadas (gestor, 21/08:
                 trafo e pai estão fisicamente juntos)

Grava data/missao/sigco_reclassificacao.json (o caso, com fonte por obra) e
dist/SIGCO_RECLASSIFICACAO.xlsx (a planilha para a controladoria).

Rodar: python3 scripts/sigco_reclassificacao.py
"""

import json
import os
import re
import sys
from collections import defaultdict

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
import taxa_falha as tf  # noqa: E402

DESTINO_JSON = os.path.join(RAIZ, "data", "missao", "sigco_reclassificacao.json")
DESTINO_XLSX = os.path.join(RAIZ, "dist", "SIGCO_RECLASSIFICACAO.xlsx")
CERTO = {"religador": "8495", "regulador": "8481"}

RE_TA_TEXTO = re.compile(
    r"TRAFO\s+AUX|TRANSFORMADOR\s+AUX|TRAFO\s+(?:DE\s+)?15\s*KVA|AUXU?ILIAR\s+(?:NO|DO)\s+RELIGADOR", re.I)
RE_COD_TRAFO = re.compile(r"\b(5[17]\d{8})\b")


def _carrega():
    with open(os.path.join(RAIZ, "data", "missao", "m2_sigco.json"), encoding="utf-8") as fh:
        m2 = json.load(fh)["obras"]
    with open(os.path.join(RAIZ, "data", "missao", "aic_rlrt.json"), encoding="utf-8") as fh:
        aic = json.load(fh)
    with open(os.path.join(RAIZ, "data", "missao", "aic_index.json"), encoding="utf-8") as fh:
        idx = json.load(fh)
    with open(os.path.join(RAIZ, "data", "missao", "ssos_min.json"), encoding="utf-8") as fh:
        base = json.load(fh)
    return m2, aic, idx, base


def montar():
    m2, aic, idx, base = _carrega()
    m2i = {o["num_obra"]: o for o in m2}
    aici = {a.get("NUM_OBRA"): a for a in aic}
    frota = tf.parque()
    caso = {}  # num_obra -> registro

    def poe(obra, familia, fonte, motivo, ativo=""):
        if obra in caso:
            return
        info = idx.get(obra) or {}
        m = m2i.get(obra) or {}
        a = aici.get(obra) or {}
        sig = a.get("NUM_PROJETO_SIGCO") or info.get("sig") or ""
        if sig in ("8481", "8495") and CERTO.get(familia) == sig:
            return  # já está no lugar certo
        ano = (a.get("DTH_ABERTURA") or "")[:4] or ("20" + obra[3:5])
        caso[obra] = {
            "obra": obra, "familia": familia, "sigco_atual": sig or "(sem)",
            "sigco_certo": CERTO.get(familia, "?"), "ano": ano,
            "status": (info.get("st") or a.get("STATUS AIC") or "")[:36],
            "valor_orcado": float(m.get("valor_orcado") or 0),
            "valor_realizado": float(m.get("valor_realizado") or 0),
            "descricao": (a.get("DESCRICAO_OBRA") or "")[:120],
            "fonte": fonte, "motivo": motivo, "ativo": ativo,
        }

    # 1) auditoria M2
    for o in m2:
        if o.get("veredito") != "sigco_errado":
            continue
        certo = o.get("sigco_certo")
        if certo not in ("8481", "8495"):
            continue
        familia = "religador" if certo == "8495" else "regulador"
        poe(o["num_obra"], familia, "auditoria",
            f"auditoria M2: tipo {o.get('tipo')} lançado no {o.get('sigco')}")

    # 2) trafo auxiliar pelo texto
    for a in aic:
        if RE_TA_TEXTO.search(a.get("DESCRICAO_OBRA") or ""):
            poe(a.get("NUM_OBRA"), "religador", "trafo_texto",
                "texto da obra: trafo auxiliar do religador — régua do gestor")

    # 3) regra do código de trafo (51/57 + 8 finais do pai)
    for a in aic:
        texto = (a.get("DESCRICAO_OBRA") or "") + " " + (a.get("DESCRICAO") or "")
        for cod in RE_COD_TRAFO.findall(texto):
            pai = next((c for c in frota if c[-8:] == cod[-8:]), None)
            if pai:
                poe(a.get("NUM_OBRA"), frota[pai]["familia"], "regra_trafo",
                    f"trafo {cod} → pai {pai} pelos 8 dígitos finais "
                    "(regra validada por código, texto e coordenadas)", ativo=pai)

    # 4) vínculo SS→obra (campo NUM_OBRA, zero à esquerda devolvido)
    decl = defaultdict(list)
    for x in base:
        o = str(x.get("NUM_OBRA") or "").strip()
        if o and o not in ("None", "nan"):
            decl[o.split(".")[0].zfill(10)].append(x)
    for obra, rows in decl.items():
        if obra not in idx:
            continue
        tem_eq = False
        familia = None
        ativo = ""
        for s in rows:
            if tf.objeto_do_fato(s) == "rede":
                continue
            cls = tf.classificar(s)
            tiposs = (s.get("TIPOSS") or "").upper()
            if cls == "falha" or tiposs in ("OBRAS (NOVOS EQUIPAMENTOS)", "COMISSIONAMENTO"):
                tem_eq = True
                ativo = str(s.get("NUM_TRAFO") or "")
                familia = frota.get(ativo, {}).get("familia") or tf.familia_pela_ss([s])
        if tem_eq and familia:
            poe(obra, familia, "vinculo_ss",
                "declarada em SS de serviço do próprio equipamento", ativo=ativo)

    # 5) obras ligadas aos ativos resolvidos pelo gestor em 2026 — a SS declarante pode
    # ser genérica (solicitação), mas o ativo é do parque e a tratativa é dele; só a
    # SS claramente de rede fica de fora.
    with open(os.path.join(RAIZ, "data", "meta.json"), encoding="utf-8") as fh:
        lista = (json.load(fh).get("entrada_mensal") or {}).get("lista") or []
    resolvidos = {str(x["ativo"]) for x in lista if x.get("resolvido")}
    for obra, rows in decl.items():
        if obra not in idx:
            continue
        for s in rows:
            ativo = str(s.get("NUM_TRAFO") or "")
            if ativo not in resolvidos or tf.objeto_do_fato(s) == "rede":
                continue
            familia = frota.get(ativo, {}).get("familia") or tf.familia_pela_ss([s])
            if familia:
                poe(obra, familia, "resolvidos",
                    "obra da tratativa de ativo resolvido pelo gestor em 2026", ativo=ativo)

    registros = sorted(caso.values(), key=lambda r: (r["ano"], r["obra"]))
    de26 = [r for r in registros if r["ano"] == "2026"]
    tot_r = sum(r["valor_realizado"] for r in registros)
    pacote = {
        "regra_do_trafo": "trafo auxiliar tem código próprio com prefixo 51 (padrão) ou 57 e os "
                          "8 dígitos finais iguais aos do equipamento pai; validada por código, "
                          "pelo texto das obras e pelas coordenadas (gestor, 21/08)",
        "total_obras": len(registros),
        "de_2026": len(de26),
        "valor_realizado_conhecido": round(tot_r, 2),
        "nota_valor": "valor só existe para obras presentes no extrato RL/RT do AIC; as achadas "
                      "pelo vínculo SS em outros projetos ficam sem valor até novo extrato",
        "nao_entra": {"0202600193": "deslocamento de regulador por obra de cliente (CIA x 3º, "
                                     "projeto 54554) — expansão, não manutenção; apenas recebeu "
                                     "célula reaproveitada do 5854566043"},
        "obras": registros,
    }
    with open(DESTINO_JSON, "w", encoding="utf-8") as fh:
        json.dump(pacote, fh, ensure_ascii=False, indent=1)
    return pacote


def planilha(p):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    TINTA, FAIXA, DEST = "1A1A1A", "E8E4DC", "C8442A"
    fina = Side(style="thin", color="8A8577")
    grossa = Side(style="medium", color=TINTA)
    G = Border(left=fina, right=fina, top=fina, bottom=fina)
    ROT_FONTE = {"auditoria": "auditoria SIGCO", "trafo_texto": "trafo auxiliar (texto)",
                 "regra_trafo": "trafo auxiliar (regra 51/57)", "vinculo_ss": "vínculo SS→obra",
                 "resolvidos": "resolvidos do gestor"}

    wb = Workbook()

    def aba(nome, registros, sub):
        ws = wb.create_sheet(nome)
        ws.sheet_view.showGridLines = False
        for i, w in enumerate([13, 11, 10, 9, 9, 12, 12, 22, 46, 42], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws["A1"] = f"Obras de RL/RT com projeto SIGCO errado — {nome}"
        ws["A1"].font = Font(size=15, bold=True, color=TINTA)
        ws.merge_cells("A1:J1")
        ws["A2"] = sub
        ws["A2"].font = Font(size=10, italic=True, color="5A5347")
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("A2:J2")
        ws.row_dimensions[2].height = 40
        r = 4
        cab = ["Obra", "Ano", "Está no", "Deveria", "Família", "Realizado", "Orçado",
               "Como foi encontrada", "Descrição", "Motivo"]
        for i, c in enumerate(cab, 1):
            x = ws.cell(row=r, column=i, value=c)
            x.font = Font(size=11, bold=True, color=TINTA)
            x.fill = PatternFill("solid", fgColor=FAIXA)
            x.border = Border(left=fina, right=fina, top=grossa, bottom=grossa)
            x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        r += 1
        tr = to = 0
        for reg in registros:
            tr += reg["valor_realizado"]; to += reg["valor_orcado"]
            vals = [reg["obra"], reg["ano"], reg["sigco_atual"], reg["sigco_certo"],
                    reg["familia"][:2].upper() + ("L" if reg["familia"] == "religador" else "T"),
                    reg["valor_realizado"] or None, reg["valor_orcado"] or None,
                    ROT_FONTE.get(reg["fonte"], reg["fonte"]), reg["descricao"], reg["motivo"]]
            for i, v in enumerate(vals, 1):
                x = ws.cell(row=r, column=i, value=v)
                x.font = Font(size=10, bold=(i == 4), color=DEST if i == 4 else TINTA)
                x.border = G
                x.alignment = Alignment(horizontal="left" if isinstance(v, str) else "center",
                                        vertical="top", wrap_text=i in (9, 10))
                if i in (6, 7) and v:
                    x.number_format = 'R$ #,##0.00'
            ws.row_dimensions[r].height = 26
            r += 1
        for i, v in enumerate(["TOTAL", "", "", "", "", tr, to, "", "", ""], 1):
            x = ws.cell(row=r, column=i, value=v if v != "" else None)
            x.font = Font(size=11, bold=True, color=TINTA); x.border = G
            if i in (6, 7):
                x.number_format = 'R$ #,##0.00'
        return ws

    de26 = [r for r in p["obras"] if r["ano"] == "2026"]
    resto = [r for r in p["obras"] if r["ano"] != "2026"]
    aba("2026 — prioridade", de26,
        "As obras do ano corrente: reclassificar as encerradas e reabrir as indeferidas no projeto "
        "certo. A 0112600611 está em andamento e sem SIGCO nenhum — pode nascer certa agora. "
        "Não entra: 0202600193 (deslocamento por obra de cliente, projeto 54554 — expansão).")
    aba("Passivo 2019–2025", resto,
        "Registro para a controladoria: serviço de RL/RT com custo carimbado fora do par ao longo "
        "dos anos. Tudo encerrado — a correção é reclassificação contábil. " + p["nota_valor"])
    wb.remove(wb["Sheet"])
    os.makedirs(os.path.dirname(DESTINO_XLSX), exist_ok=True)
    wb.save(DESTINO_XLSX)


if __name__ == "__main__":
    p = montar()
    planilha(p)
    print(f"caso: {p['total_obras']} obras | 2026: {p['de_2026']} | "
          f"realizado conhecido R$ {p['valor_realizado_conhecido']:,.2f}")
    print("gravados:", DESTINO_JSON, "e", DESTINO_XLSX)
