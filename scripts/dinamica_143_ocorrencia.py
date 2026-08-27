"""
A dinâmica dos 143 com a data da ocorrência inicial.

Pedido do gestor (27/08): ele colou a aba da dinâmica com duas colunas criadas e
vazias — «Prazo do SLA Dias» e «SS COCM» — e pediu a data de ocorrência.

A DATA DA OCORRÊNCIA INICIAL é o campo DTA_OCORRENCIA da SS que ABRIU a cadeia, não
da SS atual. Entre o defeito acontecer e a demanda chegar ao posto passam meses, e
esse pedaço é invisível em qualquer conta que comece na SS de hoje. Para achar a
raiz, a cadeia é percorrida de trás para a frente pelo carimbo de repasse.

Um ativo pode ter mais de uma SS na dinâmica (a coluna traz «A | B»); nesse caso
vale a ocorrência MAIS ANTIGA — é ela que responde há quanto tempo o equipamento
está no problema.

Grava dist/DINAMICA_143_COM_OCORRENCIA.xlsx.
Rodar: python3 scripts/dinamica_143_ocorrencia.py [base_de_repasse.xlsx]
"""

import datetime as dt
import json
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
import sla_manutencao as base  # noqa: E402
import sla_por_equipe as eq  # noqa: E402

SAIDA = os.path.join(RAIZ, "dist", "DINAMICA_143_COM_OCORRENCIA.xlsx")
HOJE = base.HOJE

TITULO = Font(bold=True, color="FFFFFF", size=10)
FUNDO = PatternFill("solid", fgColor="1F3864")
BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)

COLUNAS = [
    ("Situação", 18), ("Ativo", 13), ("Tipo", 11), ("Localidade", 24),
    ("Criticidade", 14),
    # o que o gestor pediu
    ("Data da ocorrência inicial", 14), ("SS que abriu a ocorrência", 21),
    ("Dias desde a ocorrência", 12),
    ("SS", 26), ("Prazo do SLA Dias", 11), ("SS COCM", 21), ("Equipe do COCM", 14),
    ("Desde", 12), ("Ano", 7), ("Dias", 8), ("Resolveu e voltou", 12),
    ("Como terminou", 15), ("Posto que fechou", 15), ("Realizado DCMD", 12),
    ("Onde está", 14), ("Etapa", 26), ("Nota", 60),
]


def montar(caminho=None):
    reg = base.base_de_repasse(caminho)
    anterior = eq.raizes(reg)
    ativos = eq.dados_do_ativo(caminho)
    crit = base.criticidades()
    with open(os.path.join(RAIZ, "data", "missao", "dinamica_143.json"),
              encoding="utf-8") as fh:
        d = json.load(fh)
    # a passagem pelo COCM de cada ativo, para preencher as duas colunas vazias
    # a entrega tem de ser da MESMA demanda da linha, não de outro ciclo do ativo:
    # equipamento reincidente tem várias cadeias, e pegar a última pendurava no
    # ativo uma SS de COCM de um problema antigo.
    entregas = {}
    for e in eq.entregas_ao_cocm(reg, crit, ativos):
        entregas[eq.do_comeco(e["ss_coep"], anterior)] = e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Dinâmica ({len(d['itens'])})"
    ws.append([c[0] for c in COLUNAS])
    for i, (_, larg) in enumerate(COLUNAS, 1):
        cel = ws.cell(row=1, column=i)
        cel.font, cel.fill = TITULO, FUNDO
        cel.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.row_dimensions[1].height = 32

    ordem = {"Na fila do posto": 0, "Em outra mesa": 1, "Resolvido": 2}
    sem_data = []
    for it in sorted(d["itens"], key=lambda x: (ordem.get(x["situacao"], 9),
                                                x.get("tipo", ""), x["ativo"])):
        # a ocorrência mais antiga entre as SS do ativo — o ativo pode ter mais de uma
        melhor, raiz_ss, e = None, "", None
        for pedaco in str(it.get("ss") or "").split("|"):
            k = base.norm(pedaco)
            if not k:
                continue
            inicio = eq.do_comeco(k, anterior)
            if e is None:
                e = entregas.get(inicio)
            oc = (ativos.get(inicio, {}) or ativos.get(k, {})).get("ocorrencia")
            if oc and (melhor is None or oc < melhor):
                melhor, raiz_ss = oc, inicio
        c = it.get("criticidade") or ""
        prazo = base.PRAZO.get(c, base.PRAZO_SEM_CRITICIDADE)
        if melhor is None:
            sem_data.append(it["ativo"])
        it["_cocm"] = bool(e)
        ws.append([
            it["situacao"], it["ativo"], it.get("tipo", ""), it.get("localidade", ""),
            c or "—",
            melhor.strftime("%d/%m/%Y") if melhor else "",
            raiz_ss,
            (HOJE.date() - melhor.date()).days if melhor else "",
            it.get("ss", ""), prazo,
            e["ss_gerada"] if e else "", e["equipe"] if e else "",
            it.get("desde", ""), it.get("ano", ""), it.get("dias", ""),
            "SIM" if it.get("resolvido_e_voltou") else "",
            it.get("como_terminou", ""), it.get("posto_que_fechou", ""),
            "sim" if it.get("realizado_dcmd") else "",
            it.get("onde_esta", ""), it.get("etapa", ""), it.get("nota", ""),
        ])
    for linha in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cel in linha:
            cel.border = BORDA
            cel.alignment = Alignment(vertical="top",
                                      wrap_text=cel.column in (21, 22))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:V{ws.max_row}"

    ws2 = wb.create_sheet("Como foi feito")
    ws2.column_dimensions["A"].width = 112
    for t in [
        "A DINÂMICA DOS 143 COM A DATA DA OCORRÊNCIA INICIAL.",
        "",
        "A DATA DA OCORRÊNCIA INICIAL é o campo DTA_OCORRENCIA da SS que ABRIU a cadeia, "
        "não da SS de hoje. Entre o defeito acontecer e a demanda chegar ao posto passam "
        "meses, e esse pedaço é invisível em qualquer conta que comece na SS atual. Para "
        "achar a raiz, a cadeia é percorrida de trás para a frente pelo carimbo de "
        "repasse até a primeira SS.",
        "",
        "A coluna «SS que abriu a ocorrência» mostra de onde a data veio, para a "
        "conferência ser direta no SGM.",
        "",
        "QUANDO O ATIVO TEM MAIS DE UMA SS na dinâmica (a coluna traz «A | B»), vale a "
        "ocorrência MAIS ANTIGA — é ela que responde há quanto tempo o equipamento está "
        "no problema.",
        "",
        "«Dias desde a ocorrência» conta até 18/08/2026, a posição da conta do posto — "
        "inclusive nos resolvidos, para a régua ser a mesma na coluna inteira.",
        "",
        "AS DUAS COLUNAS QUE ESTAVAM VAZIAS foram preenchidas: «Prazo do SLA Dias» sai "
        "da criticidade da operação (Muito Alta 8, Alta 15, Média 30, Baixa 50, e 26 "
        "sem classificação) e «SS COCM» traz a SS gerada na equipe de campo, com a "
        "equipe ao lado. Ativo que nunca foi ao campo fica com as duas em branco — não "
        "há SLA de manutenção a cobrar dele.",
        "",
        "O resto das colunas é a dinâmica como estava: partição disjunta dos 143, com os "
        "11 que resolveram e voltaram contando só na fila.",
        "",
        "Fonte da ocorrência: base de repasse (Eqp_joao / EQP_SS_OCORRENCIA). Posição de "
        "18/08/2026.",
    ]:
        ws2.append([t])
        ws2.cell(row=ws2.max_row, column=1).alignment = Alignment(wrap_text=True,
                                                                  vertical="top")
    ws2.cell(row=1, column=1).font = Font(bold=True, size=12)

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    return d["itens"], sem_data, entregas


if __name__ == "__main__":
    itens, sem_data, entregas = montar(
        sys.argv[1] if len(sys.argv) > 1 else None)
    print(f"gravado: {SAIDA}")
    print(f"  {len(itens)} linhas · sem data de ocorrência: {len(sem_data)} {sem_data}")
    com_cocm = sum(1 for i in itens if i.get("_cocm"))
    print(f"  com SS de COCM preenchida: {com_cocm}")
