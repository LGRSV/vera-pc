"""
Os 85 do COEP em julho e agosto — recorte de dois meses, com os gráficos.

Pedido do gestor (25/08): a visão do posto só na janela de 01/07 a 18/08, e só com
quem esteve lá dentro. «Esteve no posto» é o mesmo critério do ano: a SS cujo
intervalo [chegada, saída] cruza a janela — não só quem chegou nela. Por isso são 85
e não os 58 que chegaram no período.

A saída da SS vem da cadeia de repasse, nunca do campo de conclusão sozinho: SS
repassada sai sem conclusão, e tratar isso como «ainda no posto» arrasta SS velha
para dentro da janela.

Em dois meses o mês a mês não desenha nada, então a série é SEMANAL, de segunda a
domingo, com as pontas cortadas na janela — a primeira semana começa numa quarta e a
última tem dois dias.

Grava dist/COEP_JUL_AGO.xlsx.
Rodar: python3 scripts/coep_jul_ago_excel.py
"""

import datetime as dt
import json
import os
from collections import Counter

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FONTE = os.path.join(RAIZ, "data", "missao", "coep_2026.json")
SAIDA = os.path.join(RAIZ, "dist", "COEP_JUL_AGO.xlsx")

INICIO = dt.datetime(2026, 7, 1)
FIM = dt.datetime(2026, 8, 18, 23, 59)
NOME = {"religador": "Religador", "regulador": "Regulador"}

TITULO = Font(bold=True, color="FFFFFF", size=10)
FUNDO = PatternFill("solid", fgColor="1F3864")
BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def data(texto):
    try:
        return dt.datetime.strptime(texto, "%d/%m/%Y")
    except (TypeError, ValueError):
        return None


def saida(s):
    """Quando a SS saiu do posto — vazio quer dizer que ainda está lá."""
    return data(s["saiu"]) if s["saiu"] else None


def recorte(d):
    ss = [s for s in d["ss"]
          if data(s["chegou"]) and data(s["chegou"]) <= FIM
          and (saida(s) or FIM) >= INICIO]
    alvo = {s["ativo"] for s in ss}
    ativos = {a["ativo"]: a for a in d["ativos"] if a["ativo"] in alvo}
    resolvidos = {r["ativo"] for r in d["resolvidos_do_coep"]
                  if r["conta_como_resolvido_pelo_coep"]}
    outra = {r["ativo"] for r in d["pendentes_em_outra_mesa"]}
    return ss, ativos, resolvidos & alvo, outra & alvo


def onde_esta(a, resolvidos, outra):
    """Onde o equipamento está HOJE. Quem voltou ao posto conta como no posto —
    a demanda está de novo na mesa, mesmo tendo sido resolvida antes no ano."""
    if a["segue_no_posto"]:
        return "Segue no posto"
    if a["ativo"] in resolvidos:
        return "Saiu — resolvido pelo COEP"
    if a["ativo"] in outra:
        return "Saiu — está em outra mesa"
    return "Saiu — sem destino apurado"


def semanas(ss):
    """Segunda a domingo, com as pontas cortadas na janela."""
    fora = []
    x = INICIO - dt.timedelta(days=INICIO.weekday())
    while x <= FIM:
        ini = max(x, INICIO)
        fim = min(x + dt.timedelta(days=6, hours=23, minutes=59), FIM)
        chegaram = {s["ativo"] for s in ss if ini <= data(s["chegou"]) <= fim}
        sairam = {s["ativo"] for s in ss if saida(s) and ini <= saida(s) <= fim}
        dentro = {s["ativo"] for s in ss
                  if data(s["chegou"]) <= fim and (saida(s) or FIM) >= fim}
        fora.append({"rotulo": ini.strftime("%d/%m"), "chegaram": len(chegaram),
                     "sairam": len(sairam), "no_posto": len(dentro),
                     "ate": fim.strftime("%d/%m")})
        x += dt.timedelta(days=7)
    return fora


FAIXAS = [(0, 30, "até 30 dias"), (31, 60, "31 a 60"), (61, 90, "61 a 90"),
          (91, 180, "91 a 180"), (181, 10**6, "mais de 180")]


def faixa_de(dias):
    for a, b, nome in FAIXAS:
        if a <= dias <= b:
            return nome
    return FAIXAS[-1][2]


def cabecalho(ws, colunas):
    ws.append([c[0] for c in colunas])
    for i, (_, larg) in enumerate(colunas, 1):
        cel = ws.cell(row=1, column=i)
        cel.font, cel.fill = TITULO, FUNDO
        cel.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.row_dimensions[1].height = 32


def bordar(ws, ate, de=2):
    for linha in ws.iter_rows(min_row=de, max_row=ate):
        for cel in linha:
            cel.border = BORDA


def grafico(ws, tipo, titulo, colunas, rotulos, ancora, altura=8, largura=17):
    """Gráfico nativo. Categoria vai como texto (o openpyxl grava numRef e o Excel
    mostra 1…n), eixos declarados visíveis e o de categoria embaixo."""
    ch = LineChart() if tipo == "linha" else BarChart()
    ch.title = titulo
    ch.height, ch.width = altura, largura
    ch.style = 2
    if tipo != "linha":
        ch.type, ch.gapWidth = "col", 60
    for col in colunas:
        ch.add_data(Reference(ws, min_col=col, min_row=rotulos[0] - 1,
                              max_row=rotulos[1]), titles_from_data=True)
    ref = f"'{ws.title}'!$A${rotulos[0]}:$A${rotulos[1]}"
    for s in ch.series:
        s.cat = AxDataSource(strRef=StrRef(f=ref))
        if tipo == "linha":
            s.smooth = False
            s.marker.symbol, s.marker.size = "circle", 6
    ch.x_axis.delete = ch.y_axis.delete = False
    ch.x_axis.axPos, ch.y_axis.axPos = "b", "l"
    if len(colunas) == 1:
        ch.legend = None
    ws.add_chart(ch, ancora)


def bloco(ws, titulo, pares, coluna_valor="Equipamentos"):
    """Escreve um quadrinho rótulo/valor e devolve (primeira, última) linha."""
    ws.append([])
    ws.append([titulo, coluna_valor])
    for c in (1, 2):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
    primeira = ws.max_row + 1
    for rotulo, valor in pares:
        ws.append([rotulo, valor])
    bordar(ws, ws.max_row, primeira)
    return primeira, ws.max_row


def montar():
    with open(FONTE, encoding="utf-8") as fh:
        d = json.load(fh)
    ss, ativos, resolvidos, outra = recorte(d)
    sems = semanas(ss)
    sit = Counter(onde_esta(a, resolvidos, outra) for a in ativos.values())
    ordem_sit = ["Segue no posto", "Saiu — resolvido pelo COEP",
                 "Saiu — está em outra mesa", "Saiu — sem destino apurado"]
    voltaram = sum(1 for a in ativos.values()
                   if a["segue_no_posto"] and a["ativo"] in resolvidos)

    wb = openpyxl.Workbook()

    # ---- 1 · Os 85 -------------------------------------------------------
    ws = wb.active
    ws.title = "1 · Os 85"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.append(["OS 85 DO COEP · 1º DE JULHO A 18 DE AGOSTO DE 2026"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.append(["Equipamento cuja SS esteve no posto em algum momento da janela — "
               "não só quem chegou nela."])

    a1, a2 = bloco(ws, "Por tipo", [(NOME[t], n) for t, n in
                                    Counter(a["tipo"] for a in ativos.values()).most_common()])
    b1, b2 = bloco(ws, "Onde estão hoje", [(s, sit[s]) for s in ordem_sit if sit[s]])
    ordem_cri = ["Muito Alta", "Alta", "Média", "Baixa"]
    cri = Counter(a["criticidade"] or "Sem classificação" for a in ativos.values())
    c1, c2 = bloco(ws, "Criticidade", [(k, cri[k]) for k in ordem_cri if cri[k]]
                   + [("Sem classificação", cri["Sem classificação"])])
    dias = Counter(faixa_de(a["dias_no_posto"]) for a in ativos.values())
    d1, d2 = bloco(ws, "Tempo no posto", [(n, dias[n]) for _, _, n in FAIXAS if dias[n]],
                   "Equipamentos")
    saiu = Counter(s["foi_para"] or "Fechou no COEP" for s in ss if saida(s))
    e1, e2 = bloco(ws, "Para onde foi quem saiu", saiu.most_common(), "SS")

    ws.append([])
    ws.append([f"Dos 85, {len(resolvidos)} tiveram demanda resolvida pelo COEP no ano; "
               f"{voltaram} deles voltaram ao posto depois e por isso aparecem em "
               f"«segue no posto», não em «resolvido»."])
    ws.append([f"A partição fecha: {sit['Segue no posto']} + "
               f"{sit['Saiu — resolvido pelo COEP']} + "
               f"{sit['Saiu — está em outra mesa']} = {len(ativos)}."])

    grafico(ws, "barra", "Por tipo", [2], (a1, a2), "E3", altura=7, largura=13)
    grafico(ws, "barra", "Onde estão hoje", [2], (b1, b2), "E17", altura=7, largura=13)
    grafico(ws, "barra", "Criticidade", [2], (c1, c2), "E31", altura=7, largura=13)
    grafico(ws, "barra", "Tempo no posto", [2], (d1, d2), "N3", altura=7, largura=13)
    grafico(ws, "barra", "Para onde foi quem saiu", [2], (e1, e2), "N17", altura=7, largura=13)

    # ---- 2 · Semana a semana ---------------------------------------------
    ws = wb.create_sheet("2 · Semana a semana")
    cabecalho(ws, [("Semana de", 12), ("até", 10), ("Chegaram", 12), ("Saíram", 12),
                   ("No posto no fim da semana", 24)])
    for s in sems:
        ws.append([s["rotulo"], s["ate"], s["chegaram"], s["sairam"], s["no_posto"]])
    ultima = ws.max_row
    bordar(ws, ultima)
    for linha in ws.iter_rows(min_row=2, max_row=ultima):
        for cel in linha:
            cel.alignment = Alignment(horizontal="center")
    grafico(ws, "linha", "Chegaram e saíram, por semana", [3, 4], (2, ultima),
            f"A{ultima + 2}")
    grafico(ws, "linha", "No posto no fim de cada semana", [5], (2, ultima),
            f"J{ultima + 2}")
    ws.cell(row=ultima + 18, column=1,
            value="A primeira semana começa numa quarta (1º de julho) e a última tem "
                  "dois dias (17 e 18 de agosto) — a janela corta as pontas.")

    # ---- 3 · A lista ------------------------------------------------------
    ws = wb.create_sheet(f"3 · A lista ({len(ativos)})")
    cabecalho(ws, [("Ativo", 13), ("Tipo", 12), ("Localidade", 24), ("Criticidade", 13),
                   ("Onde está hoje", 26), ("SS no COEP", 20), ("Chegou", 12),
                   ("Dias no posto", 13), ("Parecer COEP", 30), ("Na carteira", 12)])
    for a in sorted(ativos.values(), key=lambda x: (x["tipo"], x["localidade"], x["ativo"])):
        ws.append([a["ativo"], NOME[a["tipo"]], a["localidade"], a["criticidade"] or "—",
                   onde_esta(a, resolvidos, outra), a["ss"], a["primeira_chegada"],
                   a["dias_no_posto"], (a["parecer_coep"] or "—")[:120],
                   "sim" if a["na_carteira"] else "não"])
    bordar(ws, ws.max_row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    # ---- 4 · As SS da janela ---------------------------------------------
    ws = wb.create_sheet(f"4 · SS na janela ({len(ss)})")
    cabecalho(ws, [("SS", 20), ("Ativo", 13), ("Tipo", 12), ("Situação da SS", 16),
                   ("Chegou", 12), ("Saiu", 12), ("Como apurou a saída", 24),
                   ("Foi para", 14), ("Dias no posto", 13), ("Ano da SS", 11)])
    for s in sorted(ss, key=lambda x: (data(x["chegou"]), x["ativo"])):
        ws.append([s["ss"], s["ativo"], NOME[s["tipo"]], s["status"], s["chegou"],
                   s["saiu"] or "— ainda no posto", s["como_apurou_a_saida"],
                   s["foi_para"] or "—", s["dias_no_posto"], s["ano_da_ss"]])
    bordar(ws, ws.max_row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    # ---- Como foi feito ---------------------------------------------------
    ws = wb.create_sheet("Como foi feito")
    ws.column_dimensions["A"].width = 115
    for texto in como_foi_feito(ss, ativos, resolvidos, sit, voltaram):
        ws.append([texto])
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True,
                                                                vertical="top")

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    return ss, ativos, sit, resolvidos, voltaram


def como_foi_feito(ss, ativos, resolvidos, sit, voltaram):
    return [
        "OS 85 DO COEP — janela de 1º de julho a 18 de agosto de 2026.",
        "",
        f"Quem entra: equipamento cuja SS esteve no posto do COEP em algum momento da "
        f"janela. «Esteve» é o intervalo [chegada, saída] cruzando o período — não só "
        f"quem chegou nele. São {len(ativos)} equipamentos em {len(ss)} SS. Quem chegou "
        f"dentro da janela é menos: 58.",
        "",
        "A saída da SS vem da cadeia de repasse, não do campo de conclusão sozinho: SS "
        "repassada sai da base sem data de conclusão, e quem trata isso como «ainda no "
        "posto» arrasta SS velha para dentro da janela. A saída é a conclusão, se "
        "houver; senão a abertura da SS seguinte; senão a SS ainda está no posto.",
        "",
        "Conta EQUIPAMENTO, não SS: o mesmo religador pode ter três SS no posto na "
        "janela e continua sendo um equipamento.",
        "",
        f"Onde estão hoje, na posição de 18/08: {sit['Segue no posto']} seguem no posto, "
        f"{sit['Saiu — resolvido pelo COEP']} saíram resolvidos pelo COEP e "
        f"{sit['Saiu — está em outra mesa']} saíram e estão em outra mesa. A partição "
        f"fecha em {len(ativos)}, sem sobreposição.",
        "",
        f"Cuidado com o número de resolvidos: {len(resolvidos)} dos 85 tiveram uma "
        f"demanda resolvida pelo COEP em 2026, mas {voltaram} voltaram ao posto depois. "
        f"Esses aparecem em «segue no posto», porque a demanda está de novo na mesa. "
        f"Nota nova depois de SS atendida é reincidência — demanda nova, não anula a "
        f"resolvida.",
        "",
        "Resolvido pelo COEP: a demanda passou pelo posto dentro de 2026 E a cadeia dela "
        "fechou dentro de 2026, com SS atendida ou cancelada. Quem fecha não precisa ser "
        "o COEP — o posto diagnostica e despacha, a ponta executa.",
        "",
        "A série é semanal, de segunda a domingo, porque em dois meses o mês a mês não "
        "desenha nada. As pontas ficam cortadas: a primeira semana começa numa quarta "
        "(1º de julho) e a última tem dois dias (17 e 18 de agosto).",
        "",
        "«No posto no fim da semana» é foto, não fluxo: conta quem estava dentro no "
        "último dia da semana. Por isso não é a soma de chegaram menos saíram.",
        "",
        "Fonte: base de SS/OS de 20/08/2026 e a cadeia de repasse do COEP. Posição de "
        "18/08/2026.",
    ]


if __name__ == "__main__":
    ss, ativos, sit, resolvidos, voltaram = montar()
    print(f"gravado: {SAIDA}")
    print(f"  {len(ativos)} equipamentos em {len(ss)} SS")
    for k, v in sit.items():
        print(f"  {k}: {v}")
    print(f"  resolvidos no ano entre os 85: {len(resolvidos)} "
          f"({voltaram} voltaram ao posto)")
