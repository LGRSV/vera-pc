"""
A taxa de falha POR PEÇA em 2025 e 2026 — dist/TAXA_POR_PECA.xlsx.

O rol de falhas da planilha base já traz SS, ativo, tensão e peça. O que faltava vem
dos ajustes da proteção, em GESTAO_DE_EQUIPAMENTOS.xlsx:

  Ajustes Reguladores de Tensão  →  POTÊNCIA e CONTROLADOR do RT (426 linhas)
  Ajustes RL Poste               →  TENSÃO e RELÉ do RL (1.292 linhas)

Casamento: 27 de 27 RT e 62 de 63 RL. O único RL sem ajuste (7930359149, Caseara) é
também o único com tensão «#N/A» no rol; a tensão dele fica INFERIDA pela praça — os
outros seis religadores de Caseara são todos 34.500 —, marcada como tal.

DUAS RESSALVAS que a planilha registra em aba própria, porque mudam número:

1. POTÊNCIA FORA DAS TRÊS CLASSES. O gestor diz que só existe 167, 200 e 400. O
   cadastro dos ajustes traz também 239, 250, 398 e um banco misto 167/250/167. O 398
   é claramente da classe 400; 239 e 250 não são nenhuma das três. A planilha mostra o
   valor cru E a classe pelo mais próximo, com o desvio marcado.

2. ATIVO REPETIDO NO MESMO ANO. A régua do gestor manda contar equipamento uma vez por
   ano, mas o rol tem 90 linhas para 87 pares ativo-ano. Para taxa POR PEÇA a repetição
   pode ser legítima (o 7908708116 quebrou tanque E controle na mesma SS); para taxa por
   equipamento, não. As duas contagens saem lado a lado.

Rodar: python3 scripts/taxa_por_peca.py [planilha_base.xlsx]
"""

import os
import sys
from collections import Counter, defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.drawing.colors import ColorChoice
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = os.path.join(RAIZ, "data", "raw", "GESTAO_EQUIPAMENTOS_ESPECIAIS_COEP.xlsx")
AJUSTES = os.path.join(RAIZ, "data", "raw", "GESTAO_DE_EQUIPAMENTOS.xlsx")
SAIDA = os.path.join(RAIZ, "dist", "TAXA_POR_PECA.xlsx")

TINTA, PAPEL, SINAL = "FF211D15", "FFF2EFE6", "FFBC4B0E"
COR_A, COR_B = "1F7C50", "B8480C"      # as duas séries, já validadas no dataviz
PARQUE = {"RL": 1307, "RT": 207}        # a régua oficial, vale para os três anos
CLASSES_RT = (167, 200, 400)            # o gestor: só existem estas três
PECA_ROTULO = {"completo": "Equipamento completo", "tanque": "Tanque / parte ativa",
               "controle": "Controle", "celula": "Célula", "rele": "Relé",
               "furto": "Furto"}
# a ordem é a do peso: da peça que troca o equipamento inteiro à que não troca nada
ORDEM_PECA = ["completo", "tanque", "celula", "controle", "rele", "furto"]


def texto(v):
    return "" if v is None else str(v).strip()


def classe_rt(bruto):
    """A classe de potência pelo mais próximo das três, com o desvio marcado."""
    if not bruto:
        return None, "", False
    partes = [p for p in str(bruto).replace("/", " ").split() if p]
    nums = []
    for p in partes:
        try:
            nums.append(float(p.replace(",", ".")))
        except ValueError:
            pass
    if not nums:
        return None, texto(bruto), False
    # num banco misto vale a potência que mais aparece; empate, a maior
    c = Counter(nums)
    valor = max(c, key=lambda x: (c[x], x))
    classe = min(CLASSES_RT, key=lambda k: abs(k - valor))
    return classe, texto(bruto), classe != valor


def ler():
    caminho = sys.argv[1] if len(sys.argv) > 1 else BASE
    ws = load_workbook(caminho, read_only=True, data_only=True)["Falha Equipamentos"]
    linhas = [r for r in list(ws.iter_rows(values_only=True))[1:]
              if texto(r[0]) and r[4]]

    wb = load_workbook(AJUSTES, read_only=True, data_only=True)
    rt = {}
    for r in list(wb["Ajustes Reguladores de Tensão"].iter_rows(values_only=True))[1:]:
        if r[0]:
            rt.setdefault(texto(r[0]), {"potencia": r[4], "controlador": texto(r[2]),
                                        "parte_ativa": texto(r[3]),
                                        "tensao": texto(r[8]), "corrente": r[5],
                                        "localidade": texto(r[14])})
    rl = {}
    for r in list(wb["Ajustes RL Poste"].iter_rows(values_only=True))[1:]:
        if r[0]:
            rl.setdefault(texto(r[0]), {"rele": texto(r[10]), "tensao": texto(r[12]),
                                        "localidade": texto(r[13]),
                                        "alimentador": texto(r[11])})
    return linhas, rt, rl


def faixa(bruto):
    """13,8 ou 34,5 — as duas únicas que existem."""
    t = texto(bruto).replace(".", "").replace(",", "").replace(" ", "")
    if t.startswith("138"):
        return "13,8 kV"
    if t.startswith("345"):
        return "34,5 kV"
    return ""


def montar():
    linhas, aj_rt, aj_rl = ler()
    itens, alertas = [], []
    for r in linhas:
        ativo, tipo = texto(r[4]), texto(r[1]).upper()
        aj = (aj_rt if tipo == "RT" else aj_rl).get(ativo, {})
        f = faixa(r[6]) or faixa(aj.get("tensao"))
        origem_tensao = "rol de falhas" if faixa(r[6]) else (
            "ajustes da proteção" if faixa(aj.get("tensao")) else "")
        if not f and tipo == "RL" and ativo == "7930359149":
            f, origem_tensao = "34,5 kV", "inferida pela praça (Caseara)"
            alertas.append(["Tensão inferida", ativo, tipo,
                            "sem cadastro nos Ajustes RL Poste e «#N/A» no rol; "
                            "os outros seis religadores de Caseara são 34.500"])
        pot, pot_bruta, desviou = (classe_rt(aj.get("potencia"))
                                   if tipo == "RT" else (None, "", False))
        if desviou:
            alertas.append(["Potência fora das três classes", ativo, tipo,
                            f"cadastro diz {pot_bruta}; classificado como {pot} "
                            f"pelo mais próximo de 167/200/400"])
        peca = texto(r[11]).lower()
        # o ano vem da FATIA, não da coluna Ano: numa das 90 as duas discordam
        # (7933585074, falha de 2025 com ocorrência em 27/01/2026)
        fatia = texto(r[0])
        ano_fatia = fatia.split()[-1] if fatia else texto(r[9])
        itens.append({
            "fatia": fatia, "ano": ano_fatia, "tipo": tipo, "ativo": ativo,
            "ss": texto(r[5]), "faixa": f, "origem_tensao": origem_tensao,
            "potencia": pot or "", "potencia_bruta": pot_bruta,
            "controlador": aj.get("controlador", "") if tipo == "RT" else "",
            "rele_rl": aj.get("rele", "") if tipo == "RL" else "",
            "parte_ativa": aj.get("parte_ativa", "") if tipo == "RT" else "",
            "localidade": aj.get("localidade", ""),
            "peca": peca, "peca_rotulo": PECA_ROTULO.get(peca, peca or "(sem peça)"),
            "troca": texto(r[12]), "causa": texto(r[13]),
            "data": texto(r[7]), "mes": texto(r[8]),
        })

    # a régua do gestor: ativo conta uma vez por ano
    vistos = defaultdict(list)
    for i in itens:
        vistos[(i["ano"], i["tipo"], i["ativo"])].append(i)
    for (ano, tipo, ativo), v in sorted(vistos.items()):
        if len(v) > 1:
            pecas = " + ".join(sorted(x["peca"] for x in v))
            mesma = len({x["ss"] for x in v}) == 1
            alertas.append([
                "Ativo repetido no mesmo ano", ativo, f"{tipo} {ano}",
                f"{len(v)} linhas ({pecas}) — "
                + ("mesma SS, duas peças na mesma falha" if mesma
                   else "SS diferentes, duas falhas no ano") +
                "; conta uma vez na taxa por equipamento"])
    return itens, alertas


# ------------------------------------------------------------------ a planilha
def cabeca(ws, linha, titulo, cols, larguras=None):
    ws.cell(row=linha, column=1, value=titulo).font = Font(bold=True, size=11,
                                                           color=SINAL)
    ws.append(cols)
    for i in range(1, len(cols) + 1):
        c = ws.cell(row=linha + 1, column=i)
        c.font, c.fill = Font(bold=True, color=PAPEL, size=10), \
            PatternFill("solid", fgColor=TINTA)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if larguras:
        for i, w in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return linha + 2


def fecha(ws, linha, n):
    for i in range(1, n + 1):
        ws.cell(row=linha, column=i).font = Font(bold=True)
        ws.cell(row=linha, column=i).border = Border(
            top=Side(style="medium", color=TINTA))


def grafico(ws, titulo, cab, r0, r1, onde, cols=(2, 3)):
    g = BarChart()
    g.type, g.grouping, g.overlap = "col", "stacked", 100
    g.title, g.height, g.width, g.gapWidth = titulo, 9, 22, 70
    for col, cor in zip(cols, (COR_A, COR_B)):
        s = Series(Reference(ws, min_col=col, min_row=cab, max_row=r1),
                   title_from_data=True)
        s.graphicalProperties.solidFill = ColorChoice(srgbClr=cor)
        s.graphicalProperties.line.noFill = True
        g.series.append(s)
    cats = AxDataSource(strRef=StrRef(f=f"'{ws.title}'!$A${r0}:$A${r1}"))
    for s in g.series:
        s.cat = cats
    ws.add_chart(g, onde)


def parque_por_classe():
    """O parque de cada classe, dos mesmos ajustes de onde sai a classe da falha.

    Sem isso a taxa por potência não existe: dividir a célula de 200 pelo parque
    inteiro de regulador esconde que o parque de 200 é menor que o de 400.
    """
    wb = load_workbook(AJUSTES, read_only=True, data_only=True)
    rt_pot, rt_ten, rl_ten = Counter(), Counter(), Counter()
    for r in list(wb["Ajustes Reguladores de Tensão"].iter_rows(values_only=True))[1:]:
        if not r[0]:
            continue
        classe, _, _ = classe_rt(r[4])
        if classe:
            rt_pot[classe] += 1
        f = faixa(r[8])
        if f:
            rt_ten[f] += 1
    for r in list(wb["Ajustes RL Poste"].iter_rows(values_only=True))[1:]:
        if r[0]:
            f = faixa(r[12])
            if f:
                rl_ten[f] += 1
    return rt_pot, rt_ten, rl_ten


def aba_por_classe(wb, itens):
    """A taxa de cada peça contra o parque da CLASSE, não do tipo inteiro."""
    rt_pot, rt_ten, rl_ten = parque_por_classe()
    ws = wb.create_sheet("Taxa por classe")
    ws.column_dimensions["A"].width = 26
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 14

    def bloco(titulo, chaves, parque, filtro, rotulo):
        r0 = cabeca(ws, ws.max_row + (2 if ws.max_row > 1 else 0), titulo,
                    ["Peça"] + [rotulo(k) for k in chaves] + ["Total"])
        sub = [i for i in itens if filtro(i)]
        pecas = [p for p in ORDEM_PECA if any(i["peca"] == p for i in sub)]
        for p in pecas:
            linha = [sum(1 for i in sub if i["peca"] == p and i["chave"] == k)
                     for k in chaves]
            if not sum(linha):
                continue
            ws.append([PECA_ROTULO[p]] + [x or "—" for x in linha] + [sum(linha)])
        tot = [sum(1 for i in sub if i["chave"] == k) for k in chaves]
        ws.append(["Total de falhas"] + tot + [sum(tot)])
        fecha(ws, ws.max_row, len(chaves) + 2)
        ws.append(["Parque da classe"] + [parque.get(k, 0) for k in chaves]
                  + [sum(parque.get(k, 0) for k in chaves)])
        ws.append(["TAXA no biênio"]
                  + [round(tot[n] / parque[k], 4) if parque.get(k) else 0
                     for n, k in enumerate(chaves)]
                  + [round(sum(tot) / sum(parque.get(k, 0) for k in chaves), 4)])
        for c in range(2, len(chaves) + 3):
            ws.cell(row=ws.max_row, column=c).number_format = "0.0%"
            ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        return r0

    for i in itens:
        i["chave"] = i["potencia"]
    bloco("Regulador — por potência da célula (parque dos ajustes: 190)",
          list(CLASSES_RT), rt_pot, lambda i: i["tipo"] == "RT" and i["potencia"],
          lambda k: f"{k} kVA")
    for i in itens:
        i["chave"] = i["faixa"]
    bloco("Religador — por faixa de tensão (parque dos ajustes: 1.292)",
          ["13,8 kV", "34,5 kV"], rl_ten,
          lambda i: i["tipo"] == "RL" and i["faixa"], lambda k: k)
    bloco("Regulador — por faixa de tensão (parque dos ajustes: 190)",
          ["13,8 kV", "34,5 kV"], rt_ten,
          lambda i: i["tipo"] == "RT" and i["faixa"], lambda k: k)

    # a célula isolada — a pergunta que o gestor fez em cima da mesa
    r0 = cabeca(ws, ws.max_row + 2,
                "A CÉLULA, isolada — a taxa de cada potência",
                ["Potência da célula", "Falhas no biênio", "Parque da classe",
                 "Taxa", "Índice vs média", "Amostra"],
                [26, 16, 16, 12, 15, 26])
    cel = [i for i in itens if i["tipo"] == "RT" and i["peca"] == "celula"
           and i["potencia"]]
    media = len(cel) / sum(rt_pot.get(k, 0) for k in CLASSES_RT)
    for k in CLASSES_RT:
        f = sum(1 for i in cel if i["potencia"] == k)
        n = rt_pot.get(k, 0)
        taxa = f / n if n else 0
        ws.append([f"{k} kVA", f or "—", n, round(taxa, 4),
                   round(taxa / media, 2) if media else 0,
                   "pequena demais para concluir" if n < 30 else ""])
        ws.cell(row=ws.max_row, column=4).number_format = "0.0%"
        ws.cell(row=ws.max_row, column=5).number_format = "0.00"
    ws.append(["Total / média", len(cel),
               sum(rt_pot.get(k, 0) for k in CLASSES_RT), round(media, 4), 1.0, ""])
    ws.cell(row=ws.max_row, column=4).number_format = "0.0%"
    ws.cell(row=ws.max_row, column=5).number_format = "0.00"
    fecha(ws, ws.max_row, 6)

    ws.append([])
    for t in [
        "O que a célula mostra:",
        "A de 200 kVA falha 2,5 vezes mais que as outras duas: 7 falhas em 67",
        "equipamentos contra 4 em 98 na de 400 e 1 em 25 na de 167. É o único recorte",
        "de potência em que a diferença é grande o bastante para não ser acaso.",
        "",
        "A de 167 tem parque de 25 — uma falha a mais dobra a taxa. Está marcada.",
        "",
        "Por que o parque aqui é outro:"
        "A classe de potência e a faixa de tensão só existem nos ajustes da proteção,",
        "então o denominador tem de ser o parque DESSES cadastros: 190 reguladores e",
        "1.292 religadores, não os 207 e 1.307 oficiais. A diferença são os",
        "equipamentos sem estudo de proteção cadastrado.",
        "",
        "Por isso a taxa desta aba é um pouco maior que a da aba «Taxa por peça», que",
        "usa o parque oficial. As duas estão certas; o denominador é que muda.",
        "",
        "As duas se somam de anos:",
        "É a taxa do BIÊNIO — 2025 e 2026 juntos —, porque separar por ano deixaria",
        "célula de 167 com uma falha em 25 equipamentos, número que não sustenta",
        "conclusão nenhuma.",
    ]:
        ws.append([t])
        if t.endswith(":"):
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
    return ws


def aba_modelo(wb, itens):
    """Falhas contra o PARQUE de cada modelo — sem isso o volume engana.

    Das 26 falhas de tanque, 25 são NOJA RC10. Parece problema do modelo; não é: o
    RC10 é 78,6% do parque e falha ABAIXO da média. Quem falha acima é o COOPER F6.
    """
    frota = Counter()
    for r in list(load_workbook(AJUSTES, read_only=True,
                                data_only=True)["Ajustes RL Poste"]
                  .iter_rows(values_only=True))[1:]:
        if r[0] and r[10]:
            frota[texto(r[10])] += 1
    total_frota = sum(frota.values())
    falhas = Counter(i["rele_rl"] for i in itens if i["rele_rl"])
    media = sum(falhas.values()) / total_frota

    ws = wb.create_sheet("Taxa por modelo de relé")
    r0 = cabeca(ws, 1,
                "Falhas de religador contra o parque de cada modelo — 2025 e 2026 juntos",
                ["Modelo do relé", "Parque", "% do parque", "Falhas", "Taxa no biênio",
                 "Índice vs média", "Amostra"],
                [22, 10, 12, 10, 15, 15, 26])
    for m, n in frota.most_common():
        f = falhas.get(m, 0)
        taxa = f / n if n else 0
        ws.append([m, n, round(n / total_frota, 4), f or "—", round(taxa, 4),
                   round(taxa / media, 2) if media else 0,
                   "pequena demais para concluir" if n < 30 else ""])
        for col, fmt in ((3, "0.0%"), (5, "0.0%"), (6, "0.00")):
            ws.cell(row=ws.max_row, column=col).number_format = fmt
    ws.append(["Total / média", total_frota, 1, sum(falhas.values()),
               round(media, 4), 1.0, ""])
    for col, fmt in ((3, "0.0%"), (5, "0.0%"), (6, "0.00")):
        ws.cell(row=ws.max_row, column=col).number_format = fmt
    fecha(ws, ws.max_row, 7)

    ws.append([])
    for t in [
        "Como ler o índice:",
        "1,00 é a média do parque. Acima de 1, o modelo falha mais do que o tamanho",
        "dele explicaria; abaixo, menos.",
        "",
        "O NOJA RC10 concentra 43 das 62 falhas com relé cadastrado, e 25 das 26 de",
        "tanque. Isso é volume, não qualidade: ele é 78,6% do parque e o índice dele é",
        "0,88 — falha ABAIXO da média.",
        "",
        "Quem falha acima é o COOPER F6: 10,3% do parque e 22,6% das falhas, índice",
        "2,19. É o modelo que o próprio parecer da DMSL chama de obsoleto no",
        "7900001227 de Recursolândia («Cooper Form6 de 2009, equipamento obsoleto»).",
        "",
        "ARTECHE P500 e TAVRIDA aparecem com índice alto, mas são 5 e 27 equipamentos:",
        "uma falha já joga a taxa para o teto. Estão marcados como amostra pequena.",
        "",
        "O parque aqui é o dos Ajustes RL Poste (1.292 religadores com relé",
        "cadastrado), não o parque oficial de 1.307 — a diferença são os que não têm",
        "estudo de proteção cadastrado.",
    ]:
        ws.append([t])
    for n2 in range(r0, ws.max_row + 1):
        c = ws.cell(row=n2, column=1)
        if c.value and str(c.value).endswith(":"):
            c.font = Font(bold=True, size=11)
    return ws


def planilha(itens, alertas):
    wb = Workbook()

    # 1 — a taxa por peça
    ws = wb.active
    ws.title = "Taxa por peça"
    pecas = [p for p in ORDEM_PECA if any(i["peca"] == p for i in itens)]
    for tipo in ("RL", "RT"):
        sub = [i for i in itens if i["tipo"] == tipo]
        linha = ws.max_row + (2 if ws.max_row > 1 else 0)
        r0 = cabeca(ws, linha,
                    f"Taxa de falha por peça — {'Religador' if tipo == 'RL' else 'Regulador'}"
                    f" (parque {PARQUE[tipo]})",
                    ["Peça", "2025", "2026", "Total", "Taxa 2025", "Taxa 2026"],
                    [26, 11, 11, 10, 12, 12])
        for p in pecas:
            a = sum(1 for i in sub if i["peca"] == p and i["ano"] == "2025")
            b = sum(1 for i in sub if i["peca"] == p and i["ano"] == "2026")
            if not (a or b):
                continue
            ws.append([PECA_ROTULO[p], a or "—", b or "—", a + b,
                       round(a / PARQUE[tipo], 5), round(b / PARQUE[tipo], 5)])
            for c in (5, 6):
                ws.cell(row=ws.max_row, column=c).number_format = "0.00%"
        ta = sum(1 for i in sub if i["ano"] == "2025")
        tb = sum(1 for i in sub if i["ano"] == "2026")
        ws.append(["Total", ta, tb, ta + tb,
                   round(ta / PARQUE[tipo], 5), round(tb / PARQUE[tipo], 5)])
        for c in (5, 6):
            ws.cell(row=ws.max_row, column=c).number_format = "0.00%"
        fecha(ws, ws.max_row, 6)
        grafico(ws, f"{tipo}: falhas por peça", r0 - 1, r0, ws.max_row - 1,
                f"H{r0 - 1}")

    # 2 — os cruzamentos
    ws2 = wb.create_sheet("Peça x tensão e potência")
    ws2.column_dimensions["A"].width = 26
    for c in "BCDEF":
        ws2.column_dimensions[c].width = 13

    def cruza(titulo, chave, valores, filtro=lambda i: True):
        r0 = cabeca(ws2, ws2.max_row + (2 if ws2.max_row > 1 else 0), titulo,
                    ["Peça"] + [str(v) for v in valores] + ["Total"])
        sub = [i for i in itens if filtro(i)]
        for p in pecas:
            linha = [sum(1 for i in sub if i["peca"] == p and i[chave] == v)
                     for v in valores]
            if not sum(linha):
                continue
            ws2.append([PECA_ROTULO[p]] + [x or "—" for x in linha] + [sum(linha)])
        tot = [sum(1 for i in sub if i[chave] == v) for v in valores]
        ws2.append(["Total"] + tot + [sum(tot)])
        fecha(ws2, ws2.max_row, len(valores) + 2)
        return r0

    cruza("Peça × faixa de tensão", "faixa", ["13,8 kV", "34,5 kV"])
    cruza("Peça × potência — só regulador", "potencia", list(CLASSES_RT),
          lambda i: i["tipo"] == "RT")
    ctrls = sorted({i["controlador"] for i in itens if i["controlador"]})
    cruza("Peça × controlador — só regulador", "controlador", ctrls,
          lambda i: i["tipo"] == "RT")
    reles = sorted({i["rele_rl"] for i in itens if i["rele_rl"]})
    cruza("Peça × relé — só religador", "rele_rl", reles,
          lambda i: i["tipo"] == "RL")

    # 3 — a taxa por classe: RT por potência da célula, RL por tensão
    aba_por_classe(wb, itens)

    # 4 — a taxa por modelo, contra o parque de verdade
    aba_modelo(wb, itens)

    # 5 — a base, uma falha por linha
    ws3 = wb.create_sheet("Base das falhas")
    cols = [("Ano", 7), ("SS", 21), ("Ativo", 13), ("RL/RT", 7),
            ("Faixa de tensão", 14), ("Potência (RT)", 13),
            ("Controlador (RT)", 15), ("Peça", 21), ("Troca feita", 11),
            ("Causa raiz", 40), ("Data", 12), ("Mês", 11), ("Localidade", 22),
            ("Parte ativa (RT)", 17), ("Relé (RL)", 15),
            ("Potência crua no cadastro", 20), ("Origem da tensão", 26)]
    ws3.append([c[0] for c in cols])
    ordem = {p: n for n, p in enumerate(ORDEM_PECA)}
    for i in sorted(itens, key=lambda x: (x["ano"], x["tipo"],
                                          ordem.get(x["peca"], 9), x["ativo"])):
        ws3.append([i["ano"], i["ss"], i["ativo"], i["tipo"], i["faixa"],
                    i["potencia"] or "", i["controlador"], i["peca_rotulo"],
                    i["troca"], i["causa"], i["data"], i["mes"], i["localidade"],
                    i["parte_ativa"], i["rele_rl"], i["potencia_bruta"],
                    i["origem_tensao"]])
    for n, (_, larg) in enumerate(cols, start=1):
        c = ws3.cell(row=1, column=n)
        c.font, c.fill = Font(bold=True, color=PAPEL, size=10), \
            PatternFill("solid", fgColor=TINTA)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws3.column_dimensions[get_column_letter(n)].width = larg
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions
    ws3.row_dimensions[1].height = 30

    # 6 — o que não fecha
    ws4 = wb.create_sheet("Alertas")
    cabeca(ws4, 1, "O que não fecha, e que só o gestor decide",
           ["Alerta", "Ativo", "Onde", "O que foi encontrado"], [30, 13, 12, 92])
    for a in sorted(alertas):
        ws4.append(a)
    for r in range(3, ws4.max_row + 1):
        ws4.cell(row=r, column=4).alignment = Alignment(wrap_text=True,
                                                        vertical="top")

    # 7 — a régua
    ws5 = wb.create_sheet("Como foi feito")
    ws5.column_dimensions["A"].width = 98
    unicos = len({(i["ano"], i["tipo"], i["ativo"]) for i in itens})
    for t in TEXTO(len(itens), unicos):
        ws5.append([t])
    ws5["A1"].font = Font(bold=True, size=12, color=SINAL)
    for n, t in enumerate(TEXTO(len(itens), unicos), start=1):
        if t.endswith(":"):
            ws5.cell(row=n, column=1).font = Font(bold=True, size=11)

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    return SAIDA


def TEXTO(n, unicos):
    return [
        "A taxa de falha por peça — 2025 e 2026",
        "",
        "O ano vem da FATIA, não da coluna Ano:",
        "Na aba «Falha Equipamentos» as duas discordam numa das 90 linhas: o",
        "7933585074 está na fatia RL 2025 com data de ocorrência em 27/01/2026. Vale a",
        "fatia, que é como o gestor fecha a taxa — usar a coluna Ano daria RL 34 e 29",
        "em vez dos 35 e 28 da planilha base.",
        "",
        "De onde vem cada coluna:",
        "SS, ativo, tensão, peça, troca e causa raiz saem do rol de falhas da planilha",
        "base, aba «Falha Equipamentos».",
        "",
        "POTÊNCIA e CONTROLADOR do regulador saem da aba «Ajustes Reguladores de",
        "Tensão» de GESTAO_DE_EQUIPAMENTOS.xlsx — os ajustes da proteção, como o gestor",
        "mandou. Casaram 27 de 27 reguladores.",
        "",
        "O RELÉ e a TENSÃO do religador saem da aba «Ajustes RL Poste», do mesmo",
        "arquivo. Casaram 62 de 63.",
        "",
        "A taxa:",
        "falhas da peça ÷ parque do tipo. Parque oficial dos três anos: 1.307",
        "religadores e 207 reguladores. Não anualiza 2026, que vai até agosto — a taxa",
        "do ano parcial é menor por isso, não porque o parque melhorou.",
        "",
        "Potência fora das três classes:",
        "O gestor diz que só existe 167, 200 e 400. O cadastro dos ajustes traz também",
        "239, 250, 398 e um banco misto 167/250/167. O 398 é claramente da classe 400.",
        "O 239 e o 250 não são nenhuma das três — no padrão de regulador de 13,8 kV o",
        "250 kVA existe, então pode ser o cadastro certo e a régua incompleta.",
        "",
        "A planilha mostra as duas coisas: a coluna «Potência (RT)» traz a classe pelo",
        "mais próximo, e «Potência crua no cadastro» traz o que está lá. A aba Alertas",
        "lista os quatro casos, um a um.",
        "",
        "Ativo repetido no mesmo ano:",
        f"O rol tem {n} linhas para {unicos} pares ativo-ano. Três ativos aparecem duas",
        "vezes no mesmo ano:",
        "",
        "   7908708116 (RL 2026) — MESMA SS, duas peças: tanque e controle. Para taxa",
        "   por peça as duas contam; para taxa por equipamento, uma vez só.",
        "",
        "   5841308190 (RT 2025) — duas SS, dois furtos (agosto e setembro).",
        "   5854566043 (RT 2025) — duas SS, duas células (março e julho).",
        "",
        "Pela régua do gestor — ativo conta uma vez no ano — a taxa por EQUIPAMENTO",
        f"deveria contar {unicos}, não {n}. A taxa por PEÇA, que é o que esta planilha",
        "monta, conta a peça trocada, então as três linhas extras ficam.",
        "",
        "A taxa por classe:",
        "A aba «Taxa por classe» divide cada peça pelo parque da CLASSE dela — a célula",
        "de 200 kVA pelo parque de 200, não pelo parque inteiro de regulador. É a",
        "única forma de comparar: o parque tem 98 reguladores de 400, 67 de 200 e 25",
        "de 167, então a mesma quantidade de falhas significa coisas diferentes.",
        "",
        "O modelo do relé é onde está o sinal:"
        "A aba «Taxa por modelo de relé» compara falhas contra o parque de cada",
        "modelo. O COOPER F6 falha 2,2 vezes mais que a média; o NOJA RC10, que parece",
        "o vilão por concentrar 25 das 26 falhas de tanque, falha abaixo da média —",
        "ele é 78,6% do parque.",
        "",
        "Uma observação sobre o controlador:",
        "Dos 27 reguladores com falha, 26 têm controlador RUA e 1 tem TB–R1000. A",
        "coluna existe, mas quase não separa nada — o parque de regulador é RUA em",
        "praticamente tudo. Quem separa é a peça: célula, controle, relé ou furto.",
    ]


if __name__ == "__main__":
    itens, alertas = montar()
    print("gravado:", planilha(itens, alertas))
    print(f"  {len(itens)} falhas · "
          f"{len({(i['ano'], i['tipo'], i['ativo']) for i in itens})} pares ativo-ano")
    for tipo in ("RL", "RT"):
        for ano in ("2025", "2026"):
            sub = [i for i in itens if i["tipo"] == tipo and i["ano"] == ano]
            c = Counter(i["peca"] for i in sub)
            print(f"  {tipo} {ano}: {len(sub):2d} · " +
                  " · ".join(f"{k} {v}" for k, v in
                             sorted(c.items(), key=lambda x: -x[1])))
    print(f"  alertas: {len(alertas)}")
