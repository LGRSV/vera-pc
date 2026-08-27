"""
O SLA de manutenção dos COCM's — 2025 e 2026, mês a mês e por equipe.

Régua do gestor (27/08, fechada em duas rodadas):

  UNIVERSO   toda entrega ao COCM — cada vez que o posto despachou uma demanda para
             uma equipe de campo. O compromisso nasce na entrega, independente de a
             demanda ter virado «resolvida» depois.
  JANELA     do repasse AO COCM até o repasse DO COCM para outro posto. Quando o
             campo fecha sem repassar, a janela termina na conclusão da SS.
  PRAZO      pela criticidade da operação — Muito Alta 8, Alta 15, Média 30,
             Baixa 50 — e 26 dias para quem não tem criticidade definida.
  SÉRIE      mensal pelo mês da ENTREGA (coorte de entrada), com a devolução em
             coluna própria.

Nenhuma das duas datas vem de graça. A ENTREGA é a ABERTURA da SS no posto do COCM:
o campo DTA_REPASSE é cópia byte a byte da DTA_ABERTURA e não data o repasse. A
DEVOLUÇÃO é a abertura da SS seguinte — SS repassada sai da base sem conclusão, e
quem esperar a conclusão dela espera para sempre.

A varredura de cada demanda começa na SS DO COEP, não no início dela. Muita demanda
nasce no campo, e essa passagem inicial não é entrega para executar: o posto ainda
não tinha diagnosticado nem comprado nada.

Grava dist/SLA_MANUTENCAO.xlsx.
Rodar: python3 scripts/sla_por_equipe.py [base_de_repasse.xlsx]
"""

import datetime as dt
import json
import os
import sys
from collections import Counter, defaultdict

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
import sla_manutencao as base  # noqa: E402 — a leitura da base e a régua de prazo

SAIDA = os.path.join(RAIZ, "dist", "SLA_MANUTENCAO.xlsx")
ANOS = (2025, 2026)
MESES = ["jan", "fev", "mar", "abr", "mai", "jun",
         "jul", "ago", "set", "out", "nov", "dez"]

TITULO = Font(bold=True, color="FFFFFF", size=10)
FUNDO = PatternFill("solid", fgColor="1F3864")
BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)
VERDE = PatternFill("solid", fgColor="E2EFDA")
VERMELHO = PatternFill("solid", fgColor="FCE4E4")
AMARELO = PatternFill("solid", fgColor="FFF2CC")


def eh_cocm(posto):
    return "-RD-" in (posto or "").upper()


def repasse(ss, reg):
    """(SS gerada, data do repasse). A data é a ABERTURA da SS gerada — o campo
    DTA_REPASSE da base é cópia byte a byte da DTA_ABERTURA e não data nada."""
    r = reg.get(ss)
    if not r or not r["seguinte"]:
        return None, None
    prox = reg.get(r["seguinte"])
    return r["seguinte"], (prox["abertura"] if prox else None)


def entregas_ao_cocm(reg, crit, ativos):
    """Toda entrega ao COCM, pela régua do gestor (27/08, segunda correção).

    O relógio começa no repasse da SS ORIGINAL (a do posto) para o COCM, e para no
    repasse da SS do COCM para um posto FORA DO DCMD. Repasse de COCM para COCM não
    para o relógio: continua tudo dentro do DCMD, só trocou de equipe.

    Quando o campo conclui sem repassar, a parada é a conclusão da SS.
    """
    saida, vistas = [], set()
    for ss_coep, r in reg.items():
        if "COEP" not in r["posto"]:
            continue
        # a SS ORIGINAL é quem repassou ao campo. Quase sempre é a do próprio posto,
        # mas em parte das cadeias o COEP passa por um intermediário antes — e é esse
        # que gera a SS do COCM. Andar até achar preserva as cadeias inteiras.
        ss_original, ss_gerada, entrega = ss_coep, None, None
        atual = ss_coep
        for _ in range(8):
            prox, quando = repasse(atual, reg)
            if not prox or quando is None:
                break
            p_ = reg.get(prox)
            if p_ and eh_cocm(p_["posto"]):
                ss_original, ss_gerada, entrega = atual, prox, quando
                break
            atual = prox
        if not ss_gerada:
            continue
        g = reg.get(ss_gerada)
        if ss_gerada in vistas:
            continue
        vistas.add(ss_gerada)

        # Régua do gestor (27/08, terceira correção): o relógio vai até o ÚLTIMO
        # repasse para um posto fora do DCMD, ou até a atendida. A demanda pode voltar
        # ao campo depois de passar pela TELE ou pela PROT — parar na primeira saída
        # premiaria quem devolveu cedo e recebeu de volta.
        atual, equipes, internos = ss_gerada, [g["posto"]], []
        saidas, visto = [], set()
        while atual and atual in reg and atual not in visto:
            visto.add(atual)
            r_ = reg[atual]
            prox, quando = repasse(atual, reg)
            if eh_cocm(r_["posto"]):
                if prox and quando is not None:
                    p_ = reg.get(prox)
                    if p_ and eh_cocm(p_["posto"]):
                        internos.append((prox, quando, p_["posto"]))
                        equipes.append(p_["posto"])
                    else:
                        saidas.append((prox, quando,
                                       p_["posto"] if p_ else "(fora da base)"))
                elif r_["conclusao"] is not None:
                    saidas.append(("", r_["conclusao"], "atendida no próprio COCM"))
            if not prox:
                break
            atual = prox
        if saidas:
            ss_saida, data_saida, destino = saidas[-1]
        else:
            ss_saida, data_saida, destino = "", None, ""

        cod = (ativos.get(ss_original, {}) or ativos.get(ss_coep, {})).get("equipamento", "")
        c = crit.get(cod, "")
        prazo = base.PRAZO.get(c, base.PRAZO_SEM_CRITICIDADE)
        aberto = data_saida is None
        fim = data_saida or base.HOJE
        # dias de CALENDÁRIO, não períodos de 24h: a base guarda hora, e entregue dia
        # 6 às 14h com devolução dia 8 às 9h é 2 dias de SLA, não 1.
        dias = (fim.date() - entrega.date()).days
        # 18 SS da base fecham ANTES de abrir: o campo executou e o SGM abriu depois,
        # para regularizar. Conta 0 dia e fica marcado — não é giro relâmpago.
        retroativa = dias < 0
        if retroativa:
            dias = 0
        info = ativos.get(ss_original, {}) or ativos.get(ss_coep, {})
        saida.append({
            "ano": entrega.year, "mes": entrega.month,
            "ss_original": ss_original, "entrega": entrega, "ss_gerada": ss_gerada,
            "ss_saida": ss_saida, "devolucao": data_saida, "destino": destino,
            "equipe": equipes[0], "equipe_devolveu": equipes[-1],
            "internos": internos, "ss_coep": ss_coep, "ss_cocm": ss_gerada,
            "ativo": cod,
            "tipo": "RL" if cod[:2] in ("79", "78") else ("RT" if cod[:2] == "58" else ""),
            "localidade": info.get("localidade", ""),
            "criticidade": c or "Sem classificação", "prazo": prazo,
            "dias": dias, "atraso": max(0, dias - prazo),
            # O SLA é ÍNDICE, não sim/não: dias usados sobre o prazo da criticidade.
            "indice": dias / prazo,
            "dentro_do_prazo": dias <= prazo, "em_curso": aberto,
            "veredicto": ("em curso, dentro do prazo" if dias <= prazo
                          else "em curso, ESTOURADO") if aberto
                         else ("dentro do prazo" if dias <= prazo else "ESTOURADO"),
            "retroativa": retroativa,
            "apuracao": ("ainda no COCM" if aberto else
                         ("atendida no próprio COCM" if not ss_saida
                          else "último repasse para posto fora do DCMD"))
                        + (" · SS fechada antes de abrir (serviço regularizado depois)"
                           " — conta 0 dia" if retroativa else ""),
        })
    return sorted(saida, key=lambda x: (x["entrega"], x["equipe"]))


def dados_do_ativo(caminho):
    """{SS: {equipamento, localidade}} — para a entrega saber de quem ela é."""
    d = {}
    if caminho and caminho.lower().endswith(".xlsx"):
        ws = openpyxl.load_workbook(caminho, read_only=True,
                                    data_only=True)["Exportar Planilha"]
        for r in list(ws.iter_rows(values_only=True))[1:]:
            k = base.norm(r[1])
            if k:
                d.setdefault(k, {"equipamento": str(r[3] or "").strip(),
                                 "localidade": str(r[15] or "").strip(),
                                 "ocorrencia": base.data(r[6]),
                                 "tipo_ativo": str(r[5] or "").strip()})
    return d


def cabecalho(ws, colunas):
    ws.append([c[0] for c in colunas])
    linha = ws.max_row
    for i, (_, larg) in enumerate(colunas, 1):
        cel = ws.cell(row=linha, column=i)
        cel.font, cel.fill = TITULO, FUNDO
        cel.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        if ws.column_dimensions[get_column_letter(i)].width in (None, 13):
            ws.column_dimensions[get_column_letter(i)].width = larg
    ws.row_dimensions[linha].height = 32
    return linha


def bordar(ws, de, ate):
    for linha in ws.iter_rows(min_row=de, max_row=ate):
        for cel in linha:
            cel.border = BORDA


COLS = [("ID", 9), ("Ativo", 13), ("Tipo", 7), ("Localidade", 20),
        ("Ano", 7), ("Mês nº", 7), ("Mês", 8),
        # o trio da entrada, na ordem que o gestor pediu
        ("SS ORIGINAL (repassou ao campo)", 22), ("Data do repasse", 13),
        ("SS GERADA (no COCM)", 21), ("Equipe que recebeu", 15),
        # e a saída do DCMD — o último repasse para fora, ou a atendida
        ("Repasses entre COCMs", 11), ("Equipe que devolveu", 15),
        ("SS GERADA DEPOIS (fora do DCMD)", 22),
        ("Data do repasse ao outro posto", 13), ("Posto de destino", 16),
        ("Criticidade", 14), ("Prazo SLA (dias)", 10), ("Dias no DCMD", 10),
        ("Atraso (dias)", 10), ("SLA (dias ÷ prazo)", 11), ("Dentro do prazo", 10),
        ("Em curso", 9), ("SLA de manutenção", 22), ("Como a saída foi apurada", 30)]


def linha_de(e):
    return [e["id"], e["ativo"], e["tipo"], e["localidade"], e["ano"], e["mes"],
            MESES[e["mes"] - 1],
            e["ss_original"], e["entrega"].strftime("%d/%m/%Y"), e["ss_gerada"],
            e["equipe"], len(e["internos"]) or "", e["equipe_devolveu"],
            e["ss_saida"] or ("(atendida no COCM)" if not e["em_curso"] else ""),
            e["devolucao"].strftime("%d/%m/%Y") if e["devolucao"] else "",
            e["destino"], e["criticidade"], e["prazo"], e["dias"], e["atraso"],
            e["indice"], "sim" if e["dentro_do_prazo"] else "não",
            "sim" if e["em_curso"] else "", e["veredicto"], e["apuracao"]]


def aba_entregas(wb, entregas):
    ws = wb.active
    ws.title = f"1 · Entregas ao COCM ({len(entregas)})"
    prim = cabecalho(ws, COLS) + 1
    for e in entregas:
        ws.append(linha_de(e))
        ws.cell(row=ws.max_row, column=21).number_format = "0.00"
        cel = ws.cell(row=ws.max_row, column=24)
        cel.fill = AMARELO if e["em_curso"] else (VERDE if e["dentro_do_prazo"] else VERMELHO)
    bordar(ws, prim, ws.max_row)
    ws.freeze_panes = f"A{prim}"
    ws.auto_filter.ref = f"A{prim - 1}:Y{ws.max_row}"
    return ws


def quadro(ws, titulo, cabecalhos, linhas, pct_col=None):
    ws.append([titulo])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
    ws.append(cabecalhos)
    for cel in ws[ws.max_row]:
        if cel.value is not None:
            cel.font, cel.fill = TITULO, FUNDO
            cel.alignment = Alignment(horizontal="center", wrap_text=True)
    cab = ws.max_row
    prim = cab + 1
    for l in linhas:
        ws.append(l)
        if pct_col:
            ws.cell(row=ws.max_row, column=pct_col).number_format = "0.0%"
    bordar(ws, prim, ws.max_row)
    return cab, prim, ws.max_row


def indice(grupo):
    """O SLA do grupo: dias gastos sobre prazo concedido. Ponderado de propósito —
    média de índices dá o mesmo peso a um Muito Alta de 8 dias e a um Baixa de 50."""
    prazo = sum(e["prazo"] for e in grupo)
    return (sum(e["dias"] for e in grupo) / prazo) if prazo else 0


def resumo(grupo):
    """entregas · no prazo · estourou · em curso · cumprimento · mediana · SLA."""
    n = len(grupo)
    ok = sum(1 for e in grupo if e["dentro_do_prazo"])
    curso = sum(1 for e in grupo if e["em_curso"])
    dias = sorted(e["dias"] for e in grupo)
    return [n, ok, n - ok, curso, (ok / n) if n else 0,
            dias[len(dias) // 2] if dias else 0, indice(grupo)]


def aba_mensal(wb, entregas):
    ws = wb.create_sheet("4 · SLA mensal")
    ws.column_dimensions["A"].width = 22
    for c in "BCDEFGHIJKLM":
        ws.column_dimensions[c].width = 10
    ws.append(["SLA DE MANUTENÇÃO — MÊS A MÊS, PELO MÊS DA ENTREGA AO COCM"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.append(["A demanda entra no mês em que foi entregue. «Em curso» conta os dias "
               "até 18/08/2026 e já aparece estourada quando passou do prazo."])
    faixas = []
    for ano in ANOS:
        ws.append([])
        por_mes = defaultdict(list)
        for e in entregas:
            if e["ano"] == ano:
                por_mes[e["mes"]].append(e)
        linhas = []
        for nome, i in ((n, i + 1) for i, n in enumerate(MESES)):
            g = por_mes.get(i, [])
            r = resumo(g)
            linhas.append([nome] + r)
        total = resumo([e for e in entregas if e["ano"] == ano])
        linhas.append(["ANO"] + total)
        cab, prim, fim = quadro(
            ws, f"{ano}", ["Mês", "Entregas", "No prazo", "Estourou", "Em curso",
                           "Cumprimento", "Mediana de dias", "SLA (dias ÷ prazo)"], linhas, pct_col=6)
        for cel in ws[fim]:
            cel.font = Font(bold=True)
        faixas.append((ano, cab, prim, fim - 1))
    for ano, cab, prim, fim in faixas:
        ch = LineChart()
        ch.title = f"{ano} — entregas e cumprimento do SLA"
        ch.height, ch.width, ch.style = 7.5, 17, 2
        ch.add_data(Reference(ws, min_col=2, min_row=cab, max_row=fim),
                    titles_from_data=True)
        ch.add_data(Reference(ws, min_col=3, min_row=cab, max_row=fim),
                    titles_from_data=True)
        ref = f"'{ws.title}'!$A${prim}:$A${fim}"
        for s in ch.series:
            s.cat = AxDataSource(strRef=StrRef(f=ref))
            s.smooth = False
            s.marker.symbol, s.marker.size = "circle", 6
        ch.x_axis.delete = ch.y_axis.delete = False
        ch.x_axis.axPos, ch.y_axis.axPos = "b", "l"
        ws.add_chart(ch, f"J{cab - 1}")
    return ws


def braco(posto):
    """A que braço da esteira o posto pertence — a leitura do gestor."""
    p = (posto or "").upper()
    if "COEP" in p:
        return "COEP (posto)"
    if "-RD-" in p:
        return "COCM (campo)"
    if "TELE" in p or "-SE-" in p or "SCADA" in p or "DMSL" in p:
        return "DMSL (telecom)"
    if "PROT" in p or "DEOP" in p or "COI" in p:
        return "DEOP (proteção)"
    return p or "—"


def raizes(reg):
    """Quem aponta para quem — para achar o começo de cada cadeia."""
    anterior = {}
    for ss, r in reg.items():
        if r["seguinte"]:
            anterior.setdefault(r["seguinte"], ss)
    return anterior


def do_comeco(ss, anterior, limite=25):
    """Sobe a cadeia até a SS que abriu a ocorrência."""
    visto = set()
    while ss in anterior and anterior[ss] not in visto and len(visto) < limite:
        visto.add(ss)
        ss = anterior[ss]
    return ss


def aba_cadeias(wb, entregas, reg, ativos):
    """As 131 abertas salto a salto: todo repasse e o tempo em cada posto.

    Um salto por linha, da SS que abriu a ocorrência até a última da cadeia. O tempo
    de cada posto segue a mesma régua do SLA: a saída é a conclusão da SS, ou a
    abertura da seguinte quando ela foi repassada — SS repassada sai da base sem
    conclusão, e esperar por ela é esperar para sempre.
    """
    anterior = raizes(reg)
    ws = wb.create_sheet("2 · Cadeia salto a salto", 1)
    colunas = [("ID da ocorrência", 13), ("Ativo", 13), ("Tipo", 7), ("Localidade", 20),
               ("Data da ocorrência", 13), ("SS que abriu", 20), ("Salto nº", 8),
               ("SS", 20), ("Posto", 13), ("Braço", 16), ("Entrada no posto", 13),
               ("Saída do posto", 13), ("Dias no posto", 10),
               ("Como a saída foi apurada", 24), ("Repassou para", 13),
               ("SS seguinte", 20), ("Situação da SS", 14), ("Depois do COEP", 10),
               ("É a entrega do SLA", 11), ("Prazo SLA", 9), ("Índice SLA", 10)]
    prim = cabecalho(ws, colunas) + 1
    for n, e in enumerate(sorted(entregas, key=lambda x: (x["entrega"], x["ativo"])), 1):
        ident = f"OC-{n:03d}"
        inicio = do_comeco(e["ss_coep"], anterior)
        info = ativos.get(inicio, {}) or ativos.get(e["ss_coep"], {})
        oc = info.get("ocorrencia")
        cad = base.cadeia(inicio, reg)
        passou_coep = False
        for i, (ss, r) in enumerate(cad, 1):
            if r is None:
                ws.append([ident, e["ativo"], e["tipo"], e["localidade"],
                           oc.strftime("%d/%m/%Y") if oc else "", inicio, i, ss,
                           "", "(fora da base de repasse)", "", "", "", "", "", "",
                           "", "", "", "", ""])
                continue
            saida, apurada = r["conclusao"], "conclusão da SS"
            if saida is None and r["seguinte"]:
                prox = reg.get(r["seguinte"])
                if prox and prox["abertura"]:
                    saida, apurada = prox["abertura"], "abertura da SS seguinte"
            if saida is None:
                apurada = "ainda no posto"
            dias = ""
            if r["abertura"] and saida:
                dias = max(0, (saida.date() - r["abertura"].date()).days)
            elif r["abertura"]:
                dias = (base.HOJE.date() - r["abertura"].date()).days
            e_a_entrega = ss == e["ss_cocm"]
            ws.append([
                ident, e["ativo"], e["tipo"], e["localidade"],
                oc.strftime("%d/%m/%Y") if oc else "", inicio, i, ss, r["posto"],
                braco(r["posto"]),
                r["abertura"].strftime("%d/%m/%Y") if r["abertura"] else "",
                saida.strftime("%d/%m/%Y") if saida else "", dias, apurada,
                reg[r["seguinte"]]["posto"] if r["seguinte"] in reg else
                ("(fora da base)" if r["seguinte"] else ""),
                r["seguinte"], r["status"],
                "sim" if passou_coep else "", "SIM" if e_a_entrega else "",
                e["prazo"] if e_a_entrega else "",
                e["indice"] if e_a_entrega else ""])
            if e_a_entrega:
                ws.cell(row=ws.max_row, column=19).fill = (
                    VERMELHO if e["indice"] > 1 else VERDE)
                ws.cell(row=ws.max_row, column=21).number_format = "0.00"
            if "COEP" in r["posto"]:
                passou_coep = True
    bordar(ws, prim, ws.max_row)
    ws.freeze_panes = f"A{prim}"
    ws.auto_filter.ref = f"A{prim - 1}:U{ws.max_row}"
    return ws


def aba_matriz(wb, entregas):
    """O pedido do gestor: o índice de SLA mensalizado, uma coluna por equipe.

    Cada célula é dias gastos ÷ prazo concedido naquele mês, naquela equipe. Abaixo
    de 1,00 sobrou prazo; acima, estourou. Vazio quer dizer que a equipe não recebeu
    nada no mês — zero seria mentira, porque zero é desempenho perfeito.
    """
    ws = wb.create_sheet("3 · SLA mensal por equipe", 1)
    ws.column_dimensions["A"].width = 10
    ws.append(["SLA DE MANUTENÇÃO — ÍNDICE MENSAL POR EQUIPE"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.append(["Índice = dias gastos ÷ prazo da criticidade. 2 dias num prazo de 8 dá "
               "0,25. Abaixo de 1,00 sobrou prazo; acima de 1,00 estourou."])
    ws.append(["Célula vazia = a equipe não recebeu nada no mês. O total da linha e o da "
               "coluna são ponderados: soma dos dias sobre soma dos prazos."])

    for ano in ANOS:
        doano = [e for e in entregas if e["ano"] == ano]
        equipes = sorted({e["equipe"] for e in doano})
        for rotulo, valor, fmt in (
                (f"{ano} · índice de SLA (dias ÷ prazo)",
                 lambda g: indice(g) if g else None, "0.00"),
                (f"{ano} · entregas recebidas", lambda g: len(g) or None, "0")):
            ws.append([])
            ws.append([rotulo])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
            ws.append(["Mês"] + equipes + ["TOTAL"])
            for cel in ws[ws.max_row]:
                if cel.value is not None:
                    cel.font, cel.fill = TITULO, FUNDO
                    cel.alignment = Alignment(horizontal="center", wrap_text=True)
            for i, col in enumerate(equipes, 2):
                ws.column_dimensions[get_column_letter(i)].width = 13
            ws.column_dimensions[get_column_letter(len(equipes) + 2)].width = 11
            prim = ws.max_row + 1
            for m, nome in enumerate(MESES, 1):
                domes = [e for e in doano if e["mes"] == m]
                ws.append([nome] + [valor([e for e in domes if e["equipe"] == q])
                                    for q in equipes] + [valor(domes)])
            ws.append(["TOTAL"] + [valor([e for e in doano if e["equipe"] == q])
                                   for q in equipes] + [valor(doano)])
            for cel in ws[ws.max_row]:
                cel.font = Font(bold=True)
            fim = ws.max_row
            bordar(ws, prim, fim)
            for r in range(prim, fim + 1):
                for c in range(2, len(equipes) + 3):
                    cel = ws.cell(row=r, column=c)
                    cel.number_format = fmt
                    cel.alignment = Alignment(horizontal="center")
                    if fmt == "0.00" and isinstance(cel.value, float):
                        cel.fill = VERMELHO if cel.value > 1 else (
                            VERDE if cel.value <= 0.5 else AMARELO)
    return ws


def aba_equipe(wb, entregas):
    ws = wb.create_sheet("5 · SLA por equipe")
    ws.column_dimensions["A"].width = 20
    for c in "BCDEFGHI":
        ws.column_dimensions[c].width = 12
    ws.append(["SLA DE MANUTENÇÃO — POR EQUIPE DE CAMPO"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.append(["Equipe é o posto do COCM que recebeu a demanda. A mediana diz o ritmo "
               "normal; o pior atraso diz a cauda."])
    faixas = []
    for ano in list(ANOS) + ["2025 + 2026"]:
        ws.append([])
        sel = [e for e in entregas
               if (e["ano"] == ano if isinstance(ano, int) else True)]
        equipes = sorted({e["equipe"] for e in sel})
        linhas = []
        for q in equipes:
            g = [e for e in sel if e["equipe"] == q]
            if not g:
                continue
            r = resumo(g)
            linhas.append([q] + r + [max(e["atraso"] for e in g)])
        linhas.sort(key=lambda l: (-l[1], l[0]))
        linhas.append(["TOTAL"] + resumo(sel) + [max((e["atraso"] for e in sel), default=0)])
        cab, prim, fim = quadro(
            ws, f"{ano}", ["Equipe", "Entregas", "No prazo", "Estourou", "Em curso",
                           "Cumprimento", "Mediana de dias", "SLA (dias ÷ prazo)",
                           "Pior atraso"],
            linhas, pct_col=6)
        for cel in ws[fim]:
            cel.font = Font(bold=True)
        if isinstance(ano, str):
            faixas.append((cab, prim, fim - 1))
    for cab, prim, fim in faixas:
        ch = BarChart()
        ch.type, ch.gapWidth, ch.style = "bar", 40, 2
        ch.title = "Cumprimento do SLA por equipe · 2025 + 2026"
        ch.height, ch.width = 10, 15
        ch.add_data(Reference(ws, min_col=6, min_row=cab, max_row=fim),
                    titles_from_data=True)
        ch.series[0].cat = AxDataSource(strRef=StrRef(f"'{ws.title}'!$A${prim}:$A${fim}"))
        ch.x_axis.delete = ch.y_axis.delete = False
        ch.y_axis.numFmt = "0%"
        ch.legend = None
        ws.add_chart(ch, f"K{cab - 1}")
    return ws


def aba_criticidade(wb, entregas):
    ws = wb.create_sheet("6 · SLA por criticidade")
    ws.column_dimensions["A"].width = 20
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 12
    ws.append(["SLA DE MANUTENÇÃO — POR CRITICIDADE DA OPERAÇÃO"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ordem = ["Muito Alta", "Alta", "Média", "Baixa", "Sem classificação"]
    for ano in list(ANOS) + ["2025 + 2026"]:
        ws.append([])
        sel = [e for e in entregas if (e["ano"] == ano if isinstance(ano, int) else True)]
        linhas = []
        for c in ordem:
            g = [e for e in sel if e["criticidade"] == c]
            if g:
                linhas.append([c, base.PRAZO.get(c, base.PRAZO_SEM_CRITICIDADE)] + resumo(g))
        linhas.append(["TOTAL", ""] + resumo(sel))
        cab, prim, fim = quadro(
            ws, f"{ano}", ["Criticidade", "Prazo", "Entregas", "No prazo", "Estourou",
                           "Em curso", "Cumprimento", "Mediana de dias",
                           "SLA (dias ÷ prazo)"],
            linhas, pct_col=7)
        for cel in ws[fim]:
            cel.font = Font(bold=True)
    return ws


def aba_resolvidos(wb, entregas, caminho_base):
    """A aba de resolvidos de 2026, com as colunas de SLA para a dinâmica."""
    reg = base.base_de_repasse(caminho_base)
    crit = base.criticidades()
    with open(base.COEP, encoding="utf-8") as fh:
        cp = json.load(fh)
    por_ss = {e["ss_coep"]: e for e in entregas}
    res = [r for r in cp["resolvidos_do_coep"] if r["conta_como_resolvido_pelo_coep"]]
    ws = wb.create_sheet(f"7 · Resolvidos 2026 ({len(res)})")
    colunas = [("Ativo", 13), ("Tipo", 7), ("Localidade", 22), ("SS no COEP", 20),
               ("Como terminou", 14), ("Fechou em", 12), ("Posto que fechou", 15),
               ("Passou por COCM", 11), ("Ano da entrega", 8), ("Mês nº", 7),
               ("Mês", 8), ("Equipe (COCM)", 14), ("Criticidade", 14),
               ("Prazo SLA (dias)", 10), ("Entrega ao COCM", 13), ("Devolução", 12),
               ("Dias de manutenção", 11), ("Atraso (dias)", 10),
               ("SLA (dias ÷ prazo)", 11), ("Dentro do prazo", 10),
               ("SLA de manutenção", 22)]
    prim = cabecalho(ws, colunas) + 1
    for r in sorted(res, key=lambda x: x["ativo"]):
        k = base.norm(r["ss_no_coep"])
        e = por_ss.get(k)
        c = crit.get(r["ativo"], "") or "Sem classificação"
        if e:
            ws.append([r["ativo"], "RL" if r["tipo"] == "religador" else "RT",
                       r["localidade"], k, r["como_terminou"], r["data_do_fechamento"],
                       r["posto_que_fechou"], "sim", e["ano"], e["mes"],
                       MESES[e["mes"] - 1], e["equipe"], e["criticidade"], e["prazo"],
                       e["entrega"].strftime("%d/%m/%Y"),
                       e["devolucao"].strftime("%d/%m/%Y") if e["devolucao"] else "",
                       e["dias"], e["atraso"], e["indice"],
                       "sim" if e["dentro_do_prazo"] else "não", e["veredicto"]])
            ws.cell(row=ws.max_row, column=19).number_format = "0.00"
            cel = ws.cell(row=ws.max_row, column=21)
            cel.fill = AMARELO if e["em_curso"] else (VERDE if e["dentro_do_prazo"] else VERMELHO)
        else:
            ws.append([r["ativo"], "RL" if r["tipo"] == "religador" else "RT",
                       r["localidade"], k, r["como_terminou"], r["data_do_fechamento"],
                       r["posto_que_fechou"], "não", "", "", "", "", c,
                       base.PRAZO.get(c, base.PRAZO_SEM_CRITICIDADE), "", "", "", "",
                       "", "", "não passou por COCM depois do posto"])
    bordar(ws, prim, ws.max_row)
    ws.freeze_panes = f"A{prim}"
    ws.auto_filter.ref = f"A{prim - 1}:U{ws.max_row}"
    return ws, sum(1 for r in res if base.norm(r["ss_no_coep"]) in por_ss)


def como_foi_feito(entregas, com_cocm):
    por_ano = Counter(e["ano"] for e in entregas)
    return [
        "SLA DE MANUTENÇÃO — 2025 e 2026, mês a mês e por equipe.",
        "",
        "O UNIVERSO (escolha do gestor, 27/08): toda entrega ao COCM — cada vez que o "
        "posto despachou uma demanda para uma equipe de campo. O compromisso nasce na "
        f"entrega, independente de a demanda ter virado «resolvida» depois. São "
        f"{por_ano.get(2025, 0)} entregas em 2025 e {por_ano.get(2026, 0)} em 2026.",
        "",
        "A JANELA (régua do gestor): do repasse AO COCM até o repasse DO COCM para outro "
        "posto. Quando o campo fecha sem repassar, a janela termina na conclusão da SS.",
        "",
        "AS DUAS DATAS, e por que nenhuma vem de graça: a ENTREGA é a ABERTURA da SS no "
        "posto do COCM — o campo DTA_REPASSE da base é cópia byte a byte da DTA_ABERTURA "
        "e não data o repasse. A DEVOLUÇÃO é a abertura da SS seguinte, porque SS "
        "repassada sai da base sem data de conclusão: quem esperar a conclusão dela "
        "espera para sempre.",
        "",
        "O PRAZO, pela criticidade da operação (aba de mapeamento por criticidade da "
        "Relação de Indisponíveis): Muito Alta 8 dias, Alta 15, Média 30, Baixa 50. Sem "
        "criticidade definida, 26 dias — o prazo médio.",
        "",
        "O SLA É ÍNDICE, não sim/não (régua do gestor, 27/08): dias gastos ÷ prazo da "
        "criticidade. Dois dias num prazo de oito dá 0,25 — sobrou três quartos do "
        "prazo. Acima de 1,00 estourou. O índice de um grupo é PONDERADO (soma dos dias "
        "sobre soma dos prazos), não média de índices: média daria o mesmo peso a um "
        "Muito Alta de 8 dias e a um Baixa de 50.",
        "",
        "A SÉRIE MENSAL é pelo mês da ENTREGA (coorte de entrada): a demanda entra no mês "
        "em que o compromisso foi assumido. A data de devolução está em coluna própria, "
        "para quem quiser a leitura pelo outro eixo.",
        "",
        "EM CURSO ENTRAM NA CONTA, com os dias corridos até 18/08/2026, e já aparecem "
        "estouradas quando passaram do prazo. Tirá-las faria o mês recente parecer bom "
        "só porque o que atrasou ainda não fechou — viés de sobrevivência.",
        "",
        "POR QUE A VARREDURA COMEÇA NA SS DO COEP: muita demanda nasce no campo — a "
        "equipe abre a SS e repassa para o posto —, e essa passagem inicial não é entrega "
        "para executar: o posto ainda não tinha diagnosticado nem comprado nada. Contar a "
        "cadeia inteira inflaria o número com serviço que ninguém mandou fazer.",
        "",
        "EQUIPE é o posto do COCM que recebeu (ETO-RD-*, ENC-RD-*, DOLP-RD-*, DG-RD-*, "
        "ESO-RD-*). Cada entrega conta uma vez, pela primeira equipe que recebeu depois "
        "do posto.",
        "",
        "SS FECHADA ANTES DE ABRIR: 18 SS da base inteira têm data de conclusão anterior "
        "à de abertura — o campo executou e o SGM abriu a SS depois, para regularizar. "
        "Contam 0 dia e ficam marcadas na coluna «Como a devolução foi apurada», para "
        "ninguém ler como giro relâmpago.",
        "",
        "DIAS DE CALENDÁRIO, não períodos de 24 horas: a base guarda hora, e entregue "
        "dia 6 às 14h com devolução dia 8 às 9h é 2 dias de SLA, não 1.",
        "",
        "PREMISSA REGISTRADA: a criticidade é a de hoje, aplicada também a 2025 — não "
        "existe histórico de criticidade na base. Quem foi reclassificado desde então "
        "carrega o prazo de agora.",
        "",
        f"A ABA 5 traz as 82 resolvidas de 2026 com as colunas de SLA anexadas "
        f"({com_cocm} delas passaram por COCM depois do posto); as demais ficam marcadas "
        "«não passou por COCM», porque fecharam na TELE ou na PROT — execução de outro "
        "braço, sem SLA de manutenção a cobrar do campo.",
        "",
        "A ABA 2 abre as mesmas entregas SALTO A SALTO: da SS que abriu a ocorrência até "
        "a última da cadeia, um posto por linha, com o tempo parado em cada um. O ID da "
        "ocorrência (OC-001, OC-002…) amarra as linhas de uma mesma cadeia, e o número "
        "do ativo vai em toda linha. A coluna «É a entrega do SLA» marca o salto que o "
        "SLA de manutenção cobra — o primeiro COCM depois do COEP.",
        "",
        "TODAS AS ABAS DE LISTA TÊM FILTRO E COLUNAS CRUAS (ano, mês nº, mês, equipe, "
        "criticidade, prazo, dias, atraso, dentro do prazo) — é só selecionar e inserir "
        "tabela dinâmica para conferir qualquer corte.",
        "",
        "Fonte das datas: base de repasse (Eqp_joao / EQP_SS_OCORRENCIA), a única que "
        "traz a cadeia SS a SS. Posição de 18/08/2026.",
    ]


def montar(caminho=None):
    reg = base.base_de_repasse(caminho)
    crit = base.criticidades()
    ativos = dados_do_ativo(caminho)
    todas = entregas_ao_cocm(reg, crit, ativos)
    entregas = [e for e in todas if e["ano"] in ANOS]
    for n, e in enumerate(entregas, 1):
        e["id"] = f"OC-{n:03d}"

    wb = openpyxl.Workbook()
    aba_entregas(wb, entregas)
    aba_cadeias(wb, entregas, reg, ativos)
    aba_matriz(wb, entregas)
    aba_mensal(wb, entregas)
    aba_equipe(wb, entregas)
    aba_criticidade(wb, entregas)
    _, com_cocm = aba_resolvidos(wb, entregas, caminho)
    ws = wb.create_sheet("Como foi feito")
    ws.column_dimensions["A"].width = 112
    for t in como_foi_feito(entregas, com_cocm):
        ws.append([t])
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)

    # as abas são criadas fora de ordem (cada insert empurra a anterior); reordena
    # pelo número do próprio nome, para a numeração bater com a posição
    wb._sheets.sort(key=lambda w: int(w.title.split("\u00b7")[0].strip())
                    if w.title[0].isdigit() else 99)

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    return entregas, todas


if __name__ == "__main__":
    caminho = sys.argv[1] if len(sys.argv) > 1 else None
    entregas, todas = montar(caminho)
    print(f"gravado: {SAIDA}")
    print(f"  entregas na base inteira: {len(todas)} · em 2025-2026: {len(entregas)}")
    for ano in ANOS:
        g = [e for e in entregas if e["ano"] == ano]
        ok = sum(1 for e in g if e["dentro_do_prazo"])
        curso = sum(1 for e in g if e["em_curso"])
        dias = sorted(e["dias"] for e in g)
        print(f"  {ano}: {len(g)} entregas · {ok} no prazo ({100*ok/len(g):.1f}%) · "
              f"{curso} em curso · mediana {dias[len(dias)//2]}d")
