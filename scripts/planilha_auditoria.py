"""
A planilha da auditoria de 29/08 — dist/AUDITORIA_29082026.xlsx.

Cinco agentes leram a descrição de todas as SS da cadeia dos 143, com a régua da PEÇA
do gestor por cima do rótulo TIPOSS. Sai de data/missao/auditoria_29082026.json.

Rodar: python3 scripts/planilha_auditoria.py
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FONTE = os.path.join(RAIZ, "data", "missao", "auditoria_29082026.json")
SAIDA = os.path.join(RAIZ, "dist", "AUDITORIA_29082026.xlsx")

TINTA, PAPEL, SINAL = "FF211D15", "FFF2EFE6", "FFBC4B0E"
CAB = Font(bold=True, color=PAPEL, size=10)
FILL = PatternFill("solid", fgColor=TINTA)


def aba(wb, nome, linhas, larguras):
    ws = wb.create_sheet(nome)
    for r in linhas:
        ws.append(r)
    for c in range(1, len(linhas[0]) + 1):
        cel = ws.cell(row=1, column=c)
        cel.font, cel.fill = CAB, FILL
        cel.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=len(larguras)).alignment = Alignment(wrap_text=True,
                                                                  vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30
    return ws


def montar():
    with open(FONTE, encoding="utf-8") as fh:
        d = json.load(fh)
    wb = Workbook()
    wb.remove(wb.active)

    aba(wb, "Voltam para a conta",
        [["Ativo", "Praça", "TIPOSS que excluiu", "Peça que o texto pede",
          "A prova, no texto da SS"]]
        + [[x["ativo"], x["praca"], x["tiposs"], x["peca"], x["prova"]]
           for x in d["voltam_para_a_conta"]],
        [13, 26, 32, 24, 78])

    aba(wb, "Saem — não é peça",
        [["Ativo", "Praça", "Balde onde estava", "O que o texto pede de verdade",
          "A prova, no texto da SS"]]
        + [[x["ativo"], x["praca"], x["balde"], x["peca_real"], x["prova"]]
           for x in d["saem_por_nao_ser_peca"]],
        [13, 26, 18, 32, 78])

    v = d["voltam_para_a_fila"]
    aba(wb, "Voltam para a fila",
        [["Ativo", "De onde volta", "Motivo"]]
        + [[a, "encerrados", v["motivo"]] for a in v["dos_encerrados"]]
        + [[a, "despachados", v["motivo"]] for a in v["dos_despachados"]],
        [13, 16, 78])

    aba(wb, "Pendências de decisão",
        [["Ativo", "Praça", "Grupo", "O que precisa ser decidido"]]
        + [[x["ativo"], x["praca"], "muda de ano (" + str(x["ano"]) + ")", x["prova"]]
           for x in d["mudam_de_ano"]]
        + [[x["ativo"], x["praca"], "em suspenso", x["questao"]]
           for x in d["em_suspenso"]]
        + [[x["ativo"], x["praca"], "precisa do gestor", x["questao"]]
           for x in d["precisa_do_gestor"]]
        + [[x["ativo"], x["praca"], "sem lastro na base", x["questao"]]
           for x in d["sem_lastro"]],
        [13, 26, 22, 92])

    c = d["conta_pela_peca"]
    ws = wb.create_sheet("A conta pela peça")
    for r in [
        ["A conta pela PEÇA, não pelo rótulo"], [""],
        ["Demanda encerrada em 2026 com peça grande trocada", c["encerrada_em_2026_com_peca_grande"]],
        ["Despachados com peça trocada e obra/OS nomeada", c["despachados_com_peca_trocada"]],
        ["TRABALHO DO COEP COM PEÇA EFETIVAMENTE TROCADA", c["trabalho_do_coep_com_peca"]],
        [""],
        ["Contra o que a régua do TIPOSS dizia:"],
        ["   trabalho do COEP concluído", c["contra_o_que_eu_dizia"]["trabalho_do_coep_concluido"]],
        ["   troca confirmada", c["contra_o_que_eu_dizia"]["troca_confirmada"]],
        [""], ["O achado central"], [""], [d["achado_central"]], [""],
        ["Erro no insumo, corrigido"], [""], [d["erro_meu_no_insumo"]], [""],
        ["Defeito na régua da SS mais pesada"], [""],
        [d["defeito_da_minha_regua"]["o_que"]],
        ["Atinge " + str(d["defeito_da_minha_regua"]["tamanho"]) + " ativos: "
         + ", ".join(d["defeito_da_minha_regua"]["ativos"])],
    ]:
        ws.append(r)
    ws.column_dimensions["A"].width = 104
    ws.column_dimensions["B"].width = 10
    ws["A1"].font = Font(bold=True, size=12, color=SINAL)
    for r in (3, 4, 5, 7, 11, 15, 19):
        ws.cell(row=r, column=1).font = Font(bold=True)
    ws.cell(row=5, column=1).font = Font(bold=True, size=11, color=SINAL)
    ws.cell(row=5, column=2).font = Font(bold=True, size=11, color=SINAL)
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    return SAIDA


if __name__ == "__main__":
    print("gravado:", montar())
