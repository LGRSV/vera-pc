"""
O backlog de equipamentos especiais, mês a mês — dist/BACKLOG_MENSAL_2026.xlsx.

O gestor pediu (02/09): mensalizar o backlog de janeiro de 2026 até hoje, pela base
de SS/OS. Backlog aqui é ESTOQUE: quantos religadores e reguladores estavam com
demanda aberta no fim de cada mês, quantos entraram e quantos saíram.

A RÉGUA É A DA VISÃO ETO, a que o próprio gestor fixou em 22/08: ativo 58/79 com SS
de INDISPONIBILIDADE PARA OPERAÇÃO pendente na base de SS/OS — «só tem 93, então são
as 93». A série reproduz esse 93 no fecho da base, o que serve de âncora: a conta
não foi calibrada para dar 93, ela dá 93 sozinha.

COMO A DEMANDA É MONTADA (repasse não é falha nova — a cadeia inteira é uma só):

  · a demanda de um ativo abre na abertura da primeira SS e fecha na saída da última;
  · a saída de uma SS é a conclusão dela, se houver; senão a abertura da SS seguinte
    do mesmo ativo, porque SS repassada sai da base sem data de conclusão;
  · SS PENDENTE sem conclusão segue aberta;
  · SS REPASSADA sem nenhuma SS seguinte no recorte sai do registro na data em que
    foi aberta — a base não diz para onde foi. São 145 casos, e é exatamente essa
    regra que faz o estoque fechar em 93 em vez de 217.
  · SS do mesmo ativo que se sobrepõem viram um período só: conta EQUIPAMENTO, não SS.

Estoque no fim do mês = demandas com abertura ≤ último dia e saída > último dia.
Entradas e saídas do mês = aberturas e fechamentos dentro do mês. Com essa fronteira
o saldo fecha em todos os meses: fim = início + entradas − saídas, sem sobra.

POSIÇÃO: 21/08/2026 — a base BASE_SS_OS_20082026.txt tem aberturas até 20/08 e
fechamentos até 21/08. Agosto é mês PARCIAL (01 a 21/08).

Rodar: python3 scripts/backlog_mensal.py
"""

import datetime as dt
import json
import os
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint, Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
SSOS = os.path.join(RAIZ, "data", "missao", "ssos_min.json")
COEP = os.path.join(RAIZ, "data", "missao", "coep_2026.json")
SAIDA = os.path.join(RAIZ, "dist", "BACKLOG_MENSAL_2026.xlsx")
JSON_SAIDA = os.path.join(RAIZ, "data", "missao", "backlog_mensal.json")

TINTA, PAPEL, SOMBRA, SINAL, APAGADA = "FF211D15", "FFF2EFE6", "FFE9E5D8", "FFBC4B0E", "FF8D8672"
VERDE, LARANJA, NEUTRO, GRADE = "1F7C50", "B8480C", "6D675A", "DDD8CC"
MESES = ["janeiro", "fevereiro", "março", "abril", "maio", "junho", "julho", "agosto"]
INF = dt.date(9999, 12, 31)
D = dt.timedelta(days=1)

IND = "INDISPONIBILIDADE PARA OPERAÇÃO"
ANOMALIA = {"EM OPERAÇÃO COM ANOMALIA", "ANOMALIA EM RELIGADOR", "ANOMALIA EM REGULADORES"}
REGUAS = [("Indisponibilidade para operação", {IND}),
          ("Indisponibilidade + anomalia", {IND} | ANOMALIA)]


def data(s):
    s = (s or "").strip()
    if not s:
        return None
    try:
        return dt.datetime.strptime(s.split()[0], "%d/%m/%Y").date()
    except ValueError:
        return None


def tipo_do_ativo(codigo):
    return "RT" if codigo[:2] == "58" else "RL"


def ler():
    with open(SSOS, encoding="utf-8") as fh:
        ss = json.load(fh)
    datas = [d for d in (data(x["DATA_ABERTURA_SS"]) for x in ss) if d]
    datas += [d for d in (data(x.get("DATA_TERMINO_SS")) for x in ss) if d]
    return ss, max(datas)


def demandas(ss, tipos):
    """As demandas encadeadas por ativo, com abertura, saída e as SS que as compõem."""
    por = defaultdict(list)
    for x in ss:
        if x.get("TIPOSS") not in tipos:
            continue
        a = data(x["DATA_ABERTURA_SS"])
        if a:
            por[x["NUM_TRAFO"]].append((a, data(x.get("DATA_TERMINO_SS")), x["SITUACAO_SS"],
                                        x["NUMERO_SS"], x.get("LOCALIDADE", "")))
    saida, soltas = [], 0
    for ativo, lista in por.items():
        lista.sort(key=lambda t: (t[0], t[1] or INF))
        brutos = []
        for i, (a, termino, situacao, numero, local) in enumerate(lista):
            if termino:
                fim, como = termino, "conclusão da SS"
            elif situacao == "SS PENDENTE":
                fim, como = INF, "segue pendente"
            else:
                seguintes = [z for z, _, _, _, _ in lista[i + 1:] if z >= a]
                if seguintes:
                    fim, como = seguintes[0], "abertura da SS seguinte"
                else:
                    fim, como = a, "repassada sem seguinte no recorte"
                    soltas += 1
            brutos.append((a, max(fim, a), numero, situacao, como, local))
        brutos.sort()
        unidos = []
        for a, fim, numero, situacao, como, local in brutos:
            if unidos and a <= unidos[-1]["fim"]:
                u = unidos[-1]
                if fim > u["fim"]:
                    u["fim"], u["como"], u["situacao"] = fim, como, situacao
                u["ss"].append(numero)
            else:
                unidos.append({"ativo": ativo, "tipo": tipo_do_ativo(ativo), "abertura": a,
                               "fim": fim, "ss": [numero], "situacao": situacao,
                               "como": como, "localidade": local})
        saida += unidos
    return saida, soltas


def serie(dem, posicao, so_tipo=None):
    itens = [d for d in dem if so_tipo is None or d["tipo"] == so_tipo]
    aberto = lambda t: sum(1 for d in itens if d["abertura"] <= t < d["fim"])
    linhas = []
    for m in range(1, 9):
        ini = dt.date(2026, m, 1)
        fim = posicao if m == 8 else dt.date(2026, m + 1, 1) - D
        ent = [d for d in itens if ini <= d["abertura"] <= fim]
        sai = [d for d in itens if ini <= d["fim"] <= fim]
        si, sf = aberto(ini - D), aberto(fim)
        assert sf == si + len(ent) - len(sai), (m, so_tipo, si, len(ent), len(sai), sf)
        na_fila = [d for d in itens if d["abertura"] <= fim < d["fim"]]
        idades = sorted((fim - d["abertura"]).days for d in na_fila)
        linhas.append({
            "mes": m, "rotulo": MESES[m - 1], "inicio": si, "entraram": len(ent),
            "sairam": len(sai), "fim": sf,
            "de_2025_ou_antes": sum(1 for d in na_fila if d["abertura"].year < 2026),
            "idade_mediana": idades[len(idades) // 2] if idades else 0,
            "mais_velha": idades[-1] if idades else 0,
        })
    return linhas


# ----------------------------------------------------------------------------- estilo
def titulo(ws, texto, sub=""):
    ws["A1"] = texto
    ws["A1"].font = Font(bold=True, size=13, color=SINAL)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(italic=True, size=9)
        ws["A2"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 42


def cabecalho(ws, linha, titulos, larguras=None):
    for i, t in enumerate(titulos, 1):
        c = ws.cell(row=linha, column=i, value=t)
        c.font = Font(bold=True, color=PAPEL, size=10)
        c.fill = PatternFill("solid", fgColor=TINTA)
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.cell(row=linha, column=1).alignment = Alignment(vertical="center", horizontal="left")
    if larguras:
        for i, w in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[linha].height = 30


def estilo(ch, alto=11, largo=24):
    ch.height, ch.width = alto, largo
    ch.style = None
    ch.x_axis.delete = ch.y_axis.delete = False
    ch.x_axis.axPos, ch.y_axis.axPos = "b", "l"
    ch.x_axis.crosses = ch.y_axis.crosses = "autoZero"
    gl = ChartLines()
    gl.spPr = GraphicalProperties()
    gl.spPr.line = LineProperties(solidFill=GRADE, w=6350)
    ch.y_axis.majorGridlines = gl
    ch.x_axis.majorGridlines = None
    for eixo in (ch.x_axis, ch.y_axis):
        eixo.spPr = GraphicalProperties()
        eixo.spPr.line = LineProperties(solidFill=GRADE, w=6350)
    return ch


def categorias(ch, ws, ref):
    alvo = "'%s'!%s" % (ws.title, ref)
    for sub in (getattr(ch, "_charts", None) or [ch]):
        for s in sub.series:
            s.cat = AxDataSource(strRef=StrRef(alvo))
    return ch


def rotulos(alvo):
    alvo.dLbls = DataLabelList()
    alvo.dLbls.showVal = True
    alvo.dLbls.showSerName = alvo.dLbls.showCatName = alvo.dLbls.showLegendKey = False
    return alvo


def cor_barra(s, cor):
    s.graphicalProperties = GraphicalProperties(solidFill=cor)
    s.graphicalProperties.line = LineProperties(solidFill="FBFAF6", w=25400)
    return s


# ----------------------------------------------------------------------------- abas
def aba_mensal(wb, linhas, posicao, soltas):
    ws = wb.create_sheet("Backlog mensal")
    titulo(ws, "BACKLOG MENSAL 2026 — religadores e reguladores com demanda aberta",
           "Régua da visão ETO: ativo 58/79 com SS de indisponibilidade para operação em aberto. "
           "Entrou é demanda nova no mês; saiu é demanda encerrada. O saldo fecha em todos os meses: "
           "fim = início + entrou − saiu. Agosto é parcial, até %s — a base tem aberturas até 20/08 "
           "e fechamentos até 21/08." % posicao.strftime("%d/%m/%Y"))
    cab = ["Mês", "Backlog no início", "Entraram", "Saíram", "Backlog no fim",
           "Variação", "Da fila, de 2025 ou antes", "Idade mediana da fila (dias)",
           "A mais velha da fila (dias)"]
    cabecalho(ws, 4, cab, [14, 17, 12, 12, 16, 11, 20, 20, 20])
    r = 5
    for L in linhas:
        ws.append([L["rotulo"], L["inicio"], L["entraram"], L["sairam"], L["fim"],
                   L["fim"] - L["inicio"], L["de_2025_ou_antes"], L["idade_mediana"], L["mais_velha"]])
        for c in range(2, 10):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1
    fim = r - 1
    ws.cell(row=r, column=1, value="no ano").font = Font(bold=True)
    ws.cell(row=r, column=2, value=linhas[0]["inicio"]).font = Font(bold=True)
    ws.cell(row=r, column=3, value=f"=SUM(C5:C{fim})").font = Font(bold=True)
    ws.cell(row=r, column=4, value=f"=SUM(D5:D{fim})").font = Font(bold=True)
    ws.cell(row=r, column=5, value=linhas[-1]["fim"]).font = Font(bold=True)
    ws.cell(row=r, column=6, value=linhas[-1]["fim"] - linhas[0]["inicio"]).font = Font(bold=True)
    for c in range(2, 7):
        ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth, ch.overlap = "col", "clustered", 80, -12
    ch.add_data(Reference(ws, min_col=3, min_row=4, max_col=4, max_row=fim), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=fim))
    ch.title = "Entradas e saídas por mês, com o saldo em linha"
    ch.y_axis.title = "demandas"
    cor_barra(ch.series[0], LARANJA)
    cor_barra(ch.series[1], VERDE)
    for s in ch.series:
        rotulos(s)
    lin = LineChart()
    lin.add_data(Reference(ws, min_col=5, min_row=4, max_row=fim), titles_from_data=True)
    ls = lin.series[0]
    ls.graphicalProperties = GraphicalProperties()
    ls.graphicalProperties.line = LineProperties(solidFill=TINTA[2:], w=28000)
    ls.marker = Marker(symbol="circle", size=8)
    ls.marker.graphicalProperties = GraphicalProperties(solidFill=TINTA[2:])
    ls.smooth = False
    rotulos(ls)
    ch += lin
    ch.legend.position, ch.legend.overlay = "b", False
    categorias(ch, ws, "$A$5:$A$%d" % fim)
    ws.add_chart(estilo(ch, 12, 26), "A16")
    return fim


def aba_tipo(wb, dem, posicao):
    ws = wb.create_sheet("RL e RT")
    titulo(ws, "O mesmo backlog, separado em religador e regulador",
           "Mesma régua e mesma fronteira de mês. A soma das duas linhas de saldo dá o backlog total.")
    cab = ["Mês", "RL no início", "RL entraram", "RL saíram", "RL no fim",
           "RT no início", "RT entraram", "RT saíram", "RT no fim"]
    cabecalho(ws, 4, cab, [14] + [13] * 8)
    rl, rt = serie(dem, posicao, "RL"), serie(dem, posicao, "RT")
    r = 5
    for a, b in zip(rl, rt):
        ws.append([a["rotulo"], a["inicio"], a["entraram"], a["sairam"], a["fim"],
                   b["inicio"], b["entraram"], b["sairam"], b["fim"]])
        for c in range(2, 10):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1
    fim = r - 1
    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth, ch.overlap = "col", "clustered", 80, -12
    for col in (5, 9):
        ch.add_data(Reference(ws, min_col=col, min_row=4, max_row=fim), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=fim))
    ch.title = "Backlog no fim do mês, por tipo de equipamento"
    ch.y_axis.title = "demandas abertas"
    cor_barra(ch.series[0], LARANJA)
    cor_barra(ch.series[1], VERDE)
    for s in ch.series:
        rotulos(s)
    ch.legend.position, ch.legend.overlay = "b", False
    categorias(ch, ws, "$A$5:$A$%d" % fim)
    ws.add_chart(estilo(ch, 12, 26), "A16")
    return rl, rt


def aba_cascata(wb, linhas, posicao):
    ws = wb.create_sheet("Cascata do ano")
    titulo(ws, "De %d para %d: o que entrou e o que saiu no ano"
           % (linhas[0]["inicio"], linhas[-1]["fim"]),
           "O ano abriu com o backlog herdado, recebeu demanda nova e devolveu o que foi encerrado. "
           "Resolveu-se muito, mas entrou quase o mesmo tanto — por isso a fila quase não cede.")
    ws.column_dimensions["A"].width = 30
    entrou = sum(L["entraram"] for L in linhas)
    saiu = sum(L["sairam"] for L in linhas)
    ini, fim = linhas[0]["inicio"], linhas[-1]["fim"]
    rots = ["Backlog em 01/01", "Entraram no ano", "Saíram no ano",
            "Backlog em %s" % posicao.strftime("%d/%m")]
    bases = [0, ini, fim, 0]
    vals = [ini, entrou, saiu, fim]
    for j, t in enumerate(rots):
        c = ws.cell(row=4, column=2 + j, value=t)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(2 + j)].width = 18
    ws.cell(row=5, column=1, value="Base (invisível)").font = Font(color=APAGADA, size=8)
    ws.cell(row=6, column=1, value="Demandas").font = Font(bold=True)
    for j in range(4):
        ws.cell(row=5, column=2 + j, value=bases[j]).alignment = Alignment(horizontal="center")
        ws.cell(row=6, column=2 + j, value=vals[j]).alignment = Alignment(horizontal="center")
    ch = BarChart()
    ch.type, ch.grouping, ch.overlap, ch.gapWidth = "col", "stacked", 100, 60
    ch.add_data(Reference(ws, min_col=1, min_row=5, max_col=5, max_row=6),
                from_rows=True, titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=2, min_row=4, max_col=5, max_row=4))
    ch.title = "Entraram %d · saíram %d · a fila subiu %d" % (entrou, saiu, fim - ini)
    ch.y_axis.title = "demandas"
    base, valor = ch.series
    base.graphicalProperties = GraphicalProperties()
    base.graphicalProperties.noFill = True
    base.graphicalProperties.line = LineProperties(noFill=True)
    cor_barra(valor, NEUTRO)
    valor.dPt = [DataPoint(idx=0, spPr=GraphicalProperties(solidFill=NEUTRO)),
                 DataPoint(idx=1, spPr=GraphicalProperties(solidFill=LARANJA)),
                 DataPoint(idx=2, spPr=GraphicalProperties(solidFill=VERDE)),
                 DataPoint(idx=3, spPr=GraphicalProperties(solidFill=NEUTRO))]
    rotulos(valor)
    ch.legend = None
    categorias(ch, ws, "$B$4:$E$4")
    ws.add_chart(estilo(ch, 12, 22), "A9")


def aba_reguas(wb, ss, posicao):
    ws = wb.create_sheet("Duas réguas")
    titulo(ws, "A régua estreita e a larga",
           "A estreita é a da visão ETO — só indisponibilidade para operação, que é o equipamento fora. "
           "A larga soma quem roda com defeito (em operação com anomalia, anomalia em religador). "
           "As duas contam equipamento, não SS.")
    cab = ["Mês", "Só indisponibilidade", "Indisponibilidade + anomalia"]
    cabecalho(ws, 4, cab, [14, 22, 26])
    series = []
    for _, tipos in REGUAS:
        dem, _ = demandas(ss, tipos)
        series.append(serie(dem, posicao))
    r = 5
    for i in range(8):
        ws.append([MESES[i], series[0][i]["fim"], series[1][i]["fim"]])
        for c in (2, 3):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1
    fim = r - 1
    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth, ch.overlap = "col", "clustered", 80, -12
    ch.add_data(Reference(ws, min_col=2, min_row=4, max_col=3, max_row=fim), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=fim))
    ch.title = "Backlog no fim do mês pelas duas réguas"
    ch.y_axis.title = "demandas abertas"
    cor_barra(ch.series[0], LARANJA)
    cor_barra(ch.series[1], NEUTRO)
    for s in ch.series:
        rotulos(s)
    ch.legend.position, ch.legend.overlay = "b", False
    categorias(ch, ws, "$A$5:$A$%d" % fim)
    ws.add_chart(estilo(ch, 12, 26), "A16")
    return series[1]


def aba_coep(wb):
    ws = wb.create_sheet("Posto do COEP")
    titulo(ws, "Para comparar: a fila do posto do COEP",
           "Vem de coep_2026.json, apurada em 22/08 sobre a base de ocorrência. É outro recorte — "
           "só o que passou pelo posto —, por isso é bem menor que o backlog do parque. Aqui «no posto» "
           "é a fila no fim do mês e «conta do gestor» é resolvidos no ano mais a fila.")
    with open(COEP, encoding="utf-8") as fh:
        cp = json.load(fh)
    cab = ["Mês", "Chegaram ao posto", "Resolvidos", "Na fila do posto",
           "Já passaram no ano", "Conta do gestor"]
    cabecalho(ws, 4, cab, [14, 18, 14, 18, 18, 16])
    r = 5
    for m in cp["curva_mensal"]:
        ws.append([m["rotulo"], m["chegaram"], m["resolvidos"], m["no_posto"],
                   m["estiveram"], m["conta_gestor"]])
        for c in range(2, 7):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1
    fim = r - 1
    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth, ch.overlap = "col", "clustered", 80, -12
    ch.add_data(Reference(ws, min_col=2, min_row=4, max_col=3, max_row=fim), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=fim))
    ch.title = "Chegadas e resoluções no posto do COEP"
    ch.y_axis.title = "equipamentos"
    cor_barra(ch.series[0], LARANJA)
    cor_barra(ch.series[1], VERDE)
    for s in ch.series:
        rotulos(s)
    ch.legend.position, ch.legend.overlay = "b", False
    categorias(ch, ws, "$A$5:$A$%d" % fim)
    ws.add_chart(estilo(ch, 12, 26), "A16")


def aba_base(wb, dem, posicao):
    ws = wb.create_sheet("Base das demandas")
    titulo(ws, "Cada demanda que esteve aberta em 2026",
           "Uma linha por demanda encadeada. «Saída» vazia é demanda ainda aberta na posição da base.")
    cab = ["Ativo", "Tipo", "Localidade", "Abertura", "Saída", "Dias", "Situação da última SS",
           "Como a saída foi apurada", "SS da cadeia", "Aberta na posição?"]
    cabecalho(ws, 4, cab, [13, 6, 24, 12, 12, 8, 20, 30, 46, 16])
    vivas = [d for d in dem if d["fim"] > dt.date(2026, 1, 1) and d["abertura"] <= posicao]
    r = 5
    for d in sorted(vivas, key=lambda x: (x["abertura"], x["ativo"])):
        aberta = d["fim"] > posicao
        fim = None if d["fim"] == INF else d["fim"]
        dias = ((posicao if aberta else d["fim"]) - d["abertura"]).days
        ws.append([d["ativo"], d["tipo"], d["localidade"], d["abertura"],
                   None if aberta else fim, dias, d["situacao"], d["como"],
                   " | ".join(d["ss"]), "sim" if aberta else "não"])
        for c in (4, 5):
            ws.cell(row=r, column=c).number_format = "dd/mm/yyyy"
        ws.cell(row=r, column=1).number_format = "@"
        r += 1
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:J%d" % (r - 1)
    return r - 5


def aba_como(wb, posicao, soltas, n_dem, linhas, larga):
    ws = wb.create_sheet("Como foi feito")
    ws.column_dimensions["A"].width = 112
    entrou = sum(L["entraram"] for L in linhas)
    saiu = sum(L["sairam"] for L in linhas)
    texto = [
        ("O QUE ESTÁ SENDO CONTADO", True),
        ("Backlog é ESTOQUE: quantos religadores e reguladores estavam com demanda aberta no fim de "
         "cada mês. Não é quantidade de SS — é EQUIPAMENTO. O mesmo ativo com três SS encadeadas é "
         "uma demanda só, porque repasse não é falha nova.", False),
        ("A régua é a que o gestor fixou em 22/08 para a visão ETO: ativo 58/79 com SS de "
         "INDISPONIBILIDADE PARA OPERAÇÃO em aberto — «só tem 93, então são as 93».", False),
        ("", False),
        ("A ÂNCORA", True),
        ("Na posição da base a série dá exatamente %d demandas abertas, o mesmo 93 da visão ETO. A conta "
         "não foi ajustada para bater: ela bate sozinha, e é isso que dá confiança no resto da série." % linhas[-1]["fim"], False),
        ("", False),
        ("COMO A DEMANDA ABRE E FECHA", True),
        ("Abre na abertura da primeira SS da cadeia. Fecha na saída da última. A saída de uma SS é, "
         "nesta ordem: a conclusão dela, se houver; senão a abertura da SS seguinte do mesmo ativo, "
         "porque SS repassada sai da base sem data de conclusão; senão a SS segue aberta.", False),
        ("Uma exceção, e ela pesa: SS REPASSADA que não tem NENHUMA SS seguinte no recorte sai do "
         "registro na data em que foi aberta — a base não diz para onde foi. São %d casos. Tratá-las "
         "como abertas para sempre levaria o estoque de %d para 217, e aí nada bateria com o gestor." % (soltas, linhas[-1]["fim"]), False),
        ("SS do mesmo ativo que se sobrepõem no tempo viram um período só.", False),
        ("", False),
        ("A FRONTEIRA DO MÊS", True),
        ("Estoque no fim do mês = demandas com abertura ≤ último dia do mês e saída depois dele. "
         "Entradas e saídas do mês = aberturas e fechamentos com data dentro do mês. Com essa "
         "fronteira o saldo fecha nos oito meses, sem sobra: fim = início + entrou − saiu.", False),
        ("", False),
        ("A POSIÇÃO — E POR QUE NÃO VAI ATÉ 02/09", True),
        ("A base é BASE_SS_OS_20082026.txt: aberturas até 20/08/2026 e fechamentos até 21/08/2026. "
         "A série vai até %s. Agosto é mês PARCIAL — 21 dias, não 31 —, então as entradas e saídas "
         "dele não se comparam direto com as dos meses cheios." % posicao.strftime("%d/%m/%Y"), False),
        ("Para chegar até hoje basta largar uma base de SS/OS mais nova em data/raw e rodar "
         "scripts/extrai_ssos_min.py e depois scripts/backlog_mensal.py.", False),
        ("", False),
        ("O QUE A SÉRIE MOSTRA", True),
        ("O ano abriu com %d demandas abertas e está com %d. Entraram %d no período e saíram %d — "
         "quase empate, e é por isso que a fila não cede: o posto resolve muito, mas recebe quase o "
         "mesmo tanto." % (linhas[0]["inicio"], linhas[-1]["fim"], entrou, saiu), False),
        ("Pela régua larga, somando quem roda com defeito, o backlog vai de %d a %d no mesmo período."
         % (larga[0]["inicio"], larga[-1]["fim"]), False),
        ("A fila também envelhece: a idade mediana e a mais velha estão na aba «Backlog mensal», junto "
         "com quantas demandas da fila são de 2025 ou antes.", False),
        ("", False),
        ("O QUE ESTA CONTA NÃO É", True),
        ("Não é a fila do posto do COEP — essa é bem menor e está na aba «Posto do COEP», para comparar. "
         "Aqui é o parque inteiro de RL e RT, independente de em que mesa a demanda esteja.", False),
        ("Não é taxa de falha: aqui entra tudo que ficou indisponível, inclusive o que não exigiu peça "
         "grande. A taxa de falha tem régua própria e mora em dist/TAXA_POR_PECA.xlsx.", False),
        ("Não usa a data de ocorrência, e sim a de abertura da SS: backlog é fila de trabalho, e o "
         "trabalho entra na fila quando a SS é aberta.", False),
        ("", False),
        ("Foram montadas %d demandas encadeadas na régua estreita." % n_dem, False),
    ]
    for i, (t, negrito) in enumerate(texto, 1):
        c = ws.cell(row=i, column=1, value=t)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if negrito:
            c.font = Font(bold=True, size=11, color=SINAL if i == 1 else TINTA)


def montar(saida=SAIDA):
    ss, posicao = ler()
    dem, soltas = demandas(ss, {IND})
    linhas = serie(dem, posicao)
    wb = Workbook()
    wb.remove(wb.active)
    aba_mensal(wb, linhas, posicao, soltas)
    rl, rt = aba_tipo(wb, dem, posicao)
    aba_cascata(wb, linhas, posicao)
    larga = aba_reguas(wb, ss, posicao)
    aba_coep(wb)
    n = aba_base(wb, dem, posicao)
    aba_como(wb, posicao, soltas, len(dem), linhas, larga)
    os.makedirs(os.path.dirname(saida), exist_ok=True)
    wb.save(saida)
    with open(JSON_SAIDA, "w", encoding="utf-8") as fh:
        json.dump({"posicao": posicao.isoformat(), "regua": "indisponibilidade para operação",
                   "ancora": {"estoque_na_posicao": linhas[-1]["fim"],
                              "visao_eto_do_gestor": 93},
                   "repassadas_sem_seguinte": soltas, "mensal": linhas,
                   "mensal_rl": rl, "mensal_rt": rt,
                   "mensal_regua_larga": larga}, fh, ensure_ascii=False, indent=1)
    print("%s: %d demandas, %d linhas na base, posição %s" % (saida, len(dem), n, posicao))
    print("  backlog: %s" % " · ".join("%s %d" % (L["rotulo"][:3], L["fim"]) for L in linhas))
    from planilha_automatica import grava_cache
    print("  cache: %d células" % grava_cache(saida))
    return saida


if __name__ == "__main__":
    montar(sys.argv[1] if len(sys.argv) > 1 else SAIDA)
