"""
Confere dist/GESTAO_AUTOMATICA.xlsx calculando as fórmulas com a biblioteca `formulas`
(LibreOffice não roda aqui). O que se checa:

  - Gestão: o total em fórmula bate com o total da carteira nos 53 (coluna «Bate?»);
  - Lançamento: os quatro exemplos dão código, descrição, preço e total esperados, e
    seis cenários de borda escritos numa cópia (ativo fora do cadastro, prefixo 59,
    tensão manual como número, regulador completo de 200 kVA, quantidade em texto,
    linha vazia) dão as mensagens e os valores certos;
  - nenhuma descrição traz «· 0» nem «cód. 0»;
  - Falha Equipamentos: furtos alinhados com a Gestão, 7930359149 com tensão;
  - Resumo: 35 · 28 · 18 · 9 falhas por fatia;
  - estrutura: listas de validação, nomes definidos, tabelas, completos em fórmula,
    valores em cache no XML.

Rodar: python3 scripts/confere_planilha_automatica.py [arquivo.xlsx]
"""

import os
import re
import shutil
import sys
import zipfile

import formulas
from openpyxl import load_workbook

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ARQ = sys.argv[1] if len(sys.argv) > 1 else os.path.join(RAIZ, "dist", "GESTAO_AUTOMATICA.xlsx")
SCRATCH = os.environ.get("SCRATCH", os.path.join(RAIZ, "dist", "_confere_cenarios.xlsx"))


class Solucao:
    def __init__(self, caminho):
        self.nome = os.path.basename(caminho)
        self.sol = formulas.ExcelModel().loads(caminho).finish().calculate()
        self.idx = {}
        for k, v in self.sol.items():
            m = re.match(r"^'\[(.+?)\](.+?)'!([A-Z]+\d+)$", k, re.I)
            if m and m.group(1).lower() == self.nome.lower():
                x = v.value
                try:
                    x = x[0][0]
                except (TypeError, IndexError):
                    pass
                self.idx[(m.group(2).upper(), m.group(3).upper())] = x

    def __call__(self, aba, ref):
        return self.idx[(aba.upper(), ref.upper())]


def num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def main():
    erros, ok = [], []

    def checa(cond, msg):
        (ok if cond else erros).append(msg)

    wb = load_workbook(ARQ)
    # ---- estrutura
    nomes = set(wb.defined_names.keys())
    for n in ("Pecas_RL", "Pecas_RT", "Pecas_Todas", "Tensoes", "Potencias", "Criticidades",
              "Status_DCMD", "Sim_Nao", "Causas"):
        checa(n in nomes, f"nome definido {n}")
    for aba, n_dv in (("Lançamento", 4), ("Gestão", 6), ("Falha Equipamentos", 6)):
        q = len(wb[aba].data_validations.dataValidation)
        checa(q >= n_dv, f"{aba}: {q} validações (esperado ≥ {n_dv})")
        for dv in wb[aba].data_validations.dataValidation:
            checa(not str(dv.formula1).startswith("="), f"{aba}: validação {dv.sqref} sem «=» inicial")
            checa(dv.showErrorMessage, f"{aba}: validação {dv.sqref} com mensagem de erro ativa")
    for aba, tab in (("Lançamento", "Lancamento"), ("Gestão", "Gestao"), ("Falha Equipamentos", "Falhas")):
        checa(tab in wb[aba].tables, f"tabela {tab} em {aba}")
    checa(any("Pecas_" in str(dv.formula1) for dv in wb["Lançamento"].data_validations.dataValidation),
          "lista de peça dependente do tipo (INDIRECT) no Lançamento")
    cat = wb["Catálogo"]
    completos = [r for r in range(5, cat.max_row + 1) if cat[f"C{r}"].value == "Completo"]
    checa(completos and all(str(cat[f"H{r}"].value).startswith("=") for r in completos),
          f"Catálogo: {len(completos)} linhas «Completo» com material em fórmula")
    for c in ("N", "O"):
        checa(wb["Lançamento"][f"{c}5"].number_format == "@", f"Lançamento {c}5 em formato texto")
    # cache de valores no XML
    with zipfile.ZipFile(ARQ) as z:
        xmls = "".join(z.read(n).decode("utf-8") for n in z.namelist() if n.startswith("xl/worksheets/sheet"))
    # <v></v> vazio é cache legítimo (a fórmula deu ""); o que não pode sobrar é o <v /> do openpyxl
    n_vazios = len(re.findall(r"<f>[^<]*</f><v />", xmls))
    n_cheios = len(re.findall(r"<f>[^<]*</f><v>[^<]*</v>", xmls))
    checa(n_vazios == 0 and n_cheios > 6000, f"cache de valores: {n_cheios} fórmulas com valor, {n_vazios} sem")

    # ---- cálculo do arquivo entregue
    print("calculando as fórmulas do arquivo…", flush=True)
    v = Solucao(ARQ)

    # Gestão: total novo × total carteira
    ws = wb["Gestão"]
    bate, nao = 0, []
    for r in range(5, 5 + 53):
        p, x = num(v("Gestão", f"O{r}")), num(ws[f"X{r}"].value)
        if p is not None and x is not None and abs(p - x) < 0.01:
            bate += 1
        else:
            nao.append((ws[f"A{r}"].value, ws[f"U{r}"].value, p, x))
    checa(bate == 53, f"Gestão: {bate} de 53 totais batem com a carteira; divergem: {nao}")
    checa(str(v("Gestão", "Y3")).startswith("53 de 53"), f"Gestão Y3 = {v('Gestão', 'Y3')!r}")
    checa(abs(num(v("Gestão", "O3")) - num(v("Gestão", "X3"))) < 0.01,
          f"Gestão totais O3 = {v('Gestão', 'O3')} × X3 = {v('Gestão', 'X3')}")
    checa(ws["Q54"].value in (None, ""), "Gestão Q54 (7930359149) sem tensão manual redundante")

    # Lançamento: exemplos
    esperado = {
        5: dict(G="690005", I=38151.48, J=38151.48, K=13209.34, L=51360.82, D="RL", E="34,5", M="sim",
                H="Tanque / Parte ativa — Religador 34,5 kV · cód. 690005 · ativo NOJA RC10"),
        6: dict(G="690241", I=90230.0, J=180460.0, K=80318.5, L=260778.5, D="RT", E="34,5", F="400",
                H="Célula — Regulador 400 kVA (cód. 690241, 34,5 kV) · cód. 690241 · ativo ITB / RUA"),
        7: dict(G="690001 + 690916", I=55602.54, J=55602.54, K=11016.94, L=66619.48, D="RL", E="13,8"),
        8: dict(G="651638", I=23259.6, J=23259.6, K=20000.0, L=43259.6, D="RT", E="13,8"),
    }
    for r, campos in esperado.items():
        for c, e in campos.items():
            x = v("Lançamento", f"{c}{r}")
            certo = (abs(num(x) - e) < 0.01) if isinstance(e, float) else (str(x) == e)
            checa(certo, f"Lançamento {c}{r}: {x!r} (esperado {e!r})")
    for c in ("D", "E", "H", "J", "L", "V"):
        x = v("Lançamento", f"{c}20")
        checa(x in ("", None), f"Lançamento {c}20 vazia = {x!r}")

    # nenhum «· 0» / «cód. 0»
    ruins = []
    for aba, col, r0, r1 in (("Gestão", "K", 5, 57), ("Falha Equipamentos", "P", 5, 94), ("Lançamento", "H", 5, 8)):
        for r in range(r0, r1 + 1):
            s = str(v(aba, f"{col}{r}"))
            if "· 0" in s or "cód. 0" in s or s.endswith(" 0"):
                ruins.append((aba, r, s))
    checa(not ruins, f"descrições sem «0» de célula vazia: {ruins}")

    # Falha Equipamentos
    ws = wb["Falha Equipamentos"]
    linha = {}
    for r in range(5, 95):
        linha.setdefault(ws[f"C{r}"].value, []).append(r)
    r = linha["7930359149"][0]
    checa(v("Falha Equipamentos", f"M{r}") == "34,5", f"7930359149 (Caseara) tensão = {v('Falha Equipamentos', f'M{r}')!r}")
    r = linha["5836786094"][0]
    checa(abs(num(v("Falha Equipamentos", f"T{r}")) - 170548.5) < 0.01,
          f"5836786094 furto = célula 400 ×1 como na carteira: {v('Falha Equipamentos', f'T{r}')}")
    r = linha["5858783119"][0]
    checa(abs(num(v("Falha Equipamentos", f"T{r}")) - 494586.6) < 0.01,
          f"5858783119 furto = regulador completo 400 como na carteira: {v('Falha Equipamentos', f'T{r}')}")
    for r in linha["5841308190"]:
        checa(v("Falha Equipamentos", f"T{r}") in ("", None) and "definir" in str(ws[f"Z{r}"].value),
              f"5841308190 linha {r}: sem custo e marcada para definir")
    r = linha["7933585074"][0]
    checa(num(v("Falha Equipamentos", f"G{r}")) == 2025, f"7933585074 ano da fatia = {v('Falha Equipamentos', f'G{r}')}")
    sem = [r for r in range(5, 95) if v("Falha Equipamentos", f"M{r}") not in ("13,8", "34,5")]
    checa(not sem, f"Falha Equipamentos: linhas sem tensão: {sem}")

    # Resumo: totais por fatia (linha «Falhas (linhas)»)
    wr = wb["Resumo"]
    r_tot = [r for r in range(8, 40) if wr[f"A{r}"].value == "Falhas (linhas)"][0]
    for c, e in (("B", 35), ("C", 28), ("D", 18), ("E", 9)):
        x = num(v("Resumo", f"{c}{r_tot}"))
        checa(x == e, f"Resumo {c}{r_tot} = {x} (esperado {e})")

    # ---- cenários de borda numa cópia
    shutil.copy(ARQ, SCRATCH)
    wc = load_workbook(SCRATCH)
    wl = wc["Lançamento"]
    cen = {
        9: ("7999999999", "Tanque", None, None, None),        # fora do cadastro
        10: ("5912345678", "Célula", 2, None, None),           # capacitor
        11: ("7925744087", "Tanque", 1, 13.8, None),           # tensão manual como NÚMERO
        12: ("5800440256", "Completo", 1, None, None),         # RT completo 200 kVA 34,5
        13: ("7925744087", "Tanque", "2 un", None, None),      # quantidade em texto
        14: ("5800440256", "Célula", 2, None, "400"),          # potência manual manda
    }
    for r, (a, p, q, tm, km) in cen.items():
        wl[f"A{r}"], wl[f"B{r}"], wl[f"C{r}"], wl[f"N{r}"], wl[f"O{r}"] = a, p, q, tm, km
    wc.save(SCRATCH)
    print("calculando os cenários de borda…", flush=True)
    s = Solucao(SCRATCH)
    checa(str(s("Lançamento", "H9")).startswith("ATIVO NÃO ESTÁ NO CADASTRO"), f"H9 = {s('Lançamento', 'H9')!r}")
    checa(str(s("Lançamento", "H10")).startswith("PREFIXO FORA DO ESCOPO"), f"H10 = {s('Lançamento', 'H10')!r}")
    checa(s("Lançamento", "E11") == "13,8" and abs(num(s("Lançamento", "K11")) - 11016.94) < 0.01
          and abs(num(s("Lançamento", "L11")) - 32802.66) < 0.01,
          f"tensão manual numérica: E11={s('Lançamento', 'E11')!r} K11={s('Lançamento', 'K11')} L11={s('Lançamento', 'L11')}")
    checa(abs(num(s("Lançamento", "J12")) - (3 * 51705.75 + 23259.6)) < 0.01
          and abs(num(s("Lançamento", "K12")) - 200637.0) < 0.01,
          f"RT completo 200 kVA 34,5: J12={s('Lançamento', 'J12')} K12={s('Lançamento', 'K12')}")
    checa(abs(num(s("Lançamento", "J13")) - 38151.48) < 0.01, f"quantidade em texto vale 1: J13={s('Lançamento', 'J13')}")
    checa(s("Lançamento", "F14") == "400" and abs(num(s("Lançamento", "J14")) - 180460.0) < 0.01,
          f"potência manual: F14={s('Lançamento', 'F14')!r} J14={s('Lançamento', 'J14')}")
    if os.path.exists(SCRATCH):
        os.remove(SCRATCH)

    print(f"\n{len(ok)} conferências ok · {len(erros)} erro(s)")
    for e in erros:
        print("  ERRO:", e)
    return 1 if erros else 0


if __name__ == "__main__":
    sys.exit(main())
