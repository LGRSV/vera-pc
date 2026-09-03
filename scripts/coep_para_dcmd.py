"""
Quanto tempo a demanda leva depois que o COEP entrega para o DCMD.

O gestor entrega o serviço para as equipes de campo — as que têm RD no nome
(ETO-RD-AR, ETO-RD-PS, DOLP-RD-PA e as demais). A pergunta é o que acontece
depois da entrega.

Como o relógio é lido, e é aqui que a maioria erra: a base não registra quando
a SS saiu do posto. SS repassada tem a data de conclusão vazia. Quem quiser o
tempo de uma equipe precisa montar a cadeia:

    entrega ao DCMD   =  abertura da SS no posto RD
    saída do DCMD     =  conclusão dessa SS, se houver
                         senão, abertura da SS seguinte
                         senão, ainda está com a equipe
    fim da demanda    =  conclusão da última SS da cadeia

Daí saem três tempos diferentes, e eles não são a mesma coisa:

    dias no DCMD       quanto a equipe segurou antes de passar adiante ou fechar
    dias até o fim     do repasse até a demanda morrer, em qualquer posto
    voltou ao COEP     a cadeia trouxe o serviço de volta para a mesa do gestor

Grava data/missao/coep_para_dcmd.json e dist/COEP_PARA_DCMD.xlsx.

Rodar: python3 scripts/coep_para_dcmd.py
"""

import datetime
import json
import os
import re
from collections import Counter, defaultdict

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ARQ_SS = os.path.join(RAIZ, "data", "missao", "ss_ocorrencia.json")
SAIDA_JSON = os.path.join(RAIZ, "data", "missao", "coep_para_dcmd.json")
SAIDA_XLSX = os.path.join(RAIZ, "dist", "COEP_PARA_DCMD.xlsx")

POSTO = "ETO-COEP"
HOJE = datetime.datetime(2026, 8, 18, 23, 59)
RE_SS = re.compile(r"([A-Z][A-Z-]*)\s+0*(\d+)/(\d{4})")
RE_DCMD = re.compile(r"-RD-")          # a régua do gestor: equipe de campo tem RD no nome


def norm(numero):
    m = RE_SS.match((numero or "").strip().upper())
    return f"{m.group(1)} {int(m.group(2))}/{m.group(3)}" if m else (numero or "").strip().upper()


def dia(texto):
    try:
        return datetime.datetime.strptime((texto or "")[:19], "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def percentis(valores):
    v = sorted(valores)
    if not v:
        return {}
    def p(q):
        return v[min(len(v) - 1, int(len(v) * q))]
    return {"n": len(v), "mediana": p(.5), "p75": p(.75), "p90": p(.9), "maximo": v[-1],
            "media": round(sum(v) / len(v), 1)}


def montar():
    with open(ARQ_SS, encoding="utf-8") as fh:
        base = json.load(fh)
    idx = {}
    for x in base:
        x["_id"] = norm(x["SS_ORIGINAL"])
        x["_abriu"] = dia(x.get("DTA_ABERTURA"))
        x["_concluiu"] = dia(x.get("DTA_CONCLUSAO"))
        antes = idx.get(x["_id"])
        if antes is None or (x["_abriu"] and antes["_abriu"] and x["_abriu"] > antes["_abriu"]):
            idx[x["_id"]] = x
    seguinte = {x["_id"]: norm(x["SS_APOS_REPASSE"]) for x in idx.values()
                if x.get("SS_APOS_REPASSE")}

    def saida(x):
        if x["_concluiu"]:
            return x["_concluiu"], "conclusão da SS"
        prox = seguinte.get(x["_id"])
        if prox and prox in idx and idx[prox]["_abriu"]:
            return idx[prox]["_abriu"], "repasse — abertura da SS seguinte"
        return None, "ainda com a equipe"

    def cadeia(x):
        cur, caminho, visto = x, [x], set()
        while cur["_id"] in seguinte and seguinte[cur["_id"]] in idx \
                and seguinte[cur["_id"]] not in visto:
            visto.add(cur["_id"])
            cur = idx[seguinte[cur["_id"]]]
            caminho.append(cur)
        return caminho

    casos = []
    for a in idx.values():
        if a["POSTO_SGM"] != POSTO:
            continue
        prox = seguinte.get(a["_id"])
        b = idx.get(prox) if prox else None
        if not b or not RE_DCMD.search(b["POSTO_SGM"]) or not b["_abriu"]:
            continue
        quando, como = saida(b)
        resto = cadeia(b)
        depois = resto[1] if len(resto) > 1 else None
        fim = resto[-1]
        fechou = (fim["_concluiu"] if fim["STATUS"] in ("SS ATENDIDA", "SS CANCELADA")
                  and fim["_concluiu"] else None)
        casos.append({
            "ativo": a["EQUIPAMENTO"], "tipo": a["TIPO_ATIVO"].lower(),
            "equipe": b["POSTO_SGM"],
            "ss_do_coep": a["SS_ORIGINAL"], "ss_da_equipe": b["SS_ORIGINAL"],
            "entregue_em": b["_abriu"].strftime("%d/%m/%Y"),
            "ano_da_entrega": b["_abriu"].year,
            "saiu_da_equipe": quando.strftime("%d/%m/%Y") if quando else "",
            "como_apurou_a_saida": como,
            "dias_no_dcmd": ((quando or HOJE) - b["_abriu"]).days,
            "ainda_com_a_equipe": quando is None,
            "fim_da_demanda": fechou.strftime("%d/%m/%Y") if fechou else "",
            "dias_ate_o_fim": (fechou - b["_abriu"]).days if fechou else None,
            "repassou_para": depois["POSTO_SGM"] if depois else "",
            "ss_do_destino": depois["SS_ORIGINAL"] if depois else "",
            "postos_depois_da_equipe": len(resto) - 1,
            "resolveu_ali_mesmo": len(resto) == 1,
            "voltou_ao_coep": any(y["POSTO_SGM"] == POSTO for y in resto[1:]),
            "status_da_ss_da_equipe": b["STATUS"], "status_final": fim["STATUS"],
            "posto_que_fechou": fim["POSTO_SGM"],
        })

    por_equipe = defaultdict(list)
    por_ano = defaultdict(list)
    for c in casos:
        por_equipe[c["equipe"]].append(c)
        por_ano[str(c["ano_da_entrega"])].append(c)

    def bloco(grupo):
        return {
            "entregas": len(grupo),
            "no_dcmd": percentis([c["dias_no_dcmd"] for c in grupo]),
            "ate_o_fim": percentis([c["dias_ate_o_fim"] for c in grupo
                                    if c["dias_ate_o_fim"] is not None]),
            "resolveu_ali_mesmo": sum(1 for c in grupo if c["resolveu_ali_mesmo"]),
            "voltou_ao_coep": sum(1 for c in grupo if c["voltou_ao_coep"]),
            "ainda_com_a_equipe": sum(1 for c in grupo if c["ainda_com_a_equipe"]),
        }

    # o recorte que o gestor pediu: 2026, e o tempo até a equipe passar adiante
    de26 = [c for c in casos if c["ano_da_entrega"] == 2026]
    repassaram = [c for c in de26 if c["repassou_para"]]
    destinos = defaultdict(list)
    for c in repassaram:
        destinos[c["repassou_para"]].append(c["dias_no_dcmd"])
    faixas = [("no mesmo dia", 0, 0), ("1 a 7 dias", 1, 7), ("8 a 30 dias", 8, 30),
              ("31 a 90 dias", 31, 90), ("mais de 90 dias", 91, 10 ** 6)]
    ano_2026 = {
        "entregas": len(de26),
        "repassaram_adiante": len(repassaram),
        "fecharam_na_equipe": sum(1 for c in de26 if c["resolveu_ali_mesmo"]),
        "ainda_com_a_equipe": sum(1 for c in de26 if c["ainda_com_a_equipe"]),
        "tempo_ate_repassar": percentis([c["dias_no_dcmd"] for c in repassaram]),
        "faixas": [{"faixa": nome,
                    "qtd": sum(1 for c in repassaram if lo <= c["dias_no_dcmd"] <= hi)}
                   for nome, lo, hi in faixas],
        "para_onde": [{"destino": k, "qtd": len(v), **percentis(v)}
                      for k, v in sorted(destinos.items(), key=lambda kv: -len(kv[1]))],
        "por_equipe": [{"equipe": e, "qtd": len(g),
                        **percentis([c["dias_no_dcmd"] for c in g])}
                       for e, g in sorted(
                           {e: [c for c in repassaram if c["equipe"] == e]
                            for e in {c["equipe"] for c in repassaram}}.items(),
                           key=lambda kv: -len(kv[1]))],
        "casos": de26,
    }

    pacote = {
        "gerado_em": "2026-08-22", "posicao": "18/08/2026",
        "ano_2026": ano_2026,
        "fonte": "EQP_SS_OCORRENCIA_11082026 — cadeia de repasse pelo campo SS_APOS_REPASSE",
        "premissas": PREMISSAS,
        "geral": bloco(casos),
        "por_equipe": {k: bloco(v) for k, v in sorted(por_equipe.items(),
                                                     key=lambda kv: -len(kv[1]))},
        "por_ano": {k: bloco(v) for k, v in sorted(por_ano.items())},
        "casos": casos,
    }
    with open(SAIDA_JSON, "w", encoding="utf-8") as fh:
        json.dump(pacote, fh, ensure_ascii=False, indent=1)
    return pacote


PREMISSAS = [
    "Equipe do DCMD é a que tem RD no nome do posto — ETO-RD-AR, ETO-RD-PS, DOLP-RD-PA e as "
    "demais. É a régua do gestor para separar campo de escritório.",
    "Entrega ao DCMD é a abertura da SS no posto RD. O SGM não move a SS: ele fecha a do COEP "
    "como repassada e abre uma nova no destino, gravando o número em SS_APOS_REPASSE.",
    "Saída da equipe: conclusão da SS, se houver; senão a abertura da SS seguinte; senão a "
    "demanda ainda está com ela. A data de conclusão vem VAZIA em SS repassada — sem montar a "
    "cadeia, o tempo de equipe não existe na base.",
    "Dias no DCMD é quanto a equipe segurou antes de fechar ou passar adiante. Quem ainda está "
    "com o serviço conta até 18/08/2026, a posição do relatório.",
    "Dias até o fim é do repasse até a demanda morrer, em qualquer posto — inclui o tempo de "
    "quem veio depois. Só conta onde a cadeia terminou com SS atendida ou cancelada.",
    "«Resolveu ali mesmo» é a cadeia que acaba na própria equipe, sem mais nenhum repasse.",
    "«Voltou ao COEP» é a cadeia que, depois da equipe, passa de novo pelo posto do gestor.",
    "Uma entrega por repasse, não por equipamento: o mesmo ativo entregue duas vezes conta duas.",
]


def planilha(pacote):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    tit = Font(bold=True, color="FFFFFF", size=10)
    fundo = PatternFill("solid", fgColor="1F3864")
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)

    def cabecalho(ws, colunas, larguras):
        ws.append(colunas)
        for c, (col, larg) in enumerate(zip(colunas, larguras), 1):
            cel = ws.cell(row=1, column=c)
            cel.font, cel.fill = tit, fundo
            cel.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[cel.column_letter].width = larg
        ws.freeze_panes = "A2"

    def fechar(ws):
        for linha in ws.iter_rows(min_row=2):
            for cel in linha:
                cel.border = borda
                cel.alignment = Alignment(vertical="top", wrap_text=True)

    def linhas_bloco(rotulo, b):
        n, f = b["no_dcmd"], b["ate_o_fim"]
        return [rotulo, b["entregas"], n.get("mediana"), n.get("p75"), n.get("p90"),
                n.get("maximo"), b["resolveu_ali_mesmo"],
                f"{round(100 * b['resolveu_ali_mesmo'] / b['entregas'])}%" if b["entregas"] else "",
                b["voltou_ao_coep"], b["ainda_com_a_equipe"], f.get("mediana"), f.get("p90")]

    COLS = ["Equipe", "Entregas recebidas", "Dias no DCMD — mediana", "p75", "p90", "Máximo",
            "Resolveu ali mesmo", "% que resolveu ali", "Voltou ao COEP",
            "Ainda com a equipe", "Até o fim — mediana", "Até o fim — p90"]
    LARG = [15, 12, 14, 8, 8, 9, 12, 11, 11, 11, 13, 12]

    a = pacote["ano_2026"]
    t = a["tempo_ate_repassar"]

    ws = wb.active
    ws.title = "2026 · Para onde foi"
    cabecalho(ws, ["Para onde a equipe passou", "Quantas", "Dias até passar — mediana",
                   "p75", "p90", "Máximo"], [22, 10, 15, 8, 8, 9])
    for x in a["para_onde"]:
        ws.append([x["destino"], x["qtd"], x.get("mediana"), x.get("p75"), x.get("p90"),
                   x.get("maximo")])
    ws.append([])
    ws.append(["TODAS que passaram adiante", a["repassaram_adiante"], t.get("mediana"),
               t.get("p75"), t.get("p90"), t.get("maximo")])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append([])
    ws.append(["Em quanto tempo passaram", "Quantas"])
    for c in ws[ws.max_row]:
        c.font, c.fill = tit, fundo
    for x in a["faixas"]:
        ws.append([x["faixa"], x["qtd"]])
    ws.append([])
    ws.append(["O resto das entregas de 2026", "Quantas"])
    for c in ws[ws.max_row]:
        c.font, c.fill = tit, fundo
    ws.append(["Fecharam na própria equipe, sem passar adiante", a["fecharam_na_equipe"]])
    ws.append(["Ainda com a equipe em 18/08", a["ainda_com_a_equipe"]])
    ws.append(["Total entregue em 2026", a["entregas"]])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    fechar(ws)

    ws = wb.create_sheet("2026 · Por equipe")
    cabecalho(ws, ["Equipe", "Passou adiante", "Dias até passar — mediana", "p75", "p90",
                   "Máximo"], [16, 12, 15, 8, 8, 9])
    for x in a["por_equipe"]:
        ws.append([x["equipe"], x["qtd"], x.get("mediana"), x.get("p75"), x.get("p90"),
                   x.get("maximo")])
    fechar(ws)

    ws = wb.create_sheet("2026 · Cada entrega")
    cabecalho(ws, ["Ativo", "Tipo", "Equipe que recebeu", "Entregue em", "Passou em",
                   "Dias até passar", "Passou para", "SS da equipe", "SS do destino",
                   "Fechou na equipe", "Ainda com a equipe", "Situação final"],
              [14, 12, 15, 13, 13, 12, 15, 22, 22, 11, 11, 15])
    sn2 = lambda v: "sim" if v else "não"
    for c in sorted(pacote["ano_2026"]["casos"], key=lambda x: -x["dias_no_dcmd"]):
        ws.append([c["ativo"], c["tipo"], c["equipe"], c["entregue_em"],
                   c["saiu_da_equipe"] if c["repassou_para"] else "",
                   c["dias_no_dcmd"], c["repassou_para"] or "—", c["ss_da_equipe"],
                   c["ss_do_destino"], sn2(c["resolveu_ali_mesmo"]),
                   sn2(c["ainda_com_a_equipe"]), c["status_final"]])
    fechar(ws)

    ws = wb.create_sheet("Série · Por equipe")
    cabecalho(ws, COLS, LARG)
    ws.append(linhas_bloco("TODAS", pacote["geral"]))
    for equipe, b in pacote["por_equipe"].items():
        ws.append(linhas_bloco(equipe, b))
    fechar(ws)
    for cel in ws[2]:
        cel.font = Font(bold=True)

    ws = wb.create_sheet("Série · Por ano")
    cabecalho(ws, ["Ano"] + COLS[1:], LARG)
    for ano, b in pacote["por_ano"].items():
        ws.append(linhas_bloco(ano, b))
    fechar(ws)

    ws = wb.create_sheet("Série · Cada entrega")
    cabecalho(ws, ["Ativo", "Tipo", "Equipe", "SS do COEP", "SS da equipe", "Entregue em",
                   "Saiu da equipe", "Como se apurou a saída", "Dias no DCMD",
                   "Repassou para", "SS do destino", "Ainda com a equipe",
                   "Resolveu ali mesmo", "Postos depois da equipe", "Voltou ao COEP",
                   "Fim da demanda", "Dias até o fim", "Posto que fechou", "Situação final"],
              [14, 12, 14, 22, 22, 13, 13, 30, 11, 14, 22, 11, 11, 12, 11, 13, 11, 14, 15])
    sn = lambda v: "sim" if v else "não"
    for c in sorted(pacote["casos"], key=lambda x: -x["dias_no_dcmd"]):
        ws.append([c["ativo"], c["tipo"], c["equipe"], c["ss_do_coep"], c["ss_da_equipe"],
                   c["entregue_em"], c["saiu_da_equipe"], c["como_apurou_a_saida"],
                   c["dias_no_dcmd"], c["repassou_para"], c["ss_do_destino"],
                   sn(c["ainda_com_a_equipe"]), sn(c["resolveu_ali_mesmo"]),
                   c["postos_depois_da_equipe"], sn(c["voltou_ao_coep"]), c["fim_da_demanda"],
                   c["dias_ate_o_fim"], c["posto_que_fechou"], c["status_final"]])
    fechar(ws)

    ws = wb.create_sheet("Como foi feito")
    cabecalho(ws, ["Passo", "O que foi feito"], [8, 130])
    for n, texto in enumerate(pacote["premissas"], 1):
        ws.append([n, texto])
    fechar(ws)

    os.makedirs(os.path.dirname(SAIDA_XLSX), exist_ok=True)
    wb.save(SAIDA_XLSX)


def main():
    pacote = montar()
    g = pacote["geral"]
    print(f"entregas do COEP para equipe RD.... {g['entregas']}")
    print(f"dias no DCMD...................... mediana {g['no_dcmd']['mediana']}d | "
          f"p75 {g['no_dcmd']['p75']}d | p90 {g['no_dcmd']['p90']}d | máx {g['no_dcmd']['maximo']}d")
    print(f"do repasse até o fim da demanda... mediana {g['ate_o_fim']['mediana']}d | "
          f"p90 {g['ate_o_fim']['p90']}d  (n={g['ate_o_fim']['n']})")
    print(f"resolveu ali mesmo................ {g['resolveu_ali_mesmo']} "
          f"({round(100 * g['resolveu_ali_mesmo'] / g['entregas'])}%)")
    print(f"voltou ao COEP.................... {g['voltou_ao_coep']}")
    print(f"ainda com a equipe................ {g['ainda_com_a_equipe']}")
    print("\npor equipe (5+ entregas):")
    for equipe, b in pacote["por_equipe"].items():
        if b["entregas"] < 5:
            continue
        print(f"  {equipe:<12} {b['entregas']:>4} entregas | mediana {b['no_dcmd']['mediana']:>4}d"
              f" | p90 {b['no_dcmd']['p90']:>4}d | resolveu ali "
              f"{round(100 * b['resolveu_ali_mesmo'] / b['entregas']):>3}%"
              f" | voltou {b['voltou_ao_coep']:>3}")
    a = pacote["ano_2026"]
    t = a["tempo_ate_repassar"]
    print(f"\n=== 2026, o recorte pedido ===")
    print(f"entregues às equipes RD............ {a['entregas']}")
    print(f"  passaram adiante................. {a['repassaram_adiante']}")
    print(f"  fecharam na própria equipe....... {a['fecharam_na_equipe']}")
    print(f"  ainda com a equipe em 18/08...... {a['ainda_com_a_equipe']}")
    print(f"dias até passar adiante............ mediana {t['mediana']}d | p75 {t['p75']}d "
          f"| p90 {t['p90']}d | máx {t['maximo']}d")
    print("para onde passaram:")
    for x in a["para_onde"]:
        print(f"  {x['qtd']:>3}  {x['destino']:<12} mediana {x['mediana']:>4}d | "
              f"máx {x['maximo']:>4}d")
    print(f"\ngravado: {SAIDA_JSON}")
    planilha(pacote)
    print(f"gravado: {SAIDA_XLSX}")


if __name__ == "__main__":
    main()
