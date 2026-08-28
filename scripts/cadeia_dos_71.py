"""
A cadeia completa de cada um dos 71 — dist/CADEIA_DOS_71.xlsx.

Os 71 são a conta de 28/08: 82 ativos com a régua de resolvido do gestor, menos os 11
que resolveram no ano e voltaram para a fila do posto («se voltaram eu não resolvi»).

Para cada um, a cadeia inteira de SS — do primeiro posto ao último —, com número da SS,
posto, TIPOSS, situação, abertura, término, dias parados naquele posto e a descrição
inteira. A descrição do SGM é cumulativa: parecer novo colado por cima do antigo, sem
separador, então vale sempre o trecho mais recente.

Rodar: python3 scripts/cadeia_dos_71.py
"""

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
import cadeia_obra as co  # noqa: E402
import sla_manutencao as base  # noqa: E402
import sla_por_equipe as eq  # noqa: E402

COEP = os.path.join(RAIZ, "data", "missao", "coep_2026.json")
SAIDA = os.path.join(RAIZ, "dist", "CADEIA_DOS_71.xlsx")

TINTA, PAPEL, SINAL = "FF211D15", "FFF2EFE6", "FFBC4B0E"
CAB = Font(bold=True, color=PAPEL, size=10)
FILL = PatternFill("solid", fgColor=TINTA)
ZEBRA = PatternFill("solid", fgColor="FFEFEBDF")


def base_ss():
    """{SS: campos} da base de SS/OS, com a descrição inteira."""
    d = {}
    for r in co.registros():
        n = (r.get("NUMERO_SS") or "").strip()
        if n:
            d[n] = {
                "tiposs": (r.get("TIPOSS") or "").strip(),
                "situacao": (r.get("SITUACAO_SS") or "").strip(),
                "abertura": (r.get("DATA_ABERTURA_SS") or "")[:10],
                "termino": (r.get("DATA_TERMINO_SS") or "")[:10],
                "posto": (r.get("COD_EQUIPE") or "").strip(),
                "obra": (r.get("NUM_OBRA") or "").strip(),
                "os": (r.get("NUMERO_OS") or "").strip(),
                "descricao": (r.get("DESCRICAO") or "").strip(),
            }
    return d


def dias(a, b):
    from datetime import date
    try:
        d1 = date(*map(int, a.split("-")))
        d2 = date(*map(int, b.split("-")))
        return (d2 - d1).days
    except (ValueError, TypeError, AttributeError):
        return ""


def montar():
    ss_info = base_ss()
    porchave = {base.norm(k): k for k in ss_info}
    reg = base.base_de_repasse(None)
    anterior = eq.raizes(reg)
    with open(COEP, encoding="utf-8") as fh:
        cp = json.load(fh)

    at = {a["ativo"]: a for a in cp["ativos"]}
    fila = {a for a, v in at.items() if v["segue_no_posto"]}
    res = {r["ativo"]: r for r in cp["resolvidos_do_coep"]
           if r["conta_como_resolvido_pelo_coep"]}
    voltaram = set(res) & fila
    os71 = sorted(set(res) - voltaram)
    assert len(res) == 82 and len(voltaram) == 11 and len(os71) == 71, \
        f"esperado 82-11=71, veio {len(res)}-{len(voltaram)}={len(os71)}"

    linhas, resumo = [], []
    for ativo in os71:
        v, r = at[ativo], res[ativo]
        sigla = "RL" if v["tipo"] == "religador" else "RT"
        coep_ss = [s.strip() for s in (v["ss"] or "").split("|") if s.strip()]
        # a cadeia inteira, do começo, sem repetir SS
        cadeia = []
        for s in coep_ss:
            try:
                for c, _ in base.cadeia(eq.do_comeco(base.norm(s), anterior), reg):
                    k = porchave.get(base.norm(c))
                    if k and k not in cadeia:
                        cadeia.append(k)
            except Exception:  # noqa: BLE001 — cadeia quebrada não derruba a planilha
                pass
        fechou = porchave.get(base.norm(r.get("ss_que_fechou") or ""), "")
        for k in (coep_ss + [fechou]):
            kk = porchave.get(base.norm(k)) if k else None
            if kk and kk not in cadeia:
                cadeia.append(kk)

        for i, s in enumerate(cadeia, start=1):
            m = ss_info[s]
            papel = []
            if any(base.norm(s) == base.norm(c) for c in coep_ss):
                papel.append("SS do COEP")
            if fechou and s == fechou:
                papel.append("fechou a cadeia")
            if "-RD-" in m["posto"]:
                papel.append("equipe de campo")
            linhas.append([
                ativo, sigla, v["localidade"], i, s, m["posto"],
                " · ".join(papel) or "—",
                m["tiposs"], m["situacao"], m["abertura"], m["termino"] or "",
                dias(m["abertura"], m["termino"]) if m["termino"] else "",
                str(m["obra"]).zfill(10) if m["obra"] else "", m["os"], m["descricao"],
            ])
        resumo.append([
            ativo, sigla, v["localidade"], v.get("criticidade") or "",
            r.get("ocorrencia_da_demanda") or "", r.get("ano_da_demanda") or "",
            " | ".join(coep_ss), len(coep_ss),
            r.get("ss_que_fechou") or "", r.get("posto_que_fechou") or "",
            r.get("como_terminou") or "", r.get("data_do_fechamento") or "",
            r.get("dias_da_demanda") or "", len(cadeia), r.get("prova") or "",
        ])
    return linhas, resumo, len(voltaram), sorted(voltaram)


def planilha(linhas, resumo, qtd_voltaram, voltaram):
    wb = Workbook()
    wb.remove(wb.active)

    cols_r = [("Ativo", 13), ("RL/RT", 7), ("Praça", 24), ("Criticidade", 13),
              ("Ocorrência da demanda", 17), ("Ano da demanda", 12),
              ("SS no COEP", 40), ("Qtd SS no COEP", 12), ("SS que fechou", 20),
              ("Posto que fechou", 15), ("Como terminou", 14),
              ("Data do fechamento", 15), ("Dias da demanda", 13),
              ("SS na cadeia", 11), ("Prova registrada", 44)]
    ws = wb.create_sheet("Os 71")
    ws.append([c[0] for c in cols_r])
    for r in resumo:
        ws.append(r)

    cols_c = [("Ativo", 13), ("RL/RT", 7), ("Praça", 22), ("Ordem", 7), ("SS", 21),
              ("Posto", 13), ("Papel na cadeia", 26), ("TIPOSS", 30),
              ("Situação", 15), ("Abertura", 11), ("Término", 11),
              ("Dias no posto", 12), ("Obra", 12), ("OS", 21), ("Descrição", 120)]
    ws2 = wb.create_sheet("Cadeia SS a SS")
    ws2.append([c[0] for c in cols_c])
    ant = None
    faixa = False
    for r in linhas:
        ws2.append(r)
        if r[0] != ant:
            faixa = not faixa
            ant = r[0]
        if faixa:
            for c in range(1, len(cols_c) + 1):
                ws2.cell(row=ws2.max_row, column=c).fill = ZEBRA

    for w, cols in ((ws, cols_r), (ws2, cols_c)):
        for c in range(1, len(cols) + 1):
            cel = w.cell(row=1, column=c)
            cel.font, cel.fill = CAB, FILL
            cel.alignment = Alignment(horizontal="left", vertical="center",
                                      wrap_text=True)
        for i, (_, larg) in enumerate(cols, start=1):
            w.column_dimensions[get_column_letter(i)].width = larg
        w.freeze_panes = "A2"
        w.auto_filter.ref = w.dimensions
        w.row_dimensions[1].height = 30
    ultima = get_column_letter(len(cols_c))
    for r in range(2, ws2.max_row + 1):
        ws2.cell(row=r, column=len(cols_c)).alignment = Alignment(wrap_text=False,
                                                                 vertical="top")
    ws2.column_dimensions[ultima].width = 120

    ws3 = wb.create_sheet("Como foi feito")
    for t in [
        ["A cadeia dos 71"], [""],
        ["Quem são os 71"], [""],
        ["82 ativos passaram pelo posto do COEP em 2026 e fecharam a demanda dentro do"],
        ["ano, com SS atendida ou cancelada — a régua de resolvido do gestor."],
        [""],
        ["Desses, 11 resolveram uma demanda no ano e VOLTARAM para a fila do posto. Pela"],
        ["decisão do gestor de 28/08 («se voltaram eu não resolvi») eles contam como"],
        ["pendentes, não como resolvidos. 82 - 11 = 71."],
        [""],
        ["Os 11 que saíram: " + ", ".join(voltaram)],
        [""],
        ["O que a aba Cadeia mostra"], [""],
        ["Uma linha por SS, na ordem da cadeia — do primeiro posto que abriu até o"],
        ["último. A coluna 'Papel na cadeia' marca qual SS é do COEP, qual fechou a"],
        ["cadeia e quais estiveram com equipe de campo (postos com -RD- no código)."],
        [""],
        ["Armadilhas da base que você precisa ter na cabeça ao ler"], [""],
        ["1. A DESCRIÇÃO É CUMULATIVA. O SGM cola parecer novo por cima do antigo, sem"],
        ["   separador. Vale sempre o parecer MAIS RECENTE — normalmente no fim do texto,"],
        ["   mas confira as datas citadas, porque às vezes um parecer antigo fica no topo."],
        [""],
        ["2. SS REPASSADA NÃO TEM TÉRMINO. O campo sai vazio, e isso é normal: ela saiu"],
        ["   do posto no dia em que a SS seguinte foi aberta, não foi abandonada."],
        [""],
        ["3. TEXTO DE TERCEIRO. Às vezes o laudo colado é de OUTRO ativo. Confira se o"],
        ["   código citado no texto é o da linha antes de acreditar."],
        [""],
        ["4. O SGM NÃO EXPORTA O MOTIVO DO CANCELAMENTO. Numa SS cancelada sem texto"],
        ["   explicativo não dá para afirmar que a demanda foi resolvida."],
        [""],
        ["Ressalva sobre 'resolvido'"], [""],
        ["Resolvido aqui é a régua de 28/08: a cadeia fechou no ano. Não quer dizer que"],
        ["houve conserto — dos 71, 41 fecharam por cancelamento. A auditoria de 29/08"],
        ["mostrou que só 16 têm peça grande trocada com prova (obra do AIC com conclusão"],
        ["física e valor realizado, ou OS de campo nomeada)."],
    ]:
        ws3.append(t)
    ws3.column_dimensions["A"].width = 100
    ws3["A1"].font = Font(bold=True, size=12, color=SINAL)
    for r in (3, 13, 19, 34):
        ws3.cell(row=r, column=1).font = Font(bold=True, size=11)

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    return SAIDA


if __name__ == "__main__":
    linhas, resumo, qtd, voltaram = montar()
    print(f"gravado: {planilha(linhas, resumo, qtd, voltaram)}")
    print(f"  {len(resumo)} ativos · {len(linhas)} SS na cadeia · "
          f"{len(linhas)/len(resumo):.1f} SS por ativo")
    print(f"  os {qtd} que voltaram e saíram: {', '.join(voltaram)}")
