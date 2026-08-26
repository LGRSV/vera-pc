"""
A planilha da visão ETO — o filtro do gestor em Excel, com o balde de cada um.

Lê data/missao/visao_consolidada.json (rodar scripts/visao_consolidada.py antes)
e grava dist/VISAO_ETO.xlsx com duas abas: a lista dos ativos e o passo a passo
de como o filtro foi feito, para conferência direta no SGM.

Rodar: python3 scripts/planilha_visao_eto.py
"""

import json
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
import cadeia_obra as co  # noqa: E402 — o remontador de registros da base crua
import extrai_ssos_min as em  # noqa: E402 — o normalizador de 64 campos

SAIDA = os.path.join(RAIZ, "dist", "VISAO_ETO.xlsx")
CACHE_DESC = os.path.join(RAIZ, "data", "missao", "descricao_ss_pendentes.json")
COL_DESCRIPTION_SS = 27  # a descrição cumulativa da SS na base crua
RE_ILEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def descricoes_das_ss(numeros):
    """A DESCRIPTION_SS de cada SS pedida, lida da base crua — com cache em
    data/missao para a planilha sair igual mesmo sem a base (gitignored) por perto."""
    numeros = set(numeros)
    base = next((p for p in co.PARTES if os.path.exists(p)), None)
    if base is None:
        if os.path.exists(CACHE_DESC):
            with open(CACHE_DESC, encoding="utf-8") as fh:
                return json.load(fh)
        return {}
    achadas = {}

    def registra(bruto):
        campos = em._normaliza(bruto.split("@"))
        num = campos[0].strip()
        if num in numeros:
            achadas[num] = RE_ILEGAL.sub("", campos[COL_DESCRIPTION_SS].strip())

    buffer = None
    with open(base, encoding="latin-1") as fh:
        for i, linha in enumerate(fh):
            linha = linha.rstrip("\r\n")
            if i == 0 and linha.startswith("NUMERO_SS@"):
                continue
            if co.RE_INICIO.match(linha):
                if buffer is not None:
                    registra(buffer)
                buffer = linha
            elif buffer is not None:
                buffer += "\n" + linha
        if buffer is not None:
            registra(buffer)
    with open(CACHE_DESC, "w", encoding="utf-8") as fh:
        json.dump(achadas, fh, ensure_ascii=False, indent=1)
    return achadas

BALDE_NOME = {
    "ajuste_de_protecao": "Em fase de ajuste de proteção",
    "comissionamento": "Aguardando comissionamento",
    "dcmd_execucao": "DCMD — em execução (COCM's)",
    "dcmd_logistica": "DCMD — em logística",
    "dcmd_aquisicao": "DCMD — em processo de aquisição",
    "dmsl_novos": "1º ataque do DMSL",
}

def como_foi_feito(v):
    return [
    f"O FILTRO — na base de SS/OS ({v['posicao']}), três cortes, nesta ordem:",
    "1. Código do equipamento (NUM_TRAFO) começando com 79 (religador) ou 58 (regulador).",
    "2. Tipo da SS (TIPOSS) = INDISPONIBILIDADE PARA OPERAÇÃO.",
    "3. Situação (SITUACAO_SS) = SS PENDENTE.",
    f"Resultado: {v['total']} SS pendentes, uma por ativo — {v['total']} equipamentos.",
    "",
    "Por que REPASSADA não conta: quando a SS é repassada, o SGM abre outra SS no posto "
    "seguinte — a viva é a nova, e é ela que aparece como pendente. Atendida e cancelada "
    "são cadeia fechada.",
    "",
    "O BALDE — sai do POSTO onde a SS pendente está (o prefixo do número da SS), que é a "
    "régua da esteira:",
    "• ETO-PROT → em fase de ajuste de proteção.",
    "• ETO-TELE / ETO-SE com CRITICIDADE DEFINIDA na aba de mapeamento por criticidade → "
    "aguardando comissionamento.",
    "• ETO-TELE / ETO-SE SEM criticidade definida (fora da aba, ou «Sem classificação») → "
    "1º ataque do DMSL — régua do gestor, 22/08: são os novos, que ainda não passaram "
    "pelo COEP.",
    "• ETO-RD-* (equipe de campo) → DCMD em execução, com os COCM's.",
    "• ETO-COEP → DCMD: quem a aba marca «Em logística» está em logística (material "
    "comprado, esperando chegar); todo o restante pendente no posto está em processo de "
    "aquisição.",
    "",
    "A aba de mapeamento por criticidade (Relação de Indisponíveis, ATUALIZADA 16) entra "
    "como ANOTAÇÃO: dá a criticidade, a etapa que ela dizia e mostra quem ela nem lista — "
    f"{v['fora_da_carteira']} dos {v['total']} estão fora dela.",
    "",
    "Os de aquisição foram cruzados com o plano de compras de 17/07 pelo código do ativo.",
    "",
    "A régua de quem entra é a BASE, não a carteira (gestor, 22/08): «só tem 93, então são "
    "as 93». Quem está na carteira sem SS de indisponibilidade pendente fica de fora — e "
    "quem tem SS pendente fica dentro mesmo sem estar na carteira.",
    "",
    "DECISÕES DO GESTOR — por cima da esteira. A esteira diz onde a SS está pendurada; o "
    "gestor diz onde a bola está. Cada decisão vale com motivo e data, e aparece na coluna "
    "«Decisão do gestor» da lista:",
    *[f"• {i['ativo']} ({i['localidade']}) — saía como «{i['de']}». {i['motivo']}"
      for i in v.get("decisoes_do_gestor", {}).get("itens", [])],
    "",
    "PARA REFAZER COM BASE NOVA: largar a BASE_SS_OS_ddmmaaaa.txt em data/raw e rodar "
    "python3 scripts/atualiza_visao_eto.py — extrai o recorte, refaz a visão, esta "
    "planilha e o painel. A mesma régua está no site, na home, em «Como esta visão é "
    "montada».",
]


def montar():
    with open(os.path.join(RAIZ, "data", "missao", "visao_consolidada.json"),
              encoding="utf-8") as fh:
        vc = json.load(fh)
    v = vc["visao_eto"]

    wb = openpyxl.Workbook()
    tit = Font(bold=True, color="FFFFFF", size=10)
    fundo = PatternFill("solid", fgColor="1F3864")
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)

    todos = [i for b in BALDE_NOME for i in v["baldes"][b]["ativos"]]
    descricao = descricoes_das_ss(i["ss_pendente"] for i in todos)

    ws = wb.active
    ws.title = f"Visão ETO ({v['total']})"
    colunas = ["Balde", "Ativo", "Tipo", "Localidade", "SS pendente", "Posto da SS",
               "Criticidade (aba de mapeamento)", "Etapa na aba", "Está na aba",
               "No plano de compras", "Decisão do gestor",
               "Descrição da SS (cumulativa — vale o parecer mais recente)"]
    larguras = [30, 14, 8, 24, 22, 14, 16, 24, 10, 12, 60, 90]
    ws.append(colunas)
    for c, larg in enumerate(larguras, 1):
        cel = ws.cell(row=1, column=c)
        cel.font, cel.fill = tit, fundo
        cel.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[cel.column_letter].width = larg
    ws.freeze_panes = "A2"

    sn = lambda b: "sim" if b else "não"
    for balde in BALDE_NOME:
        for i in v["baldes"][balde]["ativos"]:
            desc = descricao.get(i["ss_pendente"], "")
            if len(desc) > 32000:
                desc = desc[:32000] + " (…cortado)"
            ws.append([
                BALDE_NOME[balde], i["ativo"], i["tipo"], i["localidade"],
                i["ss_pendente"], i["ss_pendente"].split()[0],
                i["criticidade"] or "—", i["etapa_da_planilha"], sn(i["na_carteira"]),
                (sn(i["no_plano_de_compras"]) if "no_plano_de_compras" in i else "—"),
                i.get("decisao_do_gestor", "—"),
                desc or "—",
            ])
    for linha in ws.iter_rows(min_row=2):
        for cel in linha:
            cel.border = borda
            cel.alignment = Alignment(vertical="top")
        linha[-1].alignment = Alignment(vertical="top", wrap_text=True)
        linha[-2].alignment = Alignment(vertical="top", wrap_text=True)

    ws2 = wb.create_sheet("Como foi feito")
    ws2.column_dimensions["A"].width = 110
    for txt in como_foi_feito(v):
        ws2.append([txt])
        ws2.cell(row=ws2.max_row, column=1).alignment = Alignment(wrap_text=True,
                                                                  vertical="top")

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    print(f"gravado: {SAIDA} — {v['total']} ativos")


if __name__ == "__main__":
    montar()
