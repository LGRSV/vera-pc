"""
A visão do posto de 2026 separada em religador e regulador.

Pergunta do gestor (24/08): dos 143 que passaram pelo COEP, quantos são RL e
quantos são RT. Este script refaz todas as contas do posto com o corte por tipo e
grava dist/COEP_2026_POR_TIPO.xlsx.

A identidade se mantém dentro de cada tipo: passaram = resolvidos + pendentes no
posto + despachados para outra mesa − os que aparecem dos dois lados (resolvido
cuja reincidência já voltou para a fila).

Rodar: python3 scripts/coep_por_tipo_excel.py
"""

import json
import os
from collections import Counter, defaultdict

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SAIDA = os.path.join(RAIZ, "dist", "COEP_2026_POR_TIPO.xlsx")
NOME = {"RL": "Religador", "RT": "Regulador"}

TIT = Font(bold=True, color="FFFFFF", size=10)
FUNDO = PatternFill("solid", fgColor="1F3864")
BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)
NEG = Font(bold=True)


def tipo(v):
    return "RT" if str(v).lower().startswith("regul") else "RL"


def cabecalho(ws, colunas, larguras):
    ws.append(colunas)
    for c, (nome, larg) in enumerate(zip(colunas, larguras), 1):
        cel = ws.cell(row=1, column=c)
        cel.font, cel.fill = TIT, FUNDO
        cel.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = larg
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def fechar(ws, centro_a_partir_de=2):
    for linha in ws.iter_rows(min_row=2):
        for cel in linha:
            cel.border = BORDA
            cel.alignment = Alignment(vertical="top", wrap_text=True)
        for cel in linha[centro_a_partir_de:]:
            cel.alignment = Alignment(vertical="top", horizontal="center")


def montar():
    with open(os.path.join(RAIZ, "data", "missao", "coep_2026.json"), encoding="utf-8") as fh:
        cp = json.load(fh)
    ativos = cp["ativos"]
    resolvidos = [r for r in cp["resolvidos_do_coep"] if r["conta_como_resolvido_pelo_coep"]]
    nao = [r for r in cp["resolvidos_do_coep"] if not r["conta_como_resolvido_pelo_coep"]]
    pendentes = [a for a in ativos if a["segue_no_posto"]]
    fora = cp["pendentes_em_outra_mesa"]

    wb = openpyxl.Workbook()

    # ---------- 1 · o resumo, que é a resposta da pergunta
    ws = wb.active
    ws.title = "1 · RL e RT"
    cabecalho(ws, ["A conta do posto em 2026", "Religador", "Regulador", "Total",
                   "Como se lê"],
              [42, 12, 12, 10, 76])
    def n(lista, t, chave=lambda x: x["tipo"]):
        return sum(1 for x in lista if tipo(chave(x)) == t)
    sobre = {t: len({r["ativo"] for r in resolvidos if tipo(r["tipo"]) == t}
                    & {a["ativo"] for a in pendentes if tipo(a["tipo"]) == t})
             for t in ("RL", "RT")}
    linhas = [
        ("Passaram pelo posto no ano", [n(ativos, t) for t in ("RL", "RT")],
         "a SS esteve no COEP em algum momento de 2026 — não só a que chegou no ano"),
        ("Resolvidos pelo COEP", [n(resolvidos, t) for t in ("RL", "RT")],
         "a demanda passou pelo posto dentro de 2026 E a cadeia fechou dentro de 2026"),
        ("Ainda no posto em 18/08", [n(pendentes, t) for t in ("RL", "RT")],
         "a fila de hoje, esperando decisão ou material no COEP"),
        ("Conta do posto (resolvidos + fila)", [n(resolvidos, t) + n(pendentes, t)
                                                for t in ("RL", "RT")],
         "é o número que fecha em 136 no total — era 125 antes da régua da reincidência"),
        ("Despachados, pendentes em outra mesa", [n(fora, t) for t in ("RL", "RT")],
         "saíram do COEP e esperam na PROT, na TELE/SE ou com os COCMs"),
        ("Aparecem dos dois lados", [sobre[t] for t in ("RL", "RT")],
         "resolvidos cuja REINCIDÊNCIA já voltou para a fila — contam uma vez em cada"),
        ("Tirados: cancelada que voltou ao COEP",
         [n([r for r in nao if "voltou para o COEP" in r["porque_nao"]], t) for t in ("RL", "RT")],
         "cancelou e abriram nota nova no posto: a demanda voltou para a mesa"),
        ("Tirados: primeiro ataque do DMSL",
         [n([r for r in nao if "primeiro ataque" in r["porque_nao"].lower()], t)
          for t in ("RL", "RT")],
         "a demanda morreu na mão da DMSL — o posto não trabalhou nela"),
    ]
    for nome, (rl, rt), leitura in linhas:
        ws.append([nome, rl, rt, rl + rt, leitura])
    fechar(ws)
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).font = NEG
        ws.cell(row=r, column=4).font = NEG
    ws.append([])
    ws.append(["A identidade, dentro de cada tipo:"])
    ws.cell(row=ws.max_row, column=1).font = NEG
    for t in ("RL", "RT"):
        ws.append([f"{NOME[t]}: {n(resolvidos, t)} resolvidos + {n(pendentes, t)} na fila + "
                   f"{n(fora, t)} em outra mesa − {sobre[t]} contados duas vezes = "
                   f"{n(ativos, t)} que passaram pelo posto"])

    graf = BarChart()
    graf.type, graf.style, graf.height, graf.width = "col", 2, 8, 15
    graf.title = "O posto em 2026 — religador e regulador"
    graf.add_data(Reference(ws, min_col=2, max_col=3, min_row=1, max_row=7), titles_from_data=True)
    graf.set_categories(Reference(ws, min_col=1, min_row=2, max_row=7))
    ws.add_chart(graf, "A16")

    # ---------- 2 · a curva mensal por tipo
    ws = wb.create_sheet("2 · Mês a mês")
    cabecalho(ws, ["Mês", "Chegaram · RL", "Chegaram · RT", "Resolvidos · RL",
                   "Resolvidos · RT", "Conta do posto · RL", "Conta do posto · RT"],
              [12, 14, 14, 15, 15, 18, 18])
    cheg, res_m = defaultdict(int), defaultdict(int)
    primeira = {}
    for a in ativos:
        d = a["primeira_chegada"]
        if len(d) >= 10 and d[6:10] == "2026":
            primeira[a["ativo"]] = (f"{d[6:10]}-{d[3:5]}", tipo(a["tipo"]))
    for _, (m, t) in primeira.items():
        cheg[(t, m)] += 1
    for r in resolvidos:
        d = r["data_do_fechamento"]
        res_m[(tipo(r["tipo"]), f"{d[6:10]}-{d[3:5]}")] += 1
    meses = [f"2026-{i:02d}" for i in range(1, 9)]
    rot = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago"]
    ac = {"RL": 0, "RT": 0}
    fila_final = {t: sum(1 for a in pendentes if tipo(a["tipo"]) == t) for t in ("RL", "RT")}
    for i, m in enumerate(meses):
        for t in ("RL", "RT"):
            ac[t] += res_m[(t, m)]
        ws.append([rot[i], cheg[("RL", m)], cheg[("RT", m)], res_m[("RL", m)], res_m[("RT", m)],
                   ac["RL"] + fila_final["RL"], ac["RT"] + fila_final["RT"]])
    fechar(ws, centro_a_partir_de=1)
    ws.append(["total", sum(cheg[("RL", m)] for m in meses), sum(cheg[("RT", m)] for m in meses),
               ac["RL"], ac["RT"], "", ""])
    for c in range(1, 8):
        ws.cell(row=ws.max_row, column=c).font = NEG

    # ---------- 3 · os que passaram
    ws = wb.create_sheet(f"3 · Passaram ({len(ativos)})")
    cabecalho(ws, ["Tipo", "Ativo", "Localidade", "SS no COEP em 2026", "Quantas SS",
                   "Primeira chegada", "Dias no posto", "Ainda no posto", "Na carteira",
                   "Parecer COEP", "Criticidade"],
              [7, 13, 20, 38, 10, 14, 12, 13, 11, 26, 12])
    sn = lambda v: "sim" if v else "não"
    for a in sorted(ativos, key=lambda x: (tipo(x["tipo"]), -x["dias_no_posto"])):
        ws.append([tipo(a["tipo"]), a["ativo"], a["localidade"], a["ss"],
                   a["ss_no_coep_em_2026"], a["primeira_chegada"], a["dias_no_posto"],
                   sn(a["segue_no_posto"]), sn(a["na_carteira"]), a["parecer_coep"],
                   a["criticidade"]])
    fechar(ws)

    # ---------- 4 · resolvidos
    ws = wb.create_sheet(f"4 · Resolvidos ({len(resolvidos)})")
    cabecalho(ws, ["Tipo", "Ativo", "Localidade", "Ano da demanda", "Fechou em",
                   "Posto que fechou", "Como terminou", "A prova", "Nota nova no COEP",
                   "Nota pendente em outro posto"],
              [7, 13, 20, 14, 13, 16, 15, 46, 20, 22])
    for r in sorted(resolvidos, key=lambda x: (tipo(x["tipo"]), x["data_do_fechamento"][-4:],
                                               x["data_do_fechamento"][3:5])):
        ws.append([tipo(r["tipo"]), r["ativo"], r["localidade"], r["ano_da_demanda"],
                   r["data_do_fechamento"], r["posto_que_fechou"], r["como_terminou"],
                   r["prova"], r.get("nota_nova_no_coep") or "—",
                   r.get("nota_nova_em_outro_posto") or "—"])
    fechar(ws)

    # ---------- 5 · fila no posto
    ws = wb.create_sheet(f"5 · Fila no posto ({len(pendentes)})")
    cabecalho(ws, ["Tipo", "Ativo", "Localidade", "Dias no posto", "Primeira chegada",
                   "SS no COEP em 2026", "Parecer COEP", "Criticidade"],
              [7, 13, 20, 13, 15, 38, 30, 12])
    for a in sorted(pendentes, key=lambda x: (tipo(x["tipo"]), -x["dias_no_posto"])):
        ws.append([tipo(a["tipo"]), a["ativo"], a["localidade"], a["dias_no_posto"],
                   a["primeira_chegada"], a["ss"], a["parecer_coep"], a["criticidade"]])
    fechar(ws)

    # ---------- 6 · em outra mesa
    ws = wb.create_sheet(f"6 · Em outra mesa ({len(fora)})")
    cabecalho(ws, ["Tipo", "Ativo", "Localidade", "Etapa da esteira", "Onde está",
                   "Última SS", "Desde"],
              [7, 13, 20, 34, 16, 24, 13])
    for x in sorted(fora, key=lambda y: (tipo(y["tipo"]), y["etapa_da_esteira"])):
        ws.append([tipo(x["tipo"]), x["ativo"], x["localidade"], x["etapa_da_esteira"],
                   x.get("posto_atual", ""), x.get("ss_atual", ""), x.get("desde", "")])
    fechar(ws)

    # ---------- 7 · como foi feito
    ws = wb.create_sheet("Como foi feito")
    ws.column_dimensions["A"].width = 112
    for t in cp["premissas"]:
        ws.append([t])
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.append([])
    ws.append(["CORTE POR TIPO (24/08): religador é o código que começa com 79 — e com 78 no "
               "monofásico, que o cadastro recodificou; regulador começa com 58."])
    ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    print(f"gravado: {SAIDA}")
    for t in ("RL", "RT"):
        print(f"  {NOME[t]}: passaram {n(ativos,t)} · resolvidos {n(resolvidos,t)} · "
              f"fila {n(pendentes,t)} · outra mesa {n(fora,t)}")


if __name__ == "__main__":
    montar()
