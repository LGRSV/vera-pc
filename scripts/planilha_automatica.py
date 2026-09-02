"""
A planilha base com fórmulas automáticas e validação — dist/GESTAO_AUTOMATICA.xlsx.

O pedido do gestor (01/09), como foi dito: «quero que arranje um jeito daquela planilha com
fórmulas de ficarem automáticas com validação, por exemplo se eu selecionar que o defeito é
o Tanque ele já traz Tanque da 34,5 Noja Código 69001 (exemplo, não me recordo se é isso
mesmo); se eu disser que é uma célula de 400 kVA 90.000, se for 2, 180.000, usando valores
reais». Com os valores reais: o tanque de 34,5 kV é o código 690005 (R$ 38.151,48) e a
célula de 400 kVA é a 690241 a R$ 90.230 — duas dão R$ 180.460.

A planilha base (GESTAO_EQUIPAMENTOS_ESPECIAIS_COEP.xlsx) faz isso hoje com XLOOKUP em
DOIS arquivos externos: a carteira do SharePoint (Defeito, MAT e MO da aba «Criticidade
por Equipamento», colunas RL Auto / Valor Material / Valor mão de obra estimado) e a
GESTÃO DE EQUIPAMENTOS.xlsx (tensão pelos Ajustes). Fora da rede da Energisa esses
vínculos quebram e ficam os valores em cache. Esta planilha é autocontida: as tabelas de
apoio moram dentro dela.

  Catálogo   — a peça com código de material, descrição, preço de material e mão de
               obra. É o único lugar para mexer em preço; os «completos» são fórmula
               sobre as peças.
  Cadastro   — os religadores e reguladores dos ajustes da proteção (1.292 RL + 189 RT
               de um parque de 1.307 e 207), com tensão, modelo e potência. Quem não
               está lá usa as colunas manuais.
  Lançamento — a folha de entrada: digita o ativo, escolhe a peça na lista, dá a
               quantidade. O resto é fórmula.
  Gestão     — os 53 pendentes do DCMD, refeitos com as mesmas fórmulas e conferidos
               contra o que a carteira dizia.
  Falha Equipamentos — o rol de 90 falhas de 2025 e 2026 com peça em lista e custo.
  Resumo     — contagem por peça e taxa, em fórmula (sem dinâmica para atualizar).

Preço: vale a CARTEIRA DE 27/08 (é a que o gestor usa). Onde a carteira não tem linha,
vale «Premissas e Preços» do orçamento de 16/07. Os códigos de material vêm do orçamento
e do plano de compras (690001, 690005, 690916, 692263, 690236, 690240, 690241, 690669,
651638, 625510, 647641, 90556). A chave do catálogo é Tipo|Peça|Tensão|kVA — a célula
de regulador é chaveada por tensão E potência, nas três classes do gestor (167, 200,
400); o cadastro traz também 239, 250, 398, 667, que vão à classe mais próxima com
aviso.

Depois de gravar, as fórmulas são calculadas pela biblioteca `formulas` e os valores
entram como cache no arquivo, para a planilha não abrir em branco no Modo de Exibição
Protegido do Excel. `--sem-cache` pula esse passo.

Rodar: python3 scripts/planilha_automatica.py [saida.xlsx] [--sem-cache]
"""

import os
import re
import sys
import zipfile
from collections import Counter
from datetime import datetime
from xml.sax.saxutils import escape

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = os.path.join(RAIZ, "data", "raw", "GESTAO_EQUIPAMENTOS_ESPECIAIS_COEP.xlsx")
AJUSTES = os.path.join(RAIZ, "data", "raw", "GESTAO_DE_EQUIPAMENTOS.xlsx")
SAIDA = os.path.join(RAIZ, "dist", "GESTAO_AUTOMATICA.xlsx")

# Prontuário Industrial — papel, tinta, laranja-sinal
TINTA, PAPEL, SOMBRA, SINAL, APAGADA = "FF211D15", "FFF2EFE6", "FFE9E5D8", "FFBC4B0E", "FF8D8672"
MOEDA = '#,##0.00'
N_LANC = 300            # linhas prontas na folha de lançamento
CAD_MAX = 3000          # teto do intervalo de busca no cadastro (1.483 hoje; sobra para incluir)
CAT_MAX = 200           # teto do intervalo de busca no catálogo (36 linhas hoje)
FALHA_MAX = 1000        # teto das contagens do Resumo sobre a aba de falhas

MO_RL = {"13,8": 11016.94, "34,5": 13209.34}
MO_RT = {"13,8": 51402.14, "34,5": 80318.50}
MO_CONTROLE_RT = 20000.0
CLASSES_KVA = (167, 200, 400)           # as três do gestor
TENSOES = ["13,8", "34,5"]

F_ORC = "Premissas e Preços (ORCAMENTO_EQ_ESPECIAIS, 16/07)"
F_CART = "carteira 27/08, colunas RL Auto / Valor Material / Valor mão de obra"
F_PLANO = "plano de compras MA+Alta (16/07)"

# preço da célula por tensão e classe: (código, material, fonte, observação)
CELULAS = {
    ("13,8", "167"): ("690669", 27616.62, f"{F_ORC}, item 4", "a carteira 27/08 não tem linha de 167 kVA"),
    ("13,8", "200"): ("690236", 62113.20, f"{F_CART} («Célula 200 13,8», ativo 5844630060)",
                      "código 690236 é «13,8 / 239 kVA» no orçamento (R$ 61.098,29); o gestor chama de 200"),
    ("13,8", "400"): ("690241", 90230.00, f"{F_CART} («Célula 400 34,5»)",
                      "célula de 400 em 13,8 não tem linha própria — mesmo código e preço da de 34,5"),
    ("34,5", "167"): ("690669", 27616.62, f"{F_ORC}, item 4",
                      "célula de 167 em 34,5 não tem linha própria — mesmo código e preço da de 13,8"),
    ("34,5", "200"): ("690240", 51705.75, f"{F_CART} («Célula 200 34,5»)", "o orçamento de 16/07 trazia R$ 57.720,41"),
    ("34,5", "400"): ("690241", 90230.00, f"{F_CART} («Célula 400 34,5»)",
                      "o orçamento de 16/07 e o plano de compras traziam R$ 126.893,80"),
}


def chave(tipo, peca, tensao, kva=""):
    return f"{tipo}|{peca}|{tensao}|{kva}"


def monta_catalogo():
    """Linhas do catálogo. `material` e `mo` podem ser número ou ('=', função(linhas)) para
    virar fórmula sobre outras linhas — os completos são soma das peças."""
    L = []

    def rl(peca, tensao, cod, desc, mat, taxa, fonte, obs=""):
        L.append(dict(tipo="RL", peca=peca, tensao=tensao, kva="", cod=cod, desc=desc, mat=mat,
                      mo=MO_RL[tensao], taxa=taxa, fonte=fonte, obs=obs))

    def rt(peca, tensao, kva, cod, desc, mat, mo, taxa, fonte, obs=""):
        L.append(dict(tipo="RT", peca=peca, tensao=tensao, kva=kva, cod=cod, desc=desc, mat=mat,
                      mo=mo, taxa=taxa, fonte=fonte, obs=obs))

    for t in TENSOES:
        cod_t, cod_c = ("690001", "690916") if t == "13,8" else ("690005", "692263")
        mat_t, mat_c = (21785.72, 33816.82) if t == "13,8" else (38151.48, 38094.72)
        fonte3 = f"{F_ORC} · {F_CART} · {F_PLANO} — as três batem no centavo"
        rl("Completo", t, f"{cod_t} + {cod_c}", f"Religador completo (tanque + controle) — {t} kV",
           ("=", lambda r, t=t: f"={r[('RL', 'Tanque', t, '')]}+{r[('RL', 'Controle', t, '')]}"), "sim",
           fonte3, "fórmula: tanque + controle")
        rl("Tanque", t, cod_t, f"Tanque / Parte ativa — Religador {t} kV", mat_t, "sim", fonte3)
        rl("Controle", t, cod_c, f"Controle — Religador {t} kV", mat_c, "sim", fonte3)
        rl("Relé", t, "625510", "Relé de proteção multifunção 50/51 (relé do religador)", 10454.01, "sim",
           f"{F_ORC}, item granular (VALOR EM ESTOQUE)", "confiança VERIFICAR no orçamento; relé do religador conta como controle")
        rl("Placa de alimentação CA", t, "647641", "Placa de circuito impresso — religador RC10 (placa de alimentação CA)",
           1999.00, "sim", f"{F_ORC}, item granular (VALOR EM ESTOQUE)",
           "confiança VERIFICAR no orçamento; placa de alimentação CA é sinônimo de controle (régua do gestor)")
        rl("Bateria + rádio", t, "647877 + 649399", "Bateria de lítio + rádio GE Orbit", 14612.08, "não",
           f"{F_ORC}, item granular · {F_PLANO}", "confiança VERIFICAR no orçamento; telecom, fora da taxa")
        rl("Chave faca", t, "90556", "Chave seccionadora de distribuição tipo faca 36,2 kV 630 A", 736.43, "não",
           f"{F_ORC}, item granular · {F_PLANO}", "fora da taxa")
        L.append(dict(tipo="RL", peca="Acessório", tensao=t, kva="", cod="",
                      desc="Acessório / serviço sem código (valor fechado da carteira)", mat=15000.00, mo=0.0,
                      taxa="não", fonte=f"{F_CART} («Acessório 34,5», ativos 7942572059 e 7931219078)",
                      obs="valor fechado de dois ativos da carteira, sem mão de obra"
                          + ("; a carteira só orça acessório em 34,5 — linha de 13,8 inferida" if t == "13,8" else "")))
    for t in TENSOES:
        for k in map(str, CLASSES_KVA):
            cod, mat, fonte, obs = CELULAS[(t, k)]
            rt("Célula", t, k, cod, f"Célula — Regulador {k} kVA (cód. {cod}, {'13,8' if cod in ('690669', '690236') else '34,5'} kV)",
               mat, MO_RT[t], "sim", fonte, obs)
        rt("Controle", t, "", "651638", "Controle de regulador c/ nobreak", 23259.60, MO_CONTROLE_RT, "sim",
           f"{F_CART} («Controle 167/200 13,8/34,5»)",
           "o orçamento de 16/07 trazia R$ 29.445,39 e MO 0; relé de sincronismo conta como controle")
        for k in map(str, CLASSES_KVA):
            obs = ("material = 3 × célula + controle; mão de obra = 2 × (célula + controle), que é como a "
                   "carteira chega aos R$ 200.637 do «RT Completo 400 34,5»")
            if not (t == "34,5" and k == "400"):
                obs += " — nesta classe a mão de obra é derivada dessa regra, a confirmar com o gestor"
            rt("Completo", t, k, f"3 × {CELULAS[(t, k)][0]} + 651638",
               f"Regulador completo {k} kVA (3 células + controle)",
               ("=", lambda r, t=t, k=k: f"=3*{r[('RT', 'Célula', t, k)]}+{r[('RT', 'Controle', t, '')]}"),
               ("=", lambda r, t=t, k=k: f"=2*({r[('RT', 'Célula', t, k, 'mo')]}+{r[('RT', 'Controle', t, '', 'mo')]})"),
               "sim", f"{F_CART} («RT Completo 400 34,5»: R$ 293.949,60 + R$ 200.637)", obs)
        L.append(dict(tipo="RT", peca="Acessório", tensao=t, kva="", cod="",
                      desc="Acessório / serviço sem código (valor fechado da carteira)", mat=15000.00, mo=0.0,
                      taxa="não", fonte=f"{F_CART} («Acessório 400 34,5», ativos 5846328094 e 5853360007)",
                      obs="valor fechado de dois ativos da carteira, sem mão de obra"
                          + ("; a carteira só orça acessório em 34,5 — linha de 13,8 inferida" if t == "13,8" else "")))
    return L


PECAS_RL = ["Completo", "Tanque", "Controle", "Relé", "Placa de alimentação CA",
            "Bateria + rádio", "Chave faca", "Acessório"]
PECAS_RT = ["Célula", "Controle", "Completo", "Acessório"]
PECAS_TODAS = PECAS_RL + [p for p in PECAS_RT if p not in PECAS_RL]
CRITICIDADES = ["Muito Alta", "Alta", "Média", "Baixa", "Falta definir", "Sem classificação"]
STATUS = ["Avaliar compra", "Gerado PMA", "Em compra", "Em logistica (N1>N3)",
          "A realizar (COCM)", "Em execução", "Reforma", "Realizado"]
SIM_NAO = ["sim", "não"]


# ----------------------------------------------------------------------------- leitura
def texto(v):
    return "" if v is None else str(v).strip()


def faixa(bruto):
    t = texto(bruto).replace(".", "").replace(",", "").replace(" ", "")
    if t.startswith("138"):
        return "13,8"
    if t.startswith("345"):
        return "34,5"
    return ""


def classe_kva(bruto):
    """A classe do gestor mais próxima; num banco misto, a potência que mais aparece
    (empate: a maior). Devolve (classe, desviou)."""
    partes = [p for p in texto(bruto).replace("/", " ").split() if p]
    nums = []
    for p in partes:
        try:
            nums.append(float(p.replace(",", ".")))
        except ValueError:
            pass
    if not nums:
        return "", False
    c = Counter(nums)
    valor = max(c, key=lambda x: (c[x], x))
    classe = min(CLASSES_KVA, key=lambda k: (abs(k - valor), -k))
    return str(classe), classe != valor


def ler_cadastro():
    wb = load_workbook(AJUSTES, read_only=True, data_only=True)
    cad = {}
    for r in list(wb["Ajustes RL Poste"].iter_rows(values_only=True))[1:]:
        a = texto(r[0])
        if a[:2] in ("78", "79") and a not in cad:
            cad[a] = {"tipo": "RL", "tensao": faixa(r[12]), "modelo": texto(r[10]),
                      "potencia": "", "classe": "", "desvio": False,
                      "localidade": texto(r[13]), "alimentador": texto(r[11]),
                      "origem": "Ajustes RL Poste", "obs": ""}
    for r in list(wb["Ajustes Reguladores de Tensão"].iter_rows(values_only=True))[1:]:
        a = texto(r[0])
        if a[:2] == "58" and a not in cad:
            classe, desviou = classe_kva(r[4])
            modelo = " / ".join(x for x in (texto(r[3]), texto(r[2])) if x)
            cad[a] = {"tipo": "RT", "tensao": faixa(r[8]), "modelo": modelo,
                      "potencia": texto(r[4]), "classe": classe, "desvio": desviou,
                      "localidade": texto(r[14]), "alimentador": texto(r[7]),
                      "origem": "Ajustes Reguladores de Tensão", "obs": ""}
    for a, d in cad.items():
        avisos = []
        if d["desvio"]:
            avisos.append(f"potência {d['potencia']} kVA no cadastro, fora das classes 167/200/400 — "
                          f"orçada como {d['classe']}; conferir a peça")
        if not d["tensao"]:
            avisos.append("sem tensão no cadastro — preencha Tensão manual")
        d["obs"] = "; ".join(avisos)
    return cad


def ler_modelos_carteira():
    """Modelo por ativo na Planilha1 da GESTÃO DE EQUIPAMENTOS (carteira de 12/08)."""
    wb = load_workbook(AJUSTES, read_only=True, data_only=True)
    m = {}
    for r in list(wb["Planilha1"].iter_rows(values_only=True))[1:]:
        if r[1] and texto(r[2]) not in ("", "#N/A"):
            m.setdefault(texto(r[1]), texto(r[2]))
    return m


def ler_base():
    wb = load_workbook(BASE, data_only=True)
    ws = wb["Gestão"]
    gestao = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[0]:
            continue
        gestao.append({"ativo": texto(r[0]), "ss": texto(r[2]), "criticidade": texto(r[3]),
                       "status": texto(r[4]), "defeito": texto(r[5]),
                       "mo": r[6] or 0, "mat": r[7] or 0, "total": r[8] or 0,
                       "esteira": list(r[9:19]), "indice": r[19], "sla": r[20],
                       "dias": r[21], "atendimento": texto(r[22]), "resp": texto(r[23]),
                       "prazo": texto(r[24]), "alimentador": texto(r[26]),
                       "municipio": texto(r[27]), "polo": texto(r[28]), "regional": texto(r[29])})
    ws = wb["Falha Equipamentos"]
    falhas = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not (texto(r[0]) and r[4]):
            continue
        falhas.append({"fatia": texto(r[0]), "ativo": texto(r[4]), "ss": texto(r[5]),
                       "tensao": faixa(r[6]), "data": texto(r[7]), "peca": texto(r[11]).lower(),
                       "troca": texto(r[12]), "causa": texto(r[13]), "citacao": texto(r[14]),
                       "nota": texto(r[15]), "revisao": texto(r[16])})
    return gestao, falhas


def interpreta_defeito(defeito):
    """«Célula 400 34,5» → (peça, classe kVA, tensão). É o texto da carteira."""
    d = defeito.strip()
    tensao = "13,8" if d.endswith("13,8") else ("34,5" if d.endswith("34,5") else "")
    m = re.search(r"\b(167|200|400)\b", d)
    kva = m.group(1) if m else ""
    if d.startswith("RL Completo") or d.startswith("RT Completo"):
        peca = "Completo"
    elif d.startswith("Tanque"):
        peca = "Tanque"
    elif d.startswith("Controle"):
        peca = "Controle"
    elif d.startswith("Célula"):
        peca = "Célula"
    elif d.startswith("Acessório"):
        peca = "Acessório"
    else:
        peca = ""
    return peca, kva, tensao


PECA_ROL = {"completo": "Completo", "tanque": "Tanque", "controle": "Controle",
            "celula": "Célula", "rele": "Relé"}


def data_pt(s):
    try:
        return datetime.strptime(s, "%d/%m/%Y")
    except ValueError:
        return s


# ----------------------------------------------------------------------------- estilo
def cabecalho(ws, linha, titulos, larguras=None):
    for i, t in enumerate(titulos, 1):
        c = ws.cell(row=linha, column=i, value=t)
        c.font = Font(bold=True, color=PAPEL, size=10)
        c.fill = PatternFill("solid", fgColor=TINTA)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    if larguras:
        for i, w in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[linha].height = 30


def marca_entrada(ws, col, r0, r1, texto_=False):
    """Campo de formulário: fundo papel-sombreado, é onde o gestor digita. `texto_` fixa
    o formato Texto (@) — a escolha «13,8» na lista fica texto, não número."""
    for r in range(r0, r1 + 1):
        c = ws.cell(row=r, column=col)
        c.fill = PatternFill("solid", fgColor=SOMBRA)
        if texto_:
            c.number_format = "@"


def marca_apoio(ws, col, r0, r1):
    for r in range(r0, r1 + 1):
        ws.cell(row=r, column=col).font = Font(color=APAGADA, size=8)


def moeda(ws, col, r0, r1):
    for r in range(r0, r1 + 1):
        ws.cell(row=r, column=col).number_format = MOEDA


def col_de(ws, letra):
    return ws[letra + "1"].column


def lista(ws, ref, sqref, titulo=None):
    # no XML a fórmula da validação vai SEM o «=» inicial
    dv = DataValidation(type="list", formula1=ref.lstrip("="), allow_blank=True, showDropDown=False)
    dv.error = "Escolha um valor da lista"
    dv.errorTitle = titulo or "Valor fora da lista"
    dv.errorStyle = "stop"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(sqref)
    return dv


def valida_qtd(ws, sqref):
    dv = DataValidation(type="whole", operator="greaterThan", formula1="0", allow_blank=True)
    dv.error, dv.errorTitle = "Quantidade inteira, maior que zero", "Quantidade"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(sqref)


# ----------------------------------------------------------------------------- abas
def aba_catalogo(wb, catalogo):
    ws = wb.active
    ws.title = "Catálogo"
    ws["A1"] = "CATÁLOGO DE PEÇAS — o único lugar para mexer em preço e código"
    ws["A1"].font = Font(bold=True, size=12, color=SINAL)
    ws["A2"] = ("Chave = Tipo | Peça | Tensão | kVA (kVA só na célula e no regulador completo). Mão de obra "
                "é por serviço, uma vez por lançamento. Os «completos» são fórmula sobre as peças: mude o "
                "preço da peça e o completo acompanha. «Conta na taxa» diz se a peça entra na taxa de falha.")
    ws["A2"].font = Font(italic=True, size=9)
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 40
    cab = ["Chave", "Tipo", "Peça", "Tensão (kV)", "kVA", "Código de material", "Descrição",
           "Material unitário (R$)", "Mão de obra (R$)", "Conta na taxa", "Fonte", "Observação"]
    cabecalho(ws, 4, cab, [26, 6, 22, 10, 6, 20, 52, 16, 15, 10, 60, 60])
    # primeiro os números de linha, para os completos apontarem para as peças
    linhas = {}
    r = 5
    for it in catalogo:
        k = (it["tipo"], it["peca"], it["tensao"], it["kva"])
        linhas[k] = f"H{r}"
        linhas[k + ("mo",)] = f"I{r}"
        r += 1
    r = 5
    for it in catalogo:
        mat = it["mat"][1](linhas) if isinstance(it["mat"], tuple) else it["mat"]
        mo = it["mo"][1](linhas) if isinstance(it["mo"], tuple) else it["mo"]
        ws.append([chave(it["tipo"], it["peca"], it["tensao"], it["kva"]), it["tipo"], it["peca"],
                   it["tensao"], it["kva"] or None, it["cod"] or None, it["desc"], mat, mo, it["taxa"],
                   it["fonte"], it["obs"] or None])
        for col in (8, 9):
            ws.cell(row=r, column=col).number_format = MOEDA
            if not isinstance((it["mat"] if col == 8 else it["mo"]), tuple):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=SOMBRA)
        ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor=SOMBRA)
        for col in (4, 5, 6):
            ws.cell(row=r, column=col).number_format = "@"
        r += 1
    fim = r - 1
    ws.freeze_panes = "A5"

    # listas de validação, num bloco à direita
    col0 = 14
    listas = [("Pecas_RL", "Peças RL", PECAS_RL), ("Pecas_RT", "Peças RT", PECAS_RT),
              ("Pecas_Todas", "Peças (todas)", PECAS_TODAS), ("Tensoes", "Tensão", TENSOES),
              ("Potencias", "Potência kVA", [str(k) for k in CLASSES_KVA]),
              ("Criticidades", "Criticidade", CRITICIDADES), ("Status_DCMD", "Status", STATUS),
              ("Sim_Nao", "Sim / não", SIM_NAO)]
    ws.cell(row=3, column=col0, value="LISTAS DE VALIDAÇÃO (as caixas de seleção leem daqui)").font = \
        Font(bold=True, size=10, color=SINAL)
    refs = {}
    for j, (nome, rotulo, valores) in enumerate(listas):
        col = col0 + j
        letra = get_column_letter(col)
        c = ws.cell(row=4, column=col, value=rotulo)
        c.font, c.fill = Font(bold=True, color=PAPEL, size=10), PatternFill("solid", fgColor=TINTA)
        for i, v in enumerate(valores):
            ws.cell(row=5 + i, column=col, value=v).number_format = "@"
        ws.column_dimensions[letra].width = max(14, len(rotulo) + 2)
        ref = f"'Catálogo'!${letra}$5:${letra}${4 + len(valores)}"
        refs[nome] = ref
        wb.defined_names[nome] = DefinedName(nome, attr_text=ref)
    refs["_col_causas"] = col0 + len(listas)
    return fim, refs


def aba_cadastro(wb, cad, extras):
    ws = wb.create_sheet("Cadastro")
    cab = ["Ativo", "Tipo", "Tensão (kV)", "Modelo", "Potência (kVA, cadastro)",
           "Classe kVA (167/200/400)", "Localidade", "Alimentador", "Origem", "Aviso"]
    cabecalho(ws, 1, cab, [13, 6, 11, 26, 16, 12, 24, 14, 30, 60])
    r = 2
    for a in sorted(cad, key=lambda x: (cad[x]["tipo"], x)):
        d = cad[a]
        ws.append([a, d["tipo"], d["tensao"] or None, d["modelo"] or None, d["potencia"] or None,
                   d["classe"] or None, d["localidade"] or None, d["alimentador"] or None,
                   d["origem"], d["obs"] or None])
        r += 1
    for a, d in extras:
        ws.append([a, d["tipo"], d["tensao"] or None, d["modelo"] or None, d["potencia"] or None,
                   d["classe"] or None, d["localidade"] or None, d["alimentador"] or None,
                   d["origem"], d["obs"] or None])
        r += 1
    fim = r - 1
    for rr in range(2, fim + 1):
        for col in (1, 3, 5, 6):
            ws.cell(row=rr, column=col).number_format = "@"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{fim}"
    return fim


def formulas_auto(r, col):
    """As fórmulas do bloco automático. `col` dá a letra de cada campo nesta aba:
    ativo, peca, qtd, tipo, tensao, kva, codigo, desc, preco, mat, mo, total, taxa,
    tensao_manual, kva_manual, modelo, aviso, localidade, chave."""
    A = col["ativo"]
    CAD, CAT = "'Cadastro'!", "'Catálogo'!"
    cadA = f"{CAD}$A$2:$A${CAD_MAX}"
    catA = f"{CAT}$A$5:$A${CAT_MAX}"

    def cad_txt(letra):
        # T() devolve texto ou "" — nunca o 0 que INDEX dá em célula vazia
        return (f'IFERROR(T(INDEX({CAD}${letra}$2:${letra}${CAD_MAX},MATCH({A}{r}&"",{cadA},0))),"")')

    def cat_txt(letra):
        return (f'IFERROR(T(INDEX({CAT}${letra}$5:${letra}${CAT_MAX},MATCH({col["chave"]}{r},{catA},0))),"")')

    def cat_num(letra):
        return (f'IFERROR(INDEX({CAT}${letra}$5:${letra}${CAT_MAX},MATCH({col["chave"]}{r},{catA},0)),"")')

    tipo, peca, tensao, kva, chave_, codigo, preco, mat, mo, modelo = (
        col["tipo"], col["peca"], col["tensao"], col["kva"], col["chave"], col["codigo"],
        col["preco"], col["mat"], col["mo"], col["modelo"])
    f = {}
    f["tipo"] = (f'=IF({A}{r}="","",IF(LEFT({A}{r}&"",2)="58","RT",'
                 f'IF(OR(LEFT({A}{r}&"",2)="79",LEFT({A}{r}&"",2)="78"),"RL","FORA DO ESCOPO")))')
    f["tensao"] = (f'=IF({col["tensao_manual"]}{r}<>"",SUBSTITUTE({col["tensao_manual"]}{r}&"",".",","),'
                   f'{cad_txt("C")})')
    f["kva"] = (f'=IF({tipo}{r}<>"RT","",IF({col["kva_manual"]}{r}<>"",{col["kva_manual"]}{r}&"",{cad_txt("F")}))')
    f["modelo"] = f"={cad_txt('D')}"
    f["aviso"] = f"={cad_txt('J')}"
    f["localidade"] = f"={cad_txt('G')}"
    f["chave"] = (f'=IF(OR({A}{r}="",{peca}{r}="",{tipo}{r}="",{tipo}{r}="FORA DO ESCOPO"),"",'
                  f'{tipo}{r}&"|"&{peca}{r}&"|"&{tensao}{r}&"|"&'
                  f'IF(AND({tipo}{r}="RT",OR({peca}{r}="Célula",{peca}{r}="Completo")),{kva}{r},""))')
    f["codigo"] = f'=IF({chave_}{r}="","",{cat_txt("F")})'
    f["desc"] = (
        f'=IF({A}{r}="","",'
        f'IF({tipo}{r}="FORA DO ESCOPO","PREFIXO FORA DO ESCOPO — só 58 (regulador), 78 e 79 (religador)",'
        f'IF({peca}{r}="","",'
        f'IF({tensao}{r}="","ATIVO NÃO ESTÁ NO CADASTRO — preencha Tensão manual (e Potência manual se for regulador)",'
        f'IF(AND({tipo}{r}="RT",OR({peca}{r}="Célula",{peca}{r}="Completo"),{kva}{r}=""),'
        f'"FALTA A POTÊNCIA — preencha Potência manual",'
        f'IF({cat_txt("G")}="","SEM LINHA NO CATÁLOGO PARA ESTA PEÇA / TENSÃO / POTÊNCIA",'
        f'{cat_txt("G")}&IF({codigo}{r}<>""," · cód. "&{codigo}{r},"")'
        f'&IF({modelo}{r}<>""," · ativo "&{modelo}{r},"")))))))')
    f["preco"] = f'=IF({chave_}{r}="","",{cat_num("H")})'
    f["mat"] = f'=IF({preco}{r}="","",{preco}{r}*IF(ISNUMBER({col["qtd"]}{r}),{col["qtd"]}{r},1))'
    f["mo"] = f'=IF({chave_}{r}="","",{cat_num("I")})'
    f["total"] = f'=IF({mat}{r}="","",{mat}{r}+IF(ISNUMBER({mo}{r}),{mo}{r},0))'
    f["taxa"] = f'=IF({chave_}{r}="","",{cat_txt("J")})'
    return f


def escreve_auto(ws, r, col):
    f = formulas_auto(r, col)
    for campo, letra in col.items():
        if campo in f:
            ws[f"{letra}{r}"] = f[campo]


def aba_lancamento(wb, refs, exemplos):
    ws = wb.create_sheet("Lançamento")
    ws["A1"] = "LANÇAMENTO — digite o ativo, escolha a peça, dê a quantidade. O resto é fórmula."
    ws["A1"].font = Font(bold=True, size=12, color=SINAL)
    ws["A2"] = ("Campos sombreados são de entrada. Tensão e potência saem do Cadastro; se faltar ou estiver "
                "errado, «Tensão manual» e «Potência manual» mandam. Quantidade vazia vale 1. Mão de obra "
                "entra uma vez por linha. Se a planilha abrir em Modo de Exibição Protegido, clique em "
                "«Habilitar Edição» para as fórmulas recalcularem.")
    ws["A2"].font = Font(italic=True, size=9)
    cab = ["Ativo", "Peça", "Qtd", "Tipo", "Tensão (kV)", "kVA", "Código de material", "Descrição",
           "Preço unitário (R$)", "Material (R$)", "Mão de obra (R$)", "Total (R$)", "Conta na taxa",
           "Tensão manual", "Potência manual", "Modelo do ativo", "Aviso do cadastro", "Localidade",
           "SS", "Data", "Observação", "Chave"]
    cabecalho(ws, 4, cab, [13, 22, 6, 9, 10, 6, 18, 64, 16, 15, 15, 15, 9, 11, 11, 20, 40, 22, 20, 11, 34, 26])
    col = dict(ativo="A", peca="B", qtd="C", tipo="D", tensao="E", kva="F", codigo="G", desc="H",
               preco="I", mat="J", mo="K", total="L", taxa="M", tensao_manual="N", kva_manual="O",
               modelo="P", aviso="Q", localidade="R", chave="V")
    r0, r1 = 5, 4 + N_LANC
    for r in range(r0, r1 + 1):
        escreve_auto(ws, r, col)
    for i, (ativo, peca, qtd, obs) in enumerate(exemplos):
        r = r0 + i
        ws[f"A{r}"], ws[f"B{r}"], ws[f"C{r}"], ws[f"U{r}"] = ativo, peca, qtd, obs
    for c in ("A", "B", "C", "S", "T", "U"):
        marca_entrada(ws, col_de(ws, c), r0, r1, texto_=(c == "A"))
    for c in ("N", "O"):
        marca_entrada(ws, col_de(ws, c), r0, r1, texto_=True)
    marca_apoio(ws, col_de(ws, "V"), r0, r1)
    for c in ("I", "J", "K", "L"):
        moeda(ws, col_de(ws, c), r0, r1)
    for r in range(r0, r1 + 1):
        ws[f"T{r}"].number_format = "dd/mm/yyyy"
    ws.freeze_panes = "C5"
    lista(ws, f'=INDIRECT(IF($D5="","Pecas_Todas","Pecas_"&$D5))', f"B{r0}:B{r1}", "Peça")
    lista(ws, refs["Tensoes"], f"N{r0}:N{r1}", "Tensão")
    lista(ws, refs["Potencias"], f"O{r0}:O{r1}", "Potência")
    valida_qtd(ws, f"C{r0}:C{r1}")
    tab = Table(displayName="Lancamento", ref=f"A4:V{r1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
    ws.add_table(tab)
    return ws


def aba_gestao(wb, gestao, cad, refs):
    ws = wb.create_sheet("Gestão")
    ws["A1"] = "GESTÃO — os 53 pendentes do DCMD, com a peça em lista e o orçamento em fórmula"
    ws["A1"].font = Font(bold=True, size=12, color=SINAL)
    ws["A2"] = ("Peça e Qtd vieram do campo «Defeito» da carteira de 27/08 (quantidade 1 em todos, "
                "que é como a carteira orçou). As colunas «carteira» guardam o que ela dizia; "
                "«Bate?» confere o total novo contra o dela. Totais na linha 3.")
    ws["A2"].font = Font(italic=True, size=9)
    cab = ["Ativo", "Tipo", "SS SGM", "Criticidade", "Status", "Peça", "Qtd", "Tensão (kV)", "kVA",
           "Código de material", "Descrição", "Preço unitário (R$)", "Orçamento MAT (R$)",
           "Orçamento MO (R$)", "Orçamento Total (R$)", "Conta na taxa", "Tensão manual", "Potência manual",
           "Modelo do ativo", "Aviso do cadastro", "Defeito (carteira 27/08)", "MAT carteira (R$)",
           "MO carteira (R$)", "Total carteira (R$)", "Bate?",
           "PMA", "Entregue N1?", "Gerado Obra?", "Gerado EMD?", "Entregue N3?", "Concluído COCM?",
           "Enviado para Cadastro", "Estudo Proteção", "Repassado ao DMSL?", "Comissionado?",
           "Índice", "SLA_Total", "Dias Pendente", "Status Atendimento", "Responsável",
           "Status Prazo", "Alimentador", "Município", "Polo", "Regional", "Localidade (cadastro)", "Chave"]
    larg = [13, 8, 21, 12, 20, 20, 5, 10, 6, 18, 60, 15, 15, 15, 15, 9, 11, 11, 20, 40, 22, 14, 14, 14, 7] + \
           [10] * 10 + [7, 9, 11, 26, 12, 24, 30, 20, 18, 10, 22, 26]
    cabecalho(ws, 4, cab, larg)
    col = dict(ativo="A", peca="F", qtd="G", tipo="B", tensao="H", kva="I", codigo="J", desc="K",
               preco="L", mat="M", mo="N", total="O", taxa="P", tensao_manual="Q", kva_manual="R",
               modelo="S", aviso="T", localidade="AT", chave="AU")
    r = 5
    for g in gestao:
        peca, kva, tensao = interpreta_defeito(g["defeito"])
        ws[f"A{r}"], ws[f"C{r}"], ws[f"D{r}"], ws[f"E{r}"] = g["ativo"], g["ss"], g["criticidade"], g["status"]
        ws[f"F{r}"], ws[f"G{r}"] = peca, 1
        c = cad.get(g["ativo"])
        # só quando o cadastro EXISTE e discorda da carteira, a carteira manda pela coluna manual
        if c and tensao and c["tensao"] != tensao:
            ws[f"Q{r}"] = tensao
        if c and kva and peca in ("Célula", "Completo") and c["classe"] != kva:
            ws[f"R{r}"] = kva
        ws[f"U{r}"], ws[f"V{r}"], ws[f"W{r}"], ws[f"X{r}"] = g["defeito"], g["mat"], g["mo"], g["total"]
        ws[f"Y{r}"] = f'=IF(O{r}="","",IF(ABS(O{r}-X{r})<0.01,"sim","NÃO"))'
        for i, v in enumerate(g["esteira"]):
            ws.cell(row=r, column=26 + i, value=v)
        for i, v in enumerate([g["indice"], g["sla"], g["dias"], g["atendimento"], g["resp"],
                               g["prazo"], g["alimentador"], g["municipio"], g["polo"], g["regional"]]):
            ws.cell(row=r, column=36 + i, value=v)
        escreve_auto(ws, r, col)
        r += 1
    r1 = r - 1
    for c in ("D", "E", "F", "G"):
        marca_entrada(ws, col_de(ws, c), 5, r1)
    for c in ("Q", "R"):
        marca_entrada(ws, col_de(ws, c), 5, r1, texto_=True)
    for i in range(26, 36):
        marca_entrada(ws, i, 5, r1)
    for c in ("L", "M", "N", "O", "V", "W", "X"):
        moeda(ws, col_de(ws, c), 5, r1)
    for rr in range(5, r1 + 1):
        ws[f"A{rr}"].number_format = "@"
    marca_apoio(ws, col_de(ws, "AU"), 5, r1)
    ws.freeze_panes = "C5"
    lista(ws, f'=INDIRECT(IF($B5="","Pecas_Todas","Pecas_"&$B5))', f"F5:F{r1}", "Peça")
    lista(ws, refs["Criticidades"], f"D5:D{r1}", "Criticidade")
    lista(ws, refs["Status_DCMD"], f"E5:E{r1}", "Status")
    lista(ws, refs["Tensoes"], f"Q5:Q{r1}", "Tensão")
    lista(ws, refs["Potencias"], f"R5:R{r1}", "Potência")
    valida_qtd(ws, f"G5:G{r1}")
    # totais na linha 3, acima do cabeçalho, sobre um intervalo folgado (linhas novas entram)
    ws["K3"] = "TOTAL dos pendentes →"
    ws["K3"].font = Font(bold=True)
    ws["K3"].alignment = Alignment(horizontal="right")
    for c in ("M", "N", "O", "V", "W", "X"):
        ws[f"{c}3"] = f"=SUM({c}5:{c}{FALHA_MAX})"
        ws[f"{c}3"].number_format = MOEDA
        ws[f"{c}3"].font = Font(bold=True)
    ws["Y3"] = f'=COUNTIF(Y5:Y{FALHA_MAX},"sim")&" de "&COUNTA(A5:A{FALHA_MAX})'
    ws["Y3"].font = Font(bold=True)
    tab = Table(displayName="Gestao", ref=f"A4:AU{r1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
    ws.add_table(tab)
    return r1


def aba_falhas(wb, falhas, gestao, refs):
    ws = wb.create_sheet("Falha Equipamentos")
    ws["A1"] = "FALHA EQUIPAMENTOS — o rol de 2025 e 2026 com a peça em lista e o custo em fórmula"
    ws["A1"].font = Font(bold=True, size=12, color=SINAL)
    ws["A2"] = ("«Peça (rol)» é a classificação original; «Peça (custo)» é a que se orça e alimenta o Resumo. "
                "Furto é decidido pela peça: onde a carteira já orçou o ativo, a peça veio de lá (Gestão); onde "
                "não, ficou em branco para definir. Relé de regulador virou Controle (relé de sincronismo é "
                "controle). O Ano é o da fatia, não o da data — o rol fixou o 7933585074 em 2025.")
    ws["A2"].font = Font(italic=True, size=9)
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 40
    cab = ["Fatia (ano/tipo)", "Tipo", "Ativo", "SS", "Data", "Mês", "Ano (fatia)", "Peça (rol)",
           "Peça (custo)", "Qtd", "Troca feita", "Causa raiz", "Tensão (kV)", "kVA", "Código de material",
           "Descrição", "Preço unitário (R$)", "Material (R$)", "Mão de obra (R$)", "Total (R$)",
           "Conta na taxa", "Tensão manual", "Potência manual", "Modelo do ativo", "Aviso do cadastro",
           "Observação", "Citação do texto da SS", "Nota do analista", "Revisão", "Localidade (cadastro)", "Chave"]
    larg = [11, 6, 13, 21, 11, 6, 8, 11, 22, 5, 10, 34, 10, 6, 18, 60, 15, 15, 15, 15, 9, 11, 11, 20, 40,
            50, 60, 60, 40, 22, 26]
    cabecalho(ws, 4, cab, larg)
    col = dict(ativo="C", peca="I", qtd="J", tipo="B", tensao="M", kva="N", codigo="O", desc="P",
               preco="Q", mat="R", mo="S", total="T", taxa="U", tensao_manual="V", kva_manual="W",
               modelo="X", aviso="Y", localidade="AD", chave="AE")
    carteira = {g["ativo"]: g for g in gestao}
    # nos furtos, a peça vem da carteira só na falha mais recente do ativo
    ultima_falha = {}
    for i, f in enumerate(falhas):
        if f["peca"] == "furto":
            ultima_falha[f["ativo"]] = i
    r = 5
    causas = []
    for i, f in enumerate(falhas):
        ws[f"A{r}"], ws[f"C{r}"], ws[f"D{r}"] = f["fatia"], f["ativo"], f["ss"]
        ws[f"A{r}"].number_format = ws[f"C{r}"].number_format = "@"
        ws[f"B{r}"] = f'=LEFT(A{r},2)'
        d = data_pt(f["data"])
        ws[f"E{r}"] = d
        if isinstance(d, datetime):
            ws[f"E{r}"].number_format = "dd/mm/yyyy"
            ws[f"F{r}"] = f"=MONTH(E{r})"
        ws[f"G{r}"] = f"=VALUE(RIGHT(A{r},4))"
        ws[f"H{r}"] = f["peca"]
        tipo = f["fatia"][:2]
        obs = []
        if f["peca"] == "furto":
            g = carteira.get(f["ativo"])
            if g and ultima_falha.get(f["ativo"]) == i:
                peca_custo, kva, tensao = interpreta_defeito(g["defeito"])
                ws[f"I{r}"], ws[f"J{r}"] = peca_custo, 1
                obs.append(f"furto: peça pela carteira de 27/08 («{g['defeito']}», SS {g['ss']})")
            else:
                obs.append("furto: a carteira não orçou esta demanda — definir a peça e a quantidade")
        else:
            peca_custo = PECA_ROL.get(f["peca"], "")
            if f["peca"] == "rele" and tipo == "RT":
                peca_custo = "Controle"
                obs.append("relé de regulador orçado como controle (relé de sincronismo é controle)")
            ws[f"I{r}"], ws[f"J{r}"] = peca_custo, 1
        if f["ativo"] == "7947203070":
            obs.append("a carteira orça «RL Completo 34,5» (DMSL: substituir o conjunto); o rol diz controle")
        if f["ativo"] == "7933585074":
            obs.append("ocorrência em 27/01/2026, mas o rol fixou a falha em 2025 pela fatia")
        ws[f"K{r}"], ws[f"L{r}"] = f["troca"], f["causa"]
        if f["causa"] and f["causa"] not in causas:
            causas.append(f["causa"])
        ws[f"Z{r}"] = "; ".join(obs) or None
        ws[f"AA{r}"], ws[f"AB{r}"], ws[f"AC{r}"] = f["citacao"], f["nota"], f["revisao"]
        escreve_auto(ws, r, col)
        r += 1
    r1 = r - 1
    for c in ("I", "J", "K", "L"):
        marca_entrada(ws, col_de(ws, c), 5, r1)
    for c in ("V", "W"):
        marca_entrada(ws, col_de(ws, c), 5, r1, texto_=True)
    for c in ("Q", "R", "S", "T"):
        moeda(ws, col_de(ws, c), 5, r1)
    marca_apoio(ws, col_de(ws, "AE"), 5, r1)
    ws.freeze_panes = "D5"
    wc = wb["Catálogo"]
    colc = refs["_col_causas"]
    c = wc.cell(row=4, column=colc, value="Causa raiz")
    c.font, c.fill = Font(bold=True, color=PAPEL, size=10), PatternFill("solid", fgColor=TINTA)
    for i, v in enumerate(sorted(causas)):
        wc.cell(row=5 + i, column=colc, value=v)
    wc.column_dimensions[get_column_letter(colc)].width = 44
    ref_causas = f"'Catálogo'!${get_column_letter(colc)}$5:${get_column_letter(colc)}${4 + len(causas)}"
    wb.defined_names["Causas"] = DefinedName("Causas", attr_text=ref_causas)
    lista(ws, f'=INDIRECT(IF($B5="","Pecas_Todas","Pecas_"&$B5))', f"I5:I{r1}", "Peça")
    lista(ws, refs["Sim_Nao"], f"K5:K{r1}", "Troca feita")
    lista(ws, ref_causas, f"L5:L{r1}", "Causa raiz")
    lista(ws, refs["Tensoes"], f"V5:V{r1}", "Tensão")
    lista(ws, refs["Potencias"], f"W5:W{r1}", "Potência")
    valida_qtd(ws, f"J5:J{r1}")
    tab = Table(displayName="Falhas", ref=f"A4:AE{r1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
    ws.add_table(tab)
    return r1


def aba_resumo(wb):
    ws = wb.create_sheet("Resumo")
    ws["A1"] = "RESUMO — contagem por peça e taxa, em fórmula sobre a aba Falha Equipamentos"
    ws["A1"].font = Font(bold=True, size=12, color=SINAL)
    ws["A2"] = ("Conta linhas da aba Falha Equipamentos pela «Peça (custo)» (a coluna com lista), pelo Tipo "
                "e pelo Ano da fatia. Por peça; por equipamento, ativo repetido no mesmo ano conta uma vez — "
                "90 linhas são 87 pares ativo-ano. Parque é editável.")
    ws["A2"].font = Font(italic=True, size=9)
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 40
    ws["A4"], ws["B4"] = "Parque RL", 1307
    ws["A5"], ws["B5"] = "Parque RT", 207
    for r in (4, 5):
        ws[f"B{r}"].fill = PatternFill("solid", fgColor=SOMBRA)
        ws[f"A{r}"].font = Font(bold=True)
    cab = ["Peça (custo)", "RL 2025", "RL 2026", "RT 2025", "RT 2026"]
    cabecalho(ws, 7, cab, [26, 12, 12, 12, 12])
    F = "'Falha Equipamentos'!"
    rB, rG, rI, rT = (f"{F}$B$5:$B${FALHA_MAX}", f"{F}$G$5:$G${FALHA_MAX}",
                      f"{F}$I$5:$I${FALHA_MAX}", f"{F}$T$5:$T${FALHA_MAX}")
    fatias = [("RL", 2025), ("RL", 2026), ("RT", 2025), ("RT", 2026)]
    r = 8
    for p in PECAS_TODAS:
        ws[f"A{r}"] = p
        for j, (tipo, ano) in enumerate(fatias):
            ws.cell(row=r, column=2 + j, value=f'=COUNTIFS({rB},"{tipo}",{rG},{ano},{rI},"{p}")')
        r += 1
    ws[f"A{r}"] = "(sem peça definida)"
    for j, (tipo, ano) in enumerate(fatias):
        ws.cell(row=r, column=2 + j, value=f'=COUNTIFS({rB},"{tipo}",{rG},{ano},{rI},"")')
    r += 1
    r_tot = r
    ws[f"A{r}"] = "Falhas (linhas)"
    for j in range(4):
        L = get_column_letter(2 + j)
        ws[f"{L}{r}"] = f"=SUM({L}8:{L}{r - 1})"
        ws[f"{L}{r}"].font = Font(bold=True)
    ws[f"A{r}"].font = Font(bold=True)
    r += 1
    ws[f"A{r}"] = "Taxa sobre o parque"
    for j, (tipo, _) in enumerate(fatias):
        L = get_column_letter(2 + j)
        parque = "$B$4" if tipo == "RL" else "$B$5"
        ws[f"{L}{r}"] = f'=IF(ISNUMBER({parque}),IF({parque}>0,{L}{r_tot}/{parque},""),"")'
        ws[f"{L}{r}"].number_format = "0.0%"
    r += 1
    ws[f"A{r}"] = "Custo estimado (R$)"
    for j, (tipo, ano) in enumerate(fatias):
        L = get_column_letter(2 + j)
        ws[f"{L}{r}"] = f'=SUMIFS({rT},{rB},"{tipo}",{rG},{ano})'
        ws[f"{L}{r}"].number_format = MOEDA
    return r_tot


def aba_como(wb, avisos, n_cad_rl, n_cad_rt):
    ws = wb.create_sheet("Como foi feito")
    ws.column_dimensions["A"].width = 110
    linhas = [
        ("COMO ESTA PLANILHA FUNCIONA", True),
        ("As tabelas de apoio moram aqui dentro — Catálogo, Cadastro e as listas de validação —, então nada "
         "depende de arquivo externo. A planilha base fazia o mesmo com XLOOKUP na carteira do SharePoint e "
         "na GESTÃO DE EQUIPAMENTOS.xlsx; fora da rede esses vínculos ficam congelados.", False),
        ("Se o Excel abrir em Modo de Exibição Protegido (anexo de e-mail, download), os valores mostrados são "
         "os gravados no arquivo; clique em «Habilitar Edição» para as fórmulas recalcularem.", False),
        ("", False),
        ("O FLUXO, EM QUALQUER ABA COM PEÇA", True),
        ("1) O ativo diz o tipo pelo prefixo: 58 regulador; 78 e 79 religador. Outro prefixo (59 capacitor, BR "
         "reator) dá «PREFIXO FORA DO ESCOPO».", False),
        (f"2) O Cadastro dá tensão, modelo e potência do ativo (ajustes da proteção). Cobre {n_cad_rl} de 1.307 "
         f"religadores e {n_cad_rt} de 207 reguladores. Ativo fora dele dá «ATIVO NÃO ESTÁ NO CADASTRO»: "
         "preencha «Tensão manual» (e «Potência manual» se for regulador) — as manuais sempre mandam.", False),
        ("3) A peça escolhida na lista (a lista muda conforme o tipo), mais a tensão e, na célula ou no regulador "
         "completo, a potência, montam a chave «RL|Tanque|34,5|» ou «RT|Célula|34,5|400» e buscam no Catálogo "
         "o código, a descrição, o preço de material e a mão de obra.", False),
        ("4) Material = preço unitário × quantidade (vazia ou não numérica vale 1). Mão de obra entra uma vez "
         "por linha. Total = material + mão de obra.", False),
        ("5) A descrição sai como «Tanque / Parte ativa — Religador 34,5 kV · cód. 690005 · ativo NOJA RC10». O "
         "código é o item de estoque, que não muda com a marca do que falhou; «ativo …» é o modelo do "
         "equipamento no cadastro (o tanque comprado para um Cooper é o mesmo 690005).", False),
        ("6) «Aviso do cadastro» aparece quando a potência do ativo está fora das classes 167/200/400 ou falta "
         "tensão — conferir a peça antes de fechar o orçamento.", False),
        ("", False),
        ("DE ONDE VÊM OS PREÇOS", True),
        ("Vale a carteira de 27/08 (coluna RL Auto / Valor Material / Valor mão de obra estimado da aba "
         "«Criticidade por Equipamento», lida do cache da planilha base). É a que o gestor usa: célula de "
         "400 kVA a R$ 90.230, célula de 200 a R$ 51.705,75 (34,5) e R$ 62.113,20 (13,8), controle de regulador "
         "a R$ 23.259,60 com R$ 20.000 de mão de obra, regulador completo de 400 a R$ 293.949,60 + R$ 200.637.", False),
        ("Onde a carteira não tem linha vale «Premissas e Preços» do ORCAMENTO_EQ_ESPECIAIS (16/07): célula "
         "de 167 kVA (690669, R$ 27.616,62), relé 50/51 (625510, R$ 10.454,01), placa RC10 (647641, R$ 1.999), "
         "bateria + rádio (647877 + 649399, R$ 14.612,08), chave faca (90556, R$ 736,43).", False),
        ("No religador as duas fontes e o plano de compras batem no centavo. No regulador a carteira de 27/08 "
         "é mais barata que o orçamento de 16/07 (célula de 400: 90.230 contra 126.893,80; de 200 em 34,5: "
         "51.705,75 contra 57.720,41; controle: 23.259,60 contra 29.445,39). A coluna Observação do Catálogo "
         "guarda o valor antigo.", False),
        ("Mão de obra por serviço: religador 13,8 kV R$ 11.016,94 e 34,5 kV R$ 13.209,34; célula de regulador "
         "13,8 kV R$ 51.402,14 e 34,5 kV R$ 80.318,50; controle de regulador R$ 20.000; acessório R$ 0. Regulador "
         "completo: 2 × (célula + controle) — é assim que a carteira chega aos R$ 200.637 do banco de 400 em "
         "34,5; nas outras classes o valor é derivado dessa regra, a confirmar.", False),
        ("Os «completos» do Catálogo são fórmula (tanque + controle; 3 × célula + controle): mudar o preço da "
         "peça atualiza o completo.", False),
        ("", False),
        ("CLASSES DE POTÊNCIA", True),
        ("A célula de regulador é chaveada por tensão e potência nas três classes do gestor: 167, 200 e 400 kVA. "
         "Códigos: 690669 (13,8 / 167), 690236 (13,8 / 239 nominal — que o gestor chama de 200), 690240 "
         "(34,5 / 200), 690241 (34,5 / 400). O cadastro traz também 239, 250, 300, 398, 667 e bancos mistos; "
         "a coluna «Classe kVA» do Cadastro leva cada um para a classe mais próxima (398 → 400, 239 e 250 → 200) "
         "e a coluna Aviso marca o desvio — 667 e 300 kVA são os casos a conferir.", False),
        ("", False),
        ("O QUE FOI PRÉ-PREENCHIDO", True),
        ("Gestão: a peça e a quantidade (1) vieram do campo «Defeito» da carteira; onde o Cadastro discorda da "
         "carteira em tensão ou potência, a carteira foi posta na coluna manual, para o total bater. «Bate?» "
         "compara o total novo com o da carteira.", False),
        ("Falha Equipamentos: «Peça (custo)» vem da «Peça (rol)». Furto: na falha mais recente de ativo que a "
         "carteira orçou, a peça é a da carteira (5836786094 célula; 5856070091, 5856156091 e 5858783119 "
         "regulador completo); nas outras (5841308190, dois furtos; a falha de 2025 do 5856070091) ficou em "
         "branco para definir. Relé de regulador virou Controle. O 7947203070 aparece como controle no rol e "
         "como completo na carteira — a linha está anotada.", False),
        ("", False),
        ("PARA LEVAR PARA A PLANILHA BASE", True),
        ("Selecione as abas Catálogo, Cadastro e Lançamento juntas (Ctrl + clique), botão direito, «Mover ou "
         "copiar», destino GESTAO_EQUIPAMENTOS_ESPECIAIS_COEP.xlsx, marque «Criar uma cópia». Movidas juntas, as "
         "fórmulas entre elas continuam válidas e as listas de validação (nomes Pecas_RL, Pecas_RT, Tensoes, "
         "Potencias…) vão junto. Linha nova no Catálogo ou no Cadastro entra sozinha nas buscas (os intervalos "
         "vão até a linha 200 e 3.000).", False),
        ("", False),
        ("AVISOS", True),
    ] + [(a, False) for a in avisos] + [
        ("", False),
        ("As fórmulas foram conferidas por um motor de cálculo em Python (biblioteca formulas), não pelo "
         "Excel — LibreOffice não roda neste ambiente. Os valores em cache no arquivo vieram desse cálculo; "
         "ao habilitar a edição, o Excel recalcula tudo.", False),
    ]
    for i, (t, negrito) in enumerate(linhas, 1):
        c = ws.cell(row=i, column=1, value=t)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if negrito:
            c.font = Font(bold=True, size=11, color=SINAL if i == 1 else TINTA)
    return ws


# ----------------------------------------------------------------------------- cache de valores
def grava_cache(caminho):
    """Calcula as fórmulas com a biblioteca `formulas` e grava o resultado como valor em
    cache (<v>) ao lado de cada <f>, para o arquivo não abrir em branco no Modo de
    Exibição Protegido. openpyxl grava <v /> vazio em toda célula de fórmula."""
    import formulas
    sol = formulas.ExcelModel().loads(caminho).finish().calculate()
    nome = os.path.basename(caminho)
    valores = {}
    for k, v in sol.items():
        m = re.match(r"^'\[(.+?)\](.+?)'!([A-Z]+\d+)$", k, re.I)
        if not m or m.group(1).lower() != nome.lower():
            continue
        x = v.value
        try:
            x = x[0][0]
        except (TypeError, IndexError):
            pass
        valores[(m.group(2).upper(), m.group(3).upper())] = x

    def xml_valor(x):
        if x is None or (isinstance(x, str) and x == ""):
            return ' t="str"', ""
        if isinstance(x, bool):
            return ' t="b"', "1" if x else "0"
        if isinstance(x, (int, float)):
            return "", repr(round(float(x), 10)) if isinstance(x, float) else str(x)
        s = str(x)
        if s.startswith("#"):
            return ' t="e"', escape(s)
        return ' t="str"', escape(s)

    with zipfile.ZipFile(caminho) as z:
        partes = {n: z.read(n) for n in z.namelist()}
    wbxml = partes["xl/workbook.xml"].decode("utf-8")
    rels = partes["xl/_rels/workbook.xml.rels"].decode("utf-8")
    rid_alvo = dict(re.findall(r'<Relationship[^>]*Id="([^"]+)"[^>]*Target="([^"]+)"', rels))
    rid_alvo.update({k: v for v, k in re.findall(r'<Relationship[^>]*Target="([^"]+)"[^>]*Id="([^"]+)"', rels)})
    folhas = re.findall(r'<sheet [^>]*name="([^"]+)"[^>]*r:id="([^"]+)"', wbxml)
    trocados = 0
    for nome_aba, rid in folhas:
        alvo = rid_alvo[rid]
        parte = "xl/" + alvo if not alvo.startswith("/") else alvo.lstrip("/")
        xml = partes[parte].decode("utf-8")
        aba = nome_aba.replace("&amp;", "&").upper()

        def troca(m):
            nonlocal trocados
            ref = m.group(1).upper()
            if (aba, ref) not in valores:
                return m.group(0)
            t, v = xml_valor(valores[(aba, ref)])
            trocados += 1
            return f'<c r="{m.group(1)}"{m.group(2)}{t}><f>{m.group(3)}</f><v>{v}</v></c>'

        xml = re.sub(r'<c r="([A-Z]+\d+)"((?: [a-z]+="[^"]*")*)><f>(.*?)</f><v />(?:</c>)', troca, xml)
        partes[parte] = xml.encode("utf-8")
    with zipfile.ZipFile(caminho, "w", zipfile.ZIP_DEFLATED) as z:
        for n, b in partes.items():
            z.writestr(n, b)
    return trocados


# ----------------------------------------------------------------------------- montagem
def montar(saida=SAIDA, cache=True):
    cad = ler_cadastro()
    modelos = ler_modelos_carteira()
    gestao, falhas = ler_base()
    catalogo = monta_catalogo()
    avisos = []

    # ativos da planilha base que o cadastro dos ajustes não tem
    extras = []
    vistos = set()
    for g in gestao:
        a = g["ativo"]
        if a in cad or a in vistos:
            continue
        vistos.add(a)
        peca, kva, tensao = interpreta_defeito(g["defeito"])
        tipo = "RT" if a[:2] == "58" else "RL"
        obs = "sem cadastro nos ajustes da proteção; tensão e potência pelo texto do Defeito da carteira"
        if a == "7930359149":
            obs = ("sem cadastro nos ajustes; tensão pelo alimentador LD03414149 (34,5 kV, como os 11 RL do "
                   "trecho e os outros 6 de Caseara); modelo desconhecido")
        extras.append((a, {"tipo": tipo, "tensao": tensao, "modelo": modelos.get(a, ""),
                           "potencia": kva, "classe": kva, "localidade": g["municipio"],
                           "alimentador": g["alimentador"],
                           "origem": "planilha base (Gestão, cache da carteira 27/08)", "obs": obs}))
    for f in falhas:
        a = f["ativo"]
        if a in cad or a in vistos:
            continue
        vistos.add(a)
        tipo = f["fatia"][:2]
        tensao = f["tensao"]
        obs = "sem cadastro nos ajustes da proteção; tensão pelo rol de falhas"
        if not tensao and a == "7930359149":
            tensao, obs = "34,5", ("sem cadastro nos ajustes e «#N/A» no rol; tensão pelo alimentador LD03414149 "
                                   "(34,5 kV, como os 11 RL do trecho); modelo desconhecido")
        extras.append((a, {"tipo": tipo, "tensao": tensao, "modelo": modelos.get(a, ""),
                           "potencia": "", "classe": "", "localidade": "", "alimentador": "",
                           "origem": "planilha base (Falha Equipamentos)", "obs": obs}))
    if extras:
        avisos.append(f"{len(extras)} ativo(s) da planilha base não estão nos ajustes da proteção e "
                      f"entraram no Cadastro com origem marcada: " + ", ".join(a for a, _ in extras) + ".")
    desvios = [a for a, d in cad.items() if d["desvio"]]
    graves = [f"{a} ({cad[a]['potencia']})" for a in desvios
              if not any(x in cad[a]["potencia"] for x in ("398", "239", "250"))]
    avisos.append(f"{len(desvios)} reguladores têm potência fora das classes 167/200/400 e foram levados à "
                  f"classe mais próxima (398 → 400; 239 e 250 → 200), com aviso no Cadastro que aparece na "
                  f"coluna «Aviso do cadastro». Os mais distantes, a conferir: " + ", ".join(graves) + ".")
    sem_tensao = [a for a, d in cad.items() if not d["tensao"]]
    if sem_tensao:
        avisos.append(f"{len(sem_tensao)} ativo(s) do cadastro sem tensão: " + ", ".join(sem_tensao))
    n_rl = sum(1 for d in cad.values() if d["tipo"] == "RL") + sum(1 for _, d in extras if d["tipo"] == "RL")
    n_rt = sum(1 for d in cad.values() if d["tipo"] == "RT") + sum(1 for _, d in extras if d["tipo"] == "RT")
    avisos.append(f"O Cadastro cobre {n_rl} de 1.307 religadores e {n_rt} de 207 reguladores do parque; os "
                  f"demais dão «ATIVO NÃO ESTÁ NO CADASTRO» e usam as colunas manuais.")
    avisos.append("Acessório a R$ 15.000 sem mão de obra é o valor fechado de quatro ativos da carteira "
                  "(7942572059, 7931219078, 5846328094, 5853360007), não um preço de estoque.")

    wb = Workbook()
    cat_fim, refs = aba_catalogo(wb, catalogo)
    cad_fim = aba_cadastro(wb, cad, extras)
    assert cad_fim < CAD_MAX and cat_fim < CAT_MAX
    exemplos = [("7925744087", "Tanque", 1, "exemplo do gestor: tanque de religador 34,5 kV"),
                ("5840227063", "Célula", 2, "exemplo do gestor: duas células de 400 kVA"),
                ("7923673004", "Completo", 1, "exemplo: religador completo 13,8 kV (ativo ARTECHE P500)"),
                ("5803327001", "Controle", 1, "exemplo: controle de regulador")]
    aba_lancamento(wb, refs, exemplos)
    g1 = aba_gestao(wb, gestao, cad, refs)
    f1 = aba_falhas(wb, falhas, gestao, refs)
    assert f1 < FALHA_MAX
    aba_resumo(wb)
    aba_como(wb, avisos, n_rl, n_rt)
    os.makedirs(os.path.dirname(saida), exist_ok=True)
    wb.save(saida)
    print(f"{saida}: catálogo {cat_fim - 4} linhas · cadastro {cad_fim - 1} ativos "
          f"({len(extras)} fora dos ajustes) · gestão {g1 - 4} · falhas {f1 - 4} · "
          f"lançamento {N_LANC} linhas prontas")
    if cache:
        n = grava_cache(saida)
        print(f"cache de valores gravado em {n} células de fórmula")
    return saida


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    montar(args[0] if args else SAIDA, cache="--sem-cache" not in sys.argv)
