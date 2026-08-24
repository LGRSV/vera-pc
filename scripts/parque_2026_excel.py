"""
A visão de 2026 em Excel — uma aba por tipo, com os gráficos nativos.

Mesma leitura da página: o mês a mês e o acumulado com o acervo herdado. Aqui os
gráficos são do próprio Excel, então dá para mexer, refazer escala e colar em
apresentação.

Grava dist/PARQUE_2026.xlsx.
Rodar: python3 scripts/parque_2026_excel.py
"""

import json
import os

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SAIDA = os.path.join(RAIZ, "dist", "PARQUE_2026.xlsx")
NOME = {"RL": "Religador", "RT": "Regulador"}


def br(v, casas=0):
    """Número no jeito daqui: ponto no milhar, vírgula no decimal."""
    return f"{v:,.{casas}f}".replace(",", "·").replace(".", ",").replace("·", ".")

COLUNAS = [
    ("Mês", 10), ("Parque", 11), ("Expansão do mês", 15),
    ("Entrantes fora de operação", 24), ("Resolvidos no mês", 17),
    ("Falhas · peça grande", 19), ("Taxa de falha do mês (%)", 21),
    ("Indisponibilidade acumulada (só 2026)", 30), ("Entrantes com o acervo", 20),
    ("Resolvidos acumulados", 20), ("Resolvidos pelo COEP", 19), ("Fila", 9),
]


def escrever(ws, tipo, serie, cum):
    tit = Font(bold=True, color="FFFFFF", size=10)
    fundo = PatternFill("solid", fgColor="1F3864")
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    acervo = cum["acervo_em_janeiro"]

    ws.append([c[0] for c in COLUNAS])
    for i, (nome, larg) in enumerate(COLUNAS, 1):
        cel = ws.cell(row=1, column=i)
        cel.font, cel.fill = tit, fundo
        cel.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 34

    for d, m in zip(serie, cum["meses"]):
        ws.append([d["rotulo"], d["parque"], d["expansao"], d["entrantes"],
                   m["resolvidos_no_mes"], d["falhas"], d["taxa_mes_pct"] / 100,
                   m["entraram_acumulado"] - acervo, m["entraram_acumulado"],
                   m["resolvidos_acumulado"], m["resolvidos_coep_acumulado"], m["fila"]])
    ultima = ws.max_row
    for linha in ws.iter_rows(min_row=2, max_row=ultima):
        for cel in linha:
            cel.border = borda
            cel.alignment = Alignment(horizontal="center")
    for r in range(2, ultima + 1):
        ws.cell(row=r, column=7).number_format = "0,00%"

    # a linha do acervo, que é o ponto de partida do acumulado
    ws.cell(row=ultima + 2, column=1, value="Acervo em 1º de janeiro").font = Font(bold=True)
    ws.cell(row=ultima + 2, column=2, value=acervo).font = Font(bold=True)
    ws.cell(row=ultima + 3, column=1,
            value=f"{NOME[tipo]}es que já estavam fora de operação na virada do ano, "
                  f"herdados de 2024 e 2025. A conta do ano fecha: {acervo} + "
                  f"{cum['entraram_no_ano']} − {cum['resolvidos_no_ano']} = "
                  f"{cum['meses'][-1]['fila']}, o número de cadeias abertas hoje.")
    ws.cell(row=ultima + 3, column=1).alignment = Alignment(wrap_text=False)

    # a leitura das duas figuras, a mesma da página — para a planilha se explicar
    # sozinha em quem não abre o painel
    pico_ent = max(serie, key=lambda d: d["entrantes"])
    pico_taxa = max(serie, key=lambda d: d["taxa_mes_pct"])
    leitura = [
        f"Mês a mês — o parque sobe de {br(serie[0]['parque'])} para "
        f"{br(serie[-1]['parque'])} com a expansão. Quem saiu de operação contra quem foi "
        f"resolvido: pico de entrada em {pico_ent['rotulo']} ({pico_ent['entrantes']}). A "
        f"taxa de falha do mês é só peça grande sobre o parque daquele mês, com máxima de "
        f"{br(pico_taxa['taxa_mes_pct'], 2)}% em {pico_taxa['rotulo']}.",
        f"Acumulado — o ano não começa do zero: {acervo} equipamentos já estavam fora de "
        f"operação em 1º de janeiro. «Entrantes com o acervo» é a fila inteira; "
        f"«indisponibilidade acumulada» é só o que entrou em 2026, e a distância entre as "
        f"duas é exatamente o acervo. Resolvidos são {cum['resolvidos_no_ano']} cadeias "
        f"fechadas por qualquer posto, das quais "
        f"{cum['meses'][-1]['resolvidos_coep_acumulado']} pelo COEP.",
    ]
    for i, texto in enumerate(leitura):
        ws.cell(row=ultima + 4 + i, column=1, value=texto)
        ws.cell(row=ultima + 4 + i, column=1).alignment = Alignment(wrap_text=False)

    def grafico(titulo, colunas, ancora, altura=8, largura=17, pct=False):
        ch = LineChart()
        ch.title = titulo
        ch.height, ch.width = altura, largura
        ch.style = 2
        for col in colunas:
            ref = Reference(ws, min_col=col, min_row=1, max_row=ultima)
            ch.add_data(ref, titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ultima))
        ch.y_axis.majorGridlines = None if pct else ch.y_axis.majorGridlines
        if pct:
            ch.y_axis.numFmt = "0,00%"
        for s in ch.series:
            s.smooth = False
            s.marker.symbol = "circle"
            s.marker.size = 6
        ws.add_chart(ch, ancora)

    topo = ultima + 7
    grafico(f"{NOME[tipo]} · parque, mês a mês", [2], f"A{topo}")
    grafico(f"{NOME[tipo]} · entrantes fora de operação e resolvidos", [4, 5], f"J{topo}")
    grafico(f"{NOME[tipo]} · taxa de falha do mês (% do parque)", [7], f"A{topo + 16}", pct=True)
    grafico(f"{NOME[tipo]} · falhas do mês · peça grande", [6], f"J{topo + 16}")
    grafico(f"{NOME[tipo]} · acumulado no ano", [8, 9, 10, 11], f"A{topo + 32}")
    grafico(f"{NOME[tipo]} · fila no fim do mês", [12], f"J{topo + 32}")


COMO = [
    "PARQUE E FALHAS 2026 — religador e regulador, mês a mês e acumulado.",
    "",
    "Parque: base de janeiro do gestor (1.281 religadores e 180 reguladores) mais a "
    "expansão realizada, somada no próprio mês. Agosto repete julho porque a expansão "
    "do mês ainda não fechou.",
    "",
    "Entrantes fora de operação: equipamento com SS de INDISPONIBILIDADE PARA OPERAÇÃO, "
    "pela DATA DA OCORRÊNCIA — não a da abertura, que atrasa 39 dias em média e em 9,8% "
    "dos casos cai em outro ano.",
    "",
    "Resolvidos: a cadeia de SS fechada, atendida ou cancelada, por qualquer posto. "
    "Repasse não é demanda nova — a cadeia inteira conta uma vez. Cadeia cancelada não "
    "tem data de conclusão no SGM, então é datada pela abertura da última SS dela: é "
    "aproximação, e é o que existe.",
    "",
    "Resolvidos pelo COEP: a régua do posto — a demanda passou pelo COEP dentro de 2026 "
    "E a cadeia fechou dentro de 2026. São 82 no total: 55 religadores e 27 reguladores.",
    "",
    "Indisponibilidade acumulada (só 2026): o que entrou fora de operação no ano, "
    "somado mês a mês, sem o acervo.",
    "",
    "Entrantes com o acervo: a mesma soma partindo do que já estava fora de operação em "
    "1º de janeiro. A diferença entre as duas colunas é exatamente o acervo herdado — "
    "123 religadores e 32 reguladores.",
    "",
    "Fila: entrantes com acervo menos resolvidos. Fecha agosto em 158 religadores e 49 "
    "reguladores, que é o número de cadeias abertas hoje.",
    "",
    "Taxa de falha do mês: só PEÇA GRANDE pela régua do gestor — controle, tanque ou "
    "equipamento completo no religador; célula, relé, banco completo ou furto no "
    "regulador —, dividida pelo parque daquele mês. Fecha o ano em 31 religadores e 12 "
    "reguladores, os mesmos números da página da taxa.",
    "",
    "Cuidado ao somar a taxa: no ano, o ativo que falha duas vezes conta uma vez só, "
    "então somar os meses passa do total anual.",
    "",
    "Fontes: base de SS/OS de 20/08/2026, base de ocorrência (EQP_SS_OCORRENCIA), "
    "cadeia de repasse do COEP e a leitura das SS revisada da taxa de falha.",
]


def montar():
    with open(os.path.join(RAIZ, "data", "missao", "parque_2026.json"),
              encoding="utf-8") as fh:
        p = json.load(fh)
    wb = openpyxl.Workbook()
    for i, t in enumerate(("RL", "RT")):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = f"{NOME[t]} 2026"
        escrever(ws, t, p["series"][t], p["cumulativo"][t])
    ws = wb.create_sheet("Como foi feito")
    ws.column_dimensions["A"].width = 115
    for texto in COMO:
        ws.append([texto])
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    print(f"gravado: {SAIDA}")


if __name__ == "__main__":
    montar()
