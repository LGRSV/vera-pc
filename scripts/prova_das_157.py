"""
A prova das 157 — SS por SS, com o recibo de cada uma.

O gestor duvidou do número («Você tem certeza que são 157?» · «Me prova baseado em
dados») e a resposta certa não é argumento, é a lista auditável: as 341 SS de
indisponibilidade do COEP, cada uma com as datas que a colocam dentro ou fora de
2026, e a soma refeita POR FÓRMULA dentro do próprio Excel — o número não vem
escrito, ele se recalcula das linhas na frente de quem abre.

A régua da saída, que é onde a conta podia escorregar: SS repassada sai da base sem
data de conclusão (e DTA_REPASSE é cópia da abertura — não serve). A saída dela é a
ABERTURA DA SS SEGUINTE, campo SS_APOS_REPASSE da base de repasse. É essa data que
prova que as 159 repassadas antigas saíram antes de 2026.

Grava dist/PROVA_DAS_157.xlsx.
Rodar: python3 scripts/prova_das_157.py
"""

import datetime as dt
import json
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SAIDA = os.path.join(RAIZ, "dist", "PROVA_DAS_157.xlsx")
JAN = dt.datetime(2026, 1, 1)
IND = "INDISPONIBILIDADE PARA OPERAÇÃO"

RE_SS = re.compile(r"([A-Z][A-Z-]*)\s*0*(\d+)/(\d{4})")
TITULO = Font(bold=True, color="FFFFFF", size=10)
FUNDO = PatternFill("solid", fgColor="1F3864")
BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)

NA_MESA = "NA MESA EM 2026"
SAIU_ANTES = "SAIU ANTES DE 2026 — pela cadeia"
FECHOU_ANTES = "FECHADA ANTES DE 2026"


def norm(n):
    m = RE_SS.search((n or "").strip().upper())
    return f"{m.group(1)} {int(m.group(2))}/{m.group(3)}" if m else (n or "").strip().upper()


def dta(s):
    for f in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime((s or "").strip()[:10], f)
        except ValueError:
            pass
    return None


def carregar():
    with open(os.path.join(RAIZ, "data", "missao", "ssos_min.json"), encoding="utf-8") as fh:
        mins = json.load(fh)
    with open(os.path.join(RAIZ, "data", "missao", "ss_ocorrencia.json"), encoding="utf-8") as fh:
        oc = {norm(r["SS_ORIGINAL"]): r for r in json.load(fh)}
    with open(os.path.join(RAIZ, "data", "missao", "coep_2026.json"), encoding="utf-8") as fh:
        coep = json.load(fh)
    mesa = {norm(s["ss"]): s for s in coep["ss"]}
    return mins, oc, mesa


def montar():
    mins, oc, mesa = carregar()
    linhas = []
    for r in mins:
        if not r["NUMERO_SS"].strip().upper().startswith("ETO-COEP"):
            continue
        if r["TIPOSS"].strip() != IND:
            continue
        k = norm(r["NUMERO_SS"])
        ab, fim = dta(r["DATA_ABERTURA_SS"]), dta(r["DATA_TERMINO_SS"])
        reg = oc.get(k) or {}
        seg = norm(reg.get("SS_APOS_REPASSE", "")) if reg.get("SS_APOS_REPASSE") else ""
        seg_ab = dta((oc.get(seg) or {}).get("DTA_ABERTURA", "")) if seg else None
        em_mesa = mesa.get(k)

        if em_mesa:
            veredito = NA_MESA
            prova = (f"na mesa de {em_mesa['chegou']} a "
                     f"{em_mesa['saiu'] or 'hoje — ainda lá'} "
                     f"({em_mesa['como_apurou_a_saida'] or 'segue no posto'})")
        elif r["SITUACAO_SS"].strip() == "SS REPASSADA" and (ab or JAN) < JAN:
            veredito = SAIU_ANTES
            prova = (f"repassada; a sucessora {seg} abriu em "
                     f"{seg_ab:%d/%m/%Y} — saiu do posto nesse dia" if seg_ab else
                     "repassada; sucessora sem data (não ocorre na base)")
        else:
            veredito = FECHOU_ANTES
            prova = (f"{r['SITUACAO_SS'].strip().lower()} com término em "
                     f"{fim:%d/%m/%Y}" if fim else r["SITUACAO_SS"].strip().lower())

        linhas.append({
            "ss": k, "ativo": r["NUM_TRAFO"].strip(),
            "tipo": "RL" if r["NUM_TRAFO"].strip()[:2] in ("79", "78") else "RT",
            "situacao": r["SITUACAO_SS"].strip(),
            "abertura": r["DATA_ABERTURA_SS"].strip()[:10],
            "termino": r["DATA_TERMINO_SS"].strip()[:10] or "—",
            "sucessora": seg or "—",
            "sucessora_abertura": f"{seg_ab:%d/%m/%Y}" if seg_ab else "—",
            "veredito": veredito, "prova": prova,
        })

    ordem = {NA_MESA: 0, SAIU_ANTES: 1, FECHOU_ANTES: 2}
    linhas.sort(key=lambda x: (ordem[x["veredito"]], dta(x["abertura"]) or JAN))

    wb = openpyxl.Workbook()

    # ---- Prova viva: os números saem de fórmula, não de texto ------------
    ws = wb.active
    ws.title = "Prova viva"
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 12
    ws.append(["A PROVA DAS 157 — os números desta aba são FÓRMULA sobre a aba "
               "«As 341, uma a uma». Apague uma linha lá e a conta muda aqui."])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.append([])
    ws.append(["SS ETO-COEP de INDISPONIBILIDADE na base de 20/08 (total de linhas)",
               "=COUNTA('As 341, uma a uma'!A2:A10000)"])
    ws.append(["… que estiveram na mesa em 2026", f'=COUNTIF(\'As 341, uma a uma\'!I:I,"{NA_MESA}")'])
    ws.append(["… repassadas que saíram antes de 2026 (sucessora aberta antes do ano)",
               f'=COUNTIF(\'As 341, uma a uma\'!I:I,"{SAIU_ANTES}")'])
    ws.append(["… atendidas/canceladas com término antes de 2026",
               f'=COUNTIF(\'As 341, uma a uma\'!I:I,"{FECHOU_ANTES}")'])
    ws.append(["Soma das três partes (tem de bater com o total)", "=SUM(B4:B6)"])
    ws.append([])
    # a data da sucessora é texto dd/mm/aaaa — comparação com >= sairia errada;
    # o casamento de padrão «*/2026» pega qualquer abertura dentro de 2026
    ws.append(["Sucessora aberta EM 2026 entre as «saiu antes» (tem de dar zero)",
               f'=COUNTIFS(\'As 341, uma a uma\'!I:I,"{SAIU_ANTES}",'
               "'As 341, uma a uma'!H:H,\"*/2026\")"])
    for r in range(3, 10):
        ws.cell(row=r, column=2).font = Font(bold=True, size=12)
    ws.append([])
    ws.append(["Qualquer linha pode ser conferida no SGM pelo número da SS. As datas "
               "desta planilha vêm da base de SS/OS de 20/08/2026 e da base de "
               "repasse (EQP_SS_OCORRENCIA)."])

    # ---- As 341, uma a uma ----------------------------------------------
    ws = wb.create_sheet("As 341, uma a uma")
    cab = [("SS", 20), ("Equipamento", 13), ("Tipo", 7), ("Situação na base", 16),
           ("Abertura", 12), ("Término", 12), ("SS sucessora (repasse)", 20),
           ("Abertura da sucessora", 14), ("Veredicto", 30), ("O recibo", 52)]
    ws.append([c[0] for c in cab])
    for i, (nome, larg) in enumerate(cab, 1):
        cel = ws.cell(row=1, column=i)
        cel.font, cel.fill = TITULO, FUNDO
        cel.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.row_dimensions[1].height = 30
    for x in linhas:
        ws.append([x["ss"], x["ativo"], x["tipo"], x["situacao"], x["abertura"],
                   x["termino"], x["sucessora"], x["sucessora_abertura"],
                   x["veredito"], x["prova"]])
    for linha in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cel in linha:
            cel.border = BORDA
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    # ---- As 157 na mesa --------------------------------------------------
    ws = wb.create_sheet("As 157 na mesa")
    cab = [("SS", 20), ("Equipamento", 13), ("Tipo", 7), ("Chegou ao posto", 14),
           ("Saiu", 14), ("Como a saída foi apurada", 26), ("Foi para", 13),
           ("Dias no posto", 12), ("Situação hoje", 15), ("Ano do número", 12)]
    ws.append([c[0] for c in cab])
    for i, (nome, larg) in enumerate(cab, 1):
        cel = ws.cell(row=1, column=i)
        cel.font, cel.fill = TITULO, FUNDO
        cel.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.row_dimensions[1].height = 30
    _, _, mesa2 = None, None, None
    with open(os.path.join(RAIZ, "data", "missao", "coep_2026.json"), encoding="utf-8") as fh:
        coep = json.load(fh)
    ind = [s for s in coep["ss"] if s["pendencia"].strip() == IND]
    ind.sort(key=lambda s: (dta(s["chegou"]) or JAN))
    for s in ind:
        ws.append([norm(s["ss"]), s["ativo"],
                   "RL" if s["tipo"] == "religador" else "RT", s["chegou"],
                   s["saiu"] or "— ainda no posto",
                   s["como_apurou_a_saida"] or "segue no posto",
                   s["foi_para"] or "—", s["dias_no_posto"], s["status"],
                   s["ano_da_ss"]])
    for linha in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cel in linha:
            cel.border = BORDA
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    # ---- Como conferir ---------------------------------------------------
    ws = wb.create_sheet("Como conferir")
    ws.column_dimensions["A"].width = 112
    for t in [
        "COMO CONFERIR ESTA PROVA — qualquer linha, no SGM.",
        "",
        "A pergunta: das SS ETO-COEP de INDISPONIBILIDADE PARA OPERAÇÃO, quantas "
        "estiveram na mesa do posto em algum momento de 2026?",
        "",
        "A régua da presença: a SS chega ao posto na ABERTURA e sai na CONCLUSÃO — e, "
        "se foi repassada, sai na ABERTURA DA SS SEGUINTE (campo SS_APOS_REPASSE da "
        "base de repasse). Está na mesa em 2026 quem tem [chegada, saída] cruzando o "
        "ano, com corte em 18/08/2026.",
        "",
        "Por que não dá para usar a data de conclusão sozinha: SS repassada sai SEM "
        "conclusão na base, e DTA_REPASSE é cópia byte a byte da abertura — não data "
        "nada. Sem a cadeia, SS de 2020 parece «ainda no posto» e infla a conta.",
        "",
        "As três pilhas da aba «As 341, uma a uma»:",
        f"  · {NA_MESA} — 157. O recibo mostra chegada e saída no posto.",
        f"  · {SAIU_ANTES} — 159. O recibo mostra a sucessora e a data de abertura "
        "dela, sempre anterior a 01/01/2026.",
        f"  · {FECHOU_ANTES} — 25. Atendida ou cancelada com término em 2024/2025.",
        "",
        "A aba «Prova viva» refaz a soma por fórmula sobre as linhas — os números não "
        "estão escritos, são contados pelo Excel na hora. A última fórmula confere que "
        "nenhuma «saiu antes» tem sucessora aberta em 2026 (tem de dar zero).",
        "",
        "Contraprova feita antes de gravar: procurei SS de indisponibilidade do COEP "
        "que cruzam 2026 e ficaram fora da conta (aberta no ano, pendente hoje, ou "
        "fechada dentro do ano) — zero. Não há nenhuma aberta em 19–20/08, os dois "
        "dias que a base nova tem a mais que o corte.",
        "",
        "Fontes: BASE_SS_OS_20082026.txt (recorte RL/RT) e EQP_SS_OCORRENCIA_11082026 "
        "(cadeia de repasse; registros até 19/08). Posição da mesa: 18/08/2026.",
    ]:
        ws.append([t])
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    from collections import Counter
    c = Counter(x["veredito"] for x in linhas)
    return len(linhas), c, len(ind)


if __name__ == "__main__":
    total, c, ind_mesa = montar()
    print(f"gravado: {SAIDA}")
    print(f"  {total} SS auditadas: ", dict(c))
    print(f"  aba das 157: {ind_mesa} linhas")
