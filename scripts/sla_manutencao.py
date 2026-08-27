"""
O SLA de manutenção dos COCM's.

Régua do gestor (27/08): o relógio começa na ENTREGA ao COCM — o dia em que a demanda
chega numa equipe de campo — e o prazo vem da criticidade dada pela operação, na aba
de mapeamento por criticidade:

    Muito Alta   8 dias        Média   30 dias
    Alta        15 dias        Baixa   50 dias
    sem criticidade definida   26 dias (prazo médio)

Duas datas, e nenhuma delas é óbvia nesta base:

ENTREGA — é a ABERTURA da SS no posto do COCM, não o campo DTA_REPASSE (que é cópia
byte a byte da DTA_ABERTURA e não data nada). A demanda chega ao campo quando o SGM
abre a SS lá.

DEVOLUÇÃO — é quando a SS SAI do COCM: a conclusão dela, se houver; senão a abertura
da SS seguinte, porque SS repassada sai da base sem data de conclusão. Sair para a
TELE é comissionamento: a manutenção já foi executada, o relógio do COCM parou.

O relógio conta a passagem pelo COCM DEPOIS do COEP. Muita demanda nasce no campo —
a equipe abre a SS e manda para o posto —, e essa passagem inicial não é entrega para
executar: o COEP ainda não tinha diagnosticado nem comprado nada. Cobrar prazo dela
seria cobrar o COCM por serviço que ninguém tinha mandado fazer.

Quem nunca passou por COCM não tem SLA de manutenção — fechou na TELE ou na PROT, que
é execução de outro braço. Fica marcado, não fica com prazo estourado por engano.

Grava dist/SLA_MANUTENCAO.xlsx.
Rodar: python3 scripts/sla_manutencao.py [caminho_da_base_de_repasse.xlsx]
"""

import datetime as dt
import json
import os
import re
import sys
from collections import Counter, defaultdict

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SAIDA = os.path.join(RAIZ, "dist", "SLA_MANUTENCAO.xlsx")
CARTEIRA = os.path.join(RAIZ, "data", "raw", "EQUIPAMENTOS_INDISPONIVEIS_ATUALIZADA16.xlsx")
COEP = os.path.join(RAIZ, "data", "missao", "coep_2026.json")
BASE_PADRAO = os.path.join(RAIZ, "data", "missao", "ss_ocorrencia.json")

# PROPOSTA DCMD (gestor, 27/08) — a régua que vale agora. Mais folgada que a
# anterior em toda faixa: o Muito Alta ganha 3 dias, o Baixa ganha 10.
PRAZO = {"Muito Alta": 11, "Alta": 20, "Média": 40, "Baixa": 60}
# A régua anterior fica como referência, para medir o que a proposta muda.
PRAZO_ANTERIOR = {"Muito Alta": 8, "Alta": 15, "Média": 30, "Baixa": 50}
# Sem criticidade definida: 60 dias (gestor, 27/08) — o mesmo teto do Baixa. Faz
# sentido: sem classificação da operação, não há como exigir urgência. Antes eram 26,
# e é a mudança que mais mexe no número — 70 das 131 entregas caem nessa faixa.
PRAZO_SEM_CRITICIDADE = 60
PRAZO_SEM_CRITICIDADE_ANTERIOR = 26
HOJE = dt.datetime(2026, 8, 18, 23, 59)      # posição da conta do posto
RE_SS = re.compile(r"([A-Z][A-Z-]*)\s*0*(\d+)/(\d{4})")
RE_COCM = re.compile(r"-RD-")                 # ETO-RD-*, ENC-RD-*, DOLP-RD-*, DG-RD-*, ESO-RD-*

TITULO = Font(bold=True, color="FFFFFF", size=10)
FUNDO = PatternFill("solid", fgColor="1F3864")
BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)
VERDE = PatternFill("solid", fgColor="E2EFDA")
VERMELHO = PatternFill("solid", fgColor="FCE4E4")


def norm(n):
    m = RE_SS.search((n or "").strip().upper())
    return f"{m.group(1)} {int(m.group(2))}/{m.group(3)}" if m else (n or "").strip().upper()


def data(v):
    if isinstance(v, dt.datetime):
        return v
    if isinstance(v, dt.date):
        return dt.datetime(v.year, v.month, v.day)
    s = str(v or "").strip()[:19]
    for f in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s[:len(dt.datetime.now().strftime(f))], f)
        except ValueError:
            pass
    return None


def base_de_repasse(caminho):
    """{SS: {posto, abertura, conclusao, status, seguinte}} da base de repasse."""
    reg = {}
    if caminho and caminho.lower().endswith(".xlsx"):
        ws = openpyxl.load_workbook(caminho, read_only=True, data_only=True)["Exportar Planilha"]
        for r in list(ws.iter_rows(values_only=True))[1:]:
            k = norm(r[1])
            if k:
                reg.setdefault(k, {"posto": str(r[2] or "").strip(),
                                   "abertura": data(r[7]), "conclusao": data(r[9]),
                                   "status": str(r[12] or "").strip().upper(),
                                   "seguinte": norm(r[14]) if r[14] else ""})
    else:
        with open(caminho or BASE_PADRAO, encoding="utf-8") as fh:
            for r in json.load(fh):
                k = norm(r["SS_ORIGINAL"])
                if k:
                    reg.setdefault(k, {"posto": str(r["POSTO_SGM"] or "").strip(),
                                       "abertura": data(r["DTA_ABERTURA"]),
                                       "conclusao": data(r["DTA_CONCLUSAO"]),
                                       "status": str(r["STATUS"] or "").strip().upper(),
                                       "seguinte": norm(r["SS_APOS_REPASSE"])
                                       if r["SS_APOS_REPASSE"] else ""})
    return reg


def criticidades():
    """A criticidade da operação, das duas abas de mapeamento da carteira."""
    wb = openpyxl.load_workbook(CARTEIRA, read_only=True, data_only=True)
    c = {}
    for aba, col, largura in (("Criticidade por Equipamento", 31, 36),
                              ("Criticidade por Equipamento Joa", 23, 28)):
        for r in list(wb[aba].iter_rows(max_col=largura, values_only=True))[1:]:
            cod = str(r[0] or "").strip()
            v = str(r[col] or "").strip()
            if cod[:2] in ("79", "78", "58") and v in PRAZO:
                c.setdefault(cod, v)
    return c


def cadeia(ss, reg, limite=25):
    """A cadeia de SS a partir de uma, seguindo o carimbo de repasse."""
    saida, visto = [], set()
    while ss and ss in reg and ss not in visto and len(saida) < limite:
        visto.add(ss)
        saida.append((ss, reg[ss]))
        ss = reg[ss]["seguinte"]
    if ss and ss not in reg and ss not in visto:
        saida.append((ss, None))
    return saida


def passagem_anterior_ao_coep(ss_inicial, ss_coep, reg):
    """Houve equipe de campo ANTES do posto? Não conta para o SLA, mas explica o
    caso de quem parece nunca ter ido ao campo e foi — só que antes da hora."""
    for ss, r in cadeia(ss_inicial, reg):
        if ss == ss_coep:
            return False
        if r is not None and RE_COCM.search(r["posto"]):
            return True
    return False


def passagem_pelo_cocm(ss_coep, reg):
    """A primeira parada numa equipe de campo DEPOIS do COEP: entrada e saída.

    A varredura começa na SS do COEP de propósito. É ali que a demanda vira ordem de
    execução — antes disso, passagem por equipe é o campo abrindo o chamado.
    """
    for ss, r in cadeia(ss_coep, reg):
        if r is None or not RE_COCM.search(r["posto"]):
            continue
        entrada = r["abertura"]
        saida = r["conclusao"]
        base_da_saida = "conclusão da SS do COCM"
        if saida is None and r["seguinte"]:
            prox = reg.get(r["seguinte"])
            if prox and prox["abertura"]:
                saida, base_da_saida = prox["abertura"], f"repasse para {prox['posto']}"
        if saida is None:
            base_da_saida = "ainda no COCM"
        return {"ss": ss, "posto": r["posto"], "entrada": entrada, "saida": saida,
                "base_da_saida": base_da_saida, "status": r["status"]}
    return None


def avaliar(item, crit):
    """O SLA de uma demanda: prazo pela criticidade, dias reais, veredicto."""
    c = crit.get(item["ativo"], "")
    prazo = PRAZO.get(c, PRAZO_SEM_CRITICIDADE)
    p = item.get("cocm")
    if p is None or p["entrada"] is None:
        return {**item, "criticidade": c or "Sem classificação", "prazo": prazo,
                "entrega": "", "devolucao": "", "dias": "", "veredicto":
                "não passou por COCM" if p is None else "entrega sem data",
                "atraso": "", "base_da_saida": ""}
    fim = p["saida"] or HOJE
    dias = (fim - p["entrada"]).days
    aberto = p["saida"] is None
    dentro = dias <= prazo
    return {**item, "criticidade": c or "Sem classificação", "prazo": prazo,
            "entrega": p["entrada"].strftime("%d/%m/%Y"),
            "devolucao": "" if aberto else p["saida"].strftime("%d/%m/%Y"),
            "dias": dias,
            "veredicto": ("em curso, dentro do prazo" if dentro else "em curso, ESTOURADO")
            if aberto else ("dentro do prazo" if dentro else "ESTOURADO"),
            "atraso": max(0, dias - prazo),
            "base_da_saida": p["base_da_saida"], "ss_cocm": p["ss"], "posto": p["posto"]}


def montar(caminho_base=None):
    reg = base_de_repasse(caminho_base)
    crit = criticidades()
    with open(COEP, encoding="utf-8") as fh:
        cp = json.load(fh)

    res = [r for r in cp["resolvidos_do_coep"] if r["conta_como_resolvido_pelo_coep"]]
    avaliados = []
    for r in res:
        item = {"ativo": r["ativo"], "tipo": "RL" if r["tipo"] == "religador" else "RT",
                "localidade": r["localidade"], "ss_coep": norm(r["ss_no_coep"]),
                "como_terminou": r["como_terminou"],
                "fechou_em": r["data_do_fechamento"],
                "posto_que_fechou": r["posto_que_fechou"]}
        item["cocm"] = passagem_pelo_cocm(item["ss_coep"], reg)
        item["campo_antes_do_coep"] = passagem_anterior_ao_coep(
            norm(r["ss_que_abriu_a_demanda"]), item["ss_coep"], reg)
        avaliados.append(avaliar(item, crit))

    # quem está com os COCM's agora — SLA em curso
    em_curso = []
    for a in cp["ativos"]:
        if a["ativo"] in {x["ativo"] for x in avaliados}:
            continue
        p = passagem_pelo_cocm(norm(a["ss"]), reg)
        if p and p["saida"] is None:
            em_curso.append(avaliar({"ativo": a["ativo"],
                                     "tipo": "RL" if a["tipo"] == "religador" else "RT",
                                     "localidade": a["localidade"],
                                     "ss_coep": norm(a["ss"]), "como_terminou": "",
                                     "fechou_em": "", "posto_que_fechou": "",
                                     "cocm": p}, crit))
    return avaliados, em_curso, crit


def cabecalho(ws, colunas):
    ws.append([c[0] for c in colunas])
    for i, (_, larg) in enumerate(colunas, 1):
        cel = ws.cell(row=ws.max_row, column=i)
        cel.font, cel.fill = TITULO, FUNDO
        cel.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.row_dimensions[ws.max_row].height = 32


def bordar(ws, de, ate):
    for linha in ws.iter_rows(min_row=de, max_row=ate):
        for cel in linha:
            cel.border = BORDA


COLS_RES = [("Ativo", 13), ("Tipo", 7), ("Localidade", 22), ("SS no COEP", 20),
            ("Como terminou", 15), ("Fechou em", 12), ("Posto que fechou", 15),
            ("Criticidade", 14), ("Prazo SLA (dias)", 11), ("SS do COCM", 20),
            ("COCM", 13), ("Entrega ao COCM", 13), ("Devolução", 12),
            ("Dias de manutenção", 12), ("Atraso (dias)", 11),
            ("SLA de manutenção", 24), ("Como a saída foi apurada", 26),
            ("Houve campo antes do posto?", 14)]


def linha_de(x):
    return [x["ativo"], x["tipo"], x["localidade"], x["ss_coep"], x["como_terminou"],
            x["fechou_em"], x["posto_que_fechou"], x["criticidade"], x["prazo"],
            x.get("ss_cocm", ""), x.get("posto", ""), x["entrega"], x["devolucao"],
            x["dias"], x["atraso"], x["veredicto"], x["base_da_saida"],
            "sim" if x.get("campo_antes_do_coep") else ""]


def aba_resolvidos(wb, avaliados):
    ws = wb.active
    ws.title = f"Resolvidos ({len(avaliados)})"
    cabecalho(ws, COLS_RES)
    prim = ws.max_row + 1
    for x in sorted(avaliados, key=lambda y: (y["veredicto"], y["ativo"])):
        ws.append(linha_de(x))
        if x["veredicto"] == "dentro do prazo":
            ws.cell(row=ws.max_row, column=16).fill = VERDE
        elif "ESTOURADO" in x["veredicto"]:
            ws.cell(row=ws.max_row, column=16).fill = VERMELHO
    bordar(ws, prim, ws.max_row)
    ws.freeze_panes = f"A{prim}"
    ws.auto_filter.ref = f"A{prim - 1}:R{ws.max_row}"
    return ws


def aba_sla(wb, avaliados, em_curso):
    ws = wb.create_sheet("SLA de manutenção")
    ws.column_dimensions["A"].width = 30
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 15
    ws.append(["SLA DE MANUTENÇÃO DOS COCM'S"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.append(["O relógio corre da entrega ao COCM até a devolução. Prazo pela "
               "criticidade da operação; sem criticidade, 26 dias."])

    com = [x for x in avaliados if isinstance(x["dias"], int)]
    dentro = [x for x in com if x["veredicto"] == "dentro do prazo"]
    fora = [x for x in com if x["veredicto"] == "ESTOURADO"]
    sem = [x for x in avaliados if not isinstance(x["dias"], int)]

    ws.append([])
    ws.append(["Quadro geral", "Demandas"])
    for cel in ws[ws.max_row]:
        if cel.value:
            cel.font, cel.fill = TITULO, FUNDO
    prim = ws.max_row + 1
    ws.append(["Resolvidas que passaram por COCM", len(com)])
    ws.append(["  · dentro do prazo", len(dentro)])
    ws.append(["  · com o prazo estourado", len(fora)])
    ws.append(["Resolvidas sem passagem por COCM", len(sem)])
    ws.append(["Em curso com os COCM's hoje", len(em_curso)])
    bordar(ws, prim, ws.max_row)
    ws.append(["Cumprimento do SLA",
               f"{100 * len(dentro) / len(com):.1f}%" if com else "—"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True, size=12)

    ws.append([])
    ws.append(["Por criticidade", "Prazo (dias)", "Demandas", "No prazo",
               "Estourou", "Cumprimento", "Mediana de dias"])
    for cel in ws[ws.max_row]:
        if cel.value:
            cel.font, cel.fill = TITULO, FUNDO
    prim = ws.max_row + 1
    ordem = ["Muito Alta", "Alta", "Média", "Baixa", "Sem classificação"]
    for c in ordem:
        do_grupo = [x for x in com if x["criticidade"] == c]
        if not do_grupo:
            continue
        ok = [x for x in do_grupo if x["veredicto"] == "dentro do prazo"]
        dias = sorted(x["dias"] for x in do_grupo)
        ws.append([c, PRAZO.get(c, PRAZO_SEM_CRITICIDADE), len(do_grupo), len(ok),
                   len(do_grupo) - len(ok), len(ok) / len(do_grupo),
                   dias[len(dias) // 2]])
        ws.cell(row=ws.max_row, column=6).number_format = "0.0%"
    fim = ws.max_row
    bordar(ws, prim, fim)

    ch = BarChart()
    ch.type, ch.gapWidth, ch.style = "col", 50, 2
    ch.title = "Cumprimento do SLA por criticidade"
    ch.height, ch.width = 8, 15
    ch.add_data(Reference(ws, min_col=6, min_row=prim - 1, max_row=fim),
                titles_from_data=True)
    ch.series[0].cat = AxDataSource(strRef=StrRef(f=f"'{ws.title}'!$A${prim}:$A${fim}"))
    ch.x_axis.delete = ch.y_axis.delete = False
    ch.y_axis.numFmt = "0%"
    ch.legend = None
    ws.add_chart(ch, "I3")

    if em_curso:
        ws.append([])
        ws.append([f"EM CURSO COM OS COCM'S ({len(em_curso)}) — o relógio ainda corre"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        cabecalho(ws, COLS_RES)
        prim = ws.max_row + 1
        for x in sorted(em_curso, key=lambda y: -(y["dias"] if isinstance(y["dias"], int) else 0)):
            ws.append(linha_de(x))
            ws.cell(row=ws.max_row, column=16).fill = (
                VERMELHO if "ESTOURADO" in x["veredicto"] else VERDE)
        bordar(ws, prim, ws.max_row)

    ws.append([])
    ws.append([f"SEM PASSAGEM POR COCM ({len(sem)}) — execução de outro braço, "
               "sem SLA de manutenção"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    cabecalho(ws, COLS_RES)
    prim = ws.max_row + 1
    for x in sorted(sem, key=lambda y: y["ativo"]):
        ws.append(linha_de(x))
    bordar(ws, prim, ws.max_row)
    return ws


def como_foi_feito(avaliados, em_curso):
    com = [x for x in avaliados if isinstance(x["dias"], int)]
    return [
        "SLA DE MANUTENÇÃO — como foi calculado.",
        "",
        "A RÉGUA (gestor, 27/08): o relógio começa na ENTREGA ao COCM e o prazo vem da "
        "criticidade dada pela operação, na aba de mapeamento por criticidade — Muito "
        "Alta 8 dias, Alta 15, Média 30, Baixa 50. Quem não tem criticidade definida "
        "leva o prazo médio de 26 dias.",
        "",
        "ENTREGA AO COCM — é a ABERTURA da SS no posto da equipe de campo (qualquer "
        "posto com «-RD-» no código: ETO-RD-*, ENC-RD-*, DOLP-RD-*, DG-RD-*, ESO-RD-*). "
        "NÃO é o campo DTA_REPASSE da base: ele é cópia byte a byte da DTA_ABERTURA e "
        "não data o repasse. A demanda chega ao campo quando o SGM abre a SS lá.",
        "",
        "DEVOLUÇÃO — é quando a SS SAI do COCM: a conclusão dela, se houver; senão a "
        "abertura da SS seguinte, porque SS repassada sai da base sem data de conclusão. "
        "Sair para a TELE é comissionamento — a manutenção já foi executada e o relógio "
        "do COCM parou ali. Quem ainda não saiu conta os dias até 18/08/2026 e aparece "
        "como «em curso».",
        "",
        "QUEM FICA DE FORA: demanda que nunca passou por COCM não tem SLA de manutenção. "
        "Fechou na TELE ou na PROT, que é execução de outro braço — marcar prazo "
        "estourado nelas seria cobrar o COCM por serviço que não foi dele. Ficam listadas "
        "à parte, no fim da aba de SLA.",
        "",
        "A CADEIA: para achar a passagem pelo COCM, a demanda é percorrida A PARTIR DA SS "
        "DO COEP para a frente, pelo carimbo de repasse (SS_APOS_REPASSE), parada a "
        "parada, até a primeira equipe de campo. Vale a primeira passagem depois do "
        "posto — é ela que responde pela manutenção.",
        "",
        "POR QUE A VARREDURA COMEÇA NO COEP: muita demanda NASCE no campo — a equipe abre "
        "a SS e repassa para o posto —, e essa passagem inicial não é entrega para "
        "executar: o COEP ainda não tinha diagnosticado nem comprado nada. Contando a "
        "cadeia inteira apareceriam 43 com passagem por COCM em vez de 27, e 16 delas "
        "levariam prazo por um serviço que ninguém tinha mandado fazer. A coluna «Houve "
        "campo antes do posto?» marca esses casos, para não sumirem da vista.",
        "",
        f"COBERTURA: das {len(avaliados)} demandas resolvidas pelo posto em 2026, "
        f"{len(com)} passaram por COCM e têm SLA apurado; as demais fecharam sem campo "
        f"na cadeia. Além delas, {len(em_curso)} estão com os COCM's agora, com o "
        "relógio correndo.",
        "",
        "A CRITICIDADE vem das abas «Criticidade por Equipamento» e «Criticidade por "
        "Equipamento Joa» da Relação de Indisponíveis (ATUALIZADA 16) — 103 ativos "
        "classificados. Quem não está lá entra com os 26 dias.",
        "",
        "Fonte das datas: base de repasse (Eqp_joao / EQP_SS_OCORRENCIA), que é a única "
        "que traz a cadeia SS a SS. Posição de 18/08/2026.",
    ]


if __name__ == "__main__":
    caminho = sys.argv[1] if len(sys.argv) > 1 else None
    avaliados, em_curso, crit = montar(caminho)
    wb = openpyxl.Workbook()
    aba_resolvidos(wb, avaliados)
    aba_sla(wb, avaliados, em_curso)
    ws = wb.create_sheet("Como foi feito")
    ws.column_dimensions["A"].width = 112
    for t in como_foi_feito(avaliados, em_curso):
        ws.append([t])
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    com = [x for x in avaliados if isinstance(x["dias"], int)]
    ok = [x for x in com if x["veredicto"] == "dentro do prazo"]
    print(f"gravado: {SAIDA}")
    print(f"  resolvidas: {len(avaliados)} · com passagem por COCM: {len(com)} · "
          f"sem passagem: {len(avaliados) - len(com)}")
    print(f"  dentro do prazo: {len(ok)} · estourado: {len(com) - len(ok)} · "
          f"cumprimento: {100 * len(ok) / len(com):.1f}%")
    print(f"  em curso com os COCM's: {len(em_curso)}")
    print('  por criticidade:', dict(Counter(x["criticidade"] for x in com)))
