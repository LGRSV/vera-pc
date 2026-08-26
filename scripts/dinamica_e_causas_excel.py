"""
O Excel da noite de 25/08: a dinâmica sem os −11 e o porquê de cada falha.

Pedido do gestor, em duas partes:

1. A dinâmica dos 143 em partição DISJUNTA — cada equipamento conta uma vez, pelo
   estado atual. Os 11 que resolveram uma demanda e voltaram contam só como
   pendentes; os resolvidos ficam 71. É outra lente da mesma base: a conta oficial
   (82 resolvidos, 136 do gestor) continua valendo — aqui o gestor quis o livro sem
   dupla aparição.

2. Por que os equipamentos falharam, fatia a fatia (RL/RT × 2025/2026), com a causa
   raiz CITADA do texto da SS. A classificação foi feita por um agente por fatia e
   conferida por um revisor cético; causa sem citação não passa. O complemento por
   obra direta (equipamento que falhou provado pela obra do AIC, sem narrativa de SS)
   entra só como contagem — não há texto para dar causa, e engenheiro não chuta laudo.

Lê data/missao/causas_<fatia>_rev.json (cai para o sem _rev se a revisão faltar) e
data/missao/dinamica_143.json. Grava dist/DINAMICA_E_CAUSAS.xlsx.
Rodar: python3 scripts/dinamica_e_causas_excel.py
"""

import json
import os
from collections import Counter

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SAIDA = os.path.join(RAIZ, "dist", "DINAMICA_E_CAUSAS.xlsx")

TITULO = Font(bold=True, color="FFFFFF", size=10)
FUNDO = PatternFill("solid", fgColor="1F3864")
BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)

FATIAS = [("rl2025", "2 · RL 2025 — por quê", "Religadores 2025"),
          ("rl2026", "3 · RL 2026 — por quê", "Religadores 2026"),
          ("rt2025", "4 · RT 2025 — por quê", "Reguladores 2025"),
          ("rt2026", "5 · RT 2026 — por quê", "Reguladores 2026")]
TAXA = {"rl2025": ("religador|2025", 1307), "rl2026": ("religador|2026", 1307),
        "rt2025": ("regulador|2025", 207), "rt2026": ("regulador|2026", 207)}


def carrega(nome):
    for sufixo in ("_rev", ""):
        p = os.path.join(RAIZ, "data", "missao", f"causas_{nome}{sufixo}.json")
        if os.path.exists(p):
            with open(p, encoding="utf-8") as fh:
                return json.load(fh), bool(sufixo)
    raise SystemExit(f"classificação de {nome} não encontrada — a orquestra ainda roda?")


def cabecalho(ws, colunas):
    ws.append([c[0] for c in colunas])
    for i, (_, larg) in enumerate(colunas, 1):
        cel = ws.cell(row=ws.max_row, column=i)
        cel.font, cel.fill = TITULO, FUNDO
        cel.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.row_dimensions[ws.max_row].height = 30


def bordas(ws, de, ate, wrap_cols=()):
    for linha in ws.iter_rows(min_row=de, max_row=ate):
        for cel in linha:
            cel.border = BORDA
            cel.alignment = Alignment(vertical="top",
                                      wrap_text=cel.column in wrap_cols)


def barras(ws, titulo, faixa, ancora, largura=15):
    ch = BarChart()
    ch.type, ch.gapWidth, ch.style = "bar", 40, 2
    ch.title = titulo
    ch.height, ch.width = max(6.0, 1.1 * (faixa[1] - faixa[0] + 1)), largura
    ch.add_data(Reference(ws, min_col=2, min_row=faixa[0] - 1, max_row=faixa[1]),
                titles_from_data=True)
    ref = f"'{ws.title}'!$A${faixa[0]}:$A${faixa[1]}"
    for s in ch.series:
        s.cat = AxDataSource(strRef=StrRef(f=ref))
    ch.x_axis.delete = ch.y_axis.delete = False
    ch.legend = None
    ws.add_chart(ch, ancora)


def aba_dinamica(wb, d):
    ws = wb.active
    ws.title = "1 · Dinâmica (143)"
    c = d["contas"]
    ws.append(["A DINÂMICA DOS 143 — PARTIÇÃO DISJUNTA, SEM OS −11"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.append(["Régua do gestor (26/08): cada equipamento conta UMA vez, pelo estado de "
               "18/08. Quem resolveu uma demanda e voltou conta só como pendente."])
    ws.append([])
    ws.append(["Situação", "Equipamentos"])
    for cel in ws[ws.max_row]:
        if cel.value:
            cel.font, cel.fill = TITULO, FUNDO
    prim = ws.max_row + 1
    ws.append(["Resolvido (e não voltou)", c["resolvido"]])
    ws.append(["Na fila do posto", c["na_fila"]])
    ws.append(["Em outra mesa", c["outra_mesa"]])
    bordas(ws, prim, ws.max_row)
    ws.append(["Total", c["total"]])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    ws.append([f"Dos {c['na_fila']} na fila, {c.get('resolvidos_que_voltaram', 11)} "
               f"resolveram uma demanda no ano e voltaram — estão marcados. Dos "
               f"{c['resolvido']} resolvidos, {c.get('realizados_dcmd_entre_os_71', '—')} "
               "são realizado do DCMD (atendida com campo ou cancelada de pé)."])
    barras(ws, "Onde estão os 143", (prim, prim + 2), "D2", largura=13)

    ws.append([])
    colunas = [("Situação", 20), ("Ativo", 13), ("Tipo", 11), ("Localidade", 24),
               ("Criticidade", 13), ("SS", 22), ("Desde", 12), ("Dias", 9),
               ("Resolveu e voltou", 14), ("Como terminou", 15), ("Posto que fechou", 15),
               ("Realizado DCMD", 13), ("Onde está", 14), ("Etapa", 22), ("Nota", 52)]
    cabecalho(ws, colunas)
    prim = ws.max_row + 1
    ordem = {"Na fila do posto": 0, "Em outra mesa": 1, "Resolvido": 2}
    for i in sorted(d["itens"], key=lambda x: (ordem.get(x["situacao"], 9),
                                               x.get("tipo", ""), x["ativo"])):
        ws.append([i["situacao"], i["ativo"], i.get("tipo", ""), i.get("localidade", ""),
                   i.get("criticidade", "") or "—", i.get("ss", ""), i.get("desde", ""),
                   i.get("dias", ""), "SIM" if i.get("resolvido_e_voltou") else "",
                   i.get("como_terminou", ""), i.get("posto_que_fechou", ""),
                   ("sim" if i.get("realizado_dcmd") else "") if i["situacao"] == "Resolvido" else "",
                   i.get("onde_esta", ""), i.get("etapa", ""), i.get("nota", "")])
    bordas(ws, prim, ws.max_row, wrap_cols=(15,))
    ws.freeze_panes = f"A{prim}"
    ws.auto_filter.ref = f"A{prim - 1}:O{ws.max_row}"
    return ws


def aba_causas(wb, slug, nome_aba, rotulo, meta, taxa):
    dados, revisado = carrega(slug)
    itens = dados["itens"]
    resumo = Counter(i["causa_raiz"] for i in itens)
    chave, parque = TAXA[slug]
    falharam = taxa.get(chave, 0)
    obra = meta["complemento_obra_direta"].get(slug, 0)

    ws = wb.create_sheet(nome_aba)
    ws.append([f"POR QUE FALHARAM — {rotulo.upper()}"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.append([f"{falharam} equipamentos falharam no ano (taxa {falharam}/{parque} = "
               f"{100 * falharam / parque:.2f}%). Este rol dá causa às {len(itens)} "
               f"ocorrências com narrativa de SS; {obra} equipamentos entram na taxa "
               "por obra direta do AIC, sem texto para dar causa."])
    if not revisado:
        ws.append(["ATENÇÃO: revisão cética ainda não gravada — classificação de primeira mão."])
    ws.append([])
    ws.append(["Causa raiz", "Ocorrências"])
    for cel in ws[ws.max_row]:
        if cel.value:
            cel.font, cel.fill = TITULO, FUNDO
    prim = ws.max_row + 1
    for causa, n in resumo.most_common():
        ws.append([causa, n])
    bordas(ws, prim, ws.max_row)
    fim_resumo = ws.max_row
    barras(ws, f"{rotulo} — causa raiz", (prim, fim_resumo), "D4")

    ws.append([])
    colunas = [("Ativo", 13), ("SS", 22), ("Data", 12), ("Peça (modo)", 13),
               ("Troca feita", 10), ("Causa raiz", 34), ("Confiança", 11),
               ("Citação do texto da SS", 70), ("Nota do analista", 44),
               ("Revisão", 40)]
    cabecalho(ws, colunas)
    prim = ws.max_row + 1
    for i in sorted(itens, key=lambda x: (x["causa_raiz"], x["ativo"])):
        ws.append([i["ativo"], i["ss"], i.get("data", ""), i.get("peca", ""),
                   "sim" if i.get("troca_executada") else "não", i["causa_raiz"],
                   i.get("confianca", ""), f'«{i.get("citacao", "")}»',
                   i.get("nota", ""), i.get("revisao", "")])
    bordas(ws, prim, ws.max_row, wrap_cols=(8, 9, 10))
    ws.freeze_panes = f"A{prim}"
    ws.auto_filter.ref = f"A{prim - 1}:J{ws.max_row}"
    return resumo, revisado


def aba_como(wb, resumos, meta):
    ws = wb.create_sheet("Como foi feito")
    ws.column_dimensions["A"].width = 112
    linhas = [
        "DINÂMICA E CAUSAS — o que este arquivo é e como foi montado.",
        "",
        "ABA 1 — a dinâmica dos 143 em partição disjunta, régua do gestor de 26/08: cada "
        "equipamento conta uma vez, pelo estado de 18/08. Os 11 que resolveram uma demanda "
        "e voltaram contam só na fila (marcados na coluna própria); por isso os resolvidos "
        "aqui são 71 e não 82. A conta oficial do posto (82 resolvidos, 136 do gestor) "
        "continua valendo — esta aba é outra lente, não uma correção.",
        "",
        "ABAS 2 A 5 — por que os equipamentos falharam, na leitura de engenheiro: MODO de "
        "falha (a peça que quebrou) não é CAUSA RAIZ (por que quebrou). A causa só entra "
        "com CITAÇÃO verbatim do texto da SS que a sustente; texto que só dá o modo vira "
        "«SEM CAUSA DESCRITA NO TEXTO» — não se chuta laudo.",
        "",
        "Quem fez: um agente classificador por fatia (RL/RT × 2025/2026) leu a descrição "
        "completa de cada SS — com as armadilhas conhecidas da base: texto cumulativo "
        "(vale o parecer mais recente), laudo de terceiro colado (conferir o código) e o "
        "formulário DMSL. Por cima, um REVISOR CÉTICO por fatia conferiu citação por "
        "citação contra o texto-fonte e corrigiu o que não se sustentava — as correções "
        "estão na coluna «Revisão».",
        "",
        "O que fica fora das causas: os equipamentos provados por OBRA DIRETA do AIC "
        "(sem narrativa de SS) entram na taxa mas não têm texto para dar causa — "
        f"religador 2025: {meta['complemento_obra_direta'].get('rl2025', 0)} · "
        f"religador 2026: {meta['complemento_obra_direta'].get('rl2026', 0)} · "
        f"regulador 2025: {meta['complemento_obra_direta'].get('rt2025', 0)} · "
        f"regulador 2026: {meta['complemento_obra_direta'].get('rt2026', 0)}.",
        "",
        "A taxa por trás de cada aba (equipamento que falhou ÷ parque do ano, parque "
        "1.307 RL e 207 RT): os números estão no cabeçalho de cada aba. Falha é só peça "
        "grande, pela régua do gestor.",
        "",
        "Fontes: leitura revisada da taxa (data/missao/leitura_ss_os.json), descrições da "
        "base de SS/OS de 20/08/2026, cadeia do posto (coep_2026.json). Posição 18/08.",
    ]
    for t in linhas:
        ws.append([t])
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)


def montar():
    with open(os.path.join(RAIZ, "data", "missao", "causas_insumo.json"), encoding="utf-8") as fh:
        meta = json.load(fh)["meta"]
    with open(os.path.join(RAIZ, "data", "missao", "dinamica_143.json"), encoding="utf-8") as fh:
        dinamica = json.load(fh)
    taxa = meta["totais_de_equipamentos"]

    wb = openpyxl.Workbook()
    aba_dinamica(wb, dinamica)
    resumos = {}
    for slug, nome_aba, rotulo in FATIAS:
        resumos[slug] = aba_causas(wb, slug, nome_aba, rotulo, meta, taxa)
    aba_como(wb, resumos, meta)
    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    print(f"gravado: {SAIDA}")
    for slug, (resumo, revisado) in resumos.items():
        top = " · ".join(f"{c} {n}" for c, n in resumo.most_common(4))
        print(f"  {slug}{'' if revisado else ' (SEM revisão!)'}: {top}")


if __name__ == "__main__":
    montar()
