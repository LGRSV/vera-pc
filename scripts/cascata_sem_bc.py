"""
A cascata do ano — quadro «Sem BC» que o gestor mandou em 03/09.
dist/CASCATA_SEM_BC_2026.xlsx

O quadro: backlog de 2025 em 59, entrante e resolvidos mês a mês, pendentes fechando
dezembro em 24. Conferido nos doze meses — `pendentes = anterior + entrante − resolvidos`
bate em todos, sem sobra.

  backlog 59 → jan 65 · fev 71 · mar 76 · abr 80 · mai 65 · jun 58 · jul 56 · ago 55 ·
  set 57 · out 60 · nov 43 · dez 24

No ano entram 60 e saem 95: a fila cede 35. O ano tem duas metades — de janeiro a abril
a fila SOBE (entram 29, saem 8) e de maio em diante DESCE (entram 31, saem 87). Maio,
novembro e dezembro sozinhos respondem por 58 dos 95 resolvidos.

A CASCATA é o formato certo aqui porque o que interessa é de onde vem cada degrau: a
barra flutua entre o saldo do mês anterior e o do mês, laranja quando a fila cresce e
verde quando cai, com as duas pontas ancoradas no chão (backlog e fim do ano).

Três leituras: a cascata do saldo líquido (aba 1), a cascata detalhada com entrante e
resolvido em degraus separados (aba 2) e o ritmo em barras com a linha de pendentes
(aba 3). Todas leem a aba **Dados** — mudou lá, mudou o gráfico.

Rodar: python3 scripts/cascata_sem_bc.py
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Font, PatternFill

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
import backlog_mensal as bm  # noqa: E402  — só o estilo de gráfico

SAIDA = os.path.join(RAIZ, "dist", "CASCATA_SEM_BC_2026.xlsx")

TINTA, PAPEL, SOMBRA, SINAL = bm.TINTA, bm.PAPEL, bm.SOMBRA, bm.SINAL
VERDE, LARANJA, NEUTRO = bm.VERDE, bm.LARANJA, bm.NEUTRO

# o quadro, como ele mandou: (rótulo, pendentes no fim, entrante, resolvidos)
QUADRO = [
    ("Backlog 2025", 59, 0, 0),
    ("Janeiro", 65, 7, 1),
    ("Fevereiro", 71, 8, 2),
    ("Março", 76, 9, 4),
    ("Abril", 80, 5, 1),
    ("Maio", 65, 1, 16),
    ("Junho", 58, 2, 9),
    ("Julho", 56, 4, 6),
    ("Agosto", 55, 1, 2),
    ("Setembro", 57, 8, 6),
    ("Outubro", 60, 9, 6),
    ("Novembro", 43, 5, 22),
    ("Dezembro", 24, 1, 20),
]

FMT_MAIS = '"+"0;;;'      # esconde o zero e marca o sinal
FMT_MENOS = '"−"0;;;'
FMT_CHEIO = '0;;;'


def confere():
    """O quadro fecha sozinho nos doze meses? Se não fechar, o script para aqui."""
    erros = []
    for i in range(1, len(QUADRO)):
        _, fim, ent, res = QUADRO[i]
        ini = QUADRO[i - 1][1]
        if ini + ent - res != fim:
            erros.append((QUADRO[i][0], ini, ent, res, fim, ini + ent - res))
    assert not erros, erros
    return sum(q[2] for q in QUADRO), sum(q[3] for q in QUADRO)


def sem_preenchimento(s):
    """A série-base da cascata: existe para empurrar a barra para cima, e não se vê."""
    s.graphicalProperties = GraphicalProperties(noFill=True)
    s.graphicalProperties.line = LineProperties(noFill=True)
    return s


def rotulo(s, fmt):
    s.dLbls = DataLabelList()
    s.dLbls.showVal = True
    s.dLbls.showSerName = s.dLbls.showCatName = s.dLbls.showLegendKey = False
    s.dLbls.numFmt = fmt
    return s


def cascata(ws, ch, primeira, ultima):
    """Base invisível + total + sobe + desce, empilhados."""
    ch.type, ch.grouping, ch.overlap, ch.gapWidth = "col", "stacked", 100, 40
    ch.add_data(Reference(ws, min_col=2, min_row=primeira - 1, max_col=5, max_row=ultima),
                titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=primeira, max_row=ultima))
    sem_preenchimento(ch.series[0])
    for s, cor, fmt in ((ch.series[1], NEUTRO, FMT_CHEIO),
                        (ch.series[2], LARANJA, FMT_MAIS),
                        (ch.series[3], VERDE, FMT_MENOS)):
        bm.cor_barra(s, cor)
        rotulo(s, fmt)
    bm.categorias(ch, ws, "$A$%d:$A$%d" % (primeira, ultima))
    ch.y_axis.title = "equipamentos"
    return ch


# ------------------------------------------------------------------------------- abas
def aba_dados(wb, entrou, saiu):
    ws = wb.create_sheet("Dados")
    bm.titulo(ws, "O QUADRO «SEM BC», COMO O GESTOR MANDOU — 03/09",
              "Backlog de 2025 em %d, entrante e resolvidos mês a mês, dezembro fechando em %d. "
              "A coluna «Confere» refaz a conta (anterior + entrante − resolvidos) e bate nos doze "
              "meses, sem sobra. No ano entram %d e saem %d: a fila cede %d. Os gráficos das "
              "outras abas leem daqui — mudou aqui, mudou o gráfico."
              % (QUADRO[0][1], QUADRO[-1][1], entrou, saiu, QUADRO[0][1] - QUADRO[-1][1]))
    bm.cabecalho(ws, 4, ["Mês", "Pendentes", "Entrante", "Resolvidos", "Saldo do mês", "Confere"],
                 [16, 12, 11, 12, 13, 11])
    r = 5
    for i, (rot, fim, ent, res) in enumerate(QUADRO):
        ws.cell(row=r, column=1, value=rot)
        ws.cell(row=r, column=2, value=fim)
        ws.cell(row=r, column=3, value=ent)
        ws.cell(row=r, column=4, value=res)
        if i:
            ws.cell(row=r, column=5, value="=C%d-D%d" % (r, r))
            ws.cell(row=r, column=6, value='=IF(B%d-B%d+D%d-C%d=0,"ok","ERRO")' % (r - 1, r, r, r))
        else:
            ws.cell(row=r, column=1).font = Font(bold=True)
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=SOMBRA)
        for c in range(2, 7):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1
    fim_l = r - 1
    ws.cell(row=r, column=1, value="no ano").font = Font(bold=True)
    for col in ("C", "D"):
        c = ws.cell(row=r, column=ord(col) - 64, value="=SUM(%s6:%s%d)" % (col, col, fim_l))
        c.font, c.alignment = Font(bold=True), Alignment(horizontal="center")
    c = ws.cell(row=r, column=5, value="=C%d-D%d" % (r, r))
    c.font, c.alignment = Font(bold=True), Alignment(horizontal="center")
    return fim_l


def aba_cascata(wb, fim_dados):
    ws = wb.create_sheet("Cascata do ano")
    ws.sheet_view.showGridLines = False
    bm.titulo(ws, "A CASCATA DO ANO — de %d para %d, degrau a degrau"
              % (QUADRO[0][1], QUADRO[-1][1]),
              "Cada barra flutua entre o saldo do mês anterior e o do mês. LARANJA é mês em que a "
              "fila cresceu, VERDE é mês em que ela cedeu; as duas pontas — backlog e fim do ano — "
              "ficam ancoradas no chão. O ano tem duas metades claras: de janeiro a abril a fila "
              "sobe 21, de maio em diante desce 56.")
    bm.cabecalho(ws, 4, ["Mês", "Base (invisível)", "Total", "A fila cresceu", "A fila cedeu"],
                 [16, 15, 11, 15, 14])
    r, linha_d = 5, 6      # linha 5 do Dados é o backlog; os meses começam na 6
    for i, (rot, _, _, _) in enumerate(QUADRO):
        d_ant, d_at = 4 + i, 5 + i          # linhas da aba Dados
        ws.cell(row=r, column=1, value=rot)
        if i == 0:
            ws.cell(row=r, column=2, value=0)
            ws.cell(row=r, column=3, value="=Dados!B5")
            ws.cell(row=r, column=4, value=0)
            ws.cell(row=r, column=5, value=0)
        else:
            ws.cell(row=r, column=2, value="=MIN(Dados!B%d,Dados!B%d)" % (d_ant, d_at))
            ws.cell(row=r, column=3, value=0)
            ws.cell(row=r, column=4, value="=MAX(0,Dados!B%d-Dados!B%d)" % (d_at, d_ant))
            ws.cell(row=r, column=5, value="=MAX(0,Dados!B%d-Dados!B%d)" % (d_ant, d_at))
        for c in range(2, 6):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1
    # a barra de fecho, ancorada no chão
    ws.cell(row=r, column=1, value="Fim do ano").font = Font(bold=True)
    ws.cell(row=r, column=2, value=0)
    ws.cell(row=r, column=3, value="=Dados!B%d" % (4 + len(QUADRO)))
    ws.cell(row=r, column=4, value=0)
    ws.cell(row=r, column=5, value=0)
    for c in range(2, 6):
        ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    ultima = r

    ch = cascata(ws, BarChart(), 5, ultima)
    ch.title = "Do backlog de %d ao fim do ano em %d — o saldo de cada mês" % (
        QUADRO[0][1], QUADRO[-1][1])
    ws.add_chart(bm.estilo(ch, 13, 30), "A%d" % (ultima + 3))

    ch2 = LineChart()
    ch2.add_data(Reference(ws, min_col=2, min_row=4, max_row=fim_dados), titles_from_data=True)
    return ws, ultima


def aba_detalhada(wb):
    ws = wb.create_sheet("Cascata detalhada")
    ws.sheet_view.showGridLines = False
    bm.titulo(ws, "A MESMA CASCATA, COM ENTRANTE E RESOLVIDO SEPARADOS",
              "Aqui cada mês vira dois degraus: primeiro o que entrou (laranja, sobe), depois o "
              "que foi resolvido (verde, desce). Mostra o que a cascata do saldo esconde — em "
              "maio entrou 1 e resolveu 16; em novembro entrou 5 e resolveu 22.")
    bm.cabecalho(ws, 4, ["Passo", "Base (invisível)", "Total", "Entrou", "Resolvido"],
                 [22, 15, 11, 12, 13])
    r = 5
    ws.cell(row=r, column=1, value="Backlog 2025").font = Font(bold=True)
    ws.cell(row=r, column=2, value=0)
    ws.cell(row=r, column=3, value="=Dados!B5")
    ws.cell(row=r, column=4, value=0)
    ws.cell(row=r, column=5, value=0)
    r += 1
    saldo = QUADRO[0][1]
    for i, (rot, fim, ent, res) in enumerate(QUADRO[1:], start=1):
        d = 5 + i                                     # linha do mês na aba Dados
        ws.cell(row=r, column=1, value="%s · entrou" % rot)
        ws.cell(row=r, column=2, value=saldo)
        ws.cell(row=r, column=3, value=0)
        ws.cell(row=r, column=4, value="=Dados!C%d" % d)
        ws.cell(row=r, column=5, value=0)
        r += 1
        saldo += ent
        ws.cell(row=r, column=1, value="%s · resolvido" % rot)
        ws.cell(row=r, column=2, value=saldo - res)
        ws.cell(row=r, column=3, value=0)
        ws.cell(row=r, column=4, value=0)
        ws.cell(row=r, column=5, value="=Dados!D%d" % d)
        r += 1
        saldo -= res
    ws.cell(row=r, column=1, value="Fim do ano").font = Font(bold=True)
    ws.cell(row=r, column=2, value=0)
    ws.cell(row=r, column=3, value="=Dados!B%d" % (4 + len(QUADRO)))
    ws.cell(row=r, column=4, value=0)
    ws.cell(row=r, column=5, value=0)
    ultima = r
    for linha in range(5, ultima + 1):
        for c in range(2, 6):
            ws.cell(row=linha, column=c).alignment = Alignment(horizontal="center")

    ch = cascata(ws, BarChart(), 5, ultima)
    ch.gapWidth = 20
    ch.series[2].dLbls.numFmt = FMT_MAIS
    ch.series[3].dLbls.numFmt = FMT_MENOS
    ch.title = "Cada mês em dois degraus: o que entrou e o que foi resolvido"
    ws.add_chart(bm.estilo(ch, 13, 32), "A%d" % (ultima + 3))


def aba_ritmo(wb, fim_dados):
    ws = wb.create_sheet("Pendentes e ritmo")
    bm.titulo(ws, "O RITMO — enquanto o verde for maior que o laranja, a fila cede",
              "As barras são o que entrou e o que se resolveu em cada mês; a linha é o estoque de "
              "pendentes. De janeiro a abril o laranja ganha e a curva sobe; de maio em diante o "
              "verde ganha e a curva desce. Julho, agosto, setembro e outubro quase empatam — é "
              "quando a fila fica parada em torno de 55.")
    bm.cabecalho(ws, 4, ["Mês", "Entrante", "Resolvidos", "Pendentes"], [16, 12, 12, 12])
    r = 5
    for i in range(1, len(QUADRO)):
        d = 5 + i
        ws.cell(row=r, column=1, value="=Dados!A%d" % d)
        ws.cell(row=r, column=2, value="=Dados!C%d" % d)
        ws.cell(row=r, column=3, value="=Dados!D%d" % d)
        ws.cell(row=r, column=4, value="=Dados!B%d" % d)
        for c in range(2, 5):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1
    ultima = r - 1

    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth, ch.overlap = "col", "clustered", 80, -12
    ch.add_data(Reference(ws, min_col=2, min_row=4, max_col=3, max_row=ultima),
                titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=ultima))
    bm.cor_barra(ch.series[0], LARANJA)
    bm.cor_barra(ch.series[1], VERDE)
    for s in ch.series:
        bm.rotulos(s)
    linha = LineChart()
    linha.add_data(Reference(ws, min_col=4, min_row=4, max_row=ultima), titles_from_data=True)
    s = linha.series[0]
    s.graphicalProperties = GraphicalProperties()
    s.graphicalProperties.line = LineProperties(solidFill=NEUTRO, w=30000)
    s.marker = Marker(symbol="circle", size=7)
    s.marker.graphicalProperties = GraphicalProperties(solidFill=NEUTRO)
    s.smooth = False
    bm.rotulos(s)
    ch += linha                      # sem mexer em axId: eixo único, é tudo equipamento
    ch.title = "Entrante, resolvidos e o estoque de pendentes"
    ch.y_axis.title = "equipamentos"
    bm.categorias(ch, ws, "$A$5:$A$%d" % ultima)
    ws.add_chart(bm.estilo(ch, 13, 30), "A%d" % (r + 2))


def aba_como(wb, entrou, saiu):
    ws = wb.create_sheet("Como ler")
    ws.column_dimensions["A"].width = 112
    sobe = QUADRO[4][1] - QUADRO[0][1]
    desce = QUADRO[4][1] - QUADRO[-1][1]
    texto = [
        ("O QUE ESTÁ AQUI", True),
        ("O quadro «Sem BC» que o gestor mandou em 03/09, sem uma vírgula mudada: backlog de 2025 "
         "em %d, entrante e resolvidos mês a mês, dezembro fechando em %d."
         % (QUADRO[0][1], QUADRO[-1][1]), False),
        ("", False),
        ("O QUADRO FECHA SOZINHO", True),
        ("Refiz a conta nos doze meses — pendentes = mês anterior + entrante − resolvidos — e bate "
         "em todos, sem sobra. A coluna «Confere» da aba Dados mostra isso célula a célula. "
         "No ano entram %d e saem %d: a fila cede %d." % (entrou, saiu, entrou - saiu), False),
        ("", False),
        ("COMO SE LÊ UMA CASCATA", True),
        ("A primeira e a última barra ficam ancoradas no chão: são saldos (o backlog e o fim do "
         "ano). As do meio flutuam — cada uma começa onde a anterior parou e mostra só o degrau "
         "do mês. Laranja é mês em que a fila cresceu, verde é mês em que ela cedeu.", False),
        ("Por isso a soma visual funciona: da primeira barra até a última, subindo nos laranjas e "
         "descendo nos verdes, chega-se exatamente em %d." % QUADRO[-1][1], False),
        ("", False),
        ("O QUE A CASCATA MOSTRA", True),
        ("O ano tem duas metades. De janeiro a abril a fila SOBE %d (entram 29, saem 8). De maio em "
         "diante ela DESCE %d (entram 31, saem 87)." % (sobe, desce), False),
        ("Três meses fazem o ano: maio (16), novembro (22) e dezembro (20) somam 58 dos %d "
         "resolvidos. Fora deles o ritmo é de 3,4 por mês." % saiu, False),
        ("O meio do ano é de empate: julho, agosto, setembro e outubro entram 22 e resolvem 20 — a "
         "fila fica parada em torno de 55.", False),
        ("", False),
        ("AS TRÊS LEITURAS", True),
        ("«Cascata do ano» — o saldo líquido de cada mês, um degrau por mês. É a leitura de "
         "apresentação.", False),
        ("«Cascata detalhada» — cada mês vira dois degraus, o que entrou e o que se resolveu. "
         "Mostra o que a outra esconde: maio entrou 1 e resolveu 16.", False),
        ("«Pendentes e ritmo» — barras de entrante e resolvidos com a linha do estoque por cima, "
         "no mesmo eixo, porque é tudo equipamento.", False),
        ("", False),
        ("UM AVISO", True),
        ("Este quadro é outro recorte do que está em CURVA_TOTAL_2026.xlsx e ANO_DCMD_2026.xlsx. "
         "Aqui o ano abre em %d e resolve %d de janeiro a agosto; na visão DCMD abre em 50 e "
         "resolve 71 no mesmo período. Não é erro de nenhum dos dois — são universos diferentes —, "
         "mas vale dizer qual é qual antes de pôr os dois na mesma apresentação."
         % (QUADRO[0][1], sum(q[3] for q in QUADRO[1:9])), False),
        ("", False),
        ("Os gráficos leem a aba «Dados». Mudou lá, mudou o gráfico.", False),
    ]
    for i, (t, negrito) in enumerate(texto, 1):
        c = ws.cell(row=i, column=1, value=t)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if negrito:
            c.font = Font(bold=True, size=11, color=SINAL if i == 1 else TINTA)


def montar(saida=SAIDA):
    entrou, saiu = confere()
    wb = Workbook()
    wb.remove(wb.active)
    ws_d = wb.create_sheet("__tmp")            # a aba Dados precisa existir antes das fórmulas
    wb.remove(ws_d)
    fim_dados = None
    ws = wb.create_sheet("__ordem")
    wb.remove(ws)
    fim_dados = aba_dados(wb, entrou, saiu)
    aba_cascata(wb, fim_dados)
    aba_detalhada(wb)
    aba_ritmo(wb, fim_dados)
    aba_como(wb, entrou, saiu)
    wb.move_sheet("Dados", offset=3)           # a cascata na frente, os dados no fim
    os.makedirs(os.path.dirname(saida), exist_ok=True)
    wb.save(saida)
    print(saida)
    print("  quadro confere nos 12 meses | entram %d, saem %d, a fila cede %d"
          % (entrou, saiu, entrou - saiu))
    print("  cascata: %s" % " ".join(
        ("%s%d" % ("+" if QUADRO[i][2] - QUADRO[i][3] > 0 else "", QUADRO[i][2] - QUADRO[i][3]))
        for i in range(1, len(QUADRO))))
    from planilha_automatica import grava_cache
    print("  cache: %d células" % grava_cache(saida))
    return saida


if __name__ == "__main__":
    montar(sys.argv[1] if len(sys.argv) > 1 else SAIDA)
