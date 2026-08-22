"""
O repasse das 71 resolvidas — para onde o COEP passou e quanto tempo ficou lá.

A pergunta do gestor: das demandas que o posto resolveu em 2026, quantas ele
passou para outros postos, e quanto tempo cada uma ficou no posto que recebeu
até ser repassada para um terceiro posto que não seja o COEP.

O relógio, montado pela cadeia (a base não registra saída — SS repassada tem a
conclusão vazia):

    entrega       =  abertura da SS no posto que recebeu
    saída         =  conclusão dessa SS, ou a abertura da SS seguinte
    para onde foi =  o posto da SS seguinte

O que acontece depois separa três destinos, e eles contam histórias diferentes:

    passou adiante   foi para um terceiro posto que não é o COEP
    devolveu         voltou para a mesa do COEP
    fechou ali       a cadeia acabou no posto que recebeu

Universo: as 71 demandas que o COEP resolveu em 2026, de data/missao/coep_2026.json.
Um equipamento pode ter mais de um salto se a demanda foi e voltou.

Grava data/missao/repasse_dos_resolvidos.json e dist/REPASSE_DOS_RESOLVIDOS.xlsx.

Rodar: python3 scripts/repasse_dos_resolvidos.py
"""

import datetime
import json
import os
import re
from collections import Counter, defaultdict

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ARQ_SS = os.path.join(RAIZ, "data", "missao", "ss_ocorrencia.json")
ARQ_COEP = os.path.join(RAIZ, "data", "missao", "coep_2026.json")
SAIDA_JSON = os.path.join(RAIZ, "data", "missao", "repasse_dos_resolvidos.json")
SAIDA_XLSX = os.path.join(RAIZ, "dist", "REPASSE_DOS_RESOLVIDOS.xlsx")

POSTO = "ETO-COEP"
INICIO = datetime.datetime(2026, 1, 1)
FIM = datetime.datetime(2026, 8, 18, 23, 59)
RE_SS = re.compile(r"([A-Z][A-Z-]*)\s+0*(\d+)/(\d{4})")


def norm(numero):
    m = RE_SS.match((numero or "").strip().upper())
    return f"{m.group(1)} {int(m.group(2))}/{m.group(3)}" if m else (numero or "").strip().upper()


def dia(texto):
    try:
        return datetime.datetime.strptime((texto or "")[:19], "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def percentis(valores):
    v = sorted(valores)
    if not v:
        return {}
    def p(q):
        return v[min(len(v) - 1, int(len(v) * q))]
    return {"n": len(v), "mediana": p(.5), "p75": p(.75), "maximo": v[-1],
            "media": round(sum(v) / len(v), 1)}


def montar():
    with open(ARQ_SS, encoding="utf-8") as fh:
        base = json.load(fh)
    idx = {}
    for x in base:
        x["_id"] = norm(x["SS_ORIGINAL"])
        x["_abriu"] = dia(x.get("DTA_ABERTURA"))
        x["_concluiu"] = dia(x.get("DTA_CONCLUSAO"))
        antes = idx.get(x["_id"])
        if antes is None or (x["_abriu"] and antes["_abriu"] and x["_abriu"] > antes["_abriu"]):
            idx[x["_id"]] = x
    seguinte = {x["_id"]: norm(x["SS_APOS_REPASSE"]) for x in idx.values()
                if x.get("SS_APOS_REPASSE")}
    antecessor = {v: k for k, v in seguinte.items()}

    with open(ARQ_COEP, encoding="utf-8") as fh:
        cp = json.load(fh)
    resolvidos = {r["ativo"]: r for r in cp["resolvidos_do_coep"]
                  if r["conta_como_resolvido_pelo_coep"]}
    cidade = {a["ativo"]: a["localidade"] for a in cp["ativos"]}

    def cadeia(numero):
        """A cadeia inteira da demanda, da primeira SS à última."""
        cur = idx.get(norm(numero))
        if not cur:
            return []
        visto = set()
        while cur["_id"] in antecessor and antecessor[cur["_id"]] in idx \
                and antecessor[cur["_id"]] not in visto:
            visto.add(cur["_id"])
            cur = idx[antecessor[cur["_id"]]]
        caminho, visto = [cur], {cur["_id"]}
        while cur["_id"] in seguinte and seguinte[cur["_id"]] in idx \
                and seguinte[cur["_id"]] not in visto:
            cur = idx[seguinte[cur["_id"]]]
            visto.add(cur["_id"])
            caminho.append(cur)
        return caminho

    saltos = []
    for ativo, r in sorted(resolvidos.items()):
        caminho = cadeia(r["ss_no_coep"])
        for i, x in enumerate(caminho):
            if x["POSTO_SGM"] != POSTO or i + 1 >= len(caminho):
                continue
            b = caminho[i + 1]
            if not b["_abriu"] or not (INICIO <= b["_abriu"] <= FIM):
                continue                     # só as entregas feitas dentro de 2026
            prox = caminho[i + 2] if i + 2 < len(caminho) else None
            quando = b["_concluiu"] or (prox["_abriu"] if prox and prox["_abriu"] else None)
            destino = prox["POSTO_SGM"] if prox else ""
            if not destino:
                desfecho = "fechou no posto que recebeu"
            elif destino == POSTO:
                desfecho = "devolveu ao COEP"
            else:
                desfecho = "passou adiante para outro posto"
            saltos.append({
                "ativo": ativo, "tipo": r["tipo"], "cidade": cidade.get(ativo, ""),
                "ano_da_demanda": r["ano_da_demanda"],
                "posto_que_recebeu": b["POSTO_SGM"],
                "ss_do_coep": x["SS_ORIGINAL"], "ss_de_quem_recebeu": b["SS_ORIGINAL"],
                "entregue_em": b["_abriu"].strftime("%d/%m/%Y"),
                "saiu_em": quando.strftime("%d/%m/%Y") if quando else "",
                "dias_no_posto": ((quando or FIM) - b["_abriu"]).days,
                "foi_para": destino, "ss_do_terceiro": prox["SS_ORIGINAL"] if prox else "",
                "desfecho": desfecho,
                "status_de_quem_recebeu": b["STATUS"],
                # o cancelamento aparece nas duas pontas ao mesmo tempo: a nota que o
                # posto recebeu morreu cancelada E a demanda terminou cancelada
                "cancelada": b["STATUS"] == "SS CANCELADA",
                "como_a_demanda_terminou": r["como_terminou"],
                "prova_do_resolvido": r["prova"],
                "data_do_fechamento": r["data_do_fechamento"],
            })

    canceladas = [s for s in saltos if s["cancelada"]]
    limpos = [s for s in saltos if not s["cancelada"]]
    por_desfecho = defaultdict(list)
    for s in saltos:
        por_desfecho[s["desfecho"]].append(s)
    recebeu = Counter(s["posto_que_recebeu"] for s in saltos)
    terceiro = Counter(s["foi_para"] for s in saltos if s["desfecho"].startswith("passou"))

    pacote = {
        "gerado_em": "2026-08-22", "posicao": "18/08/2026",
        "universo": "as 71 demandas que o COEP resolveu em 2026",
        "premissas": PREMISSAS,
        "resumo": {
            "saltos": len(saltos),
            "equipamentos": len(set(s["ativo"] for s in saltos)),
            "resolvidos_no_universo": len(resolvidos),
            "sem_repasse_em_2026": len(resolvidos) - len(set(s["ativo"] for s in saltos)),
            "por_desfecho": {k: {"qtd": len(v),
                                 "canceladas": sum(1 for s in v if s["cancelada"]),
                                 **percentis([s["dias_no_posto"] for s in v])}
                             for k, v in por_desfecho.items()},
            "canceladas": {"saltos": len(canceladas),
                           "equipamentos": len(set(s["ativo"] for s in canceladas)),
                           **percentis([s["dias_no_posto"] for s in canceladas])},
            "sem_cancelamento": {"saltos": len(limpos),
                                 "equipamentos": len(set(s["ativo"] for s in limpos)),
                                 **percentis([s["dias_no_posto"] for s in limpos])},
            "geral": percentis([s["dias_no_posto"] for s in saltos]),
        },
        "quem_recebeu": [{"posto": k, "qtd": v,
                          **percentis([s["dias_no_posto"] for s in saltos
                                       if s["posto_que_recebeu"] == k])}
                         for k, v in recebeu.most_common()],
        "para_onde_foi_depois": [{"posto": k, "qtd": v} for k, v in terceiro.most_common()],
        "saltos": saltos,
    }
    with open(SAIDA_JSON, "w", encoding="utf-8") as fh:
        json.dump(pacote, fh, ensure_ascii=False, indent=1)
    return pacote


PREMISSAS = [
    "Universo: as 71 demandas que o COEP resolveu em 2026. Não é a carteira nem a base "
    "inteira — é só o que o posto fechou no ano.",
    "Salto é uma passagem do COEP para outro posto, com a entrega feita dentro de 2026. O mesmo "
    "equipamento pode ter mais de um salto se a demanda foi e voltou.",
    "Entrega é a abertura da SS no posto que recebeu. O SGM não move a SS: fecha a do COEP como "
    "repassada e abre uma nova no destino, gravando o número em SS_APOS_REPASSE.",
    "Saída é a conclusão dessa SS, se houver; senão a abertura da SS seguinte. A conclusão vem "
    "VAZIA em SS repassada — sem montar a cadeia, o tempo de posto não existe na base.",
    "Três desfechos, contados em separado porque contam coisas diferentes: passou adiante para "
    "um terceiro posto que não é o COEP; devolveu ao COEP; ou a cadeia fechou no posto que "
    "recebeu.",
    "Quem ainda estava no posto em 18/08/2026 conta até essa data.",
    "O cancelamento vai em aba própria. Nesses saltos a nota que o posto recebeu morreu "
    "cancelada E a demanda terminou cancelada — as duas pontas batem, nos dez casos, sem "
    "exceção. Não é serviço executado; é demanda encerrada sem troca de peça.",
]


def planilha(pacote):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    tit = Font(bold=True, color="FFFFFF", size=10)
    fundo = PatternFill("solid", fgColor="1F3864")
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)

    def cabecalho(ws, colunas, larguras):
        ws.append(colunas)
        for c, (col, larg) in enumerate(zip(colunas, larguras), 1):
            cel = ws.cell(row=1, column=c)
            cel.font, cel.fill = tit, fundo
            cel.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[cel.column_letter].width = larg
        ws.freeze_panes = "A2"

    def fechar(ws):
        for linha in ws.iter_rows(min_row=2):
            for cel in linha:
                cel.border = borda
                cel.alignment = Alignment(vertical="top", wrap_text=True)

    r = pacote["resumo"]
    ORDEM = ["passou adiante para outro posto", "devolveu ao COEP",
             "fechou no posto que recebeu"]

    ws = wb.active
    ws.title = "A conta"
    cabecalho(ws, ["O que aconteceu depois da entrega", "Saltos", "Dos quais cancelados",
                   "Dias no posto — mediana", "p75", "Máximo", "Média"],
              [36, 10, 12, 15, 8, 9, 9])
    for nome in ORDEM:
        b = r["por_desfecho"].get(nome)
        if b:
            ws.append([nome, b["qtd"], b.get("canceladas"), b.get("mediana"), b.get("p75"),
                       b.get("maximo"), b.get("media")])
    g = r["geral"]
    ws.append(["TODOS os saltos", r["saltos"], r["canceladas"]["saltos"], g.get("mediana"),
               g.get("p75"), g.get("maximo"), g.get("media")])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append([])
    ws.append(["Separando o cancelamento", "Saltos", "Equipamentos",
               "Dias no posto — mediana", "p75", "Máximo", "Média"])
    for c in ws[ws.max_row]:
        c.font, c.fill = tit, fundo
    for rot, b in (("Terminaram em cancelamento — vão na aba própria", r["canceladas"]),
                   ("Terminaram com serviço executado", r["sem_cancelamento"])):
        ws.append([rot, b["saltos"], b["equipamentos"], b.get("mediana"), b.get("p75"),
                   b.get("maximo"), b.get("media")])
    ws.append([])
    ws.append(["Cobertura", ""])
    for c in ws[ws.max_row]:
        c.font, c.fill = tit, fundo
    ws.append(["Equipamentos com pelo menos um salto", r["equipamentos"]])
    ws.append(["Resolvidos sem nenhum repasse do COEP em 2026", r["sem_repasse_em_2026"]])
    ws.append(["Total de resolvidos no universo", r["resolvidos_no_universo"]])
    fechar(ws)

    ws = wb.create_sheet("Quem recebeu")
    cabecalho(ws, ["Posto que recebeu do COEP", "Saltos", "Dias no posto — mediana", "p75",
                   "Máximo"], [22, 10, 15, 8, 9])
    for x in pacote["quem_recebeu"]:
        ws.append([x["posto"], x["qtd"], x.get("mediana"), x.get("p75"), x.get("maximo")])
    fechar(ws)

    ws = wb.create_sheet("Para onde foi depois")
    cabecalho(ws, ["Terceiro posto", "Saltos"], [22, 10])
    for x in pacote["para_onde_foi_depois"]:
        ws.append([x["posto"], x["qtd"]])
    fechar(ws)

    COLS_SALTO = ["Ativo", "Tipo", "Cidade", "Ano da demanda", "SS do COEP",
                  "Posto que recebeu", "SS de quem recebeu", "Entregue em", "Saiu em",
                  "Dias no posto", "O que aconteceu depois", "Foi para",
                  "SS do terceiro posto"]
    LARG_SALTO = [14, 12, 20, 11, 22, 16, 22, 12, 12, 11, 26, 14, 22]

    def linha_salto(s):
        return [s["ativo"], s["tipo"], s["cidade"], s["ano_da_demanda"], s["ss_do_coep"],
                s["posto_que_recebeu"], s["ss_de_quem_recebeu"], s["entregue_em"],
                s["saiu_em"], s["dias_no_posto"], s["desfecho"], s["foi_para"] or "—",
                s["ss_do_terceiro"]]

    limpos = [s for s in pacote["saltos"] if not s["cancelada"]]
    canc = [s for s in pacote["saltos"] if s["cancelada"]]

    ws = wb.create_sheet(f"Cada salto ({len(limpos)})")
    cabecalho(ws, COLS_SALTO, LARG_SALTO)
    for s in sorted(limpos, key=lambda x: (x["desfecho"], -x["dias_no_posto"])):
        ws.append(linha_salto(s))
    fechar(ws)

    ws = wb.create_sheet(f"Canceladas ({len(canc)})")
    cabecalho(ws, COLS_SALTO + ["Como a demanda terminou", "Fechou em",
                                "Prova do resolvido"],
              LARG_SALTO + [16, 12, 48])
    for s in sorted(canc, key=lambda x: -x["dias_no_posto"]):
        ws.append(linha_salto(s) + [s["como_a_demanda_terminou"], s["data_do_fechamento"],
                                    s["prova_do_resolvido"]])
    fechar(ws)

    ws = wb.create_sheet("Como foi feito")
    cabecalho(ws, ["Passo", "O que foi feito"], [8, 130])
    for n, texto in enumerate(pacote["premissas"], 1):
        ws.append([n, texto])
    fechar(ws)

    os.makedirs(os.path.dirname(SAIDA_XLSX), exist_ok=True)
    wb.save(SAIDA_XLSX)


def main():
    pacote = montar()
    r = pacote["resumo"]
    print(f"saltos do COEP para outro posto, em 2026, nas {r['resolvidos_no_universo']} "
          f"resolvidas: {r['saltos']}")
    print(f"  equipamentos com pelo menos um salto..: {r['equipamentos']}")
    print(f"  resolvidos sem repasse do COEP em 2026: {r['sem_repasse_em_2026']}")
    print("\no que aconteceu depois da entrega:")
    for nome in ("passou adiante para outro posto", "devolveu ao COEP",
                 "fechou no posto que recebeu"):
        b = r["por_desfecho"].get(nome)
        if b:
            print(f"  {b['qtd']:>3}  {nome:<34} mediana {b['mediana']:>4}d | "
                  f"p75 {b['p75']:>4}d | máx {b['maximo']:>4}d")
    print("\nquem recebeu do COEP:")
    for x in pacote["quem_recebeu"]:
        print(f"  {x['qtd']:>3}  {x['posto']:<12} mediana {x['mediana']:>4}d | "
              f"máx {x['maximo']:>4}d")
    print(f"\ncom cancelamento: {r['canceladas']['saltos']} saltos em "
          f"{r['canceladas']['equipamentos']} equipamentos "
          f"(mediana {r['canceladas'].get('mediana')}d, máx {r['canceladas'].get('maximo')}d)")
    print(f"sem cancelamento: {r['sem_cancelamento']['saltos']} saltos em "
          f"{r['sem_cancelamento']['equipamentos']} equipamentos "
          f"(mediana {r['sem_cancelamento'].get('mediana')}d, "
          f"máx {r['sem_cancelamento'].get('maximo')}d)")
    print(f"\ngravado: {SAIDA_JSON}")
    planilha(pacote)
    print(f"gravado: {SAIDA_XLSX}")


if __name__ == "__main__":
    main()
