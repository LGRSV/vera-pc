"""
A cadeia SS → OS → Obra → AIC, na base crua.

O gestor pediu o caminho inteiro: pega a SS na base de SS/OS, pega a OS dela,
pega o número da obra que a SS declara e cruza essa obra com o AIC.

Como é feito, em ordem:

  1. remonta os registros da base crua (dois arquivos, separador «@», latin-1).
     A descrição da SS tem quebra de linha dentro, então um registro ocupa
     várias linhas do arquivo — o começo de registro é reconhecido pelo padrão
     do número da SS (ETO-XXX 00000/0000).
  2. fica só com o que é religador ou regulador — pelo código do ativo no
     cadastro do gestor, pela ORIGEM_SS ou pela descrição do ativo.
  3. monta o trio SS · OS · Obra. O NUM_OBRA vem numérico na base, com 9
     dígitos; o AIC guarda 10 com zero à esquerda — por isso o zfill(10).
  4. cruza a obra com o AIC (extrato de 07/08/2026, aba única «Export») e traz
     projeto SIGCO, status, descrição, datas e valor.

Grava data/missao/cadeia_obra.json e dist/CADEIA_SS_OS_OBRA.xlsx.

Rodar: python3 scripts/cadeia_obra.py
"""

import json
import os
import re
import sys
from collections import Counter, defaultdict

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))

UPLOADS = "/root/.claude/uploads/c3d5c486-5de5-52ac-a54a-1691b373e364"


def _base_mais_nova():
    """A BASE_SS_OS*.txt mais nova em data/raw — pela data ddmmaaaa do nome e,
    empatando, pela hora do arquivo. Largar a base nova lá basta para todos os
    scripts passarem a usá-la. O nome não diz o horizonte do dado: conferir a
    data máxima de abertura depois de extrair."""
    import glob
    achadas = glob.glob(os.path.join(RAIZ, "data", "raw", "BASE_SS_OS*.txt"))

    def chave(p):
        m = re.search(r"(\d{2})(\d{2})(\d{4})", os.path.basename(p))
        data = (m.group(3), m.group(2), m.group(1)) if m else ("0000", "00", "00")
        return (data, os.path.getmtime(p))
    return max(achadas, key=chave) if achadas else None


# A base mais nova manda; as partes de 11/08 ficam de reserva para reproduzir o passado.
_BASE_NOVA = _base_mais_nova()
PARTES = ([_BASE_NOVA] if _BASE_NOVA else
          [os.path.join(UPLOADS, "470fbe86-BASE_SS_OS_parte1.txt"),
           os.path.join(UPLOADS, "43942150-BASE_SS_OS_parte2.txt")])
AIC_XLSX = os.path.join(UPLOADS, "fa750c58-AIC_OBRAS_07082026.xlsx")
CACHE_AIC = os.path.join(RAIZ, "data", "missao", "aic_full.json")
SAIDA_JSON = os.path.join(RAIZ, "data", "missao", "cadeia_obra.json")
SAIDA_XLSX = os.path.join(RAIZ, "dist", "CADEIA_SS_OS_OBRA.xlsx")

# O código do posto não tem forma fixa: ETO-COEP, DOLP-RD-PA, ETO-CADTOC, ETO-SCADA,
# ETO-TEC01 e até DMSLETO, sem hífen. O que identifica o começo de registro é o
# «@» colado no ano — separador de campo, que texto livre não produz.
RE_INICIO = re.compile(r"^[A-Z][A-Z0-9-]{1,14}\s+\d{5}/\d{4}@")
RE_RL = re.compile(r"RELIGADOR|RELIG\b", re.I)
RE_RT = re.compile(r"REGULADOR|REG\.?\s*DE\s*TENS", re.I)

# campos que interessam, pela posição da esquerda (antes da descrição livre)
POS = {
    "NUMERO_SS": 0, "NUMERO_OS": 1, "NUM_OBRA": 2, "ORIGEM_SS": 3, "DEFEITO_SS": 4,
    "ORIGEM": 5, "DEFEITO": 6, "ESQUEMA": 7, "TASK": 8, "ANO": 9, "ORG_SOLIC": 10,
    "COD_EQUIPE": 11, "ALIMENTADOR": 12, "NUM_TRAFO": 13, "COORD_X": 14, "COORD_Y": 15,
    "COD_GIS": 16, "CRITICIDADE_SS": 17, "SITUACAO_SS": 18, "DATA_ABERTURA_SS": 19,
    "DATA_TERMINO_SS": 20, "DATA_LIMITE_SS": 21, "DESCICAO_DO_ATIVO": 22,
    "LOCALIDADE": 23, "SOLICITANTE": 24, "RURA_URBANO": 25, "TIPOSS": 26,
}
N_CAMPOS = 64


def registros():
    """Remonta os registros da base crua, um dicionário por SS/OS."""
    for caminho in PARTES:
        buffer = None
        with open(caminho, encoding="latin-1") as fh:
            for i, linha in enumerate(fh):
                linha = linha.rstrip("\r\n")
                if i == 0 and linha.startswith("NUMERO_SS@"):
                    continue
                if RE_INICIO.match(linha):
                    if buffer is not None:
                        yield _parse(buffer)
                    buffer = linha
                elif buffer is not None:
                    buffer += "\n" + linha
            if buffer is not None:
                yield _parse(buffer)


def _parse(bruto):
    campos = bruto.split("@")
    reg = {nome: (campos[p].strip() if p < len(campos) else "") for nome, p in POS.items()}
    # a descrição da SS e da OS são o resto até a cauda; a cauda tem 35 campos
    corpo = campos[len(POS):]
    if len(campos) >= N_CAMPOS:
        cauda = campos[-(N_CAMPOS - len(POS) - 2):]
        reg["DESCRICAO"] = "@".join(corpo[:len(corpo) - len(cauda)]).strip()
    else:
        reg["DESCRICAO"] = "@".join(corpo).strip()
    reg["_n_campos"] = len(campos)
    return reg


def obra10(valor):
    """O AIC guarda 10 dígitos; a base entrega o número sem o zero à esquerda."""
    v = (valor or "").strip()
    if not v or not v.replace(".0", "").isdigit():
        return ""
    v = v.split(".")[0]
    return v.zfill(10) if v not in ("0", "") else ""


def familia(reg, frota):
    cod = (reg.get("NUM_TRAFO") or "").strip()
    if cod in frota:
        return frota[cod]["familia"]
    texto = f"{reg.get('ORIGEM_SS','')} {reg.get('DESCICAO_DO_ATIVO','')}"
    if RE_RT.search(texto):
        return "regulador"
    if RE_RL.search(texto):
        return "religador"
    return None


def aic_completo():
    """Índice do AIC inteiro por número de obra, com as colunas que interessam."""
    if os.path.exists(CACHE_AIC):
        with open(CACHE_AIC, encoding="utf-8") as fh:
            return json.load(fh)
    import openpyxl
    quero = ["NUM_OBRA", "NUM_OS", "DSC_STATUS", "DSC_OCORRENCIA", "NUM_PROJETO_SIGCO",
             "AIC", "AREA", "CLASS_OBRA", "CONST_MANUT", "DESCRICAO", "DESCRICAO_OBRA",
             "NOMELOC", "POLO", "REGIONAL", "STATUS AIC", "STATUS PRAZO", "TIPO_OBRA",
             "DTH_ABERTURA", "DATA_CONCLUSAO_FISICA", "DTH_ENCERRAMENTO_TECNICO",
             "DTH_ENCERRAMENTO", "DTH_TERMINO_FISICO", "VAL_TOTAL_ORCADO", "TOTAL_REALIZADO"]
    wb = openpyxl.load_workbook(AIC_XLSX, read_only=True, data_only=True)
    ws = wb["Export"]
    it = ws.iter_rows(values_only=True)
    hdr = [str(h) if h is not None else "" for h in next(it)]
    idx = {nome: hdr.index(nome) for nome in quero if nome in hdr}
    saida = {}
    for linha in it:
        if len(linha) <= idx["NUM_OBRA"]:
            continue
        num = linha[idx["NUM_OBRA"]]
        if num is None:
            continue
        num = str(num).split(".")[0].strip().zfill(10)
        reg = {}
        for nome, p in idx.items():
            v = linha[p] if p < len(linha) else None
            reg[nome] = "" if v is None else (str(v)[:10] if nome.startswith(("DTH_", "DATA_")) else str(v).strip())
        saida[num] = reg
    wb.close()
    with open(CACHE_AIC, "w", encoding="utf-8") as fh:
        json.dump(saida, fh, ensure_ascii=False)
    return saida


RE_SUBST_EQP = re.compile(
    r"(?:SUBST\w*|TROCA\w*|INSTALA[ÇC][ÃA]O)\s+(?:D[EO]\s+|D[AO]S?\s+)?(?:\d+\s+)?"
    r"(?:ATIVO\s+D[EO]\s+)?(RELIGADOR|REGULADOR|RECLOSER|BANCO\s+REGULADOR)", re.I)
RE_PECA_GRANDE = re.compile(
    r"TANQUE|PARTE ATIVA|C[ÉE]LULA|ARM[ÁA]RIO DE CONTROLE|RETROFIT|"
    r"PLACA DE ALIMENTA[ÇC][ÃA]O(\s+CA)?|REL[EÉÊ]\s*DE\s*SINCRONISMO|"
    r"(?:RELIGADOR|REGULADOR)\s+FURTAD|FURTO D[EO]\s+(?:RELIGADOR|REGULADOR)|"
    r"CONTROLE D[EO]\s+(?:RELIGADOR|REGULADOR)", re.I)
RE_TRAFO_AUX = re.compile(
    r"TRAFO\s+AUX|TRANSFORMADOR\s+AUX|AUXU?ILIAR\s+(?:NO|DO)\s+RELIGADOR|"
    r"TRAFO\s+(?:DE\s+)?15\s*KVA", re.I)
RE_REDE_TEXTO = re.compile(
    r"POSTE|CONDUTOR|\bCABO\b|CHAVES?\s+(?:FUS|FACA|SECC|SECCIONADORA)|CRUZETA|PODA|"
    r"ISOLADOR|P[ÁA]RA-?RAIO|JUMPER|TRAFO QUEIMADO|SOBRECARGA|ESPA[ÇC]ADOR|EMENDA|ESTAI", re.I)
CERTO = {"religador": "8495", "regulador": "8481"}

# Onde o texto da OS desmente o texto do parecer. O parecer da SS conta o defeito;
# a OS conta o que a obra pagou. Quando os dois discordam, quem manda é a OS —
# é ela que amarra o dinheiro da obra. Cada linha aqui foi lida uma a uma.
REVISAO_TEXTO = {
    "0612500378": ("acessorio", "parecer fala em tanque estourado, mas a OS desta obra "
                                "pagou 03 chaves secc faca; o tanque saiu por outra obra"),
    "0712600039": ("rede", "OS pagou estrutura N3 e chave fusível; o ativo da SS "
                           "(3300274066) é chave, não religador"),
    "0212501060": ("remanejamento", "remanejamento de religador com poste, estrutura e "
                                    "chaves — muda de lugar, não repõe peça"),
    "0612600709": ("rede", "ponto quente na CONEXÃO DE ENTRADA do tanque — é conexão, "
                           "não o tanque"),
    "0612400758": ("rede", "trafo do posto 5701410019, que não tem os 8 dígitos finais do "
                           "ativo da SS; a SS é de limpeza de faixa"),
    "0412500987": ("rede", "as duas SS são de poste; o texto de trafo vem de registro de "
                           "outro ativo colado na descrição"),
    "0662500325": ("melhoria", "rebaixar o armário de controle — melhoria de altura, "
                               "não troca de peça"),
}


def ler_objeto(texto_obra, texto_ss):
    """O que a obra fez, afinal — pelo texto da obra primeiro, pelo da SS depois.

    Ordem: peça grande do próprio equipamento manda sobre tudo; depois trafo
    auxiliar; depois rede. Texto que não diz nada fica indefinido — é o caso das
    obras cujo título é só «MANUTENÇÃO CORRET EMERGENCIAL».
    """
    for fonte, texto in (("obra", texto_obra), ("ss", texto_ss)):
        if not texto:
            continue
        if RE_PECA_GRANDE.search(texto) or RE_SUBST_EQP.search(texto):
            return "equipamento", fonte
        if RE_TRAFO_AUX.search(texto):
            return "trafo auxiliar", fonte
        if RE_REDE_TEXTO.search(texto):
            return "rede", fonte
    return "indefinido", ""


def main():
    import taxa_falha as tf
    frota = tf.parque()
    aic = aic_completo()

    total, rlrt, com_obra = 0, 0, 0
    cadeia = {}                      # obra -> registro consolidado
    sem_obra_por_familia = Counter()
    por_ss = []
    for reg in registros():
        total += 1
        fam = familia(reg, frota)
        if fam is None:
            continue
        rlrt += 1
        num = obra10(reg.get("NUM_OBRA"))
        if not num:
            sem_obra_por_familia[fam] += 1
            continue
        com_obra += 1
        ano_ss = (reg.get("DATA_ABERTURA_SS") or "")[6:10]
        item = {
            "ss": reg["NUMERO_SS"], "os": reg["NUMERO_OS"], "obra": num,
            "familia": fam, "ativo": reg.get("NUM_TRAFO", ""),
            "situacao_ss": reg.get("SITUACAO_SS", ""), "tipo_ss": reg.get("TIPOSS", ""),
            "origem_ss": reg.get("ORIGEM_SS", ""), "abertura": reg.get("DATA_ABERTURA_SS", "")[:10],
            "termino": reg.get("DATA_TERMINO_SS", "")[:10], "ano_ss": ano_ss,
            "localidade": reg.get("LOCALIDADE", ""), "equipe": reg.get("COD_EQUIPE", ""),
            "texto": re.sub(r"\s+", " ", reg.get("DESCRICAO", ""))[:4000],
            "esquema": reg.get("ESQUEMA", ""), "defeito_ss": reg.get("DEFEITO_SS", ""),
        }
        item["objeto"] = tf.objeto_do_fato({"ORIGEM_SS": reg.get("ORIGEM_SS"),
                                            "ESQUEMA": reg.get("ESQUEMA"),
                                            "TIPOSS": reg.get("TIPOSS")})
        por_ss.append(item)
        alvo = cadeia.setdefault(num, {"obra": num, "familia": fam, "ss": set(), "os": set(),
                                       "ativos": set(), "anos_ss": set()})
        alvo["ss"].add(item["ss"])
        if item["os"]:
            alvo["os"].add(item["os"])
        if item["ativo"]:
            alvo["ativos"].add(item["ativo"])
        if ano_ss:
            alvo["anos_ss"].add(ano_ss)

    # cruza com o AIC
    achadas, orfas = 0, []
    for num, alvo in cadeia.items():
        a = aic.get(num)
        alvo["ss"] = sorted(alvo["ss"]); alvo["os"] = sorted(alvo["os"])
        alvo["ativos"] = sorted(alvo["ativos"]); alvo["anos_ss"] = sorted(alvo["anos_ss"])
        if a:
            achadas += 1
            alvo["aic"] = a
        else:
            orfas.append(num)
            alvo["aic"] = None

    # veredito de cada obra: o que ela fez e se está no projeto SIGCO certo
    veredito = Counter()
    fora_do_par = []
    for num, alvo in cadeia.items():
        a = alvo["aic"] or {}
        texto_ss = " ".join(i["texto"] for i in por_ss if i["obra"] == num)
        obj, fonte = ler_objeto(a.get("DESCRICAO_OBRA", ""), texto_ss)
        # SS cujo fato já é da rede pela origem/esquema não vira obra de equipamento
        objetos_ss = {i["objeto"] for i in por_ss if i["obra"] == num}
        if obj == "indefinido" and objetos_ss == {"rede"}:
            obj, fonte = "rede", "origem da SS"
        sig = (a.get("NUM_PROJETO_SIGCO") or "").split(".")[0] or "(sem projeto)"
        if num in REVISAO_TEXTO:
            obj, motivo = REVISAO_TEXTO[num]
            alvo["revisao"] = motivo
            fonte = "leitura da OS"
        alvo["objeto"] = obj
        alvo["fonte_do_objeto"] = fonte
        alvo["sigco"] = sig
        alvo["sigco_certo"] = CERTO[alvo["familia"]]
        alvo["no_projeto_certo"] = sig == CERTO[alvo["familia"]]
        veredito[obj] += 1
        if alvo["aic"] is not None and obj in ("equipamento", "trafo auxiliar") \
                and not alvo["no_projeto_certo"]:
            fora_do_par.append(alvo)

    print(f"registros remontados............ {total}")
    print(f"  de religador/regulador........ {rlrt}")
    print(f"  com número de obra............ {com_obra}  (sem obra: {dict(sem_obra_por_familia)})")
    print(f"obras distintas declaradas...... {len(cadeia)}")
    print(f"  encontradas no AIC............ {achadas}")
    print(f"  órfãs (não estão no AIC)...... {len(orfas)}")
    print(f"o que a obra fez................ {dict(veredito)}")
    print(f"obras de equipamento/trafo fora do projeto certo: {len(fora_do_par)}")

    with open(SAIDA_JSON, "w", encoding="utf-8") as fh:
        json.dump({"obras": list(cadeia.values()), "ss": por_ss,
                   "fora_do_par": [o["obra"] for o in fora_do_par],
                   "resumo": {"registros": total, "rlrt": rlrt, "com_obra": com_obra,
                              "obras": len(cadeia), "no_aic": achadas, "orfas": orfas,
                              "objeto": dict(veredito)}},
                  fh, ensure_ascii=False)
    print(f"gravado: {SAIDA_JSON}")
    with open(SAIDA_JSON, encoding="utf-8") as fh:
        planilha(json.load(fh))


def planilha(pacote):
    """A planilha do gestor: a cadeia inteira, obra a obra, com a conta do AIC."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    obras = {o["obra"]: o for o in pacote["obras"]}
    ss_por = defaultdict(list)
    for i in pacote["ss"]:
        ss_por[i["obra"]].append(i)

    wb = openpyxl.Workbook()
    tit = Font(bold=True, color="FFFFFF", size=10)
    fundo = PatternFill("solid", fgColor="1F3864")
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)

    def cabecalho(ws, colunas, larguras):
        ws.append(colunas)
        for c, (col, larg) in enumerate(zip(colunas, larguras), 1):
            cel = ws.cell(row=1, column=c)
            cel.font, cel.fill = tit, fundo
            cel.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[cel.column_letter].width = larg
        ws.freeze_panes = "A2"

    def fechar(ws):
        for linha in ws.iter_rows(min_row=2):
            for cel in linha:
                cel.border = borda
                cel.alignment = Alignment(vertical="top", wrap_text=True)

    def num(v):
        try:
            return float(str(v).replace(",", "."))
        except (TypeError, ValueError):
            return None

    # 1) a cadeia, linha a linha
    ws = wb.active
    ws.title = "Cadeia SS-OS-Obra"
    cabecalho(ws, ["SS", "OS", "Obra", "Família", "Ativo", "Tipo da SS", "Abertura",
                   "Projeto SIGCO", "O que a obra fez", "Descrição da obra no AIC",
                   "Status no AIC", "Conclusão física", "Realizado (R$)"],
              [22, 24, 13, 11, 13, 24, 12, 13, 16, 46, 26, 14, 14])
    for i in sorted(pacote["ss"], key=lambda x: (x["obra"], x["ss"])):
        o = obras[i["obra"]]
        a = o["aic"] or {}
        ws.append([i["ss"], i["os"], i["obra"], o["familia"], i["ativo"], i["tipo_ss"],
                   i["abertura"], o["sigco"], o["objeto"], a.get("DESCRICAO_OBRA", ""),
                   a.get("DSC_STATUS", ""), a.get("DATA_CONCLUSAO_FISICA", ""),
                   num(a.get("TOTAL_REALIZADO"))])
    fechar(ws)

    # 2) uma linha por obra
    ws = wb.create_sheet("Obras no AIC")
    cabecalho(ws, ["Obra", "Família", "SS que declaram", "OS", "Ativos", "Projeto SIGCO",
                   "Projeto certo da família", "Está no projeto certo?", "O que a obra fez",
                   "Descrição da obra no AIC", "Tipo da obra", "Status no AIC",
                   "Conclusão física", "Orçado (R$)", "Realizado (R$)"],
              [13, 11, 30, 30, 24, 13, 12, 12, 16, 46, 26, 26, 14, 14, 14])
    for o in sorted(obras.values(), key=lambda x: x["obra"]):
        a = o["aic"] or {}
        ws.append([o["obra"], o["familia"], "\n".join(o["ss"]), "\n".join(o["os"]),
                   "\n".join(o["ativos"]), o["sigco"], o["sigco_certo"],
                   "sim" if o["no_projeto_certo"] else "não", o["objeto"],
                   a.get("DESCRICAO_OBRA", ""), a.get("TIPO_OBRA", ""), a.get("DSC_STATUS", ""),
                   a.get("DATA_CONCLUSAO_FISICA", ""), num(a.get("VAL_TOTAL_ORCADO")),
                   num(a.get("TOTAL_REALIZADO"))])
    fechar(ws)

    # 3) o que sai do lugar
    ws = wb.create_sheet("Fora do projeto certo")
    cabecalho(ws, ["Obra", "Família", "Está no SIGCO", "Deveria estar no", "O que a obra fez",
                   "Ano da conclusão", "Realizado (R$)", "Descrição da obra no AIC",
                   "O que a SS e a OS dizem"],
              [13, 11, 14, 14, 16, 14, 14, 44, 70])
    for num_obra in pacote["fora_do_par"]:
        o = obras[num_obra]
        a = o["aic"] or {}
        texto = " || ".join(i["texto"][:400] for i in ss_por[num_obra])
        ws.append([o["obra"], o["familia"], o["sigco"], o["sigco_certo"], o["objeto"],
                   (a.get("DATA_CONCLUSAO_FISICA") or "")[:4], num(a.get("TOTAL_REALIZADO")),
                   a.get("DESCRICAO_OBRA", ""), texto])
    fechar(ws)

    # 4) o que foi descartado e por quê
    ws = wb.create_sheet("Descartadas na leitura")
    cabecalho(ws, ["Obra", "Família", "Está no SIGCO", "Veredito", "Por que não entra",
                   "Descrição da obra no AIC"], [13, 11, 14, 16, 66, 44])
    for num_obra, (vered, motivo) in REVISAO_TEXTO.items():
        o = obras.get(num_obra)
        if not o:
            continue
        ws.append([num_obra, o["familia"], o["sigco"], vered, motivo,
                   (o["aic"] or {}).get("DESCRICAO_OBRA", "")])
    fechar(ws)

    # 5) como foi feito
    ws = wb.create_sheet("Como foi feito")
    cabecalho(ws, ["Passo", "O que foi feito"], [8, 130])
    for n, texto in enumerate(PASSOS, 1):
        ws.append([n, texto])
    fechar(ws)

    os.makedirs(os.path.dirname(SAIDA_XLSX), exist_ok=True)
    wb.save(SAIDA_XLSX)
    print(f"gravado: {SAIDA_XLSX}")


PASSOS = [
    "Base usada: BASE_SS_OS (extração de 11/08/2026, dois arquivos de texto separados por «@») "
    "e o extrato do AIC de 07/08/2026, aba única «Export». O AIC é um só arquivo.",
    "A descrição da SS tem quebra de linha dentro dela, então um registro ocupa várias linhas do "
    "arquivo. Os registros foram remontados pelo padrão do número da SS — o código do posto não "
    "tem forma fixa (ETO-COEP, DOLP-RD-PA, ETO-CADTOC, DMSLETO), o que identifica o começo é o "
    "«@» colado no ano.",
    "Ficaram os de religador e regulador — pelo código do ativo no cadastro, pela ORIGEM_SS ou pela "
    "descrição do ativo. As contagens desta extração estão na aba «Obras no AIC» e no rodapé.",
    "Dentro deles, só os que declaram número de obra no campo NUM_OBRA entram na cadeia; o resto é "
    "serviço de equipe, sem obra aberta.",
    "O NUM_OBRA vem numérico na base, com 9 dígitos; o AIC guarda 10, com zero à esquerda. Por isso "
    "o número foi completado com zero antes de cruzar. Sem isso, nenhuma obra casa.",
    "Obra declarada por SS aberta DEPOIS do extrato do AIC (07/08) pode não estar no extrato: "
    "fica como órfã, sem julgamento de SIGCO — o número dela só existe no SGM por enquanto.",
    "Cada obra foi classificada pelo que ela fez, lendo primeiro o texto da obra no AIC e depois o "
    "texto da SS e da OS: equipamento (peça grande do religador ou do regulador), trafo auxiliar, "
    "rede (poste, cabo, chave, cruzeta, poda, para-raio) ou indefinido.",
    "Quando o parecer da SS e o texto da OS discordam, vale o texto da OS: o parecer conta o defeito, "
    "a OS conta o que a obra pagou. Sete obras foram descartadas por isso — estão na aba "
    "«Descartadas na leitura», uma a uma, com o motivo.",
    "Projeto certo, pela régua do gestor: 8495 para religador, 8481 para regulador. Trafo auxiliar "
    "acompanha o projeto do equipamento pai — o código dele tem prefixo 51 (padrão) ou 57 e os 8 "
    "dígitos finais iguais aos do pai.",
    "Resultado: as obras de equipamento ou de trafo auxiliar fora do projeto certo estão na aba "
    "própria, uma a uma, com o texto da SS ao lado.",
]

if __name__ == "__main__":
    main()
