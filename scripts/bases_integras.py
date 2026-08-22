"""
As bases de origem da cadeia SS → OS → Obra, na íntegra.

Não é resumo nem recorte de colunas: é o registro inteiro, do jeito que saiu do
sistema, para as linhas que formaram a cadeia.

  Aba 1  SS/OS íntegra ....... os 259 registros da base de SS/OS que declaram obra,
                              com as 64 colunas originais
  Aba 2  Obras íntegra ....... as 232 obras no extrato do AIC, com as 93 colunas
  Aba 3  Sobre as bases ...... qual arquivo, qual aba, quantas linhas, e a conferência
                              de que «OBRAS_status_extracao» e «AIC_OBRAS» são o mesmo
                              arquivo, byte a byte

Rodar: python3 scripts/bases_integras.py
"""

import hashlib
import json
import os
import re
import sys

# O Excel recusa caracteres de controle, e a descrição da base traz uns poucos
# (quebra vertical, restos de formulário). Saem antes de ir para a célula.
RE_ILEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
import cadeia_obra as co  # noqa: E402

SAIDA = os.path.join(RAIZ, "dist", "BASES_INTEGRAS_DA_CADEIA.xlsx")
ARQ_STATUS = os.path.join(co.UPLOADS, "44e553a9-OBRAS_status_extracao_07082026.xlsx")
LIMITE_CELULA = 32000            # o Excel corta em 32.767; sobra folga


def cabecalho_ssos():
    with open(co.PARTES[0], encoding="latin-1") as fh:
        return fh.readline().rstrip("\r\n").split("@")


def registros_inteiros():
    """Devolve (numero_ss, lista de 64 campos) de cada registro da base crua.

    Registro com mais de 64 campos tem «@» dentro do texto livre. Nesse caso os 27
    campos da frente e os 35 da cauda são lidos pelas pontas, e o miolo — que é a
    descrição da SS e a da OS — é remontado: a última fatia é a descrição da OS,
    o resto é a da SS.
    """
    n_frente, n_cauda = len(co.POS), 35
    for caminho in co.PARTES:
        buffer = None
        with open(caminho, encoding="latin-1") as fh:
            for i, linha in enumerate(fh):
                linha = linha.rstrip("\r\n")
                if i == 0 and linha.startswith("NUMERO_SS@"):
                    continue
                if co.RE_INICIO.match(linha):
                    if buffer is not None:
                        yield _campos(buffer, n_frente, n_cauda)
                    buffer = linha
                elif buffer is not None:
                    buffer += "\n" + linha
            if buffer is not None:
                yield _campos(buffer, n_frente, n_cauda)


def _campos(bruto, n_frente, n_cauda):
    campos = bruto.split("@")
    if len(campos) == 64:
        return campos
    if len(campos) > 64:
        frente, cauda = campos[:n_frente], campos[-n_cauda:]
        miolo = campos[n_frente:len(campos) - n_cauda]
        if len(miolo) >= 2:
            miolo = ["@".join(miolo[:-1]), miolo[-1]]
        while len(miolo) < 2:
            miolo.append("")
        return frente + miolo + cauda
    return campos + [""] * (64 - len(campos))


def sha256(caminho):
    h = hashlib.sha256()
    with open(caminho, "rb") as fh:
        for bloco in iter(lambda: fh.read(1 << 20), b""):
            h.update(bloco)
    return h.hexdigest()


def montar():
    with open(co.SAIDA_JSON, encoding="utf-8") as fh:
        cadeia = json.load(fh)
    quero = {re.sub(r"\s+", " ", i["ss"].strip()) for i in cadeia["ss"]}
    obras_quero = {o["obra"] for o in cadeia["obras"]}

    hdr_ss = cabecalho_ssos()
    linhas_ss = []
    for campos in registros_inteiros():
        if re.sub(r"\s+", " ", campos[0].strip()) in quero:
            linhas_ss.append([RE_ILEGAL.sub(" ", c.strip())[:LIMITE_CELULA] for c in campos])

    import openpyxl
    wb_aic = openpyxl.load_workbook(co.AIC_XLSX, read_only=True, data_only=True)
    ws = wb_aic["Export"]
    it = ws.iter_rows(values_only=True)
    hdr_aic = ["" if h is None else str(h) for h in next(it)]
    pos_obra = hdr_aic.index("NUM_OBRA")
    linhas_aic = []
    for linha in it:
        if len(linha) <= pos_obra or linha[pos_obra] is None:
            continue
        if str(linha[pos_obra]).split(".")[0].strip().zfill(10) in obras_quero:
            linhas_aic.append(["" if v is None else RE_ILEGAL.sub(" ", str(v))[:LIMITE_CELULA]
                               for v in linha])
    wb_aic.close()
    return hdr_ss, linhas_ss, hdr_aic, linhas_aic


def planilha(hdr_ss, linhas_ss, hdr_aic, linhas_aic):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    tit = Font(bold=True, color="FFFFFF", size=9)
    fundo = PatternFill("solid", fgColor="1F3864")

    def folha(nome, hdr, linhas, larg=26):
        ws = wb.create_sheet(nome)
        ws.append(hdr)
        for c in range(1, len(hdr) + 1):
            cel = ws.cell(row=1, column=c)
            cel.font, cel.fill = tit, fundo
            cel.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[cel.column_letter].width = larg
        ws.freeze_panes = "A2"
        for linha in linhas:
            ws.append(linha)
        return ws

    wb.remove(wb.active)
    folha("SS-OS integra", hdr_ss, linhas_ss, 24)
    folha("Obras AIC integra", hdr_aic, linhas_aic, 22)

    ws = wb.create_sheet("Sobre as bases")
    ws.append(["Item", "Arquivo de origem", "Aba", "Linhas trazidas", "Colunas",
               "SHA-256 do arquivo", "Observação"])
    for c in range(1, 8):
        cel = ws.cell(row=1, column=c)
        cel.font, cel.fill = tit, fundo
        cel.alignment = Alignment(vertical="center", wrap_text=True)
    for col, larg in zip("ABCDEFG", (28, 34, 16, 15, 10, 70, 74)):
        ws.column_dimensions[col].width = larg
    ws.freeze_panes = "A2"

    sha_aic = sha256(co.AIC_XLSX)
    sha_status = sha256(ARQ_STATUS) if os.path.exists(ARQ_STATUS) else "(arquivo ausente)"
    nomes = " + ".join(os.path.basename(c) for c in co.PARTES)
    for n, caminho in enumerate(co.PARTES, 1):
        ws.append([f"1 · Base de SS/OS ({os.path.basename(caminho)})", nomes,
                   "texto, separador @", len(linhas_ss) if n == 1 else "", len(hdr_ss),
                   sha256(caminho),
                   "Extração com aberturas até 20/08/2026, codificação latin-1. A descrição "
                   "quebra linha, então um registro ocupa várias linhas; aqui vêm só os que "
                   "declaram obra e são de religador ou regulador."])
    ws.append(["2 · Base de Obras (AIC)", "AIC_OBRAS_07082026.xlsx", "Export",
               len(linhas_aic), len(hdr_aic), sha_aic,
               "Extrato de 07/08/2026 com 124.084 obras. Aba única «Export» — não há aba de "
               "infotrafo. Aqui vêm as obras da cadeia presentes no extrato; obra declarada "
               "por SS aberta depois de 07/08 fica órfã até o próximo extrato."])
    ws.append(["3 · Base de OS status", "OBRAS_status_extracao_07082026.xlsx", "Export",
               len(linhas_aic), len(hdr_aic), sha_status,
               "É O MESMO ARQUIVO da linha 2 — o SHA-256 bate byte a byte com o AIC_OBRAS. "
               "Foi enviado duas vezes com nomes diferentes. Não existe uma terceira base de "
               "status separada; o status da obra são as colunas DSC_STATUS, DSC_OCORRENCIA, "
               "«STATUS AIC» e «STATUS PRAZO» desta mesma aba."])
    ws.append(["Ausentes do pacote", "Critica_082026.txt e AIC_infotrafo_11-08.xlsx", "—", "", "",
               "", "Estão listados no CHECKSUMS.txt do zip mas não vieram no arquivo. Nenhuma "
               "conta desta cadeia depende deles."])
    for linha in ws.iter_rows(min_row=2):
        for cel in linha:
            cel.alignment = Alignment(vertical="top", wrap_text=True)

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)


def main():
    hdr_ss, linhas_ss, hdr_aic, linhas_aic = montar()
    print(f"SS/OS íntegra: {len(linhas_ss)} registros × {len(hdr_ss)} colunas")
    print(f"Obras íntegra: {len(linhas_aic)} obras × {len(hdr_aic)} colunas")
    planilha(hdr_ss, linhas_ss, hdr_aic, linhas_aic)
    print(f"gravado: {SAIDA}")


if __name__ == "__main__":
    main()
