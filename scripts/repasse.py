"""
Quando foi o repasse — reconstruído pela cadeia de SS.

O problema: a coluna DTA_REPASSE da base não serve. Ela é cópia byte a byte da
DTA_ABERTURA nas 10.386 linhas, sem uma única exceção. Quem lê essa coluna como
«data em que a SS foi repassada» está lendo a data em que a SS CHEGOU no posto.

A data do repasse está em outro lugar: é a DTA_ABERTURA da SS seguinte. O SGM não
move a SS de posto — ele fecha a SS como REPASSADA e abre uma nova no posto de
destino, gravando o número dela em SS_APOS_REPASSE. Então:

    repasse de A  =  abertura de B, onde B = SS_APOS_REPASSE de A
    tempo de A no posto  =  abertura de B − abertura de A

Seguindo o SS_APOS_REPASSE de ponta a ponta sai a cadeia inteira da demanda, com
o caminho dos postos e o tempo parado em cada um.

A base de obras (extrato do AIC) não entra aqui: obra não sabe de repasse. Ela
serve para o outro lado do relógio — quando a obra foi concluída fisicamente,
que é em média uns dias DEPOIS de a SS ser fechada. Onde a SS da ponta declara
obra, essa data entra como fim da linha.

Grava data/missao/repasse.json e dist/REPASSE.xlsx.

Rodar: python3 scripts/repasse.py
"""

import datetime
import json
import os
import re
from collections import Counter, defaultdict

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ARQ_SS = os.path.join(RAIZ, "data", "missao", "ss_ocorrencia.json")
ARQ_CADEIA = os.path.join(RAIZ, "data", "missao", "cadeia_obra.json")
SAIDA_JSON = os.path.join(RAIZ, "data", "missao", "repasse.json")
SAIDA_XLSX = os.path.join(RAIZ, "dist", "REPASSE.xlsx")

RE_SS = re.compile(r"([A-Z][A-Z-]*)\s+0*(\d+)/(\d{4})")


def norm(numero):
    m = RE_SS.match((numero or "").strip().upper())
    return f"{m.group(1)} {int(m.group(2))}/{m.group(3)}" if m else (numero or "").strip().upper()


def dia(texto):
    try:
        return datetime.datetime.strptime((texto or "")[:19], "%Y-%m-%d %H:%M:%S")
    except ValueError:
        try:
            return datetime.datetime.strptime((texto or "")[:10], "%Y-%m-%d")
        except ValueError:
            return None


def carregar():
    with open(ARQ_SS, encoding="utf-8") as fh:
        base = json.load(fh)
    for x in base:
        x["_id"] = norm(x["SS_ORIGINAL"])
        x["_abertura"] = dia(x.get("DTA_ABERTURA"))
        x["_conclusao"] = dia(x.get("DTA_CONCLUSAO"))
        x["_proxima"] = norm(x["SS_APOS_REPASSE"]) if x.get("SS_APOS_REPASSE") else ""
    return base


def fim_pela_obra():
    """Data de conclusão física da obra, por SS que declara obra — o fim da linha."""
    if not os.path.exists(ARQ_CADEIA):
        return {}
    with open(ARQ_CADEIA, encoding="utf-8") as fh:
        ca = json.load(fh)
    obras = {o["obra"]: o for o in ca["obras"]}
    saida = {}
    for i in ca["ss"]:
        a = (obras.get(i["obra"]) or {}).get("aic") or {}
        if a.get("DATA_CONCLUSAO_FISICA"):
            saida[norm(i["ss"])] = {"obra": i["obra"],
                                    "conclusao_fisica": a["DATA_CONCLUSAO_FISICA"][:10],
                                    "realizado": a.get("TOTAL_REALIZADO", "")}
    return saida


def montar():
    base = carregar()
    idx = {}
    for x in base:                      # 15 números repetidos: fica o de abertura mais nova
        antes = idx.get(x["_id"])
        if antes is None or (x["_abertura"] and antes["_abertura"]
                             and x["_abertura"] > antes["_abertura"]):
            idx[x["_id"]] = x
    obra_fim = fim_pela_obra()

    sucessor = {x["_id"]: x["_proxima"] for x in idx.values() if x["_proxima"]}
    apontadas = set(sucessor.values())
    quebrados = sorted(v for v in apontadas if v not in idx)

    # 1) cada repasse, com data e tempo parado
    repasses, sem_data, invertidos = [], 0, 0
    tempo_por_posto = defaultdict(list)
    for de, para in sucessor.items():
        a, b = idx[de], idx.get(para)
        if b is None:
            sem_data += 1
            continue
        if not (a["_abertura"] and b["_abertura"]):
            sem_data += 1
            continue
        dias = (b["_abertura"] - a["_abertura"]).days
        if dias < 0:
            invertidos += 1
            continue
        repasses.append({
            "de": a["SS_ORIGINAL"], "posto_de": a["POSTO_SGM"],
            "para": b["SS_ORIGINAL"], "posto_para": b["POSTO_SGM"],
            "equipamento": a["EQUIPAMENTO"], "tipo": a["TIPO_ATIVO"],
            "chegou": a["_abertura"].strftime("%d/%m/%Y"),
            "repassou": b["_abertura"].strftime("%d/%m/%Y"),
            "dias_no_posto": dias, "status_da_ss": a["STATUS"],
            "ano": a["_abertura"].year,
        })
        tempo_por_posto[a["POSTO_SGM"]].append(dias)

    # 2) a cadeia inteira, do começo ao fim
    cadeias = []
    for raiz in (x for x in idx.values() if x["_id"] not in apontadas):
        caminho, atual, visto = [], raiz, set()
        while atual is not None and atual["_id"] not in visto:
            visto.add(atual["_id"])
            caminho.append(atual)
            prox = sucessor.get(atual["_id"])
            atual = idx.get(prox) if prox else None
        if len(caminho) < 2:
            continue
        inicio, fim = caminho[0]["_abertura"], caminho[-1]
        obra = obra_fim.get(caminho[-1]["_id"])
        desfecho = fim["_conclusao"]
        if obra:
            d = dia(obra["conclusao_fisica"])
            if d and (desfecho is None or d > desfecho):
                desfecho = d
        cadeias.append({
            "ss_inicial": caminho[0]["SS_ORIGINAL"], "ss_final": fim["SS_ORIGINAL"],
            "equipamento": caminho[0]["EQUIPAMENTO"], "tipo": caminho[0]["TIPO_ATIVO"],
            "passos": len(caminho),
            "caminho": " → ".join(p["POSTO_SGM"] for p in caminho),
            "abriu": inicio.strftime("%d/%m/%Y") if inicio else "",
            "fechou": desfecho.strftime("%d/%m/%Y") if desfecho else "",
            "dias_total": (desfecho - inicio).days if (desfecho and inicio) else None,
            "status_final": fim["STATUS"],
            "obra": (obra or {}).get("obra", ""),
            "conclusao_da_obra": (obra or {}).get("conclusao_fisica", ""),
            "ano": inicio.year if inicio else None,
        })

    # 3) o tempo de cada posto
    postos = []
    for posto, v in tempo_por_posto.items():
        v = sorted(v)
        postos.append({"posto": posto, "repasses": len(v), "mediana": v[len(v) // 2],
                       "p90": v[int(len(v) * 0.9)], "maximo": v[-1],
                       "media": round(sum(v) / len(v), 1),
                       "acima_de_30d": sum(1 for x in v if x > 30)})
    postos.sort(key=lambda p: -p["repasses"])

    todos = sorted(r["dias_no_posto"] for r in repasses)
    pacote = {
        "gerado_em": "2026-08-22",
        "fonte": "EQP_SS_OCORRENCIA_11082026 (10.386 SS de religador e regulador) — "
                 "cadeia reconstruída pelo campo SS_APOS_REPASSE",
        "achado_da_coluna": {
            "coluna": "DTA_REPASSE",
            "veredito": "não serve — é cópia da DTA_ABERTURA nas 10.386 linhas, sem exceção",
            "linhas_iguais": 10386, "linhas_diferentes": 0,
        },
        "premissas": PREMISSAS,
        "cobertura": {
            "ss_na_base": len(idx),
            "ss_com_sucessor": len(sucessor),
            "sucessor_presente_na_base": len(sucessor) - len(quebrados),
            "elos_quebrados": quebrados,
            "repasses_com_data": len(repasses),
            "descartados_sem_data": sem_data,
            "descartados_data_invertida": invertidos,
            "cadeias_de_2_ou_mais": len(cadeias),
        },
        "tempo_ate_repassar": {
            "mediana": todos[len(todos) // 2], "p75": todos[int(len(todos) * .75)],
            "p90": todos[int(len(todos) * .90)], "maximo": todos[-1],
            "media": round(sum(todos) / len(todos), 1),
        },
        "por_posto": postos,
        "repasses": repasses,
        "cadeias": cadeias,
    }
    with open(SAIDA_JSON, "w", encoding="utf-8") as fh:
        json.dump(pacote, fh, ensure_ascii=False, indent=1)
    return pacote


PREMISSAS = [
    "A data do repasse não está na SS que foi repassada — está na SS que nasceu do repasse. "
    "O SGM fecha a SS de origem como REPASSADA e abre uma nova no posto de destino.",
    "Por isso: repasse de A = abertura de B, onde B é a SS gravada em SS_APOS_REPASSE de A. "
    "E o tempo de A parado no posto é a diferença entre as duas aberturas.",
    "A coluna DTA_REPASSE da base foi conferida linha a linha e é cópia da DTA_ABERTURA nas "
    "10.386 linhas. Ela não diz quando repassou; diz quando chegou. Não foi usada.",
    "Entra como repasse toda SS que aponta uma sucessora, não só as de status REPASSADA: "
    "301 canceladas e 33 atendidas também apontam sucessora — a demanda andou do mesmo jeito.",
    "Repasse com data invertida (sucessora aberta antes da origem) foi descartado, não corrigido.",
    "Elo quebrado é quando a SS aponta uma sucessora que não está nesta base — são poucos e "
    "todos apontam para postos fora do escopo de religador e regulador, como o ETO-CADTOC.",
    "Cadeia é a demanda inteira, do primeiro posto ao último. Só entram as de 2 passos ou mais; "
    "SS que nunca foi repassada não forma cadeia.",
    "O fim da linha é a conclusão da última SS. Onde essa última SS declara obra, e a obra foi "
    "concluída depois, vale a conclusão física da obra — o serviço só acabou de verdade ali.",
    "Quinze números de SS aparecem repetidos na base; ficou a linha de abertura mais recente.",
]


def planilha(pacote):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    tit = Font(bold=True, color="FFFFFF", size=10)
    fundo = PatternFill("solid", fgColor="1F3864")
    borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)

    def cabecalho(ws, colunas, larguras):
        ws.append(colunas)
        for c, (col, larg) in enumerate(zip(colunas, larguras), 1):
            cel = ws.cell(row=1, column=c)
            cel.font, cel.fill = tit, fundo
            cel.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[cel.column_letter].width = larg
        ws.freeze_panes = "A2"

    def fechar(ws):
        for linha in ws.iter_rows(min_row=2):
            for cel in linha:
                cel.border = borda
                cel.alignment = Alignment(vertical="top", wrap_text=True)

    ws = wb.active
    ws.title = "Tempo por posto"
    cabecalho(ws, ["Posto", "Repasses", "Mediana (dias)", "p90 (dias)", "Máximo (dias)",
                   "Média (dias)", "Repasses acima de 30 dias"],
              [16, 11, 14, 12, 14, 12, 16])
    for p in pacote["por_posto"]:
        ws.append([p["posto"], p["repasses"], p["mediana"], p["p90"], p["maximo"],
                   p["media"], p["acima_de_30d"]])
    fechar(ws)

    ws = wb.create_sheet("Repasses")
    cabecalho(ws, ["SS de origem", "Posto de origem", "Chegou em", "Repassou em",
                   "Dias parada no posto", "SS de destino", "Posto de destino",
                   "Equipamento", "Tipo", "Status da SS de origem"],
              [22, 16, 13, 13, 13, 22, 16, 14, 12, 18])
    for r in sorted(pacote["repasses"], key=lambda x: -x["dias_no_posto"]):
        ws.append([r["de"], r["posto_de"], r["chegou"], r["repassou"], r["dias_no_posto"],
                   r["para"], r["posto_para"], r["equipamento"], r["tipo"], r["status_da_ss"]])
    fechar(ws)

    ws = wb.create_sheet("Cadeias")
    cabecalho(ws, ["SS inicial", "SS final", "Passos", "Caminho dos postos", "Abriu",
                   "Fechou", "Dias do começo ao fim", "Situação final", "Equipamento",
                   "Tipo", "Obra", "Conclusão da obra"],
              [22, 22, 8, 46, 12, 12, 14, 18, 14, 12, 13, 14])
    for c in sorted(pacote["cadeias"], key=lambda x: -(x["dias_total"] or 0)):
        ws.append([c["ss_inicial"], c["ss_final"], c["passos"], c["caminho"], c["abriu"],
                   c["fechou"], c["dias_total"], c["status_final"], c["equipamento"],
                   c["tipo"], c["obra"], c["conclusao_da_obra"]])
    fechar(ws)

    ws = wb.create_sheet("Como foi feito")
    cabecalho(ws, ["Passo", "O que foi feito"], [8, 130])
    for n, texto in enumerate(pacote["premissas"], 1):
        ws.append([n, texto])
    fechar(ws)

    os.makedirs(os.path.dirname(SAIDA_XLSX), exist_ok=True)
    wb.save(SAIDA_XLSX)


def main():
    pacote = montar()
    c, t = pacote["cobertura"], pacote["tempo_ate_repassar"]
    print(f"SS na base.......................... {c['ss_na_base']}")
    print(f"  com sucessora declarada........... {c['ss_com_sucessor']}")
    print(f"  sucessora presente na base........ {c['sucessor_presente_na_base']}"
          f"  (elos quebrados: {len(c['elos_quebrados'])})")
    print(f"repasses com data reconstruída...... {c['repasses_com_data']}"
          f"  (sem data: {c['descartados_sem_data']}, invertidos: {c['descartados_data_invertida']})")
    print(f"cadeias de 2 passos ou mais......... {c['cadeias_de_2_ou_mais']}")
    print(f"tempo até repassar.................. mediana {t['mediana']}d | p75 {t['p75']}d "
          f"| p90 {t['p90']}d | máx {t['maximo']}d")
    print(f"gravado: {SAIDA_JSON}")
    planilha(pacote)
    print(f"gravado: {SAIDA_XLSX}")


if __name__ == "__main__":
    main()
