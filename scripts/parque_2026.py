"""
As quatro curvas de 2026, separadas por religador e regulador.

Pedido do gestor (24/08). Ele deu a base do parque e a expansão realizada mês a mês;
o resto sai das bases:

  1. PARQUE — cresce com a expansão. Base de janeiro: 1.281 religadores e 180
     reguladores, e a expansão do mês soma no próprio mês («até janeiro 1281, soma
     mais 2 em janeiro»).
  2. ENTRANTES FORA DE OPERAÇÃO — equipamento com SS de indisponibilidade para
     operação, pela DATA DE OCORRÊNCIA, contado uma vez por mês (base de ocorrência).
  3. REALIZADO — concluído pelo DCMD no mês (régua do revisor: equipe de campo na
     cadeia e demanda fechada com SS atendida).
  4. TAXA DE FALHA MENSAL — equipamentos que falharam no mês (peça grande, pela
     leitura das SS revisada, mais a troca por obra direta datada pela conclusão
     física) ÷ parque daquele mês.

Grava data/missao/parque_2026.json.
Rodar: python3 scripts/parque_2026.py
"""

import json
import os
import re
import sys
from collections import Counter, defaultdict

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
SAIDA = os.path.join(RAIZ, "data", "missao", "parque_2026.json")

MESES = [f"2026-{m:02d}" for m in range(1, 9)]
ROTULOS = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago"]

# Régua do gestor (24/08): parque de janeiro e expansão realizada por mês.
BASE_JANEIRO = {"RL": 1281, "RT": 180}
EXPANSAO = {  # jan a jul; agosto ainda sem número fechado
    "RL": [2, 0, 2, 2, 3, 1, 3],
    "RT": [0, 3, 3, 1, 1, 1, 1],
}
RE_OBRA_SUBST = re.compile(
    r"SUBSTITUI[ÇC][ÃA]O\s+D[EO]\s*(?:\d+\s+)?(?:CH\.?\s+)?(?:ATIVO\s+D[EO]\s+)?"
    r"(RELIGADOR|REGULADOR)", re.I)
RE_COD = re.compile(r"\b(\d{10})\b")


def _mes(data_br):
    """dd/mm/aaaa -> aaaa-mm."""
    d = str(data_br or "")
    return f"{d[6:10]}-{d[3:5]}" if len(d) >= 10 else ""


def parque():
    saida = {}
    for t, base in BASE_JANEIRO.items():
        exp, serie, acum = EXPANSAO[t], [], base
        for i, _ in enumerate(MESES):
            se_tem = exp[i] if i < len(exp) else 0
            acum += se_tem
            serie.append({"mes": MESES[i], "expansao": se_tem, "parque": acum,
                          "expansao_conhecida": i < len(exp)})
        saida[t] = serie
    return saida


def entrantes():
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(RAIZ, "data", "raw",
                                             "EQP_SS_OCORRENCIA_11082026.xlsx"))
    ws = wb["Exportar Planilha"]
    cab = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    i = {c: n for n, c in enumerate(cab)}
    por = defaultdict(set)
    for l in ws.iter_rows(min_row=2, values_only=True):
        mes = str(l[i["DTA_OCORRENCIA"]])[:7]
        if mes not in MESES:
            continue
        if str(l[i["PENDENCIA_DO_ATIVO"]]).strip().upper() != "INDISPONIBILIDADE PARA OPERAÇÃO":
            continue
        t = "RL" if str(l[i["TIPO_ATIVO"]]) == "RELIGADOR" else "RT"
        por[(t, mes)].add(str(l[i["EQUIPAMENTO"]]))
    wb.close()
    return {t: [len(por[(t, m)]) for m in MESES] for t in ("RL", "RT")}


def realizado():
    with open(os.path.join(RAIZ, "data", "missao", "visao_consolidada.json"),
              encoding="utf-8") as fh:
        vc = json.load(fh)
    por = defaultdict(int)
    for a in vc["concluidos_dcmd_2026"]["ativos"]:
        t = "RL" if a["tipo"].lower().startswith("relig") else "RT"
        m = _mes(a["fechou_em"])
        if m in MESES:
            por[(t, m)] += 1
    return {t: [por[(t, m)] for m in MESES] for t in ("RL", "RT")}


def falhas():
    """Equipamentos que falharam no mês, pela régua da peça grande."""
    with open(os.path.join(RAIZ, "data", "missao", "leitura_ss_os.json"),
              encoding="utf-8") as fh:
        leitura = json.load(fh)
    por = defaultdict(set)
    lidos = set()
    for d in leitura["detalhe"]:
        if d.get("ano") != 2026:
            continue
        lidos.add(d["ativo"])
        m = _mes(d.get("data"))
        if m in MESES:
            t = "RL" if d["familia"] == "religador" else "RT"
            por[(t, m)].add(d["ativo"])
    # o complemento: troca por obra direta, fora da carteira lida, datada pela
    # conclusão física da obra — é assim que o total do ano é montado
    with open(os.path.join(RAIZ, "data", "missao", "aic_rlrt.json"), encoding="utf-8") as fh:
        obras = json.load(fh)
    todos_lidos = set()
    for v in (leitura.get("equipamentos") or {}).values():
        todos_lidos |= set(v)
    comp = defaultdict(int)
    for o in obras:
        texto = o.get("DESCRICAO_OBRA") or ""
        m_ = RE_OBRA_SUBST.search(texto)
        if not m_:
            continue
        concl = str(o.get("DATA_CONCLUSAO_FISICA") or "")[:7]
        if concl not in MESES:
            continue
        if set(RE_COD.findall(texto)) & todos_lidos:
            continue
        t = "RL" if m_.group(1).upper().startswith("RELIG") else "RT"
        comp[(t, concl)] += 1
    return ({t: [len(por[(t, m)]) for m in MESES] for t in ("RL", "RT")},
            {t: [comp[(t, m)] for m in MESES] for t in ("RL", "RT")})


def cumulativo():
    """A carteira de indisponibilidade por tipo, com o acervo dos anos anteriores.

    É a conta do gestor aplicada ao equipamento: o que já estava fora de operação em
    1º de janeiro, mais o que entrou mês a mês, menos o que fechou — e a distância
    entre as duas curvas é a fila daquele mês. A demanda é a CADEIA de SS encadeada
    (repasse não é demanda nova); só entram cadeias com SS de indisponibilidade.

    Cancelamento não tem data de conclusão no SGM, então a cadeia cancelada é datada
    pela abertura da última SS dela — aproximação, e vai dita na página.
    """
    import demandas as dm
    with open(os.path.join(RAIZ, "data", "missao", "ssos_min.json"), encoding="utf-8") as fh:
        ssos = json.load(fh)
    por_ativo = defaultdict(list)
    for r in ssos:
        por_ativo[r["NUM_TRAFO"]].append(r)

    demandas = {"RL": [], "RT": []}
    for cod, linhas in por_ativo.items():
        t = "RT" if cod.startswith("58") else "RL"
        for d in dm.encadear(linhas):
            if not any(l.get("TIPOSS") == "INDISPONIBILIDADE PARA OPERAÇÃO" for l in d["ss"]):
                continue
            r = dm.resumir_demanda(d)
            fim = r["termino"]
            if r["situacao"] == "cancelada" and not fim:
                bruto = max((l.get("DATA_ABERTURA_SS", "") for l in d["ss"]),
                            key=lambda s: (s[6:10], s[3:5], s[:2]), default="")
                fim = f"{bruto[6:10]}-{bruto[3:5]}-{bruto[:2]}" if len(bruto) >= 10 else ""
            demandas[t].append({"ativo": cod, "abertura": r["abertura"], "fim": fim,
                                "situacao": r["situacao"]})

    saida = {}
    for t in ("RL", "RT"):
        L = demandas[t]
        # acervo: cadeia aberta antes de 2026 e ainda viva na virada do ano
        acervo = sum(1 for x in L if x["abertura"] < "2026-01-01"
                     and (x["situacao"] == "aberta" or (x["fim"] or "9999") >= "2026-01-01"))
        ent, fec = Counter(), Counter()
        for x in L:
            if x["abertura"][:7] in MESES:
                ent[x["abertura"][:7]] += 1
            if x["situacao"] in ("concluída", "cancelada") and x["fim"][:7] in MESES:
                fec[x["fim"][:7]] += 1
        linhas_mes, entrou, resolveu = [], acervo, 0
        for i, m in enumerate(MESES):
            entrou += ent[m]
            resolveu += fec[m]
            linhas_mes.append({"mes": m, "rotulo": ROTULOS[i],
                               "entraram_no_mes": ent[m], "resolvidos_no_mes": fec[m],
                               "entraram_acumulado": entrou, "resolvidos_acumulado": resolveu,
                               "fila": entrou - resolveu})
        saida[t] = {"acervo_em_janeiro": acervo,
                    "entraram_no_ano": sum(ent.values()),
                    "resolvidos_no_ano": sum(fec.values()),
                    "fila_em_agosto": linhas_mes[-1]["fila"],
                    "abertas_hoje": sum(1 for x in L if x["situacao"] == "aberta"),
                    "meses": linhas_mes}
    return saida


def resolvidos_pelo_coep():
    """Os resolvidos PELO POSTO, por tipo e mês — a conta dos 82 (era 71), que o
    gestor pediu subdividida entre religador e regulador."""
    with open(os.path.join(RAIZ, "data", "missao", "coep_2026.json"), encoding="utf-8") as fh:
        cp = json.load(fh)
    por = defaultdict(int)
    for r in cp["resolvidos_do_coep"]:
        if not r["conta_como_resolvido_pelo_coep"]:
            continue
        t = "RT" if r["tipo"].lower().startswith("regul") else "RL"
        m = _mes(r["data_do_fechamento"])
        if m in MESES:
            por[(t, m)] += 1
    return {t: [por[(t, m)] for m in MESES] for t in ("RL", "RT")}


def montar():
    pq = parque()
    ent = entrantes()
    real = realizado()
    lidas, comp = falhas()
    cum = cumulativo()
    coep = resolvidos_pelo_coep()
    for t in ("RL", "RT"):
        acumulado = 0
        for i, linha_mes in enumerate(cum[t]["meses"]):
            acumulado += coep[t][i]
            linha_mes["resolvidos_coep_no_mes"] = coep[t][i]
            linha_mes["resolvidos_coep_acumulado"] = acumulado
        cum[t]["resolvidos_coep_no_ano"] = acumulado

    series = {}
    for t in ("RL", "RT"):
        linhas = []
        for i, m in enumerate(MESES):
            f = lidas[t][i] + comp[t][i]
            p = pq[t][i]["parque"]
            linhas.append({
                "mes": m, "rotulo": ROTULOS[i],
                "parque": p, "expansao": pq[t][i]["expansao"],
                "expansao_conhecida": pq[t][i]["expansao_conhecida"],
                "entrantes": ent[t][i],
                "realizado": real[t][i],
                "falhas": f, "falhas_pela_leitura": lidas[t][i],
                "falhas_por_obra_direta": comp[t][i],
                "taxa_mes_pct": round(100 * f / p, 3) if p else 0.0,
            })
        series[t] = linhas

    pacote = {
        "gerado_em": "2026-08-24",
        "meses": MESES, "rotulos": ROTULOS,
        "regua": {
            "parque": "base de janeiro do gestor (RL 1.281 · RT 180) mais a expansão "
                      "realizada, somada no próprio mês; agosto repete julho porque a "
                      "expansão de agosto ainda não fechou",
            "entrantes": "equipamento com SS de indisponibilidade para operação, pela "
                         "data de OCORRÊNCIA, contado uma vez por mês",
            "realizado": "concluído pelo DCMD: equipe de campo na cadeia e demanda "
                         "fechada com SS atendida, pelo mês do fechamento",
            "taxa": "equipamentos que falharam no mês (peça grande) ÷ parque do mês; "
                    "no ano, ativo que falha duas vezes conta uma vez, então a soma "
                    "dos meses passa do total anual",
        },
        "base_janeiro": BASE_JANEIRO,
        "expansao": EXPANSAO,
        "series": series,
        "cumulativo": cum,
        "totais": {t: {
            "parque_final": series[t][-1]["parque"],
            "expansao_no_ano": sum(EXPANSAO[t]),
            "entrantes": sum(x["entrantes"] for x in series[t]),
            "realizado": sum(x["realizado"] for x in series[t]),
            "falhas_somadas": sum(x["falhas"] for x in series[t]),
        } for t in ("RL", "RT")},
    }
    with open(SAIDA, "w", encoding="utf-8") as fh:
        json.dump(pacote, fh, ensure_ascii=False, indent=1)
    return pacote


if __name__ == "__main__":
    p = montar()
    for t in ("RL", "RT"):
        print(f"\n=== {t}")
        print(f"{'mês':<6}{'parque':>8}{'exp':>5}{'entrantes':>11}{'realizado':>11}"
              f"{'falhas':>8}{'taxa %':>9}")
        for x in p["series"][t]:
            print(f"{x['rotulo']:<6}{x['parque']:>8}{x['expansao']:>5}{x['entrantes']:>11}"
                  f"{x['realizado']:>11}{x['falhas']:>8}{x['taxa_mes_pct']:>9.2f}")
        tt = p["totais"][t]
        print(f"  parque final {tt['parque_final']} (+{tt['expansao_no_ano']} no ano) · "
              f"entrantes {tt['entrantes']} · realizado {tt['realizado']} · "
              f"falhas somadas {tt['falhas_somadas']}")
    for t in ("RL", "RT"):
        c = p["cumulativo"][t]
        print(f"\n{t} cumulativo: acervo {c['acervo_em_janeiro']} + entraram "
              f"{c['entraram_no_ano']} − resolvidos {c['resolvidos_no_ano']} = "
              f"{c['fila_em_agosto']} (abertas hoje: {c['abertas_hoje']})")
    print(f"\ngravado: {SAIDA}")
