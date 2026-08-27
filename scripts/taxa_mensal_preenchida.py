"""
A tabela mensal de taxa de falha do gestor, preenchida.

Ele mandou dois prints: a aba «taxa de falha» (quatro quadros mensais, RL/RT ×
2025/2026) e uma tabela achatada com chave Concat que estava pela metade. O pedido
foi preencher a segunda com o que a primeira diz.

O que a conferência mostrou, e vale saber antes de usar: os quadros do gestor contam
OCORRÊNCIA do rol lido nas SS (35 · 18 · 28 · 9), não EQUIPAMENTO que falhou no ano
(62 · 30 · 31 · 12). A diferença é o complemento por obra direta do AIC — falha
provada pela obra, sem narrativa de SS, que não tem mês para cair. Por isso o
mensal soma menos que o anual, e as duas contas estão certas cada uma no seu lugar.

A classe de tensão sai do cadastro de ajustes da própria GESTÃO DE EQUIPAMENTOS
(coluna TENSÃO em «Ajustes RL Poste», TENSÃO PRIMÁRIA em «Ajustes Reguladores»),
que é a fonte viva do parque — 1.292 religadores e 189 reguladores cadastrados.

Grava dist/TAXA_MENSAL_PREENCHIDA.xlsx.
Rodar: python3 scripts/taxa_mensal_preenchida.py
"""

import json
import os
from collections import Counter, defaultdict

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GESTAO = os.path.join(RAIZ, "data", "raw", "GESTAO_DE_EQUIPAMENTOS.xlsx")
LEITURA = os.path.join(RAIZ, "data", "missao", "leitura_ss_os.json")
PARQUE = os.path.join(RAIZ, "data", "missao", "parque_2026.json")
SAIDA = os.path.join(RAIZ, "dist", "TAXA_MENSAL_PREENCHIDA.xlsx")

MESES = ["janeiro", "fevereiro", "março", "abril", "maio", "junho",
         "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]
CURTO = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
         "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
CLASSES = ["34.500", "13.800"]
SIG = {"religador": "RL", "regulador": "RT"}
# base de janeiro do gestor (régua de 24/08) e a expansão realizada, somada no mês
BASE_2025 = {"RL": 1281, "RT": 180}

TITULO = Font(bold=True, color="FFFFFF", size=10)
FUNDO = PatternFill("solid", fgColor="1F3864")
CINZA = PatternFill("solid", fgColor="DDE3EE")
BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def classe_de_tensao():
    """{codigo: '34.500'|'13.800'} do cadastro de ajustes — a fonte viva do parque."""
    wb = openpyxl.load_workbook(GESTAO, read_only=True, data_only=True)
    t = {}
    for aba, col, largura in (("Ajustes RL Poste", 12, 14),
                              ("Ajustes Reguladores de Tensão", 8, 20)):
        for linha in list(wb[aba].iter_rows(max_col=largura, values_only=True))[1:]:
            cod = str(linha[0] or "").strip()
            if not (cod.isdigit() and len(cod) == 10):
                continue
            v = str(linha[col] or "").strip().replace("34500", "34.500").replace("13800", "13.800")
            if v in CLASSES:
                t.setdefault(cod, v)
    return t


def parque_por_classe(tensao):
    """Quantos equipamentos de cada tipo em cada classe, pelo cadastro."""
    c = Counter()
    for cod, v in tensao.items():
        c[("RL" if cod[:2] in ("79", "78") else "RT", v)] += 1
    return c


def falhas(tensao):
    """Ocorrências do rol por tipo/ano/mês e por classe — a conta dos quadros."""
    with open(LEITURA, encoding="utf-8") as fh:
        d = json.load(fh)
    por_mes = defaultdict(int)
    por_mes_classe = defaultdict(int)
    for r in d["detalhe"]:
        if r["ano"] not in (2025, 2026) or len(r.get("data", "")) < 10:
            continue
        tipo, m = SIG[r["familia"]], int(r["data"][3:5])
        por_mes[(tipo, r["ano"], m)] += 1
        por_mes_classe[(tensao.get(r["ativo"], "SEM CADASTRO"), tipo, r["ano"], m)] += 1
    return por_mes, por_mes_classe, d


def parque_mensal():
    """Parque de cada mês: 2025 fixo na base de janeiro; 2026 com a expansão."""
    with open(PARQUE, encoding="utf-8") as fh:
        p = json.load(fh)
    parque, novos = {}, {}
    for tipo in ("RL", "RT"):
        for m in range(1, 13):
            parque[(tipo, 2025, m)] = BASE_2025[tipo]
            novos[(tipo, 2025, m)] = 0
        serie = p["series"][tipo]
        ultimo = serie[-1]["parque"]
        for i in range(12):
            if i < len(serie):
                parque[(tipo, 2026, i + 1)] = serie[i]["parque"]
                novos[(tipo, 2026, i + 1)] = serie[i]["expansao"]
            else:
                parque[(tipo, 2026, i + 1)] = ultimo
                novos[(tipo, 2026, i + 1)] = 0
    return parque, novos


def cabecalho(ws, colunas, linha=None):
    if linha is None:
        ws.append([c[0] for c in colunas])
        linha = ws.max_row
    for i, (nome, larg) in enumerate(colunas, 1):
        cel = ws.cell(row=linha, column=i, value=nome)
        cel.font, cel.fill = TITULO, FUNDO
        cel.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.row_dimensions[linha].height = 30


def bordar(ws, de, ate):
    for linha in ws.iter_rows(min_row=de, max_row=ate):
        for cel in linha:
            cel.border = BORDA


def aba_achatada(wb, por_mes_classe, pq_classe, parque, novos):
    """A tabela do print, com as mesmas colunas e a mesma chave Concat."""
    ws = wb.active
    ws.title = "1 · Tabela preenchida"
    colunas = [("TAXA DE TENSÃO", 15), ("TIPO", 8), ("Ano", 8), ("Mês", 12),
               ("Concat", 26), ("Qtd RL Fora de operação", 15), ("Parque RL", 11),
               ("RL 13,8", 10), ("RL 34,5", 10), ("Qtd RT Fora de operação", 15),
               ("Parque RT", 11), ("RT 13,8", 10), ("RT 34,5", 10),
               ("Acumulado no ano", 13), ("Taxa do mês", 12)]
    cabecalho(ws, colunas)
    prim = ws.max_row + 1
    acum = defaultdict(int)
    for classe in CLASSES:
        for tipo in ("RL", "RT"):
            for ano in (2025, 2026):
                for i, mes in enumerate(MESES, 1):
                    f_rl = por_mes_classe[(classe, "RL", ano, i)]
                    f_rt = por_mes_classe[(classe, "RT", ano, i)]
                    minha = f_rl if tipo == "RL" else f_rt
                    acum[(classe, tipo, ano)] += minha
                    base = pq_classe[(tipo, classe)]
                    ws.append([classe, tipo, ano, mes, f"{classe}{tipo}{ano}{mes}",
                               f_rl, pq_classe[("RL", classe)],
                               pq_classe[("RL", "13.800")], pq_classe[("RL", "34.500")],
                               f_rt, pq_classe[("RT", classe)],
                               pq_classe[("RT", "13.800")], pq_classe[("RT", "34.500")],
                               acum[(classe, tipo, ano)],
                               (minha / base) if base else 0])
    bordar(ws, prim, ws.max_row)
    for r in range(prim, ws.max_row + 1):
        ws.cell(row=r, column=15).number_format = "0.00%"
        for c in range(1, 16):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    ws.freeze_panes = f"A{prim}"
    ws.auto_filter.ref = f"A{prim - 1}:O{ws.max_row}"
    return ws


def aba_quadros(wb, por_mes, parque, novos):
    """Os quatro quadros da aba de taxa, no mesmo desenho — RT 2026 incluído."""
    ws = wb.create_sheet("2 · Quadros mensais")
    ws.column_dimensions["A"].width = 16
    for c in "BCDEFGHIJKLM":
        ws.column_dimensions[c].width = 9
    faixas = []
    for tipo, ano in (("RL", 2025), ("RT", 2025), ("RL", 2026), ("RT", 2026)):
        ws.append([f"Taxa de Falha Mensal — {tipo} — {ano}"] + [None] * 12)
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
        topo = ws.max_row
        ws.append([""] + CURTO)
        for cel in ws[ws.max_row][1:]:
            cel.font, cel.fill = TITULO, FUNDO
            cel.alignment = Alignment(horizontal="center")
        ws.append(["Novos ativos"] + [novos[(tipo, ano, m)] for m in range(1, 13)])
        ws.append(["Parque"] + [parque[(tipo, ano, m)] for m in range(1, 13)])
        ws.append(["Falha mês"] + [por_mes[(tipo, ano, m)] for m in range(1, 13)])
        linha_falha, linha_parque = ws.max_row, ws.max_row - 1
        ws.append(["Taxa do mês"] + [
            f"={get_column_letter(m + 1)}{linha_falha}/{get_column_letter(m + 1)}{linha_parque}"
            for m in range(1, 13)])
        for cel in ws[ws.max_row][1:]:
            cel.number_format = "0.00%"
        for r in range(topo + 1, ws.max_row + 1):
            ws.cell(row=r, column=1).font = Font(bold=True)
        bordar(ws, topo + 1, ws.max_row)
        faixas.append((tipo, ano, topo + 1, linha_falha, ws.max_row))
        ws.append([])
        ws.append([])
    for tipo, ano, cab, lf, lt in faixas:
        ch = LineChart()
        ch.title = f"{tipo} {ano} — taxa de falha do mês"
        ch.height, ch.width, ch.style = 6.5, 16, 2
        ch.add_data(Reference(ws, min_col=1, min_row=lt, max_col=13, max_row=lt),
                    titles_from_data=True, from_rows=True)
        ref = f"'{ws.title}'!$B${cab}:$M${cab}"
        for s in ch.series:
            s.cat = AxDataSource(strRef=StrRef(f=ref))
            s.smooth = False
            s.marker.symbol, s.marker.size = "circle", 6
        ch.x_axis.delete = ch.y_axis.delete = False
        ch.x_axis.axPos, ch.y_axis.axPos = "b", "l"
        ch.y_axis.numFmt = "0.00%"
        ch.legend = None
        ws.add_chart(ch, f"O{cab - 1}")
    return ws


def aba_conferencia(wb, por_mes, d, pq_classe, por_mes_classe):
    ws = wb.create_sheet("3 · Conferência")
    ws.column_dimensions["A"].width = 34
    for c in "BCDEF":
        ws.column_dimensions[c].width = 15
    ws.append(["CONFERÊNCIA — revisada duas vezes"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.append([])
    ws.append(["1) O mensal do gestor bate com o rol lido nas SS, mês a mês:"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Fatia"] + CURTO + ["Soma"])
    for cel in ws[ws.max_row]:
        if cel.value:
            cel.font, cel.fill = TITULO, FUNDO
    prim = ws.max_row + 1
    for tipo, ano in (("RL", 2025), ("RT", 2025), ("RL", 2026), ("RT", 2026)):
        linha = [por_mes[(tipo, ano, m)] for m in range(1, 13)]
        ws.append([f"{tipo} {ano}"] + linha + [sum(linha)])
    bordar(ws, prim, ws.max_row)

    ws.append([])
    ws.append(["2) Ocorrência não é equipamento, e o anual leva o complemento por obra:"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Fatia", "Ocorrências (mensal)", "Equipamentos com SS",
               "Complemento por obra", "Equipamentos no ano", "Parque"])
    for cel in ws[ws.max_row]:
        if cel.value:
            cel.font, cel.fill = TITULO, FUNDO
    prim = ws.max_row + 1
    eq = d["equipamentos"]
    tot = d["total_equipamentos_que_falharam"]
    comp = d["complemento_obra_direta"]
    for tipo, fam, ano, parque in (("RL", "religador", 2025, 1281), ("RT", "regulador", 2025, 180),
                                   ("RL", "religador", 2026, 1294), ("RT", "regulador", 2026, 190)):
        k = f"{fam}|{ano}"
        ws.append([f"{tipo} {ano}", sum(por_mes[(tipo, ano, m)] for m in range(1, 13)),
                   len(eq.get(k, [])), comp.get(k, 0), tot.get(k, 0), parque])
    bordar(ws, prim, ws.max_row)

    ws.append([])
    ws.append(["3) Classe de tensão — do cadastro de ajustes da GESTÃO (27/08):"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["", "Parque 34.500", "Parque 13.800", "Total"])
    for cel in ws[ws.max_row]:
        if cel.value:
            cel.font, cel.fill = TITULO, FUNDO
    prim = ws.max_row + 1
    for tipo in ("RL", "RT"):
        a, b = pq_classe[(tipo, "34.500")], pq_classe[(tipo, "13.800")]
        ws.append([tipo, a, b, a + b])
    bordar(ws, prim, ws.max_row)

    ws.append([])
    ws.append(["Falhas por classe (ocorrências do rol):"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Fatia", "34.500", "13.800", "Sem cadastro", "Total"])
    for cel in ws[ws.max_row]:
        if cel.value:
            cel.font, cel.fill = TITULO, FUNDO
    prim = ws.max_row + 1
    for tipo, ano in (("RL", 2025), ("RT", 2025), ("RL", 2026), ("RT", 2026)):
        a = sum(por_mes_classe[("34.500", tipo, ano, m)] for m in range(1, 13))
        b = sum(por_mes_classe[("13.800", tipo, ano, m)] for m in range(1, 13))
        s = sum(por_mes_classe[("SEM CADASTRO", tipo, ano, m)] for m in range(1, 13))
        ws.append([f"{tipo} {ano}", a, b, s, a + b + s])
    bordar(ws, prim, ws.max_row)
    return ws


COMO = [
    "COMO ESTA TABELA FOI PREENCHIDA — e o que conferi duas vezes.",
    "",
    "O PEDIDO: preencher a tabela achatada (chave Concat) com o que a aba de taxa de "
    "falha mostra. Feito na aba 1, com as mesmas colunas do seu print, mais duas que "
    "faltavam para a tabela se fechar sozinha: «Acumulado no ano» e «Taxa do mês».",
    "",
    "PRIMEIRA REVISÃO — o mensal do seu quadro bate com o nosso rol de ocorrências "
    "lidas nas SS, mês a mês, sem uma unidade de diferença: RL 2025 "
    "1·1·2·2·7·4·5·2·2·4·1·4 = 35; RT 2025 0·1·2·0·0·1·3·5·3·0·1·2 = 18; RL 2026 "
    "6·7·8·3·0·1·3·0 = 28; RT 2026 1·1·1·2·1·1·1·1 = 9. As porcentagens do seu quadro "
    "também conferem na divisão direta.",
    "",
    "SEGUNDA REVISÃO — achei uma diferença de RÉGUA, não de conta. O quadro mensal "
    "conta OCORRÊNCIA (a SS lida); a taxa anual oficial conta EQUIPAMENTO que falhou e "
    "soma o complemento por obra direta do AIC — falha provada pela obra encerrada, sem "
    "narrativa de SS, que não tem mês para cair. Por isso o mensal soma 35 e o anual "
    "diz 62 em RL 2025. Nenhum dos dois está errado; são perguntas diferentes. A aba 3 "
    "mostra as duas colunas lado a lado.",
    "",
    "CUIDADO AO SOMAR: no ano, o equipamento que falha duas vezes conta uma vez só. "
    "Somar os doze meses passa do total anual — em RT 2025 são 18 ocorrências em 16 "
    "equipamentos, e em RL 2026 são 28 ocorrências em 27 equipamentos.",
    "",
    "PARQUE: 2025 fica na base de janeiro do gestor (1.281 RL e 180 RT, sem expansão "
    "lançada); 2026 leva a expansão realizada somada no próprio mês — RL 2·0·2·2·3·1·3 "
    "e RT 0·3·3·1·1·1·1 de janeiro a julho, fechando agosto em 1.294 e 190. Setembro a "
    "dezembro repetem agosto, porque a expansão do mês ainda não fechou.",
    "",
    "CLASSE DE TENSÃO: saiu do cadastro de ajustes da própria GESTÃO DE EQUIPAMENTOS — "
    "coluna TENSÃO na aba «Ajustes RL Poste» e TENSÃO PRIMÁRIA em «Ajustes Reguladores "
    "de Tensão». Dá 832 religadores em 34.500 e 460 em 13.800 (1.292 cadastrados), e "
    "137 reguladores em 34.500 e 52 em 13.800 (189). As falhas foram divididas pelo "
    "mesmo cadastro, ativo por ativo.",
    "",
    "DIVERGÊNCIA A RESOLVER: o quadrinho lateral do seu print traz RL 34,5 = 1.081 e "
    "RL 13,8 = 200, com 83,73% e 16,27%. Três coisas não fecham: 1.081 + 200 = 1.281, "
    "mas 1.081 ÷ 1.281 dá 84,39%, não 83,73%; e o cadastro de hoje diz 832 e 460. Já o "
    "RT 34,5 = 137 do seu quadrinho bate exatamente com o cadastro — o que sugere que "
    "os números do RL vieram de fonte mais antiga. Deixei a aba 1 com o cadastro atual; "
    "se a fonte certa for outra, é trocar duas células e a tabela toda acompanha.",
    "",
    "UM ATIVO SEM CLASSE: o religador 7930359149 (falha em 2026) não está em nenhum dos "
    "dois cadastros de ajustes — aparece como «sem cadastro» na aba 3 e fica fora da "
    "divisão por classe, mas dentro do total.",
    "",
    "Fontes: GESTÃO DE EQUIPAMENTOS de 27/08 (cadastro de ajustes), leitura revisada da "
    "taxa de falha (rol de ocorrências lido nas SS e conferido por revisor) e a régua "
    "de parque do gestor de 24/08.",
]


def montar():
    tensao = classe_de_tensao()
    pq_classe = parque_por_classe(tensao)
    por_mes, por_mes_classe, d = falhas(tensao)
    parque, novos = parque_mensal()

    wb = openpyxl.Workbook()
    aba_achatada(wb, por_mes_classe, pq_classe, parque, novos)
    aba_quadros(wb, por_mes, parque, novos)
    aba_conferencia(wb, por_mes, d, pq_classe, por_mes_classe)
    ws = wb.create_sheet("Como foi feito")
    ws.column_dimensions["A"].width = 112
    for t in COMO:
        ws.append([t])
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    print(f"gravado: {SAIDA}")
    for tipo, ano in (("RL", 2025), ("RT", 2025), ("RL", 2026), ("RT", 2026)):
        linha = [por_mes[(tipo, ano, m)] for m in range(1, 13)]
        print(f"  {tipo} {ano}: {linha} = {sum(linha)}")
    print(f"  parque por classe: {dict(pq_classe)}")


if __name__ == "__main__":
    montar()
