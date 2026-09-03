"""
Quando o posto resolveu, mês a mês — dist/RESOLVIDOS_MENSAL_2026.xlsx.

Os 62 do trabalho concluído do COEP (48 encerradas + 14 despachadas), pelo mês em que
foram resolvidos. A encerrada conta no mês em que a cadeia fechou; a despachada, no mês
em que a SS chegou à mesa seguinte — que é quando a parte do COEP acabou.

Todos os 62 caem em 2026: conferido por `assert` na montagem.

Rodar: python3 scripts/planilha_mensal_62.py
"""

import json
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.colors import ColorChoice
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FONTE = os.path.join(RAIZ, "data", "missao", "mensal_62.json")
SAIDA = os.path.join(RAIZ, "dist", "RESOLVIDOS_MENSAL_2026.xlsx")

TINTA, PAPEL, SINAL = "FF211D15", "FFF2EFE6", "FFBC4B0E"
# as duas séries, conferidas com o validador do dataviz sobre o papel claro
VERDE, LARANJA = "1F7C50", "B8480C"
NOME = {"01": "jan", "02": "fev", "03": "mar", "04": "abr",
        "05": "mai", "06": "jun", "07": "jul", "08": "ago"}


def montar():
    with open(FONTE, encoding="utf-8") as fh:
        mn = json.load(fh)
    meses = mn["meses"]
    assert all(m.startswith("2026-") for m in meses), "mês fora de 2026"
    linhas, acum = [], 0
    for m in meses:
        e = mn["encerrados"].get(m, {})
        d = mn["despachados"].get(m, {})
        er, et = e.get("RL", 0), e.get("RT", 0)
        dr, dt = d.get("RL", 0), d.get("RT", 0)
        total = er + et + dr + dt
        acum += total
        linhas.append([f"{NOME[m[-2:]]}/26", er + et, dr + dt, total, acum,
                       er, et, dr, dt])
    assert sum(x[1] for x in linhas) == 48, "encerradas devem somar 48"
    assert sum(x[2] for x in linhas) == 14, "despachadas devem somar 14"
    assert linhas[-1][4] == 62, "o acumulado tem de fechar em 62"
    return linhas


def planilha(linhas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mensal 2026"
    cols = [("Mês", 10), ("Encerrada", 11), ("Despachada", 12), ("Total do mês", 13),
            ("Acumulado", 11), ("Enc. RL", 9), ("Enc. RT", 9),
            ("Desp. RL", 10), ("Desp. RT", 10)]
    ws.append([c[0] for c in cols])
    for r in linhas:
        ws.append(r)
    ws.append(["Total", sum(r[1] for r in linhas), sum(r[2] for r in linhas),
               sum(r[3] for r in linhas), "", sum(r[5] for r in linhas),
               sum(r[6] for r in linhas), sum(r[7] for r in linhas),
               sum(r[8] for r in linhas)])
    for c in range(1, len(cols) + 1):
        cel = ws.cell(row=1, column=c)
        cel.font = Font(bold=True, color=PAPEL, size=10)
        cel.fill = PatternFill("solid", fgColor=TINTA)
        cel.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
    for i, (_, larg) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    ult = len(linhas) + 1
    g = BarChart()
    g.type, g.grouping, g.overlap = "col", "stacked", 100
    g.title = "Resolvidos pelo COEP por mês — 2026"
    g.y_axis.title = "Equipamentos"
    g.height, g.width = 10, 24
    g.gapWidth = 70
    for col, cor in ((2, VERDE), (3, LARANJA)):
        s = Series(Reference(ws, min_col=col, min_row=1, max_row=ult),
                   title_from_data=True)
        s.graphicalProperties.solidFill = ColorChoice(srgbClr=cor)
        s.graphicalProperties.line.noFill = True
        g.series.append(s)
    # os meses são texto: sem StrRef o Excel numera as categorias 1..8
    cats = AxDataSource(strRef=StrRef(f=f"'{ws.title}'!$A$2:$A${ult}"))
    for s in g.series:
        s.cat = cats
    g.dataLabels = DataLabelList()
    g.dataLabels.showVal = True
    ws.add_chart(g, "K2")

    ws2 = wb.create_sheet("Como foi feito")
    for t in [
        ["Quando o posto resolveu — os 62 de 2026"], [""],
        ["O que está contado"], [""],
        ["Os 62 são o trabalho do COEP concluído: 48 demandas encerradas de ponta a"],
        ["ponta mais 14 despachadas para outra mesa, onde a peça foi trocada e o campo"],
        ["devolveu, mas a SS segue aberta na PROT, TELE ou SE."],
        [""],
        ["Em que mês cada uma entra"], [""],
        ["ENCERRADA: no mês em que a cadeia fechou — a data de conclusão da SS que"],
        ["encerrou, seja ela atendida ou cancelada, em qualquer posto."],
        [""],
        ["DESPACHADA: no mês em que a SS chegou à mesa seguinte, que é quando a parte"],
        ["do COEP acabou. Não há data de fechamento para elas: a SS ainda está aberta."],
        [""],
        ["O ano — conferido"], [""],
        ["Todos os 62 caem em 2026. As 48 encerradas têm data de fechamento e as 48"],
        ["estão dentro do ano, nenhuma sem data. As 14 despachadas chegaram à mesa"],
        ["seguinte em julho (11) e agosto (3), todas com SS de 2026."],
        [""],
        ["Não confundir com o ano da DEMANDA. Das 48 encerradas, 33 são de 2025 e 3 de"],
        ["2024; só 12 nasceram em 2026. Resolver é deste ano; o problema, não."],
        [""],
        ["Agosto vai até o dia 18, a posição do relatório."],
        [""],
        ["Ressalva"], [""],
        ["Estes 62 saem da régua do TIPOSS. A auditoria de 29/08 mostrou que, pela peça,"],
        ["só 16 encerradas e 9 despachadas têm troca comprovada por obra do AIC ou OS."],
    ]:
        ws2.append(t)
    ws2.column_dimensions["A"].width = 96
    ws2["A1"].font = Font(bold=True, size=12, color=SINAL)
    for r in (3, 9, 17, 28):
        ws2.cell(row=r, column=1).font = Font(bold=True, size=11)

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    return SAIDA


if __name__ == "__main__":
    linhas = montar()
    print("gravado:", planilha(linhas))
    for r in linhas:
        print(f"  {r[0]:8s} enc {r[1]:3d}  desp {r[2]:3d}  total {r[3]:3d}  acum {r[4]:3d}")
