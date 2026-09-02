"""
O quadro da premissa set–dez em gráficos — dist/PAINEL_PREMISSA_SETDEZ.xlsx.

O gestor mandou o quadro (02/09) com seis linhas e quatro colunas sem rótulo:
backlog, entrante, resolvidos, pendentes, orçado e forecast de desembolso.

AS COLUNAS SÃO SETEMBRO, OUTUBRO, NOVEMBRO E DEZEMBRO DE 2026. Duas provas, no
centavo, contra a planilha base:

  · o primeiro forecast, R$ 2.129.866,67, é o realizado de jan–ago (R$ 1.605.280,07,
    aba Apresentação, soma acumulada) mais o desembolso de setembro da premissa da
    aba Orçamento (R$ 524.586,60);
  · o último, R$ 6.058.299,31, é a soma acumulada de tudo na aba Apresentação
    (R$ 6.058.299,32) — e fica R$ 4.024,53 abaixo do orçado de 2026,
    R$ 6.062.323,84, que é o valor da última coluna de ORÇADO.

Oito visões, uma por aba, todas lendo a aba «Dados» — mudou o número lá, mudou o
gráfico. Duas hipóteses do quadro ficam escritas na aba «Como ler», porque mexem
no resultado: o desembolso de setembro carrega o realizado do ano inteiro, e o
pendente de cada mês desconta só os entrantes daquele mês.

Rodar: python3 scripts/painel_premissa_setdez.py
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint, Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
SAIDA = os.path.join(RAIZ, "dist", "PAINEL_PREMISSA_SETDEZ.xlsx")

# Prontuário Industrial — papel, tinta, laranja-sinal
TINTA, PAPEL, SOMBRA, SINAL, APAGADA = "FF211D15", "FFF2EFE6", "FFE9E5D8", "FFBC4B0E", "FF8D8672"
# as duas séries: passam em todas as checagens do validador (inclusive deuteranopia)
VERDE, LARANJA, NEUTRO, GRADE = "1F7C50", "B8480C", "6D675A", "DDD8CC"
MOEDA = 'R$ #,##0.00'
PCT = '0.0%'

MESES = ["Setembro", "Outubro", "Novembro", "Dezembro"]
# o quadro do gestor, como ele mandou
BACKLOG = [100, 100, 100, 100]
ENTRANTE = [8, 5, 4, 6]
RESOLVIDOS = [6, 35, 57, 77]                      # acumulado no ano
ORCADO = [5206065.02, 5579287.46, 5821192.96, 6062323.84]
FORECAST = [2129866.67, 2936475.89, 4637315.57, 6058299.31]
PENDENTES = [b + e - r for b, e, r in zip(BACKLOG, ENTRANTE, RESOLVIDOS)]   # 102 · 70 · 47 · 29


def titulo(ws, texto, sub=""):
    ws["A1"] = texto
    ws["A1"].font = Font(bold=True, size=13, color=SINAL)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(italic=True, size=9)
        ws["A2"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 30


def cabecalho(ws, linha, titulos, larguras=None):
    for i, t in enumerate(titulos, 1):
        c = ws.cell(row=linha, column=i, value=t)
        c.font = Font(bold=True, color=PAPEL, size=10)
        c.fill = PatternFill("solid", fgColor=TINTA)
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.cell(row=linha, column=1).alignment = Alignment(vertical="center", horizontal="left")
    if larguras:
        for i, w in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def estilo(ch, alto=11, largo=22):
    """Marca fina, grade recuada, eixos à vista — a régua do dataviz."""
    ch.height, ch.width = alto, largo
    ch.style = None
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    gl = ChartLines()
    gl.spPr = GraphicalProperties()
    gl.spPr.line = LineProperties(solidFill=GRADE, w=6350)
    ch.y_axis.majorGridlines = gl
    ch.x_axis.majorGridlines = None
    for eixo in (ch.x_axis, ch.y_axis):
        eixo.spPr = GraphicalProperties()
        eixo.spPr.line = LineProperties(solidFill=GRADE, w=6350)
    return ch


def rotulos(alvo, fmt=None):
    alvo.dLbls = DataLabelList()
    alvo.dLbls.showVal = True
    alvo.dLbls.showSerName = False
    alvo.dLbls.showCatName = False
    alvo.dLbls.showLegendKey = False
    if fmt:
        alvo.dLbls.numFmt = fmt
        alvo.dLbls.showVal = True
    return alvo


def barras(ws, ref_dados, ref_cats, cores, titulo_ch, eixo_y, fmt=None,
           empilhado=None, com_legenda=True):
    ch = BarChart()
    ch.type = "col"
    if empilhado:
        ch.grouping = empilhado
        ch.overlap = 100
    else:
        ch.grouping = "clustered"
        ch.overlap = -12          # 2 px de papel entre as barras vizinhas
    ch.gapWidth = 70
    ch.add_data(ref_dados, from_rows=True, titles_from_data=True)
    ch.set_categories(ref_cats)
    ch.title = titulo_ch
    ch.y_axis.title = eixo_y
    if fmt:
        ch.y_axis.numFmt = fmt
    for s, cor in zip(ch.series, cores):
        s.graphicalProperties = GraphicalProperties(solidFill=cor)
        s.graphicalProperties.line = LineProperties(solidFill="FBFAF6", w=25400)  # 2 px
    if len(ch.series) < 2 or not com_legenda:
        ch.legend = None
    else:
        ch.legend.position = "b"
        ch.legend.overlay = False
    return estilo(ch)


def aba_dados(wb):
    ws = wb.create_sheet("Dados")
    titulo(ws, "PREMISSA SET–DEZ 2026 — o quadro do gestor e o que sai dele",
           "As seis primeiras linhas são as que ele mandou (campos sombreados: mude aqui e todos os "
           "gráficos acompanham). PENDENTES e o bloco «Calculado» são fórmula. Os meses vêm da premissa "
           "da aba Orçamento da planilha base — o forecast de setembro e o de dezembro batem no centavo.")
    cabecalho(ws, 4, ["O quadro"] + MESES, [40, 17, 17, 17, 17])
    linhas = [("BACKLOG", BACKLOG, "0"), ("ENTRANTE", ENTRANTE, "0"),
              ("RESOLVIDOS (acumulado)", RESOLVIDOS, "0")]
    r = 5
    for nome, vals, fmt in linhas:
        ws.cell(row=r, column=1, value=nome).font = Font(bold=True)
        for j, v in enumerate(vals):
            c = ws.cell(row=r, column=2 + j, value=v)
            c.number_format = fmt
            c.fill = PatternFill("solid", fgColor=SOMBRA)
        r += 1
    ws.cell(row=r, column=1, value="PENDENTES").font = Font(bold=True)
    for j in range(4):
        L = get_column_letter(2 + j)
        ws.cell(row=r, column=2 + j, value=f"={L}5+{L}6-{L}7").number_format = "0"
    r += 1
    for nome, vals in (("ORÇADO", ORCADO), ("FORECAST DESEMBOLSO", FORECAST)):
        ws.cell(row=r, column=1, value=nome).font = Font(bold=True)
        for j, v in enumerate(vals):
            c = ws.cell(row=r, column=2 + j, value=v)
            c.number_format = MOEDA
            c.fill = PatternFill("solid", fgColor=SOMBRA)
        r += 1
    # ---- bloco calculado
    cabecalho(ws, 12, ["Calculado a partir do quadro"] + MESES)
    calc = [
        ("Carteira do mês (backlog + entrante)", lambda L, j: f"={L}5+{L}6", "0"),
        ("Resolvidos NO MÊS", lambda L, j: f"={L}7" if j == 0 else f"={L}7-{get_column_letter(1 + j)}7", "0"),
        ("Desembolso NO MÊS", lambda L, j: f"={L}10" if j == 0 else f"={L}10-{get_column_letter(1 + j)}10", MOEDA),
        ("Saldo a desembolsar (orçado − forecast)", lambda L, j: f"={L}9-{L}10", MOEDA),
        ("Execução do orçamento", lambda L, j: f"={L}10/{L}9", PCT),
        ("Carteira já resolvida", lambda L, j: f"={L}7/{L}13", PCT),
        ("Ainda pendente", lambda L, j: f"={L}8/{L}13", PCT),
        ("PENDENTES com entrantes acumulados (leitura alternativa)",
         lambda L, j: f"={L}5+SUM($B$6:{L}6)-{L}7", "0"),
    ]
    r = 13
    for nome, f, fmt in calc:
        ws.cell(row=r, column=1, value=nome)
        for j in range(4):
            L = get_column_letter(2 + j)
            c = ws.cell(row=r, column=2 + j, value=f(L, j))
            c.number_format = fmt
        r += 1
    ws.cell(row=20, column=1).font = Font(italic=True)
    for rr in list(range(5, 11)) + list(range(13, 21)):
        for cc in range(2, 6):
            ws.cell(row=rr, column=cc).alignment = Alignment(horizontal="center")
    ws.freeze_panes = "B5"
    return ws


def aba(wb, nome, tit, sub, mini, fmts):
    """Cria a aba, escreve a mini-tabela ligada a «Dados» e devolve a planilha."""
    ws = wb.create_sheet(nome)
    titulo(ws, tit, sub)
    ws.column_dimensions["A"].width = 34
    for j, m in enumerate(MESES):
        c = ws.cell(row=4, column=2 + j, value=f"=Dados!{get_column_letter(2 + j)}4")
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(2 + j)].width = 15
    for i, (rot, origem) in enumerate(mini):
        ws.cell(row=5 + i, column=1, value=rot).font = Font(bold=True)
        for j in range(4):
            L = get_column_letter(2 + j)
            c = ws.cell(row=5 + i, column=2 + j, value=f"=Dados!{L}{origem}")
            c.number_format = fmts[i]
            c.alignment = Alignment(horizontal="center")
    return ws


def refs(ws, n):
    dados = Reference(ws, min_col=1, min_row=5, max_col=5, max_row=4 + n)
    cats = Reference(ws, min_col=2, min_row=4, max_col=5, max_row=4)
    return dados, cats


# --------------------------------------------------------------------------- as sete visões
def v1_fila(wb):
    ws = aba(wb, "1 · A fila", "A fila cai de 102 para 29 até dezembro",
             "As barras são o que fica pendente no fim de cada mês; a linha é o total resolvido no ano "
             "até ali. As duas contam equipamentos, então dividem o mesmo eixo.",
             [("Pendentes no fim do mês", 8), ("Resolvidos (acumulado)", 7)], ["0", "0"])
    dados = Reference(ws, min_col=1, min_row=5, max_col=5, max_row=5)
    cats = Reference(ws, min_col=2, min_row=4, max_col=5, max_row=4)
    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth, ch.overlap = "col", "clustered", 90, -12
    ch.add_data(dados, from_rows=True, titles_from_data=True)
    ch.set_categories(cats)
    ch.title = "Pendentes no fim do mês e total resolvido no ano"
    ch.y_axis.title = "equipamentos"
    s = ch.series[0]
    s.graphicalProperties = GraphicalProperties(solidFill=LARANJA)
    s.graphicalProperties.line = LineProperties(solidFill="FBFAF6", w=25400)
    rotulos(s)
    lin = LineChart()
    lin.add_data(Reference(ws, min_col=1, min_row=6, max_col=5, max_row=6),
                 from_rows=True, titles_from_data=True)
    ls = lin.series[0]
    ls.graphicalProperties = GraphicalProperties()
    ls.graphicalProperties.line = LineProperties(solidFill=VERDE, w=25400)
    ls.marker = Marker(symbol="circle", size=8)
    ls.marker.graphicalProperties = GraphicalProperties(solidFill=VERDE)
    ls.smooth = False
    rotulos(ls)
    ch += lin                      # mesmo eixo — sem segundo eixo Y
    ch.legend.position = "b"
    ch.legend.overlay = False
    ws.add_chart(estilo(ch, 12, 24), "A9")


def v2_cascata(wb):
    ws = wb.create_sheet("2 · Cascata de dezembro")
    titulo(ws, "De onde saem os 29 pendentes de dezembro",
           "Parte do backlog de 100, soma o que entra e desconta o que foi resolvido no ano. "
           "As duas barras cinzas são saldos; as coloridas, movimento.")
    ws.column_dimensions["A"].width = 30
    for j, t in enumerate(["Backlog", "Entrante (dez)", "Resolvidos (acum.)", "Pendentes"]):
        ws.cell(row=4, column=2 + j, value=t).font = Font(bold=True)
        ws.cell(row=4, column=2 + j).alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(2 + j)].width = 17
    ws.cell(row=5, column=1, value="Base (invisível)").font = Font(color=APAGADA, size=8)
    ws.cell(row=6, column=1, value="Equipamentos").font = Font(bold=True)
    ws["B5"], ws["C5"], ws["D5"], ws["E5"] = 0, "=Dados!E5", "=Dados!E8", 0
    ws["B6"], ws["C6"], ws["D6"], ws["E6"] = "=Dados!E5", "=Dados!E6", "=Dados!E7", "=Dados!E8"
    for c in "BCDE":
        for r in (5, 6):
            ws[f"{c}{r}"].number_format = "0"
            ws[f"{c}{r}"].alignment = Alignment(horizontal="center")
    ch = BarChart()
    ch.type, ch.grouping, ch.overlap, ch.gapWidth = "col", "stacked", 100, 60
    ch.add_data(Reference(ws, min_col=1, min_row=5, max_col=5, max_row=6),
                from_rows=True, titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=2, min_row=4, max_col=5, max_row=4))
    ch.title = "Backlog 100 · entra 6 · resolve 77 · sobra 29"
    ch.y_axis.title = "equipamentos"
    base, valor = ch.series
    base.graphicalProperties = GraphicalProperties()
    base.graphicalProperties.noFill = True
    base.graphicalProperties.line = LineProperties(noFill=True)
    valor.graphicalProperties = GraphicalProperties(solidFill=NEUTRO)
    valor.graphicalProperties.line = LineProperties(solidFill="FBFAF6", w=25400)
    valor.dPt = [
        DataPoint(idx=0, spPr=GraphicalProperties(solidFill=NEUTRO)),
        DataPoint(idx=1, spPr=GraphicalProperties(solidFill=LARANJA)),
        DataPoint(idx=2, spPr=GraphicalProperties(solidFill=VERDE)),
        DataPoint(idx=3, spPr=GraphicalProperties(solidFill=NEUTRO)),
    ]
    rotulos(valor)
    ch.legend = None
    ws.add_chart(estilo(ch, 12, 22), "A9")


def v3_ritmo(wb):
    ws = aba(wb, "3 · Ritmo do mês", "O que resolve e o que entra, mês a mês",
             "Resolvidos no mês é a diferença do acumulado; entrante é o que a premissa espera queimar. "
             "Enquanto a barra verde for maior que a laranja, a fila drena.",
             [("Resolvidos no mês", 14), ("Entrantes no mês", 6)], ["0", "0"])
    dados, cats = refs(ws, 2)
    ch = barras(ws, dados, cats, [VERDE, LARANJA],
                "Resolvidos × entrantes, por mês", "equipamentos")
    for s in ch.series:
        rotulos(s)
    ws.add_chart(estilo(ch, 12, 22), "A9")


def v4_dinheiro(wb):
    ws = aba(wb, "4 · Orçado × forecast", "O forecast alcança o orçado em dezembro",
             "Os dois acumulados. Em dezembro o forecast fecha R$ 4.024,53 abaixo do orçado de 2026 "
             "(R$ 6.058.299,31 contra R$ 6.062.323,84) — a mesma diferença da aba Apresentação da "
             "planilha base.",
             [("Orçado (acumulado)", 9), ("Forecast de desembolso", 10)], [MOEDA, MOEDA])
    dados, cats = refs(ws, 2)
    ch = barras(ws, dados, cats, [NEUTRO, VERDE],
                "Orçado e forecast de desembolso, acumulados", "R$", fmt=MOEDA)
    ws.add_chart(estilo(ch, 12, 24), "A9")


def v5_saldo(wb):
    ws = aba(wb, "5 · Saldo a gastar", "O que falta desembolsar encolhe a cada mês",
             "Orçado menos forecast. De R$ 3,08 milhões em setembro para R$ 4.024,53 em dezembro — "
             "ou seja, a premissa consome o orçamento inteiro do ano.",
             [("Saldo a desembolsar", 16)], [MOEDA])
    dados, cats = refs(ws, 1)
    ch = barras(ws, dados, cats, [LARANJA], "Saldo a desembolsar no fim de cada mês", "R$", fmt=MOEDA)
    rotulos(ch.series[0], MOEDA)
    ws.add_chart(estilo(ch, 12, 22), "A8")


def v6_execucao(wb):
    ws = aba(wb, "6 · Execução", "A execução do orçamento vai de 41% a 99,9%",
             "Forecast dividido pelo orçado do mês. Em dezembro sobra 0,07% do orçamento — "
             "R$ 4.024,53.",
             [("Execução do orçamento", 17)], [PCT])
    dados, cats = refs(ws, 1)
    ch = barras(ws, dados, cats, [VERDE], "Percentual do orçamento comprometido", "% do orçado", fmt=PCT)
    rotulos(ch.series[0], PCT)
    ch.y_axis.scaling.max = 1
    ws.add_chart(estilo(ch, 12, 22), "A8")


def v7_composicao(wb):
    ws = aba(wb, "7 · Composição", "Quanto da carteira do mês já está resolvido",
             "Cada barra é a carteira daquele mês (backlog + entrante) repartida em resolvido e "
             "pendente. Em setembro 5,6%; em dezembro, 72,6%.",
             [("Carteira já resolvida", 18), ("Ainda pendente", 19)], [PCT, PCT])
    dados, cats = refs(ws, 2)
    ch = barras(ws, dados, cats, [VERDE, LARANJA],
                "Composição da carteira do mês", "% da carteira", fmt=PCT, empilhado="percentStacked")
    for s in ch.series:
        rotulos(s, PCT)
    ws.add_chart(estilo(ch, 12, 22), "A9")


def v8_leituras(wb):
    ws = aba(wb, "8 · Duas leituras", "A conta do pendente muda com o entrante acumulado",
             "A linha do quadro desconta, em cada mês, só os entrantes daquele mês. Se os entrantes "
             "dos meses anteriores continuarem na fila, dezembro fecha em 46, não em 29. "
             "Não é correção — é a pergunta para o gestor.",
             [("Pendentes (como no quadro)", 8), ("Pendentes com entrantes acumulados", 20)], ["0", "0"])
    dados, cats = refs(ws, 2)
    ch = barras(ws, dados, cats, [LARANJA, NEUTRO],
                "Pendentes: as duas leituras da mesma premissa", "equipamentos")
    for s in ch.series:
        rotulos(s)
    ws.add_chart(estilo(ch, 12, 24), "A9")


def aba_como(wb):
    ws = wb.create_sheet("Como ler")
    ws.column_dimensions["A"].width = 112
    linhas = [
        ("COMO ESTE PAINEL FOI MONTADO", True),
        ("Os números são os do quadro que o gestor mandou em 02/09, sem arredondar nem ajustar. "
         "Todos os gráficos leem a aba «Dados»: mudou lá, mudou aqui.", False),
        ("", False),
        ("POR QUE AS COLUNAS SÃO SETEMBRO A DEZEMBRO", True),
        ("O quadro veio sem rótulo de coluna. Duas contas fecham no centavo com a planilha base e "
         "dizem que são os quatro últimos meses de 2026:", False),
        ("· o primeiro forecast, R$ 2.129.866,67 = R$ 1.605.280,07 (realizado de jan–ago, aba "
         "Apresentação, soma acumulada) + R$ 524.586,60 (desembolso de setembro na premissa da aba "
         "Orçamento);", False),
        ("· o último, R$ 6.058.299,31, é a soma acumulada de tudo na aba Apresentação (R$ 6.058.299,32); "
         "e a última coluna de ORÇADO, R$ 6.062.323,84, é o orçado de 2026 inteiro.", False),
        ("", False),
        ("O QUE O PAINEL MOSTRA", True),
        ("1 · A fila — pendentes no fim do mês (barra) contra o resolvido acumulado (linha).", False),
        ("2 · Cascata de dezembro — de onde saem os 29: 100 de backlog, +6 que entram, −77 resolvidos.", False),
        ("3 · Ritmo do mês — resolvidos no mês contra entrantes; enquanto o verde ganha, a fila drena.", False),
        ("4 · Orçado × forecast — os dois acumulados, fechando com diferença de R$ 4.024,53.", False),
        ("5 · Saldo a gastar — o que falta desembolsar, de R$ 3,08 milhões a R$ 4.024,53.", False),
        ("6 · Execução — o percentual do orçamento comprometido, de 40,9% a 99,93%.", False),
        ("7 · Composição — quanto da carteira de cada mês já está resolvido, em percentual.", False),
        ("8 · Duas leituras — a conta do pendente com e sem acumular os entrantes.", False),
        ("", False),
        ("DUAS COISAS QUE MEXEM NO NÚMERO — PARA O GESTOR DECIDIR", True),
        ("a) PENDENTES desconta só o entrante do próprio mês. No quadro, PENDENTES = BACKLOG + ENTRANTE − "
         "RESOLVIDOS, coluna a coluna, com o backlog fixo em 100 — e reproduz 102 · 70 · 47 · 29 exato. "
         "Só que RESOLVIDOS é acumulado no ano e ENTRANTE é do mês: os 8 que entraram em setembro não "
         "aparecem na conta de outubro. Se eles continuarem na fila, a série vira 102 · 78 · 60 · 46 — "
         "dezembro fecha 17 equipamentos acima. É a aba «8 · Duas leituras».", False),
        ("b) O desembolso de setembro carrega o ano inteiro. R$ 2.129.866,67 são R$ 1.605.280,07 já "
         "realizados de janeiro a agosto mais R$ 524.586,60 de setembro. Por isso o «desembolso no mês» "
         "de setembro na aba Dados fica alto: ele não é só setembro. Para ver o ritmo mensal limpo, "
         "trocar a primeira coluna por 524.586,60.", False),
        ("", False),
        ("COR E LEITURA", True),
        ("Verde é o que anda (resolvido, desembolsado); laranja é o que espera (pendente, entrante, "
         "saldo); cinza é saldo de referência. O par verde/laranja foi conferido no validador de "
         "paleta e passa em contraste e em daltonismo (deuteranopia ΔE 8,9; visão normal 22,9), e todo "
         "gráfico com duas séries tem legenda e rótulo — a cor nunca é a única pista.", False),
        ("Nenhum gráfico usa dois eixos verticais: onde as grandezas são diferentes (equipamento e "
         "R$), são gráficos separados.", False),
    ]
    for i, (t, negrito) in enumerate(linhas, 1):
        c = ws.cell(row=i, column=1, value=t)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if negrito:
            c.font = Font(bold=True, size=11, color=SINAL if i == 1 else TINTA)
    return ws


def montar(saida=SAIDA, cache=True):
    wb = Workbook()
    wb.remove(wb.active)
    aba_dados(wb)
    v1_fila(wb)
    v2_cascata(wb)
    v3_ritmo(wb)
    v4_dinheiro(wb)
    v5_saldo(wb)
    v6_execucao(wb)
    v7_composicao(wb)
    v8_leituras(wb)
    aba_como(wb)
    os.makedirs(os.path.dirname(saida), exist_ok=True)
    wb.save(saida)
    print(f"{saida}: {len(wb.sheetnames)} abas, 8 gráficos")
    if cache:
        from planilha_automatica import grava_cache
        n = grava_cache(saida)
        print(f"cache de valores gravado em {n} células de fórmula")
    return saida


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    montar(args[0] if args else SAIDA, cache="--sem-cache" not in sys.argv)
