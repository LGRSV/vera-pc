"""
Planilha da taxa de falha de religadores e reguladores — ETO-COEP.

Consome data/missao/taxa_falha.json (a lógica em camadas) e
data/missao/verificacao_tiposs.json (a leitura das SS pelos agentes, com revisão
adversarial) e grava dist/TAXA_DE_FALHA_RL_RT.xlsx.

Abas:
  Síntese              a resposta da verificação, em texto, com a matriz
  Taxa por ano         2024 / 2025 / 2026 nas duas réguas, lado a lado
  Matriz da hipótese   INDISPONIBILIDADE × peça grande, precisão e cobertura
  Verificação SS a SS  as 183 SS lidas, com veredito, evidência e revisão
  Peça grande          a fila material por classe (convenção do Allan)
  Premissas            tudo que sustenta os números

LibreOffice não sobe neste ambiente, então a planilha grava VALORES, não fórmulas.

Rodar: python3 scripts/taxa_falha_excel.py
"""

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ARQ_TAXA = os.path.join(RAIZ, "data", "missao", "taxa_falha.json")
ARQ_VERIF = os.path.join(RAIZ, "data", "missao", "verificacao_tiposs.json")
SAIDA = os.path.join(RAIZ, "dist", "TAXA_DE_FALHA_RL_RT.xlsx")

TINTA = "1A1A1A"
PAPEL = "FFFFFF"
FAIXA = "E8E4DC"
DESTAQUE = "C8442A"
CALMO = "F4F2ED"

fina = Side(style="thin", color="B8B2A6")
grossa = Side(style="medium", color=TINTA)
GRADE = Border(left=fina, right=fina, top=fina, bottom=fina)


def _titulo(ws, texto, subtitulo=None, largura=8):
    ws["A1"] = texto
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=TINTA)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=largura)
    linha = 2
    if subtitulo:
        ws["A2"] = subtitulo
        ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="5A5347")
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=largura)
        ws.row_dimensions[2].height = 30
        linha = 3
    return linha + 1


def _cabecalho(ws, linha, colunas):
    for i, nome in enumerate(colunas, start=1):
        c = ws.cell(row=linha, column=i, value=nome)
        c.font = Font(name="Calibri", size=11, bold=True, color=TINTA)
        c.fill = PatternFill("solid", fgColor=FAIXA)
        c.border = Border(left=fina, right=fina, top=grossa, bottom=grossa)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[linha].height = 30
    return linha + 1


def _linha(ws, linha, valores, negrito=False, destaque=False, quebra=False):
    for i, v in enumerate(valores, start=1):
        c = ws.cell(row=linha, column=i, value=v)
        c.font = Font(name="Calibri", size=11, bold=negrito,
                      color=DESTAQUE if destaque else TINTA)
        c.border = GRADE
        c.alignment = Alignment(
            horizontal="left" if isinstance(v, str) else "center",
            vertical="top" if quebra else "center",
            wrap_text=quebra,
        )
    return linha + 1


def _larguras(ws, larguras):
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def aba_sintese(wb, taxa, verif):
    ws = wb.create_sheet("Síntese")
    _larguras(ws, [26, 16, 16, 16, 16, 16, 16, 16])
    linha = _titulo(
        ws,
        "Taxa de falha — religadores e reguladores de tensão · ETO",
        "Régua do gestor (21/08/2026): conta na taxa a falha que exige peça grande — tanque, "
        "controle ou equipamento completo no religador; controle, células ou banco completo no "
        "regulador. Furto conta quando leva uma dessas peças. Posição de 12/08/2026.",
    )

    m = (verif or {}).get("matriz") or {}
    if m:
        linha = _linha(ws, linha, ["A HIPÓTESE VERIFICADA"], negrito=True)
        linha = _linha(ws, linha, [
            'O gestor levantou: "quando tem que trocar algo de verdade, a SS fica com o tipo '
            'INDISPONIBILIDADE PARA OPERAÇÃO". 183 SS foram lidas na íntegra por oito agentes '
            'e os vereditos discordantes passaram por revisores adversariais.'
        ], quebra=True)
        ws.merge_cells(start_row=linha - 1, start_column=1, end_row=linha - 1, end_column=8)
        ws.row_dimensions[linha - 1].height = 46
        linha += 1
        linha = _cabecalho(ws, linha, ["", "Exigiu peça grande", "Não exigiu", "Indefinido", "Total"])
        ind_tot = m.get("ind_grande", 0) + m.get("ind_nao", 0) + m.get("ind_null", 0)
        out_tot = m.get("out_grande", 0) + m.get("out_nao", 0) + m.get("out_null", 0)
        linha = _linha(ws, linha, ["SS de INDISPONIBILIDADE", m.get("ind_grande"), m.get("ind_nao"),
                                   m.get("ind_null"), ind_tot], negrito=True)
        linha = _linha(ws, linha, ["SS de outros tipos", m.get("out_grande"), m.get("out_nao"),
                                   m.get("out_null"), out_tot])
        linha += 1
        prec, cob = verif.get("precisao"), verif.get("cobertura")
        linha = _linha(ws, linha, [
            "Precisão da regra", f"{prec:.1f}%" if prec is not None else "n/d",
            "das SS de INDISPONIBILIDADE realmente exigiram peça grande"], negrito=True, destaque=True)
        ws.merge_cells(start_row=linha - 1, start_column=3, end_row=linha - 1, end_column=8)
        linha = _linha(ws, linha, [
            "Cobertura da regra", f"{cob:.1f}%" if cob is not None else "n/d",
            "das SS que exigiram peça grande estavam classificadas como INDISPONIBILIDADE"],
            negrito=True, destaque=True)
        ws.merge_cells(start_row=linha - 1, start_column=3, end_row=linha - 1, end_column=8)
        linha += 2

    texto = (verif or {}).get("sintese") or ""
    if texto:
        linha = _linha(ws, linha, ["LEITURA DO RESULTADO"], negrito=True)
        for paragrafo in [p.strip() for p in texto.split("\n") if p.strip()]:
            c = ws.cell(row=linha, column=1, value=paragrafo)
            c.font = Font(name="Calibri", size=11, color=TINTA)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=8)
            ws.row_dimensions[linha] = ws.row_dimensions[linha]
            ws.row_dimensions[linha].height = max(16, 15 * (len(paragrafo) // 110 + 1))
            linha += 1
    return ws


def aba_taxa_por_ano(wb, taxa):
    ws = wb.create_sheet("Taxa por ano")
    _larguras(ws, [30, 14, 14, 14, 6, 14, 14, 14])
    linha = _titulo(
        ws,
        "Taxa por ano — 2024, 2025 e 2026",
        "2026 vai até 12/08 e a exposição usa o fator 0,611 do ano; sem isso a taxa sairia pela "
        "metade e pareceria queda. Duas réguas lado a lado: a chamada atribuída ao equipamento "
        "(tudo que a base registra contra o ativo) e a peça grande (a régua do gestor).",
    )
    serie = taxa.get("serie_por_ano") or {}
    regua = (taxa.get("regua_do_componente") or {}).get("por_familia_e_ano") or {}
    aic = (taxa.get("trocas_no_aic") or {}).get("por_ano_de_conclusao_fisica") or {}

    for fam, rotulo in (("religador", "RELIGADOR"), ("regulador", "REGULADOR")):
        parque = (taxa.get("parque") or {}).get(fam)
        linha = _linha(ws, linha, [f"{rotulo} — parque {parque}"], negrito=True)
        linha = _cabecalho(ws, linha, [
            "", "2024", "2025", "2026 (até 12/08)", "", "unidade", "", ""])
        anos = ("2024", "2025", "2026")

        def bloco(rot, fn, unidade, negrito=False, destaque=False):
            nonlocal linha
            linha = _linha(ws, linha, [rot] + [fn(a) for a in anos] + ["", unidade],
                           negrito=negrito, destaque=destaque)

        bloco("Parque exposto (equipamento-ano)",
              lambda a: serie.get(a, {}).get(fam, {}).get("equipamento_ano"), "eq-ano")
        linha = _linha(ws, linha, ["RÉGUA AMPLA — chamada atribuída ao equipamento"], negrito=True)
        bloco("Eventos de falha", lambda a: serie.get(a, {}).get(fam, {}).get("eventos"), "eventos")
        bloco("Taxa", lambda a: serie.get(a, {}).get(fam, {}).get("taxa_100"), "por 100 eq-ano", negrito=True)
        bloco("— falha funcional (parou)", lambda a: serie.get(a, {}).get(fam, {}).get("funcional_100"), "por 100 eq-ano")
        bloco("— anomalia (opera com defeito)", lambda a: serie.get(a, {}).get(fam, {}).get("anomalia_100"), "por 100 eq-ano")
        bloco("Ativos distintos que falharam", lambda a: serie.get(a, {}).get(fam, {}).get("ativos_distintos"), "ativos")
        bloco("Incidência sobre o parque", lambda a: serie.get(a, {}).get(fam, {}).get("incidencia_pct"), "%")

        linha = _linha(ws, linha, ["RÉGUA DO GESTOR — só o que exige peça grande"], negrito=True)
        bloco("Troca executada (obra encerrada no AIC)", lambda a: aic.get(a, {}).get(fam, 0), "obras")
        bloco("Peça grande com evidência direta",
              lambda a: regua.get(fam, {}).get(a, {}).get("com_peca_grande"), "eventos")
        bloco("Taxa na régua do gestor",
              lambda a: _taxa_gestor(taxa, fam, a), "por 100 eq-ano", negrito=True, destaque=True)
        bloco("Acessório — trazido, não somado",
              lambda a: regua.get(fam, {}).get(a, {}).get("acessorio_separado"), "eventos")
        bloco("Sem evidência de componente",
              lambda a: regua.get(fam, {}).get(a, {}).get("sem_evidencia_de_componente"), "eventos")
        linha += 2
    return ws


def _taxa_gestor(taxa, fam, ano):
    regua = (taxa.get("regua_do_componente") or {}).get("por_familia_e_ano") or {}
    aic = (taxa.get("trocas_no_aic") or {}).get("por_ano_de_conclusao_fisica") or {}
    serie = taxa.get("serie_por_ano") or {}
    eq_ano = serie.get(ano, {}).get(fam, {}).get("equipamento_ano")
    if not eq_ano:
        return None
    n = (aic.get(ano, {}).get(fam, 0) or 0) + (regua.get(fam, {}).get(ano, {}).get("com_peca_grande") or 0)
    return round(100.0 * n / eq_ano, 1)


def aba_matriz(wb, verif):
    ws = wb.create_sheet("Verificação SS a SS")
    _larguras(ws, [22, 30, 13, 22, 12, 9, 58, 10, 46])
    linha = _titulo(
        ws,
        "Verificação SS a SS — leitura integral com revisão adversarial",
        "Oito agentes leram a descrição completa de cada SS. Os vereditos que contrariam a "
        "hipótese do gestor, mais uma amostra de controle, passaram por revisores adversariais "
        "instruídos a derrubar o veredito — só ficou o que resistiu.",
        largura=9,
    )
    linha = _cabecalho(ws, linha, [
        "SS", "Tipo da SS", "Peça grande?", "Componente", "Família", "Furto",
        "Evidência no texto", "Confiança", "Revisão"])
    for item in (verif or {}).get("itens", []):
        veredito = item.get("peca_grande")
        rotulo = "SIM" if veredito is True else ("não" if veredito is False else "indefinido")
        revisao = ""
        if item.get("revisado"):
            revisao = ("manteve — " if item.get("revisao_manteve") else "DERRUBOU — ") + \
                      (item.get("revisao_motivo") or "")[:220]
        linha = _linha(ws, linha, [
            item.get("ss"), item.get("tiposs"), rotulo, item.get("componente"),
            item.get("familia"), "sim" if item.get("furto") else "",
            (item.get("evidencia") or "")[:300], item.get("confianca"), revisao,
        ], destaque=veredito is True, quebra=True)
    ws.freeze_panes = ws.cell(row=linha - len((verif or {}).get("itens", [])), column=1)
    return ws


def aba_peca_grande(wb, taxa):
    ws = wb.create_sheet("Peça grande")
    _larguras(ws, [26, 22, 14, 20, 40])
    linha = _titulo(
        ws,
        "Peça grande já levada ao campo — a fila material",
        "Convenção do Allan, que é a régua do gestor letra por letra: religador se divide em "
        "parte ativa e controle; regulador, em célula e controle. «Levado» é o que saiu do "
        "almoxarifado e ficou na obra (RMA menos DMA). O extrato foi montado só com obra NÃO "
        "concluída, então não se sobrepõe às obras encerradas do AIC.",
        largura=5,
    )
    campo = taxa.get("peca_grande_em_campo") or {}
    linha = _cabecalho(ws, linha, ["Família", "Classe da peça", "Peças levadas", "Valor", ""])
    for fam, dados in (campo.get("por_familia") or {}).items():
        for classe, qtd in (dados.get("detalhe") or {}).items():
            linha = _linha(ws, linha, [fam.title(), classe, qtd, None, ""])
        linha = _linha(ws, linha, [f"{fam.title()} — total", "", dados.get("pecas"),
                                   dados.get("valor"), ""], negrito=True)
    linha += 1
    linha = _linha(ws, linha, ["TOTAL", f"{campo.get('obras')} obras não concluídas",
                               campo.get("pecas_levadas"), campo.get("valor_levado"), ""],
                   negrito=True, destaque=True)
    for r in range(1, linha + 1):
        c = ws.cell(row=r, column=4)
        if isinstance(c.value, (int, float)):
            c.number_format = 'R$ #,##0.00'

    linha += 2
    sub = taxa.get("substituicao") or {}
    linha = _linha(ws, linha, ["PEÇA REGISTRADA NA CARTEIRA DO COEP"], negrito=True)
    linha = _cabecalho(ws, linha, ["Peça", "Ocorrências", "", "", ""])
    for peca, n in (sub.get("peca_substituida") or [])[:16]:
        linha = _linha(ws, linha, [peca, n, "", "", ""])
    return ws


def aba_premissas(wb, taxa, verif):
    ws = wb.create_sheet("Premissas")
    _larguras(ws, [4, 118])
    linha = _titulo(
        ws, "Premissas",
        "Cada número desta planilha depende do que está escrito aqui. Premissa que muda, "
        "número que muda.", largura=2,
    )
    for i, p in enumerate(taxa.get("premissas", []), start=1):
        c1 = ws.cell(row=linha, column=1, value=i)
        c1.font = Font(name="Calibri", size=11, bold=True, color=DESTAQUE)
        c1.alignment = Alignment(horizontal="center", vertical="top")
        c2 = ws.cell(row=linha, column=2, value=p)
        c2.font = Font(name="Calibri", size=11, color=TINTA)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[linha].height = max(16, 14 * (len(p) // 115 + 1))
        linha += 1

    linha += 1
    limite = (taxa.get("regua_do_componente") or {}).get("limite")
    if limite:
        ws.cell(row=linha, column=2, value="LIMITE DA RÉGUA HOJE").font = Font(bold=True, color=DESTAQUE)
        linha += 1
        c = ws.cell(row=linha, column=2, value=limite)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[linha].height = 14 * (len(limite) // 115 + 2)
    return ws


def main():
    with open(ARQ_TAXA, encoding="utf-8") as fh:
        taxa = json.load(fh)
    verif = None
    if os.path.exists(ARQ_VERIF):
        with open(ARQ_VERIF, encoding="utf-8") as fh:
            verif = json.load(fh)

    wb = Workbook()
    wb.remove(wb.active)
    aba_sintese(wb, taxa, verif)
    aba_taxa_por_ano(wb, taxa)
    if verif:
        aba_matriz(wb, verif)
    aba_peca_grande(wb, taxa)
    aba_premissas(wb, taxa, verif)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    print("gravado:", SAIDA)
    for ws in wb.worksheets:
        print(f"  aba {ws.title}: {ws.max_row} linhas")


if __name__ == "__main__":
    main()
