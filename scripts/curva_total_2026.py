"""
A curva TOTAL de 2026 — religador, regulador e banco de capacitor juntos.
dist/CURVA_TOTAL_2026.xlsx

O gestor pediu em 03/09, com o passivo dos anos anteriores dentro e o banco de
capacitor somado: «mensalize a curva da quantidade de pendentes — Total até chegar em
exatamente 102 em setembro. Depois mensaliza a curva de resolvidos — 71». E repetiu a
régua da emenda: «mensalize isso de Janeiro a Agosto; a partir de setembro os dados
têm que bater».

O QUE A BASE DÁ. Universo: quem passou pelo posto do COEP em 2026 — os 143 RL/RT da
partição fechada em 28 e 29/08, mais os bancos de capacitor que estiveram no posto pela
régua de manutenção do gestor (indisponibilidade, anomalia e aviso de anomalia).

  RL + RT   143 passaram ·  71 resolvidos ·  72 pendentes · herdados 50
  BC         46 passaram ·  11 resolvidos ·  35 pendentes · herdados 24
  TOTAL     189 passaram ·  82 resolvidos · 107 pendentes · herdados 74

A CONTA DELE E A CONTA DA BASE, LADO A LADO. Ele diz 173 passaram, 71 resolvidos e 102
pendentes; a base diz 189, 82 e 107. Os 71 resolvidos são exatamente os do RL/RT — a
base reproduz esse número no centavo, e o BC entra somando mais 11 em coluna própria,
sem mexer nele. No pendente a distância é de 5 equipamentos (107 contra 102), que vira
7 quando se compara com os 100 de agosto implícitos no gráfico dele — a diferença anda
constante o ano inteiro e está na aba «A ponte para o 102».

NÃO EXISTE CORTE DE BC QUE FECHE OS 173. Foram varridas as 32.768 combinações de TIPOSS
do banco de capacitor, no parque e no posto: nenhuma entrega ao mesmo tempo 30 que
passaram e 30 pendentes, que é o que faltaria para 143 + 30 = 173 e 72 + 30 = 102. As
que dão 30 num eixo erram no outro e não têm régua que as defenda. O 173 do gestor é
aritmética — 102 + 71 —, não um universo medido. Por isso a planilha entrega o apurado
e a ponte, e não um corte fabricado para bater.

A EMENDA COM SETEMBRO. A premissa dele (entrante 8·5·4·6, resolver 6·29·22·20) aplicada
ao apurado de agosto leva dezembro a 53, não aos 46 do gráfico dele: a diferença de 7
atravessa os quatro meses sem se fechar.

Rodar: python3 scripts/curva_total_2026.py
"""

import datetime as dt
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Font, PatternFill

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
import ano_dcmd as ad      # noqa: E402  — o apurado de RL/RT
import backlog_mensal as bm  # noqa: E402  — a máquina de demandas e o estilo

SAIDA = os.path.join(RAIZ, "dist", "CURVA_TOTAL_2026.xlsx")
JSON_SAIDA = os.path.join(RAIZ, "data", "missao", "curva_total.json")

TINTA, PAPEL, SOMBRA, SINAL = bm.TINTA, bm.PAPEL, bm.SOMBRA, bm.SINAL
VERDE, LARANJA, NEUTRO = bm.VERDE, bm.LARANJA, bm.NEUTRO
MESES = ad.MESES
D = dt.timedelta(days=1)
INI26, FIM26 = dt.date(2026, 1, 1), dt.date(2026, 12, 31)

# a régua de manutenção do gestor aplicada ao BC: saiu de operação, roda com defeito,
# ou tem aviso de anomalia. Fica de fora obra nova, comissionamento e ajuste de proteção.
MANUTENCAO = {bm.IND} | bm.ANOMALIA | {"AVISO DE ANOMALIA"}

# o que o gestor deu em 03/09 e o gráfico de forecast dele (set → dez)
GESTOR_PASSARAM, GESTOR_RESOLVIDOS, GESTOR_PENDENTES = 173, 71, 102
PREM_ENTRA = [8, 5, 4, 6]
PREM_RESOLVE = [6, 29, 22, 20]
GESTOR_CURVA = [102, 78, 60, 46]


# --------------------------------------------------------------------------- apuração
def bc_no_posto():
    """Os bancos de capacitor que estiveram no posto do COEP, pela régua de manutenção."""
    ss, posicao = bm.ler()
    posto = {x["NUMERO_SS"] for x in ss if x.get("COD_EQUIPE") == "ETO-COEP"}
    dem, _ = bm.demandas([x for x in ss if x["NUM_TRAFO"][:2] == "59"], MANUTENCAO)
    itens = [d for d in dem
             if any(n in posto for n in d["ss"])
             and d["abertura"] <= FIM26 and d["fim"] >= INI26]
    return itens, posicao


def serie_bc(itens, posicao):
    aberto = lambda t: sum(1 for d in itens if d["abertura"] <= t < d["fim"])
    linhas = []
    for m in range(1, 9):
        i = dt.date(2026, m, 1)
        f = posicao if m == 8 else dt.date(2026, m + 1, 1) - D
        ent = sum(1 for d in itens if i <= d["abertura"] <= f)
        sai = sum(1 for d in itens if i <= d["fim"] <= f)
        si, sf = aberto(i - D), aberto(f)
        assert sf == si + ent - sai, (m, si, ent, sai, sf)
        linhas.append({"mes": m, "rotulo": MESES[m - 1], "inicio": si, "entraram": ent,
                       "resolvidos": sai, "fim": sf})
    return linhas


def juntar(rlrt, bc):
    """Uma linha por mês com RL/RT, BC e o total, jan a ago."""
    linhas = []
    for a, b in zip(rlrt[:8], bc):
        linhas.append({
            "mes": a["mes"], "rotulo": a["rotulo"],
            "rlrt_inicio": a["inicio"], "rlrt_entraram": a["entraram"],
            "rlrt_resolvidos": a["resolvidos"], "rlrt_fim": a["fim"],
            "bc_inicio": b["inicio"], "bc_entraram": b["entraram"],
            "bc_resolvidos": b["resolvidos"], "bc_fim": b["fim"],
            "inicio": a["inicio"] + b["inicio"], "entraram": a["entraram"] + b["entraram"],
            "resolvidos": a["resolvidos"] + b["resolvidos"], "fim": a["fim"] + b["fim"],
        })
    for L in linhas:
        assert L["fim"] == L["inicio"] + L["entraram"] - L["resolvidos"], L["rotulo"]
    return linhas


def premissa(fim_agosto):
    """Set–dez pela premissa do gestor, ancorada no estoque que a base apurou."""
    saldo, linhas = fim_agosto, []
    for j in range(4):
        ini = saldo
        saldo += PREM_ENTRA[j] - PREM_RESOLVE[j]
        linhas.append({"mes": 9 + j, "rotulo": MESES[8 + j], "inicio": ini,
                       "entraram": PREM_ENTRA[j], "resolvidos": PREM_RESOLVE[j], "fim": saldo,
                       "gestor": GESTOR_CURVA[j]})
    return linhas


# ------------------------------------------------------------------------------- abas
def aba_curva(wb, linhas, set_dez):
    ws = wb.create_sheet("A curva total")
    passaram = linhas[0]["inicio"] + sum(L["entraram"] for L in linhas)
    bm.titulo(ws, "A CURVA TOTAL DE 2026 — religador, regulador e banco de capacitor",
              "Quem passou pelo posto do COEP em 2026, com o passivo dos anos anteriores dentro. "
              "O ano abre com %d demandas abertas, entram %d e saem %d: %d passaram pelo posto e "
              "sobram %d no fim de agosto. O gestor trabalha com 173 · 71 · 102 — os %d resolvidos "
              "de RL e RT são exatamente o 71 dele; a diferença mora no banco de capacitor e está "
              "aberta na aba «A ponte para o 102»."
              % (linhas[0]["inicio"], sum(L["entraram"] for L in linhas),
                 sum(L["resolvidos"] for L in linhas), passaram, linhas[-1]["fim"],
                 sum(L["rlrt_resolvidos"] for L in linhas)))
    cab = ["Mês", "Origem", "Pendentes no início", "Entraram", "Resolvidos", "PENDENTES NO FIM",
           "RL/RT no fim", "BC no fim", "RL/RT resolvidos", "BC resolvidos",
           "RL/RT entraram", "BC entraram"]
    bm.cabecalho(ws, 4, cab, [13, 11, 18, 11, 11, 17, 13, 11, 15, 13, 14, 12])
    r = 5
    for L in linhas:
        ws.append([L["rotulo"], "apurado", L["inicio"], L["entraram"], L["resolvidos"], L["fim"],
                   L["rlrt_fim"], L["bc_fim"], L["rlrt_resolvidos"], L["bc_resolvidos"],
                   L["rlrt_entraram"], L["bc_entraram"]])
        r += 1
    for L in set_dez:
        ws.append([L["rotulo"], "premissa", L["inicio"], L["entraram"], L["resolvidos"], L["fim"],
                   None, None, None, None, None, None])
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=SOMBRA)
        r += 1
    fim = r - 1
    for linha in range(5, fim + 1):
        for c in range(2, 13):
            ws.cell(row=linha, column=c).alignment = Alignment(horizontal="center")
        ws.cell(row=linha, column=6).font = Font(bold=True)
    ws.cell(row=r, column=1, value="no ano").font = Font(bold=True)
    ws.cell(row=r, column=3, value=linhas[0]["inicio"]).font = Font(bold=True)
    ws.cell(row=r, column=4, value="=SUM(D5:D%d)" % fim).font = Font(bold=True)
    ws.cell(row=r, column=5, value="=SUM(E5:E%d)" % fim).font = Font(bold=True)
    ws.cell(row=r, column=6, value=set_dez[-1]["fim"]).font = Font(bold=True)
    for c in range(3, 7):
        ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth, ch.overlap = "col", "stacked", 60, 100
    ch.add_data(Reference(ws, min_col=7, min_row=4, max_col=8, max_row=12), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=12))
    ch.title = "Pendentes no fim de cada mês — RL/RT embaixo, banco de capacitor em cima"
    ch.y_axis.title = "equipamentos"
    bm.cor_barra(ch.series[0], LARANJA)
    bm.cor_barra(ch.series[1], NEUTRO)
    for s in ch.series:
        bm.rotulos(s)
    bm.categorias(ch, ws, "$A$5:$A$12")
    ws.add_chart(bm.estilo(ch, 11, 26), "A%d" % (r + 3))

    ch2 = BarChart()
    ch2.type, ch2.grouping, ch2.gapWidth, ch2.overlap = "col", "clustered", 80, -12
    ch2.add_data(Reference(ws, min_col=4, min_row=4, max_col=5, max_row=fim), titles_from_data=True)
    ch2.set_categories(Reference(ws, min_col=1, min_row=5, max_row=fim))
    ch2.title = "Entraram e resolvidos, mês a mês (set–dez é a premissa do gestor)"
    ch2.y_axis.title = "equipamentos"
    bm.cor_barra(ch2.series[0], LARANJA)
    bm.cor_barra(ch2.series[1], VERDE)
    for s in ch2.series:
        bm.rotulos(s)
    bm.categorias(ch2, ws, "$A$5:$A$%d" % fim)
    ws.add_chart(bm.estilo(ch2, 11, 26), "A%d" % (r + 26))
    return ws


def aba_pendentes(wb, linhas, set_dez):
    ws = wb.create_sheet("Pendentes mes a mes")
    bm.titulo(ws, "A CURVA DE PENDENTES — o estoque no fim de cada mês",
              "É a curva que o gestor pediu. Sai de %d no começo do ano e chega a %d no fim de "
              "agosto. Aplicando a premissa dele a partir daí, setembro fecha em %d contra os %d "
              "do gráfico dele: %d equipamentos de diferença, que atravessam os quatro meses sem "
              "se fechar. A coluna «Conta do gestor» repete a curva dele para comparar."
              % (linhas[0]["inicio"], linhas[-1]["fim"], set_dez[0]["fim"], GESTOR_PENDENTES,
                 set_dez[0]["fim"] - GESTOR_PENDENTES))
    bm.cabecalho(ws, 4, ["Mês", "RL/RT", "BC", "TOTAL apurado", "Conta do gestor", "Diferença"],
                 [14, 11, 9, 15, 16, 12])
    r = 5
    ws.append(["dezembro/2025", linhas[0]["rlrt_inicio"], linhas[0]["bc_inicio"],
               linhas[0]["inicio"], None, None])
    r += 1
    for L in linhas:
        ws.append([L["rotulo"], L["rlrt_fim"], L["bc_fim"], L["fim"], None, None])
        r += 1
    for L in set_dez:
        ws.append([L["rotulo"], None, None, L["fim"], L["gestor"], L["fim"] - L["gestor"]])
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=SOMBRA)
        r += 1
    fim = r - 1
    for linha in range(5, fim + 1):
        for c in range(2, 7):
            ws.cell(row=linha, column=c).alignment = Alignment(horizontal="center")
        ws.cell(row=linha, column=4).font = Font(bold=True)

    ch = LineChart()
    ch.add_data(Reference(ws, min_col=4, min_row=4, max_row=fim), titles_from_data=True)
    ch.add_data(Reference(ws, min_col=5, min_row=4, max_row=fim), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=fim))
    ch.title = "Pendentes no fim de cada mês: o apurado e a conta do gestor"
    ch.y_axis.title = "equipamentos"
    for s, cor in zip(ch.series, (LARANJA, NEUTRO)):
        s.graphicalProperties = GraphicalProperties()
        s.graphicalProperties.line = LineProperties(solidFill=cor, w=28000)
        s.marker = Marker(symbol="circle", size=6)
        s.marker.graphicalProperties = GraphicalProperties(solidFill=cor)
        s.smooth = False
        bm.rotulos(s)
    bm.categorias(ch, ws, "$A$5:$A$%d" % fim)
    ws.add_chart(bm.estilo(ch, 12, 26), "A%d" % (fim + 3))

    ch2 = BarChart()
    ch2.type, ch2.grouping, ch2.gapWidth, ch2.overlap = "col", "stacked", 60, 100
    ch2.add_data(Reference(ws, min_col=2, min_row=4, max_col=3, max_row=13), titles_from_data=True)
    ch2.set_categories(Reference(ws, min_col=1, min_row=5, max_row=13))
    ch2.title = "De que é feito o estoque: religador e regulador embaixo, capacitor em cima"
    ch2.y_axis.title = "equipamentos"
    bm.cor_barra(ch2.series[0], LARANJA)
    bm.cor_barra(ch2.series[1], NEUTRO)
    for s in ch2.series:
        bm.rotulos(s)
    bm.categorias(ch2, ws, "$A$5:$A$13")
    ws.add_chart(bm.estilo(ch2, 11, 26), "A%d" % (fim + 27))


def aba_resolvidos(wb, linhas):
    ws = wb.create_sheet("Resolvidos mes a mes")
    rl = sum(L["rlrt_resolvidos"] for L in linhas)
    bc = sum(L["bc_resolvidos"] for L in linhas)
    bm.titulo(ws, "A CURVA DE RESOLVIDOS — os 71 do gestor, mês a mês",
              "Os %d resolvidos de religador e regulador são exatamente o 71 que o gestor "
              "reconhece, e a base os mensaliza sem ajuste: junho sozinho responde por %d. O banco "
              "de capacitor soma outros %d, em coluna própria, para não mexer no número dele — "
              "no total o posto encerrou %d demandas de janeiro a agosto."
              % (rl, max(L["rlrt_resolvidos"] for L in linhas), bc, rl + bc))
    bm.cabecalho(ws, 4, ["Mês", "RL/RT (os 71)", "BC", "TOTAL", "Acumulado RL/RT", "Acumulado total"],
                 [14, 14, 9, 11, 16, 16])
    r, ac, at = 5, 0, 0
    for L in linhas:
        ac += L["rlrt_resolvidos"]
        at += L["resolvidos"]
        ws.append([L["rotulo"], L["rlrt_resolvidos"], L["bc_resolvidos"], L["resolvidos"], ac, at])
        for c in range(2, 7):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1
    fim = r - 1
    ws.cell(row=r, column=1, value="no ano").font = Font(bold=True)
    for c, ref in ((2, "B"), (3, "C"), (4, "D")):
        cel = ws.cell(row=r, column=c, value="=SUM(%s5:%s%d)" % (ref, ref, fim))
        cel.font = Font(bold=True)
        cel.alignment = Alignment(horizontal="center")

    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth, ch.overlap = "col", "stacked", 60, 100
    ch.add_data(Reference(ws, min_col=2, min_row=4, max_col=3, max_row=fim), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=fim))
    ch.title = "Resolvidos em cada mês — os 71 de RL/RT e os do banco de capacitor"
    ch.y_axis.title = "equipamentos"
    bm.cor_barra(ch.series[0], VERDE)
    bm.cor_barra(ch.series[1], NEUTRO)
    for s in ch.series:
        bm.rotulos(s)
    bm.categorias(ch, ws, "$A$5:$A$%d" % fim)
    ws.add_chart(bm.estilo(ch, 11, 26), "A%d" % (r + 3))

    ch2 = LineChart()
    ch2.add_data(Reference(ws, min_col=5, min_row=4, max_col=6, max_row=fim), titles_from_data=True)
    ch2.set_categories(Reference(ws, min_col=1, min_row=5, max_row=fim))
    ch2.title = "Resolvidos acumulados: a linha que chega aos 71"
    ch2.y_axis.title = "equipamentos"
    for s, cor in zip(ch2.series, (VERDE, NEUTRO)):
        s.graphicalProperties = GraphicalProperties()
        s.graphicalProperties.line = LineProperties(solidFill=cor, w=28000)
        s.marker = Marker(symbol="circle", size=6)
        s.marker.graphicalProperties = GraphicalProperties(solidFill=cor)
        s.smooth = False
        bm.rotulos(s)
    bm.categorias(ch2, ws, "$A$5:$A$%d" % fim)
    ws.add_chart(bm.estilo(ch2, 11, 26), "A%d" % (r + 26))


def aba_ponte(wb, linhas, set_dez):
    ws = wb.create_sheet("A ponte para o 102")
    passaram = linhas[0]["inicio"] + sum(L["entraram"] for L in linhas)
    resolvidos = sum(L["resolvidos"] for L in linhas)
    pendentes = linhas[-1]["fim"]
    bm.titulo(ws, "A PONTE — da conta do gestor para a conta da base",
              "O gestor dá 173 que passaram, 71 resolvidos e 102 pendentes. A base dá %d, %d e %d. "
              "Nada aqui foi ajustado para bater: as duas contas ficam à vista e a diferença é "
              "nomeada linha a linha." % (passaram, resolvidos, pendentes))
    bm.cabecalho(ws, 4, ["O que se conta", "Conta do gestor", "Conta da base", "Diferença",
                         "Onde está a diferença"], [30, 16, 15, 11, 74])
    linhas_ponte = [
        ("Passaram pelo posto em 2026", GESTOR_PASSARAM, passaram,
         "Os 143 RL/RT da partição fechada em 28 e 29/08 mais %d bancos de capacitor que estiveram "
         "no posto pela régua de manutenção. O 173 do gestor é 102 + 71 — aritmética, não um "
         "universo medido." % (passaram - 143)),
        ("Resolvidos até agosto", GESTOR_RESOLVIDOS, resolvidos,
         "Os %d de RL/RT são o 71 dele, no número exato. A diferença é só o banco de capacitor, que "
         "encerrou %d demandas no posto e nunca entrou na conta dele."
         % (sum(L["rlrt_resolvidos"] for L in linhas), sum(L["bc_resolvidos"] for L in linhas))),
        ("Pendentes no fim de agosto", GESTOR_PENDENTES - 2, pendentes,
         "O gráfico dele fecha agosto em 100 para que setembro caia em 102 (100 + 8 − 6). A base "
         "fecha em %d: %d RL/RT e %d BC." % (pendentes, linhas[-1]["rlrt_fim"], linhas[-1]["bc_fim"])),
        ("Pendentes em setembro", GESTOR_PENDENTES, set_dez[0]["fim"],
         "Mesma premissa dele — entram 8, resolvem 6 — aplicada ao estoque apurado."),
        ("Pendentes em dezembro", GESTOR_CURVA[-1], set_dez[-1]["fim"],
         "A premissa inteira reancorada no apurado. A diferença de %d não se fecha: ela anda igual "
         "nos quatro meses." % (set_dez[-1]["fim"] - GESTOR_CURVA[-1])),
    ]
    r = 5
    for nome, g, b, onde in linhas_ponte:
        ws.append([nome, g, b, b - g, onde])
        for c in (2, 3, 4):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 44
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="POR QUE NÃO EXISTE UM CORTE QUE DÊ 173").font = \
        Font(bold=True, size=11, color=SINAL)
    r += 1
    for t in [
        "Para 143 + BC = 173 e 72 + BC = 102, o banco de capacitor teria de entregar ao mesmo tempo "
        "30 que passaram e 30 pendentes — ou seja, nenhum resolvido no ano inteiro.",
        "Foram varridas as 32.768 combinações de TIPOSS do BC, no parque e no posto. As que dão 30 "
        "num eixo erram no outro, e nenhuma tem régua que a defenda: são listas de tipos montadas "
        "para bater o número.",
        "Os cortes com régua, no posto: indisponibilidade dá 24 e 13; indisponibilidade + anomalia "
        "dá 26 e 15; a régua de manutenção usada aqui dá 46 e 35; tudo dá 56 e 45.",
        "O 173 é a soma 102 + 71 e o 102 é o primeiro ponto do gráfico de forecast. São números de "
        "referência, e a planilha os mantém à vista em vez de fabricar um corte para alcançá-los.",
    ]:
        c = ws.cell(row=r, column=1, value="· " + t)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 30
        r += 1


def aba_setdez(wb, set_dez):
    ws = wb.create_sheet("Setembro em diante")
    bm.titulo(ws, "A EMENDA — o que fecha agosto é o que abre setembro",
              "A régua que o gestor repetiu duas vezes. A premissa dele é a mesma nas duas colunas "
              "— entrante 8·5·4·6 e resolver 6·29·22·20 —; o que muda é o ponto de partida. Saindo "
              "do estoque apurado, dezembro fecha em %d; saindo dos 100 do gráfico dele, em %d."
              % (set_dez[-1]["fim"], GESTOR_CURVA[-1]))
    bm.cabecalho(ws, 4, ["Mês", "Entrante", "Resolver", "Ancorado no apurado",
                         "Como está no gráfico", "Diferença"], [13, 11, 11, 19, 20, 11])
    r = 5
    for L in set_dez:
        ws.append([L["rotulo"], L["entraram"], L["resolvidos"], L["fim"], L["gestor"],
                   L["fim"] - L["gestor"]])
        for c in range(2, 7):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1
    fim = r - 1
    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth, ch.overlap = "col", "clustered", 80, -12
    ch.add_data(Reference(ws, min_col=4, min_row=4, max_col=5, max_row=fim), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=fim))
    ch.title = "Set–dez: a premissa ancorada no apurado e a curva do gráfico do gestor"
    ch.y_axis.title = "equipamentos"
    bm.cor_barra(ch.series[0], LARANJA)
    bm.cor_barra(ch.series[1], NEUTRO)
    for s in ch.series:
        bm.rotulos(s)
    bm.categorias(ch, ws, "$A$5:$A$%d" % fim)
    ws.add_chart(bm.estilo(ch, 11, 26), "A%d" % (r + 3))


def aba_bc(wb, itens, posicao, bc):
    ws = wb.create_sheet("Banco de capacitor")
    bm.titulo(ws, "O BANCO DE CAPACITOR — código operativo 59",
              "Ele nunca aparecia porque a consulta da base de repasse o descarta na cláusula "
              "«COD_ELE NOT IN ('59','BR')». Indo direto na base de SS/OS são 242 SS em 102 ativos; "
              "no posto do COEP, pela régua de manutenção, são %d demandas em 2026 — %d resolvidas "
              "e %d ainda abertas na posição de %s."
              % (len(itens), sum(1 for d in itens if INI26 <= d["fim"] <= posicao),
                 sum(1 for d in itens if d["abertura"] <= posicao < d["fim"]),
                 posicao.strftime("%d/%m/%Y")))
    bm.cabecalho(ws, 4, ["Mês", "No início", "Entraram", "Resolvidos", "No fim"],
                 [14, 12, 11, 12, 11])
    r = 5
    for L in bc:
        ws.append([L["rotulo"], L["inicio"], L["entraram"], L["resolvidos"], L["fim"]])
        for c in range(2, 6):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1
    fim = r - 1
    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth = "col", "clustered", 70
    ch.add_data(Reference(ws, min_col=5, min_row=4, max_row=fim), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=fim))
    ch.title = "Bancos de capacitor com demanda aberta no posto, fim de cada mês"
    ch.y_axis.title = "bancos"
    bm.cor_barra(ch.series[0], NEUTRO)
    bm.rotulos(ch.series[0])
    ch.legend = None
    bm.categorias(ch, ws, "$A$5:$A$%d" % fim)
    ws.add_chart(bm.estilo(ch, 11, 24), "G4")

    r += 2
    ws.cell(row=r, column=1, value="AS %d DEMANDAS, UMA A UMA" % len(itens)).font = \
        Font(bold=True, size=11, color=SINAL)
    r += 1
    bm.cabecalho(ws, r, ["Ativo", "Localidade", "Abriu", "Saiu", "Situação", "SS"],
                 [14, 20, 12, 12, 22, 46])
    r += 1
    for d in sorted(itens, key=lambda d: (d["abertura"], d["ativo"])):
        aberta = d["abertura"] <= posicao < d["fim"]
        ws.cell(row=r, column=1, value=d["ativo"])
        ws.cell(row=r, column=2, value=d["localidade"])
        ws.cell(row=r, column=3, value=d["abertura"].strftime("%d/%m/%Y"))
        ws.cell(row=r, column=4, value="" if aberta else d["fim"].strftime("%d/%m/%Y"))
        ws.cell(row=r, column=5, value="ainda aberta" if aberta else "resolvida")
        ws.cell(row=r, column=6, value=" · ".join(d["ss"]))
        for c in range(3, 6):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1


def aba_base(wb, itens_rlrt, itens_bc, posicao):
    ws = wb.create_sheet("Base das demandas")
    bm.titulo(ws, "A BASE — cada equipamento que passou pelo posto em 2026, uma linha",
              "Religador e regulador vêm de coep_2026.json, na partição de 28 e 29/08; o banco de "
              "capacitor vem da base de SS/OS. Ativo nunca se repete.")
    bm.cabecalho(ws, 4, ["Ativo", "Tipo", "Localidade", "Chegou", "Saiu", "Balde"],
                 [14, 8, 22, 12, 12, 24])
    r = 5
    linhas = []
    for i in itens_rlrt:
        linhas.append((i["ativo"], i["tipo"], i["localidade"], i["chegada"], i["saida"], i["balde"]))
    for d in itens_bc:
        aberta = d["abertura"] <= posicao < d["fim"]
        linhas.append((d["ativo"], "BC", d["localidade"], d["abertura"],
                       None if aberta else d["fim"], "resolvido" if not aberta else "na fila do posto"))
    for ativo, tipo, local, ent, sai, balde in sorted(linhas, key=lambda t: (t[1], t[0])):
        ws.cell(row=r, column=1, value=ativo)
        ws.cell(row=r, column=2, value=tipo)
        ws.cell(row=r, column=3, value=local)
        ws.cell(row=r, column=4, value=ent.strftime("%d/%m/%Y") if ent else "")
        ws.cell(row=r, column=5, value=sai.strftime("%d/%m/%Y") if sai else "")
        ws.cell(row=r, column=6, value=balde)
        for c in (2, 4, 5):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1
    codigos_rl = {i["ativo"] for i in itens_rlrt}
    codigos_bc = {d["ativo"] for d in itens_bc}
    assert not (codigos_rl & codigos_bc), "o mesmo código em RL/RT e BC"
    assert len(codigos_rl) == len(itens_rlrt), "ativo repetido em RL/RT"
    return len(linhas), len(codigos_rl) + len(codigos_bc)


def aba_como(wb, linhas, set_dez, posicao, n, ativos, n_bc):
    ws = wb.create_sheet("Como foi feito")
    ws.column_dimensions["A"].width = 112
    texto = [
        ("O QUE FOI PEDIDO", True),
        ("Gestor, 03/09: «levando em consideração que eu tinha 71 resolvidos até o final de agosto, "
         "então passaram 173 esses de Jan (com o passivo dos anos anteriores, levando em "
         "consideração também banco de capacitor). Quero que mensalize a curva da quantidade de "
         "pendentes — Total até chegar em exatamente 102 em setembro. Depois mensaliza a curva de "
         "resolvidos — 71.» E logo depois: «mensalize isso de Janeiro a Agosto; a partir de "
         "setembro os dados têm que bater».", False),
        ("", False),
        ("O UNIVERSO", True),
        ("Quem passou pelo posto do COEP em 2026, com o passivo dos anos anteriores dentro. São os "
         "143 religadores e reguladores da partição fechada em 28 e 29/08 — 71 resolvidos, 54 na "
         "fila, 15 despachados, 3 em execução — mais os bancos de capacitor que estiveram no posto.", False),
        ("A régua do BC é a de manutenção do gestor: indisponibilidade para operação, operação com "
         "anomalia e aviso de anomalia. Fica de fora obra de equipamento novo, comissionamento e "
         "ajuste de proteção, que ele mandou tirar da conta em 29/08. Dá %d demandas: %d herdadas de "
         "2025 ou antes e %d abertas no ano — em %d bancos, porque dois deles abriram e fecharam "
         "duas vezes no ano (5900358003 e 5900600004)."
         % (n_bc, linhas[0]["bc_inicio"], sum(L["bc_entraram"] for L in linhas),
            n_bc - 2), False),
        ("", False),
        ("COMO A DEMANDA ABRE E FECHA", True),
        ("Religador e regulador: abre na primeira chegada ao posto e fecha na data em que a cadeia "
         "encerrou. Dois ativos têm fechamento anterior à chegada (7921040031 e 7955430075) e a "
         "saída fica travada na data de entrada, senão o saldo não fecha.", False),
        ("Banco de capacitor: a demanda abre na abertura da primeira SS e fecha na saída da última. "
         "A saída de uma SS é a conclusão dela; senão a abertura da SS seguinte do mesmo ativo, "
         "porque SS repassada sai sem data de conclusão; senão segue aberta. SS repassada sem "
         "nenhuma seguinte no recorte sai na data em que foi aberta.", False),
        ("Conta EQUIPAMENTO, não SS. Ativo com três SS encadeadas é uma demanda só, porque repasse "
         "não é falha nova. São %d demandas em %d equipamentos: religador e regulador nunca se "
         "repetem, e no BC dois bancos aparecem duas vezes porque a demanda fechou e outra abriu "
         "depois." % (n, ativos), False),
        ("", False),
        ("A FRONTEIRA DO MÊS", True),
        ("Estoque no fim do mês = demanda com abertura ≤ último dia do mês e saída depois dele. "
         "Entradas e saídas do mês = datas dentro do mês. Com essa fronteira o saldo fecha nos oito "
         "meses, um a um: fim = início + entrou − saiu. O script quebra se não fechar.", False),
        ("", False),
        ("O QUE A SÉRIE DEU", True),
        ("O ano abre com %d demandas abertas — %d de RL/RT e %d de BC. Entram %d, saem %d, e agosto "
         "fecha em %d: %d religadores e reguladores mais %d bancos de capacitor."
         % (linhas[0]["inicio"], linhas[0]["rlrt_inicio"], linhas[0]["bc_inicio"],
            sum(L["entraram"] for L in linhas), sum(L["resolvidos"] for L in linhas),
            linhas[-1]["fim"], linhas[-1]["rlrt_fim"], linhas[-1]["bc_fim"]), False),
        ("Os resolvidos de RL/RT dão %d — o 71 do gestor, sem ajuste. O BC soma %d."
         % (sum(L["rlrt_resolvidos"] for L in linhas), sum(L["bc_resolvidos"] for L in linhas)), False),
        ("", False),
        ("A DISTÂNCIA PARA O 102, DITA POR INTEIRO", True),
        ("A base fecha agosto em %d e setembro em %d; o gráfico do gestor fecha agosto em 100 e "
         "setembro em 102. São %d equipamentos de diferença, e ela anda constante até dezembro, "
         "onde a base dá %d contra os 46 dele."
         % (linhas[-1]["fim"], set_dez[0]["fim"], set_dez[0]["fim"] - 102, set_dez[-1]["fim"]), False),
        ("Não existe corte de banco de capacitor que feche os 173: seriam precisos 30 que passaram "
         "E 30 pendentes ao mesmo tempo, ou seja, nenhum BC resolvido no ano. Foram varridas as "
         "32.768 combinações de TIPOSS, no parque e no posto — as que dão 30 num eixo erram no "
         "outro, e nenhuma tem régua que a defenda. O detalhe está na aba «A ponte para o 102».", False),
        ("", False),
        ("A POSIÇÃO", True),
        ("Religador e regulador: 18/08/2026, de coep_2026.json. Banco de capacitor: %s, que é o "
         "fecho de BASE_SS_OS_20082026.txt para o código 59. Agosto é mês PARCIAL nas duas pontas, "
         "então as entradas e saídas dele não se comparam direto com as dos meses cheios."
         % posicao.strftime("%d/%m/%Y"), False),
        ("Para chegar até hoje: base de SS/OS nova em data/raw, depois scripts/extrai_ssos_min.py e "
         "scripts/curva_total_2026.py.", False),
    ]
    for i, (t, negrito) in enumerate(texto, 1):
        c = ws.cell(row=i, column=1, value=t)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if negrito:
            c.font = Font(bold=True, size=11, color=SINAL if i == 1 else TINTA)


# ----------------------------------------------------------------------------- montar
def montar(saida=SAIDA):
    itens_rlrt, _ = ad.apurar()
    herdados, rlrt = ad.mensal(itens_rlrt)
    itens_bc, posicao = bc_no_posto()
    bc = serie_bc(itens_bc, posicao)
    linhas = juntar(rlrt, bc)
    set_dez = premissa(linhas[-1]["fim"])

    assert sum(L["rlrt_resolvidos"] for L in linhas) == 71, "os 71 do gestor têm de sair inteiros"
    assert linhas[-1]["rlrt_fim"] == 72, "RL/RT tem de fechar agosto em 72"
    assert linhas[0]["inicio"] == herdados + bc[0]["inicio"]

    wb = Workbook()
    wb.remove(wb.active)
    aba_curva(wb, linhas, set_dez)
    aba_pendentes(wb, linhas, set_dez)
    aba_resolvidos(wb, linhas)
    aba_ponte(wb, linhas, set_dez)
    aba_setdez(wb, set_dez)
    aba_bc(wb, itens_bc, posicao, bc)
    n, ativos = aba_base(wb, itens_rlrt, itens_bc, posicao)
    aba_como(wb, linhas, set_dez, posicao, n, ativos, len(itens_bc))
    os.makedirs(os.path.dirname(saida), exist_ok=True)
    wb.save(saida)

    with open(JSON_SAIDA, "w", encoding="utf-8") as fh:
        json.dump({"posicao_bc": posicao.isoformat(), "posicao_rlrt": ad.POSICAO.isoformat(),
                   "mensal": linhas, "set_dez": set_dez, "bc": bc,
                   "passaram": linhas[0]["inicio"] + sum(L["entraram"] for L in linhas),
                   "resolvidos": sum(L["resolvidos"] for L in linhas),
                   "pendentes_agosto": linhas[-1]["fim"],
                   "gestor": {"passaram": GESTOR_PASSARAM, "resolvidos": GESTOR_RESOLVIDOS,
                              "pendentes_setembro": GESTOR_PENDENTES, "curva": GESTOR_CURVA}},
                  fh, ensure_ascii=False, indent=1)

    print(saida)
    print("  passaram %d (143 RL/RT + %d BC) | resolvidos %d (71 + %d) | agosto %d (%d + %d)"
          % (linhas[0]["inicio"] + sum(L["entraram"] for L in linhas), len(itens_bc),
             sum(L["resolvidos"] for L in linhas), sum(L["bc_resolvidos"] for L in linhas),
             linhas[-1]["fim"], linhas[-1]["rlrt_fim"], linhas[-1]["bc_fim"]))
    print("  pendentes: %s" % " · ".join("%s %d" % (L["rotulo"][:3], L["fim"]) for L in linhas))
    print("  set–dez:   %s" % " · ".join("%s %d (gestor %d)" % (L["rotulo"][:3], L["fim"], L["gestor"])
                                         for L in set_dez))
    from planilha_automatica import grava_cache
    print("  cache: %d células" % grava_cache(saida))
    return saida


if __name__ == "__main__":
    montar(sys.argv[1] if len(sys.argv) > 1 else SAIDA)
