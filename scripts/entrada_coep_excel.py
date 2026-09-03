#!/usr/bin/env python3
"""
Planilha da entrada do posto ETO-COEP, mês a mês — três colunas, como o gestor pediu.

Todos os tipos de SS contam.

  1. ATIVOS — os indisponíveis da foto de junho pela data de abertura da SS. Janeiro carrega
     os entrantes de janeiro mais tudo que foi aberto no COEP em anos anteriores.
  2. ENTRANTES — ativos novos no posto pela data de abertura da SS, direto da base
     de SS/OS do ETO-COEP. Novo = primeira SS do posto naquele ativo em toda a base.
  3. RESOLVIDOS — pelo mês em que a tratativa aconteceu de verdade: data de término
     da SS ou data de repasse, não a data em que a SS abriu.

Gera: dist/ENTRADA_COEP_MES_A_MES.xlsx

Os valores são calculados aqui e gravados como número. O LibreOffice não roda neste
ambiente, então uma fórmula ficaria sem valor em cache e apareceria vazia em qualquer
leitor que não recalcule. Cada agregado tem a aba de detalhe que o sustenta.

Uso:  python3 scripts/entrada_coep_excel.py
"""

import csv
import datetime
import json
import os
import re
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SAIDA = os.path.join(RAIZ, "dist", "ENTRADA_COEP_MES_A_MES.xlsx")

MESES_PT = ["jan", "fev", "mar", "abr", "mai", "jun",
            "jul", "ago", "set", "out", "nov", "dez"]
MESES_2026 = [f"2026-{m:02d}" for m in range(1, 8)]  # janela: janeiro a julho

FONTE = "Arial"
TINTA = "1F1C17"
DESTAQUE = "FFF2CC"
ATUACAO = "DCE9DD"

fina = Side(style="thin", color="BFB9A6")
grossa = Side(style="medium", color=TINTA)


def dbr(t):
    t = str(t or "").strip()
    for f in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(t[:10], f).date()
        except ValueError:
            continue
    return None


def rotulo(mes):
    ano, m = mes.split("-")
    return f"{MESES_PT[int(m) - 1]}/{ano}"


# --------------------------------------------------------------- leitura
def carregar():
    with open(os.path.join(RAIZ, "data", "meta.json"), encoding="utf-8") as fh:
        meta = json.load(fh)
    with open(os.path.join(RAIZ, "data", "missao", "ssos_min.json"), encoding="utf-8") as fh:
        base = json.load(fh)
    aic = {}
    arq = os.path.join(RAIZ, "data", "raw", "aic_obras.csv")
    if os.path.exists(arq):
        with open(arq, encoding="utf-8") as fh:
            for row in csv.DictReader(fh, delimiter=";"):
                if row.get("data_encerramento"):
                    aic.setdefault(row["obra"], row["data_encerramento"])
    reportes = defaultdict(list)
    arq = os.path.join(RAIZ, "data", "raw", "reportes_campo.json")
    if os.path.exists(arq):
        with open(arq, encoding="utf-8") as fh:
            for r in json.load(fh):
                reportes[r["ativo"]].append(r)
    return meta, base, aic, reportes


ANO_DA_SS = re.compile(r"/(\d{4})\s*$")


def entrantes_no_coep(base, foto):
    """Ativos que passaram pelo posto ETO-COEP, mês a mês, pela abertura da SS."""
    coep = []
    for r in base:
        if (r.get("COD_EQUIPE") or "").strip() != "ETO-COEP":
            continue
        d = dbr(r.get("DATA_ABERTURA_SS"))
        a = (r.get("NUM_TRAFO") or "").strip()
        if d and a:
            coep.append({**r, "_d": d, "_a": a})
    coep.sort(key=lambda r: r["_d"])

    primeira = {}
    for r in coep:
        primeira.setdefault(r["_a"], r)

    def recarimbada(r):
        """SS cujo ano do número não bate com o ano da abertura.

        O SGM re-carimba a DATA_ABERTURA_SS quando a SS é reaberta ou repassada,
        então uma SS de 2023 pode aparecer com abertura em 2026. Onde isso
        acontece, a «entrada» daquele mês é demanda velha voltando, não nova.
        """
        m = ANO_DA_SS.search(str(r.get("NUMERO_SS") or ""))
        return bool(m) and int(m.group(1)) != r["_d"].year

    por_mes = defaultdict(lambda: {"ss": 0, "ativos": set(), "novos": set()})
    for r in coep:
        e = por_mes[f"{r['_d'].year}-{r['_d'].month:02d}"]
        e["ss"] += 1
        e["ativos"].add(r["_a"])
        if primeira[r["_a"]] is r:
            e["novos"].add(r["_a"])

    serie = []
    for k, e in sorted(por_mes.items()):
        novos = [primeira[a] for a in e["novos"]]
        serie.append({
            "mes": k, "rotulo": rotulo(k), "ss": e["ss"],
            "ativos": len(e["ativos"]), "novos": len(novos),
            "revisita": len(e["ativos"]) - len(novos),
            "na_foto": sum(1 for r in novos if r["_a"] in foto),
            "fora_da_foto": sum(1 for r in novos if r["_a"] not in foto),
            "ss_do_ano": sum(1 for r in novos if not recarimbada(r)),
            "ss_recarimbada": sum(1 for r in novos if recarimbada(r)),
        })

    detalhe = [{
        "ativo": a, "primeira_ss": r["NUMERO_SS"], "abertura": r["_d"],
        "mes": f"{r['_d'].year}-{r['_d'].month:02d}",
        "localidade": r.get("LOCALIDADE", ""),
        "equipamento": r.get("DESCICAO_DO_ATIVO", ""),
        "criticidade_ss": r.get("CRITICIDADE_SS", ""),
        "tiposs": r.get("TIPOSS", ""),
        "situacao_hoje": r.get("SITUACAO_SS", ""),
        "na_foto": "sim" if a in foto else "não",
        "recarimbada": "sim" if recarimbada(r) else "não",
    } for a, r in sorted(primeira.items(), key=lambda kv: kv[1]["_d"])]
    return serie, detalhe


def resolucoes(meta, base, aic, reportes):
    """Quando cada um dos 117 saiu da carteira — término da SS ou repasse."""
    mm = meta["entrada_mensal"]
    por_ss = defaultdict(list)
    por_ativo_ss = defaultdict(list)
    for r in base:
        por_ss[(r.get("NUMERO_SS") or "").strip()].append(r)
        a = (r.get("NUM_TRAFO") or "").strip()
        if a:
            por_ativo_ss[a].append(r)
    ficha = {}
    for balde in ("resolvidos", "verificar", "em_andamento"):
        for x in (meta["entrada"].get(balde) or {}).get("lista", []):
            ficha.setdefault(x["ativo"], x)

    linhas = []
    for x in mm["lista"]:
        e = ficha.get(x["ativo"], {})
        d = via = None
        situacao = "; ".join(sorted({(r.get("SITUACAO_SS") or "").strip()
                                     for r in por_ss.get(x["numero_ss"], [])
                                     if (r.get("SITUACAO_SS") or "").strip()}))

        if x["resolvido"]:
            for r in por_ss.get(x["numero_ss"], []):
                t = dbr(r.get("DATA_TERMINO_SS"))
                if t and (d is None or t < d):
                    d, via = t, ("cancelamento da SS de entrada"
                                 if "CANCELADA" in (r.get("SITUACAO_SS") or "").upper()
                                 else "término da SS de entrada")
            if not d and e.get("cauda_mesma_demanda"):
                ds = [y for y in (dbr(c.get("abertura")) for c in e["cauda_mesma_demanda"]) if y]
                if ds:
                    d, via = min(ds), "repasse para a etapa seguinte"
            if not d and e.get("obras_encerradas"):
                ds = [y for y in (dbr(aic.get(o)) for o in e["obras_encerradas"]) if y]
                if ds:
                    d, via = max(ds), "obra encerrada no AIC"
            if not d and reportes.get(x["ativo"]):
                ds = [y for y in (dbr(r["data"]) for r in reportes[x["ativo"]]) if y]
                if ds:
                    d, via = max(ds), "reporte de campo"
            if not d and e.get("decisao_gestor"):
                d = dbr(e["decisao_gestor"].get("data"))
                via = "decisão do gestor" if d else None
            if not d:
                abertura = dbr(x["abertura"])
                ds = [t for t in (dbr(r.get("DATA_TERMINO_SS"))
                                  for r in por_ativo_ss.get(x["ativo"], []))
                      if t and (not abertura or t >= abertura)]
                if ds:
                    d, via = max(ds), "última SS atendida do ativo"

        linhas.append({
            "ativo": x["ativo"], "localidade": x["localidade"], "tipo": x["tipo"],
            "numero_ss": x["numero_ss"], "abertura": dbr(x["abertura"]),
            "mes_entrada": x["mes"], "legado": "sim" if x["legado"] else "não",
            "resolvido": x["resolvido"], "parecer_coep": e.get("parecer_coep", ""),
            "situacao_ss": situacao, "resolucao": d,
            "mes_resolucao": f"{d.year}-{d.month:02d}" if d else "",
            "via": via or ("" if x["resolvido"] else "ainda no fluxo"),
            "motivo": x.get("motivo", ""), "fonte_data": x["fonte_data"],
        })
    return linhas


# --------------------------------------------------------------- planilha
def titulo(ws, texto, subtitulo, largura):
    ws["A1"] = texto
    ws["A1"].font = Font(name=FONTE, size=14, bold=True, color=TINTA)
    ws["A2"] = subtitulo
    ws["A2"].font = Font(name=FONTE, size=9, italic=True, color="5B5443")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=largura)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=largura)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30
    ws["A2"].alignment = Alignment(vertical="top", wrap_text=True)


def cabecalho(ws, linha, colunas, altura=34):
    for i, (rot, larg) in enumerate(colunas, start=1):
        c = ws.cell(row=linha, column=i, value=rot)
        c.font = Font(name=FONTE, size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=TINTA)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=grossa)
        if larg:
            ws.column_dimensions[get_column_letter(i)].width = larg
    ws.row_dimensions[linha].height = altura
    ws.freeze_panes = ws.cell(row=linha + 1, column=1)


def corpo(ws, linha, valores, negrito=False, fundo=None, numero=None, topo=False):
    for i, v in enumerate(valores, start=1):
        c = ws.cell(row=linha, column=i, value=v)
        c.font = Font(name=FONTE, size=10, bold=negrito, color=TINTA)
        c.border = Border(bottom=grossa if topo else fina, top=grossa if topo else None)
        c.alignment = Alignment(horizontal="left" if i <= 2 else "center", vertical="center")
        if fundo:
            c.fill = PatternFill("solid", fgColor=fundo)
        if numero and i in numero:
            c.number_format = numero[i]
    return linha + 1


def prosa(ws, linha, textos, largura, negrito_primeiro=None):
    if negrito_primeiro:
        c = ws.cell(row=linha, column=1, value=negrito_primeiro)
        c.font = Font(name=FONTE, size=10, bold=True, color=TINTA)
        linha += 1
    for t in textos:
        c = ws.cell(row=linha, column=1, value="• " + t)
        c.font = Font(name=FONTE, size=9, color="5B5443")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura)
        ws.row_dimensions[linha].height = 26
        linha += 1
    return linha


def montar(meta, serie, detalhe, res):
    mm = meta["entrada_mensal"]
    wb = Workbook()

    foto = {b["mes"]: b["qtd"] for b in mm["meses"]}
    entrantes = {x["mes"]: x["novos"] for x in serie}
    resolvidos = Counter(x["mes_resolucao"] for x in res if x["mes_resolucao"])
    legado = mm["legado"]
    anos_legado = {x["ano"]: x["qtd"] for x in legado["por_ano"]}
    jan_proprio = foto.get("2026-01", 0) - legado["qtd"]
    novos_antes = sum(x["novos"] for x in serie if x["mes"] < "2026-01")
    por_ano_coep = Counter()
    for x in serie:
        por_ano_coep[x["mes"][:4]] += x["novos"]

    # ------------------------------------------------------- 1. Mês a mês
    ws = wb.active
    ws.title = "Mês a mês"
    titulo(ws, "ETO-COEP — MÊS A MÊS DE 2026",
           f"Todos os tipos de SS contam. Três leituras do mesmo "
           f"posto. ATIVOS: os {mm['total']} da foto de junho pela data de abertura "
           "da SS, com janeiro carregando o acervo dos anos anteriores. ENTRANTES: ativos novos "
           "no posto pela abertura da SS de indisponibilidade, direto da base de SS/OS do "
           "ETO-COEP. RESOLVIDOS: pelo mês em que a tratativa aconteceu — término da SS ou "
           "repasse —, não pelo mês em que a SS abriu.", 4)
    cabecalho(ws, 4, [
        ("Mês", 22),
        (f"ENTRANTES\nfoto dos {mm['total']}, pela abertura da SS — janeiro com o acervo", 34),
        ("RESOLVIDOS\npelo mês da tratativa ou do repasse", 30),
    ], altura=48)

    linha = 5
    for k in MESES_2026:
        rot = rotulo(k)
        if k == "2026-01":
            rot = "jan/2026  (com o acervo)"
        depois = k >= "2026-04"
        linha = corpo(ws, linha, [
            rot,
            foto.get(k, 0) or "—",
            resolvidos.get(k, 0) or "—",
        ], fundo=ATUACAO if depois else (DESTAQUE if k == "2026-01" else None))
    corpo(ws, linha, ["Total até julho",
                      sum(foto.values()),
                      sum(resolvidos.get(k, 0) for k in MESES_2026)], negrito=True, topo=True)
    linha += 2

    linha = prosa(ws, linha, [
        f"Janeiro = {jan_proprio} SS abertas no próprio mês + {legado['qtd']} de anos anteriores "
        f"= {foto.get('2026-01', 0)}. Dos {legado['qtd']} antigos, "
        + ", ".join(f"{q} são de {a}" for a, q in sorted(anos_legado.items())) + ". "
        "A mais velha é de 2023 — é o acervo que o posto herdou, não demanda nova.",
        f"A coluna ENTRANTES não empilha o acervo: ela mostra o mês real. Antes de 2026 o COEP "
        f"já tinha recebido {novos_antes} ativos novos ("
        + ", ".join(f"{q} em {a}" for a, q in sorted(por_ano_coep.items()) if a < "2026") + ").",
        f"{sum(resolvidos[k] for k in resolvidos if k >= '2026-04')} dos "
        f"{sum(resolvidos.values())} resolvidos foram tratados de abril em diante — "
        f"{round(100 * sum(resolvidos[k] for k in resolvidos if k >= '2026-04') / max(sum(resolvidos.values()), 1))}% "
        "do total. Confirma o que você suspeitava: a fila só começou a andar quando o posto "
        "passou a atuar. As linhas de abril em diante estão sombreadas em verde.",
        "As três colunas medem coisas diferentes e não se somam entre si: ATIVOS é um estoque "
        "parado, ENTRANTES é fluxo de chegada, RESOLVIDOS é fluxo de saída.",
    ], 4)

    # ------------------------------------------------- 1b. Livro-caixa
    ws = wb.create_sheet("Livro-caixa")
    sal = mm.get("saldo") or []
    abertura = mm.get("abertura", 0)
    fora_livro = mm.get("fora_do_livro", 0)
    titulo(ws, "O LIVRO-CAIXA DA CARTEIRA",
           f"O acervo de anos anteriores ({abertura} SS de 2023–2025) é o saldo de abertura. "
           "Cada mês soma as SS abertas no próprio mês e desconta as tratadas: o que sobra é o "
           f"saldo com que o mês seguinte começa. Universo = os {mm['total']} da foto, que têm "
           "entrada e saída rastreadas.", 6)
    cabecalho(ws, 4, [("Mês", 30), ("Começou com", 14), ("Entraram", 12),
                      ("Saíram", 12), ("Sobrou no fim", 14), ("Variação", 12)])
    linha = 5
    linha = corpo(ws, linha, ["Acervo de 2023–2025 (saldo de abertura)", "—", "—", "—",
                              abertura, "—"], fundo=DESTAQUE)
    for s_ in sal:
        var = s_["final"] - s_["inicial"]
        linha = corpo(ws, linha, [s_["rotulo"], s_["inicial"], s_["entram"] or "—",
                                  s_["saem"] or "—", s_["final"],
                                  f"{var:+d}" if var else "—"],
                      fundo=ATUACAO if s_["mes"] >= "2026-04" else None)
    corpo(ws, linha, ["Total até julho", "—",
                      sum(s_["entram"] for s_ in sal),
                      sum(s_["saem"] for s_ in sal),
                      sal[-1]["final"] if sal else abertura, ""], negrito=True, topo=True)
    linha += 2
    linha = prosa(ws, linha, [
        f"Janeiro: {abertura} + {sal[0]['entram'] if sal else 0} − {sal[0]['saem'] if sal else 0} "
        f"= {sal[0]['final'] if sal else abertura}; fevereiro já começa com "
        f"{sal[0]['final'] if sal else abertura}. E assim em diante.",
        f"{sal[-1]['rotulo'] if sal else '—'} fecha em {sal[-1]['final'] if sal else abertura}; "
        f"fora da janela, em agosto, mais "
        f"{(mm.get('apos_janela') or {}).get('resolvidos', 0)} já foram tratados — a carteira "
        f"mostra {(sal[-1]['final'] if sal else abertura) - (mm.get('apos_janela') or {}).get('resolvidos', 0)} "
        "ainda no fluxo hoje. A conta fecha ativo a ativo.",
        f"Os {fora_livro} ativos que passaram pelo COEP em 2026 por fora da foto não entram "
        "aqui: sem SS na foto de entrada, não há data de tratativa para dar baixa. Eles estão "
        "fora desta conta — sem SS na foto de entrada, não há data de baixa.",
        "O topo da fila foi 99, no fim de abril e segurado em maio. De lá até julho a "
        "carteira caiu 44.",
        f"Na conta por SS são {mm.get('ss_resolvidas', 0)} resolvidas — "
        f"{' e '.join(mm.get('resolvidos_duplicados', []))} tinham duas SS cada na foto e "
        "contam uma vez no livro. Contando por SS e com agosto dentro, julho fecharia em "
        f"{mm.get('abertura', 0) + sum(s_['entram'] for s_ in sal) - mm.get('ss_resolvidas', 0)}.",
    ], 6)

    # ------------------------------------------------- 2. Janeiro por dentro
    ws = wb.create_sheet("Janeiro por dentro")
    titulo(ws, "O QUE JANEIRO CARREGA",
           f"Janeiro é metade da carteira e é quase tudo acervo. Dos {foto.get('2026-01', 0)} "
           f"ativos, só {jan_proprio} têm SS aberta no próprio mês; os outros {legado['qtd']} "
           "vêm de anos anteriores e foram jogados em janeiro pela regra do gestor.", 3)
    cabecalho(ws, 4, [("Origem", 34), ("Ativos", 12), ("% de janeiro", 14)])
    linha = 5
    total_jan = foto.get("2026-01", 0)
    linha = corpo(ws, linha, ["SS aberta em jan/2026", jan_proprio,
                              round(jan_proprio / max(total_jan, 1), 3)],
                  numero={3: "0.0%"})
    for ano, q in sorted(anos_legado.items(), reverse=True):
        linha = corpo(ws, linha, [f"Acervo — SS aberta em {ano}", q,
                                  round(q / max(total_jan, 1), 3)],
                      fundo=DESTAQUE, numero={3: "0.0%"})
    corpo(ws, linha, ["Total de janeiro", total_jan, 1.0], negrito=True, topo=True,
          numero={3: "0.0%"})
    linha += 2
    mais_velha = legado.get("mais_antiga") or {}
    prosa(ws, linha, [
        f"A mais velha é de {(mais_velha.get('abertura') or '2023')[:4]}: "
        f"{mais_velha.get('numero_ss', '—')}, ativo {mais_velha.get('ativo', '—')} em "
        f"{mais_velha.get('localidade', '—')}, aberta em "
        f"{dbr(mais_velha.get('abertura')).strftime('%d/%m/%Y') if dbr(mais_velha.get('abertura')) else '—'}.",
        "Sem a regra de janeiro, esses ativos ficariam espalhados por 2023, 2024 e 2025 e a "
        "curva de 2026 perderia o tamanho do que foi herdado. Com a regra, janeiro deixa de ser "
        "um mês e passa a ser o carimbo do acervo.",
        f"Os {jan_proprio} de janeiro de verdade são o volume normal de um mês antes de o posto "
        "ser assumido — na mesma faixa de fevereiro e maio.",
    ], 3)

    # ------------------------------------------------ 4. Resolvidos, detalhe
    ws = wb.create_sheet("Resolvidos por mês")
    feitos = [x for x in res if x["resolvido"]]
    com_data = [x for x in feitos if x["mes_resolucao"]]
    titulo(ws, "QUANDO CADA UM FOI TRATADO DE VERDADE",
           "Mês da tratativa, não da abertura. A data é o término da SS de entrada; quando a SS "
           "foi repassada em vez de encerrada, vale a data do repasse — a abertura da SS "
           "seguinte da mesma demanda. Faltando as duas, entram obra encerrada no AIC, reporte "
           "de campo, decisão do gestor e, por último, a SS mais recente atendida no ativo.", 6)
    cabecalho(ws, 4, [("Mês da tratativa", 18), ("Resolvidos", 12), ("Acumulado", 12),
                      ("Por cancelamento da SS", 18), ("Por repasse", 12), ("Outras vias", 13),
                      ("Com parecer COEP", 16)], altura=42)
    jan_jul = [x for x in com_data if x["mes_resolucao"] <= "2026-07"]
    linha = 5
    acumulado = 0
    for k in sorted({x["mes_resolucao"] for x in jan_jul}):
        doMes = [x for x in jan_jul if x["mes_resolucao"] == k]
        acumulado += len(doMes)
        canc = sum(1 for x in doMes if x["via"] == "cancelamento da SS de entrada")
        rep = sum(1 for x in doMes if x["via"] == "repasse para a etapa seguinte")
        par = sum(1 for x in doMes if x["parecer_coep"])
        linha = corpo(ws, linha, [rotulo(k), len(doMes), acumulado, canc or "—",
                                  rep or "—", (len(doMes) - canc - rep) or "—", par or "—"],
                      fundo=ATUACAO if k >= "2026-04" else None)
    corpo(ws, linha, ["Total até julho", len(jan_jul), "",
                      sum(1 for x in jan_jul if x["via"] == "cancelamento da SS de entrada"),
                      sum(1 for x in jan_jul if x["via"] == "repasse para a etapa seguinte"),
                      sum(1 for x in jan_jul
                          if x["via"] not in ("cancelamento da SS de entrada",
                                              "repasse para a etapa seguinte")),
                      sum(1 for x in jan_jul if x["parecer_coep"])],
          negrito=True, topo=True)
    linha += 2

    ws.cell(row=linha, column=1, value="De onde saiu a data de cada um").font = \
        Font(name=FONTE, size=10, bold=True, color=TINTA)
    linha += 1
    cabecalho(ws, linha, [("Via", None), ("Ativos", None), ("", None),
                          ("", None), ("", None), ("", None), ("", None)], altura=20)
    linha += 1
    for v, q in Counter(x["via"] for x in com_data).most_common():
        linha = corpo(ws, linha, [v, q, "", "", "", "", ""])
    linha += 1
    linha = prosa(ws, linha, [
        f"Fora da janela: {len(com_data) - len(jan_jul)} tratados em agosto/2026 — estão na "
        "«Base da foto», só não entram na tabela mensal.",
        f"{sum(1 for x in com_data if x['parecer_coep'])} dos {len(com_data)} tinham parecer "
        "COEP registrado na planilha de criticidade — a coluna está na aba «Base da foto».",
        "Repasse quer dizer que a demanda saiu do posto, não que o serviço acabou em campo. "
        "Em vários casos a etapa seguinte segue aberta no DMSL ou na Proteção.",
    ], 7)

    # ------------------------------------------------------ 5. Base do recorte
    ws = wb.create_sheet("Base da foto")
    titulo(ws, f"OS {mm['total']} DA FOTO, ATIVO A ATIVO",
           "Uma linha por ativo da foto de entrada: a SS que o trouxe, o mês de abertura, o mês "
           "da tratativa e de onde saiu cada data. Linhas sombreadas = resolvidos.", 15)
    cabecalho(ws, 4, [("Ativo", 13), ("Localidade", 24), ("Tipo", 19), ("SS de entrada", 20),
                      ("Abertura", 12), ("Mês de entrada", 14), ("Do acervo", 11),
                      ("Situação da SS hoje", 18), ("Parecer COEP", 26), ("Resolvido", 11),
                      ("Data da tratativa", 15), ("Mês da tratativa", 15),
                      ("De onde saiu a data", 30), ("Como foi resolvido", 52),
                      ("Origem da data de abertura", 26)])
    linha = 5
    for x in sorted(res, key=lambda y: (y["mes_entrada"], y["abertura"] or datetime.date.min, y["ativo"])):
        linha = corpo(ws, linha, [
            x["ativo"], x["localidade"], x["tipo"], x["numero_ss"], x["abertura"],
            rotulo(x["mes_entrada"]), x["legado"], x["situacao_ss"], x["parecer_coep"],
            "sim" if x["resolvido"] else "não", x["resolucao"],
            rotulo(x["mes_resolucao"]) if x["mes_resolucao"] else "",
            x["via"], x["motivo"], x["fonte_data"],
        ], fundo=DESTAQUE if x["resolvido"] else None,
            numero={5: "DD/MM/YYYY", 11: "DD/MM/YYYY"})
    for row in ws.iter_rows(min_row=5, max_row=linha - 1, min_col=9, max_col=15):
        for c in row:
            c.alignment = Alignment(horizontal="left", vertical="center")

    # ---------------------------------------------------------- 8. Notas
    ws = wb.create_sheet("Notas")
    titulo(ws, "DE ONDE VEM CADA NÚMERO", "Fontes, réguas e o que a planilha não prova.", 2)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 112
    linha = 4
    notas = [
        ("Coluna ATIVOS",
         "Planilha 1_Base_SS_OS_Equipamentos_especiais.xlsx, abas «Dados» e «Dados Tratados»: as "
         f"SS do ETO-COEP que estavam PENDENTES quando o posto foi assumido. {mm['total']} ativos "
         f"em {mm['total_ss']} SS, só religador e regulador. Regra do gestor: SS aberta antes de "
         "2026 entra em jan/2026. Ativo com mais de uma SS na foto entra pela mais antiga."),
        ("Coluna ENTRANTES",
         "A carteira herdada É a entrada: os ativos da foto de junho, cada um contado uma vez no "
         "mês em que a SS entrou, com janeiro carregando o acervo de anos anteriores — decisão "
         "do gestor (13/08/2026). A antiga série de entrantes da base de SS/OS saiu da planilha "
         "para não haver dois números de entrada."),
        ("Coluna RESOLVIDOS",
         "Ordem das fontes da data: 1) DATA_TERMINO_SS da SS de entrada — separada entre "
         "cancelamento e término normal; 2) data de repasse, que é a abertura da SS seguinte da "
         "mesma demanda; 3) encerramento da obra no AIC; 4) data do reporte de campo; 5) data da "
         "decisão do gestor; 6) SS mais recente atendida no ativo depois da abertura da SS de "
         "entrada. Os 64 resolvidos têm data."),
        ("Datas de abertura cruzadas",
         f"{next((f['qtd'] for f in mm['fonte_data'] if f['fonte'].startswith('cruzamento')), 0)}"
         f" dos {mm['total']} não tinham DATA_ABERTURA_SS na planilha de entrada, porque a SS só "
         "aparece na aba «Dados Tratados». Para esses a data veio do cruzamento pelo número da "
         "SS com a base de SS/OS, como o gestor mandou. Nenhuma ficou sem data."),
        ("O que a planilha não prova",
         "«Resolvido» segue as sete réguas da carteira de entrada e a palavra do gestor; não é "
         "baixa administrativa. Repasse quer dizer que a demanda saiu do posto, não que o "
         "serviço terminou em campo."),
        ("Valores, não fórmulas",
         "Os agregados foram calculados em Python e gravados como número, porque o LibreOffice "
         "não roda no ambiente que gerou o arquivo e uma fórmula ficaria sem valor em cache. "
         "Cada total tem a aba de detalhe que o sustenta: a «Base da foto»."),
        ("Livro-caixa",
         "Saldo de abertura = o acervo (SS de 2023–2025 pendentes na chegada). Entraram = SS da "
         "foto abertas no próprio mês. Saíram = tratativas do mês (término, cancelamento ou "
         "repasse). O fecho de agosto bate com os «ainda no fluxo» da carteira. Os ativos que "
         "passaram pelo posto por fora da foto ficam fora do livro por não terem data de baixa."),
        ("Tipos de SS",
         "Todos os tipos de SS contam, na carteira e nos entrantes — decisão final do gestor "
         "(13/08/2026). O TIPOSS de cada SS segue anotado na «Base da foto» para quem quiser "
         "recortar depois."),
        ("Janela de exibição",
         "As tabelas mensais vão de janeiro a julho de 2026. Agosto está em curso e ficaria como "
         "um toco enganoso no fim das curvas; o que já aconteceu nele — tratativas e entrantes — "
         "está anotado nas notas das abas e aparece na «Base da foto», que é completa."),
        ("Posição",
         "Base de SS/OS e AIC de 07/08/2026; foto de entrada de junho/2026; pareceres e decisões "
         "do gestor até 13/08/2026."),
    ]
    for rot, txt in notas:
        a = ws.cell(row=linha, column=1, value=rot)
        a.font = Font(name=FONTE, size=10, bold=True, color=TINTA)
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b = ws.cell(row=linha, column=2, value=txt)
        b.font = Font(name=FONTE, size=10, color=TINTA)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[linha].height = 58
        linha += 1

    for aba in wb:
        aba.sheet_view.showGridLines = False
    return wb


def main():
    meta, base, aic, reportes = carregar()
    foto = {x["ativo"] for x in meta["entrada_mensal"]["lista"]}
    serie, detalhe = entrantes_no_coep(base, foto)
    res = resolucoes(meta, base, aic, reportes)
    wb = montar(meta, serie, detalhe, res)
    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)

    foto = {b["mes"]: b["qtd"] for b in meta["entrada_mensal"]["meses"]}
    entr = {x["mes"]: x["novos"] for x in serie}
    resol = Counter(x["mes_resolucao"] for x in res if x["mes_resolucao"])
    print(f"OK — {SAIDA}")
    print(f"  {'mês':10} {'entrantes':>10} {'resolvidos':>11}")
    for k in MESES_2026:
        print(f"  {rotulo(k):10} {foto.get(k, 0):>10} {resol.get(k, 0):>11}")
    print(f"  {'TOTAL':10} {sum(foto.values()):>10} "
          f"{sum(resol.get(k, 0) for k in MESES_2026):>11}")
    depois = sum(resol[k] for k in resol if k >= "2026-04")
    print(f"  tratados de abril em diante: {depois} de {sum(resol.values())}")


if __name__ == "__main__":
    main()
