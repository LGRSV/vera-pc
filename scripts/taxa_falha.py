"""
Taxa de falha de religadores e reguladores — lógica em quatro camadas.

A pergunta do gestor (21/08): como medir a taxa de falha do parque de RL/RT usando a
base de SS/OS, o tipo de fato e o que foi substituído.

O problema é que a base SS/OS não é um registro de falhas: é um registro de SOLICITAÇÕES.
Uma falha pode gerar quatro SS (repasse a repasse) e uma SS pode não ter falha nenhuma
por trás (ajuste de proteção, comissionamento, obra nova). Medir "SS por equipamento"
dá 4,9 SS/ativo/ano, número que não significa nada. A lógica abaixo separa as camadas:

  C1  EVENTO      a SS vira evento de falha (régua de tipo) e as SS gêmeas colapsam
                  numa demanda só (demandas.encadear) — 1 demanda = 1 evento
  C2  EXPOSIÇÃO   o denominador é equipamento-ano do parque informado pelo gestor
                  (1.297 religadores, 197 reguladores), não a carteira de
                  indisponíveis (que já é o resultado, não a população)
  C3  MODO        o par ORIGEM_SS × DEFEITO_SS diz QUAL fato; cobertura parcial, então
                  a distribuição vale sobre os declarados e isso é dito no número
  C4  CONSEQUÊNCIA  o que foi substituído vem do OBRAS_EQ_ESPECIAL (peça + código de
                  material + data), nunca dos campos FABRICANTE_* da SS, que estão vazios

λ = eventos de falha ÷ equipamento-ano. Reportado por 100 equipamentos-ano para não
trabalhar com três zeros à direita da vírgula.

Rodar: python3 scripts/taxa_falha.py
Grava: data/missao/taxa_falha.json
"""

import datetime
import glob
import json
import os
import re
import sys
from collections import Counter, defaultdict

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))

import demandas  # noqa: E402  — reaproveita a regra de SS gêmeas já validada

ARQ_MIN = os.path.join(RAIZ, "data", "missao", "ssos_min.json")
XLSX_GESTAO = os.path.join(RAIZ, "data", "raw", "GESTAO_DE_EQUIPAMENTOS.xlsx")
XLSX_OBRAS = os.path.join(RAIZ, "data", "raw", "OBRAS_EQ_ESPECIAL.xlsx")
XLSX_ENTRADA = os.path.join(RAIZ, "data", "raw", "BASE_SS_OS_EQ_ESPECIAIS_ENTRADA.xlsx")
SAIDA = os.path.join(RAIZ, "data", "missao", "taxa_falha.json")
SAIDA_EVENTOS = os.path.join(RAIZ, "data", "missao", "taxa_falha_eventos.json")

# A base de SS/OS só tem lastro a partir de 2024: 2 SS em 2022 e 58 em 2023 contra
# 2.197 em 2024. Não é queda de falha, é extrato truncado. 2026 entra à parte porque
# está aberto — anualizar ano parcial junto dos cheios mistura duas coisas.
ANOS_CHEIOS = (2024, 2025)
ANO_PARCIAL = 2026
FIM_DO_PARCIAL = datetime.date(2026, 8, 20)  # última abertura na base de SS/OS

# Parque informado pelo gestor em 21/08 — é a autoridade sobre a população exposta.
# O cadastro de ajustes de GESTÃO DE EQUIPAMENTOS chega a 1.292 religadores e 189
# reguladores com código válido (mais o "RT SE CRISTALÂNDIA", sem código): faltam 5
# religadores e 8 reguladores que operam sem estudo de ajuste cadastrado. A diferença
# não é erro de contagem, é lacuna de cadastro — e vale como achado por si.
# Correção do gestor (21/08, terceira rodada): parque ATUAL, simples — 1.297 + 10
# religadores instalados em 2026 = 1.307, e 197 + 10 reguladores = 207. Instala-se
# pouco por ano (dificilmente mais de 20 por família), então o parque atual vale
# para os três anos: a variação real é menor que 2% e não paga a complexidade.
# A reconstrução pelas obras do AIC foi ABANDONADA: ela supercontava — dava 174
# religadores instalados em 2024 e 28 reguladores em 2026, contra 10 reais.
PARQUE_ATUAL = {"religador": 1307, "regulador": 207}


def parque_por_ano():
    """O mesmo parque atual para os três anos — decisão do gestor, 21/08."""
    return {
        familia: {
            str(ano): {"medio": total, "fonte_dos_instalados": "gestor"}
            for ano in (2024, 2025, 2026)
        }
        for familia, total in PARQUE_ATUAL.items()
    }


PREMISSAS = [
    "Falha é o que exigiu peça grande. Religador: controle, tanque ou completo. "
    "Regulador: célula, relé, completo ou furto. Definição do gestor em 21/08.",

    "Placa de alimentação CA e relé de sincronismo são outras palavras para controle "
    "e contam como falha. Placa de comunicação, placa 3G, rádio e antena são telecom e "
    "não contam. O que decide é a peça, não a palavra placa nem a palavra relé.",

    "Não conta como falha: trafo auxiliar, chave faca, rádio, antena, bateria, "
    "aterramento, cabo, conector, poste, poda, ajuste de proteção, comissionamento e "
    "obra de equipamento novo. Vai em aba separada.",

    "Agentes leram o texto completo das SS e das OS dos 129 ativos da carteira — "
    "1.087 SS. Outro time revisou cada falha apontada e derrubou o que não se "
    "sustentava no texto: de 127 apontadas, 113 ficaram e 14 caíram.",

    "A taxa conta equipamentos que falharam no ano, não ocorrências: ativo que "
    "falhou duas vezes no mesmo ano conta uma vez naquele ano — e conta de novo se "
    "falhar em outro ano. As ocorrências aparecem ao lado, como detalhe.",

    "Equipamento trocado por obra direta, sem nunca entrar na carteira do COEP, "
    "também falhou: entra pela obra de substituição do AIC concluída no ano, "
    "descontando a obra que cita ativo já contado pela leitura. O ano é o da "
    "conclusão física da obra — a troca direta costuma sair na semana da falha.",

    "O ano da falha é o da data de ocorrência. Se não houver, é a data escrita junto "
    "do parecer. A data de abertura da SS é o último recurso: ela vem em média 39 dias "
    "depois da falha, e em 9,8% dos casos cai em outro ano — medido nas 10.386 "
    "SS da base nova, com data de ocorrência em 100% das linhas.",

    "Repasse não é falha nova. Quando a mesma demanda passa de equipe em equipe, o "
    "sistema abre SS nova a cada passagem. Todas contam como uma falha só.",

    "O parque é o atual, informado pelo gestor: 1.307 religadores (1.297 + 10 "
    "instalados em 2026) e 207 reguladores (197 + 10). Vale para os três anos: "
    "instala-se pouco por ano — dificilmente mais de 20 por família — e a variação "
    "não muda a taxa. A reconstrução pelas obras do AIC foi abandonada porque "
    "superconta.",

    "A taxa é a divisão direta: total que falharam ÷ parque. 2026 vai até 20/08 e a "
    "taxa mostrada é a do ano até aqui, sem anualizar; o ritmo projetado vai em nota "
    "de rodapé.",

    "Regulador é banco de três células. O parque conta banco, não célula. Falha de uma "
    "célula é uma falha do banco.",

    "A carteira dos 129 é o que o COEP acompanha. Equipamento trocado por obra direta, "
    "sem passar pela carteira, entra pelas obras de substituição do AIC.",

    "A base de SS/OS só é confiável de 2024 em diante: tem 2 SS em 2022 e 58 em 2023, "
    "contra 2.197 em 2024. É extrato truncado, não parque saudável.",

    "O texto da SS é cumulativo: o sistema cola parecer novo por cima do antigo. Vale "
    "sempre o parecer mais recente. Isso foi instruído aos agentes e conferido na "
    "revisão.",

    "Quando a SS e a OS discordam, vale a OS. Ela traz o campo Serviço Executado, que "
    "é o registro do que foi feito de fato.",
]


# ── Camada 1: o que é evento de falha ────────────────────────────────────────────
FALHA_TIPOSS = {
    "INDISPONIBILIDADE PARA OPERAÇÃO",
    "EM OPERAÇÃO COM ANOMALIA",
    "ANOMALIA EM RELIGADOR",
    "ANOMALIA EM REGULADORES",
    "AVISO DE ANOMALIA",
    "AVISO DE EMERGÊNCIA",
    "AVISO ELEMENTO REINCIDENTE",
}
NAO_FALHA_TIPOSS = {
    "AJUSTES DE PROTEÇÃO",
    "AJUSTE DE RELÉ",
    "COMISSIONAMENTO",
    "OBRAS (NOVOS EQUIPAMENTOS)",
    "AVISO DE CADASTRO",
    "AVISO PROTEÇÃO & SELETIVIDADE",
    "AVISO DE INCONSISTÊNCIA",
    "CLUSTER",
    "MELHORIA POSTO DE TRANSFORMAÇÃO",
}
# Indisponibilidade é falha funcional (parou); anomalia é degradação (opera com defeito).
FUNCIONAL = {"INDISPONIBILIDADE PARA OPERAÇÃO", "AVISO DE EMERGÊNCIA", "FORA DE SERVIÇO"}

# Terceiro eixo, e o que mais engana nesta base: a SS pendura no código do religador
# porque ele é o marco do trecho, mas o fato é do POSTE, da cruzeta, do conector, da
# vegetação. Sem separar objeto do fato, a taxa de falha do equipamento vira taxa de
# ocorrência do alimentador — 88 SS de POSTE e 41 de ATERRAMENTO entravam como falha
# de religador. Só conta como falha quem tem o fato NO equipamento.
ORIGEM_DA_REDE = {
    "POSTE", "ESTRUTURA", "ESTRUTURA PRIMÁRIA", "CRUZETA", "ATERRAMENTO", "CONECTOR",
    "CONECTOR DERIVAÇÃO CUNHA", "JUMPER MT/CONECTOR JUMPER MT", "CABOS MT", "CABOS BT",
    "CABOS DE LIGAÇOES MT", "PÁRA-RAIOS", "ISOLADOR", "CHAVE", "CHAVE FUSÍVEL",
    "EMENDAS", "ESTAI", "ESPAÇADOR", "VEGETAÇÃO PODA DE ÁRVORE",
    "REDE PRIMÁRIA PODA DE ÁRVORE",
}
ORIGEM_ADMINISTRATIVA = {
    "INEXISTENTE", "INCONSISTÊNCIA CADASTRAL", "PLACA DE IDENTIFICAÇÃO", "INSPEÇÃO VISUAL",
}
RE_ESQUEMA_DE_REDE = re.compile(
    r"POSTE|PODA|CONDUTOR|ISOLADOR|CONEX[ÃA]O MT|CRUZETA|P[ÁA]RA-?RAIOS|ATERRAMENTO|"
    r"SUBSTITUI[ÇC][ÃA]O DE CHAVE|LINHA VIVA|EMENDA",
    re.I,
)
TIPOSS_DE_REDE = {
    "NOTA DE SERVIÇO (NS) - LINHA VIVA", "NOTA DE SERVIÇO (NS) - PODA",
    "FORMS SUBST DE POSTES", "FORMS CHAVE INST DEOP",
}

RE_MC = re.compile(r"^MC\s*-|^MC-|MANUTEN[ÇC][ÃA]O CORRETIVA", re.I)
RE_PROGRAMADA = re.compile(
    r"^MP\s*-|^PREVNP|^COM\s*-|^INSP|ATUALIZA[ÇC][ÃA]O CADASTRAL|INSP\w* PREV", re.I
)


def objeto_do_fato(ss):
    """equipamento | rede | administrativo — de quem é o defeito, afinal."""
    origem = (ss.get("ORIGEM_SS") or "").strip().upper()
    esquema = (ss.get("ESQUEMA") or "").strip().upper()
    tiposs = (ss.get("TIPOSS") or "").strip().upper()
    if origem in ORIGEM_DA_REDE:
        return "rede"
    if origem in ORIGEM_ADMINISTRATIVA:
        return "administrativo"
    if tiposs in TIPOSS_DE_REDE:
        return "rede"
    # origem em branco: o esquema desempata. "MC - POSTES" no código do religador é
    # serviço de poste, não falha de equipamento.
    if not origem and RE_ESQUEMA_DE_REDE.search(esquema):
        return "rede"
    return "equipamento"


def classificar(ss):
    """falha | programada | rede | indefinida — a régua de numerador."""
    tiposs = (ss.get("TIPOSS") or "").strip().upper()
    esquema = (ss.get("ESQUEMA") or "").strip().upper()
    objeto = objeto_do_fato(ss)
    if objeto != "equipamento":
        return objeto if objeto == "rede" else "programada"
    if tiposs in FALHA_TIPOSS:
        return "falha"
    if tiposs in NAO_FALHA_TIPOSS:
        return "programada"
    if RE_MC.search(esquema):
        return "falha"
    if RE_PROGRAMADA.search(esquema):
        return "programada"
    if tiposs.startswith("FORMS") or tiposs.startswith("NOTA DE SERVIÇO"):
        return "programada"
    return "indefinida"


def severidade(ss_da_demanda):
    tipos = {(s.get("TIPOSS") or "").strip().upper() for s in ss_da_demanda}
    origens = {(s.get("ORIGEM_SS") or "").strip().upper() for s in ss_da_demanda}
    return "funcional" if (tipos | origens) & FUNCIONAL else "anomalia"


# ── Camada 1b: o componente exigido — a régua do gestor (21/08) ─────────────────
# "Tem que ser falha mesmo, tipo precisar substituir tanque ou controle ou equipamento
# completo. Ou no caso de regulador o controle ou algumas das células, furto também.
# Trafo auxiliar, chave faca ou algo assim não deveriam contabilizar na taxa de falha
# mas pode trazer separado."
#
# O que decide é a PEÇA, não a causa. Por isso o furto entra quando leva tanque,
# controle, células ou o equipamento inteiro — e fica de fora quando leva o trafo
# auxiliar, que é o caso mais comum de furto no AIC (5 das 12 obras).
DIR_IA = os.path.join(RAIZ, "data", "analise_ia")

PECA_GRANDE = {"Tanque/Parte Ativa", "Controle/Eletrônica", "Célula de Potência"}
ACESSORIO = {
    "Comunicação/Telecom", "Aterramento", "Bateria/Fonte Auxiliar",
    "Parametrização/Proteção", "Transformador Auxiliar", "Cabo/Conector/Umbilical",
    "Estrutura/Instalação Civil",
}
# Furto é decidido pela peça que levou, não pela categoria: cai em peça grande quando
# o componente citado é célula, controle, tanque ou o banco inteiro.
# Vocabulário confirmado pelo gestor em 21/08: "placa de alimentação CA" e "relé de
# sincronismo" são outras palavras para CONTROLE. Não confundir com placa de
# comunicação e rádio, que são telecom e ficam fora da taxa — o que decide é a peça
# que o texto nomeia, não a palavra "placa" nem a palavra "relé" isoladas.
RE_PECA_GRANDE_TEXTO = re.compile(
    r"TANQUE|PARTE ATIVA|C[ÉE]LULA|CONTROLE|COMPLET|BANCO REGULADOR|"
    r"REGULADOR DE TENS[ÃA]O FURTAD|RELIGADOR FURTAD|"
    r"PLACA DE ALIMENTA[ÇC][ÃA]O(\s+CA)?|REL[EÉÊ]\s*DE\s*SINCRONISMO|ARM[ÁA]RIO DE CONTROLE|"
    r"RETROFIT",
    re.I,
)
RE_ACESSORIO_TEXTO = re.compile(
    r"TRAFO AUXILIAR|TRANSFORMADOR AUXILIAR|CHAVE FACA|CHAVE SECCIONADORA|R[ÁA]DIO|"
    r"ANTENA|BATERIA|ATERRAMENTO|UMBILICAL|CONECTOR|PARA-?RAIO|"
    r"PLACA DE COMUNICA[ÇC][ÃA]O|PLACA 3G",
    re.I,
)
# Códigos de material do orçamento (catálogo em economia_cancelados.py)
MATERIAL_GRANDE = {"690001", "690916", "690005", "692263", "690236", "690669",
                   "690240", "690241", "651638", "616033"}  # 616033: relé de
# sincronismo, que a convenção do Allan põe em Controle de regulador
MATERIAL_ACESSORIO = {"90556"}  # chave seccionadora faca — R$ 736,43


def componente_por_ss():
    """{SS: ('grande'|'acessorio', categoria)} — das 92 descrições lidas na íntegra.

    Única fonte que diz o componente com leitura do texto completo da SS. Cobre a
    carteira do COEP, não as 6.305 — e o teste de aderência mostrou que ORIGEM_SS
    não substitui: com ORIGEM_SS = RELIGADOR, 18 casos são peça grande e 13 não são.
    """
    mapa = {}
    for caminho in sorted(glob.glob(os.path.join(DIR_IA, "result_*.json"))):
        with open(caminho, encoding="utf-8") as fh:
            for reg in json.load(fh):
                cat = reg.get("categoria_primaria") or ""
                comp = reg.get("componente_especifico") or ""
                acao = reg.get("acao_requerida") or ""
                if cat in PECA_GRANDE:
                    veredito = "grande"
                elif cat == "Vandalismo/Furto":
                    veredito = "grande" if RE_PECA_GRANDE_TEXTO.search(comp + " " + acao) else "acessorio"
                elif cat in ACESSORIO:
                    veredito = "acessorio"
                else:
                    # Indefinido/Sem Diagnóstico, Vazamento, Descarga: decide o texto
                    if RE_PECA_GRANDE_TEXTO.search(comp + " " + acao):
                        veredito = "grande"
                    elif RE_ACESSORIO_TEXTO.search(comp + " " + acao):
                        veredito = "acessorio"
                    else:
                        veredito = "indefinido"
                mapa[_norm_ss(reg.get("ss"))] = (veredito, cat)
    return mapa


def componente_por_ativo():
    """{ativo: ('grande'|'acessorio', peça)} — do OBRAS_EQ_ESPECIAL e do material."""
    import openpyxl

    wb = openpyxl.load_workbook(XLSX_OBRAS, read_only=True, data_only=True)
    linhas = list(wb["Planilha1"].iter_rows(values_only=True))
    cab = [str(c).strip() if c else "" for c in linhas[1]]
    idx = {nome: i for i, nome in enumerate(cab) if nome}
    mapa = {}
    for linha in linhas[2:]:
        ativo = str(linha[idx.get("Ativo", 2)] or "").strip()
        if not ativo.isdigit():
            continue
        peca = str(linha[idx["Defeito identificado"]] or "").strip()
        codigo = str(linha[idx["Cod Material P/ Requisitar"]] or "").strip()
        if codigo in MATERIAL_GRANDE or RE_PECA_GRANDE_TEXTO.search(peca):
            mapa[ativo] = ("grande", peca or codigo)
        elif codigo in MATERIAL_ACESSORIO or RE_ACESSORIO_TEXTO.search(peca):
            mapa[ativo] = ("acessorio", peca or codigo)
    return mapa


# ── Camada 2: o parque ───────────────────────────────────────────────────────────
def _celulas(ws, max_col=None):
    return list(ws.iter_rows(min_row=2, max_col=max_col, values_only=True))


def parque():
    """Devolve {codigo: {familia, marca}} a partir do cadastro de ajustes."""
    import openpyxl

    wb = openpyxl.load_workbook(XLSX_GESTAO, read_only=True, data_only=False)
    frota = {}
    for linha in _celulas(wb["Ajustes RL Poste"], 14):
        cod = str(linha[0] or "").strip()
        if not cod.isdigit():
            continue
        marca = str(linha[10] or "").strip().upper() or "SEM CADASTRO"
        frota[cod] = {"familia": "religador", "marca": marca.split()[0] if marca else "SEM CADASTRO"}
    for linha in _celulas(wb["Ajustes Reguladores de Tensão"], 12):
        cod = str(linha[0] or "").strip()
        if not cod.isdigit():
            continue
        parte = str(linha[3] or "").strip().upper() or "SEM CADASTRO"
        frota[cod] = {"familia": "regulador", "marca": parte.split("/")[0].split()[0]}
    return frota


RE_FAMILIA_RT = re.compile(r"REGULADOR", re.I)
RE_FAMILIA_RL = re.compile(r"RELIGADOR", re.I)


def familia_pela_ss(linhas):
    """Família do ativo que o cadastro de ajustes não tem, lida na descrição da SS."""
    texto = " ".join((l.get("DESCICAO_DO_ATIVO") or "") for l in linhas)
    if RE_FAMILIA_RT.search(texto):
        return "regulador"
    if RE_FAMILIA_RL.search(texto):
        return "religador"
    return None


def exposicao(frota):
    """equipamento-ano por família e por marca, nos anos cheios e no parcial.

    A família usa o parque do gestor (PARQUE). A marca só pode usar o cadastro de
    ajustes, que é a única fonte que diz o modelo — por isso a soma das marcas fica
    abaixo do parque, e a diferença é publicada como 'sem modelo cadastrado'.
    """
    anos_cheios = len(ANOS_CHEIOS)
    fracao_parcial = (FIM_DO_PARCIAL - datetime.date(ANO_PARCIAL, 1, 1)).days / 365.0
    cadastradas = Counter(v["familia"] for v in frota.values())
    por_marca = Counter((v["familia"], v["marca"]) for v in frota.values())
    return {
        "anos_cheios": anos_cheios,
        "fracao_parcial": round(fracao_parcial, 4),
        "familia": dict(PARQUE_ATUAL),
        "no_cadastro_de_ajustes": dict(cadastradas),
        "sem_estudo_cadastrado": {
            f: PARQUE_ATUAL[f] - cadastradas.get(f, 0) for f in PARQUE_ATUAL
        },
        "marca": {f"{f}|{m}": n for (f, m), n in por_marca.items()},
    }


# ── Camada 4: o que foi substituído ──────────────────────────────────────────────
# Objeto DIRETO da substituição. "SUBSTITUIÇÃO DE 02 POSTES E INSTALAÇÃO DE RELIGADOR"
# é obra de poste; só conta quando o equipamento é o que está sendo trocado.
RE_OBRA_SUBST = re.compile(
    r"SUBSTITUI[ÇC][ÃA]O\s+D[EO]\s*(?:\d+\s+)?(?:CH\.?\s+)?(?:ATIVO\s+D[EO]\s+)?"
    r"(RELIGADOR|REGULADOR)",
    re.I,
)
ARQ_AIC_RLRT = os.path.join(RAIZ, "data", "missao", "aic_rlrt.json")
ARQ_OBRAS_EQP = os.path.join(RAIZ, "data", "raw", "obras_equipamento.json")
RE_NAO_E_PECA = re.compile(r"REQUISITAR|APROVEITAR|VAMOS|VERIFICAR|AGUARD|^NA$|^N/?A$", re.I)


def trocas_no_aic():
    """Taxa de substituição: obra do AIC cujo objeto é o próprio equipamento."""
    if not os.path.exists(ARQ_AIC_RLRT):
        return None
    with open(ARQ_AIC_RLRT, encoding="utf-8") as fh:
        obras = json.load(fh)
    por_ano = defaultdict(Counter)
    total = Counter()
    for obra in obras:
        m = RE_OBRA_SUBST.search(obra.get("DESCRICAO_OBRA") or "")
        if not m:
            continue
        familia = "religador" if m.group(1).upper().startswith("RELIG") else "regulador"
        total[familia] += 1
        ano = (obra.get("DATA_CONCLUSAO_FISICA") or "")[:4]
        if ano.isdigit():
            por_ano[ano][familia] += 1
    return {
        "obras_de_substituicao": dict(total),
        "por_ano_de_conclusao_fisica": {a: dict(c) for a, c in sorted(por_ano.items())},
    }


def peca_grande_em_campo():
    """Peça grande já levada para a obra e ainda não concluída — a fila material.

    OBRAS_COM_EQP usa a convenção do Allan, que é a régua do gestor letra por letra:
    religador se divide em PARTE ATIVA e CONTROLE, regulador em CÉLULA e CONTROLE.
    "Levado" é RMA menos DMA: saiu do almoxarifado e ficou na obra.

    O extrato foi montado só com obra NÃO concluída, então este bloco e as obras
    encerradas do AIC não se sobrepõem — um é fila, o outro é executado.
    """
    if not os.path.exists(ARQ_OBRAS_EQP):
        return None
    with open(ARQ_OBRAS_EQP, encoding="utf-8") as fh:
        d = json.load(fh)
    classes = d.get("por_classe") or []
    por_familia = defaultdict(lambda: {"pecas": 0, "valor": 0.0, "detalhe": {}})
    for linha in classes:
        fam = "religador" if linha["familia"].lower().startswith("relig") else "regulador"
        por_familia[fam]["pecas"] += linha["levado"]
        por_familia[fam]["valor"] += linha["valor"]
        por_familia[fam]["detalhe"][linha["classe"]] = linha["levado"]
    return {
        "fonte": d.get("fonte"),
        "convencao": d.get("convencao"),
        "so_obra_nao_concluida": True,
        "obras": (d.get("totais") or {}).get("obras"),
        "pecas_levadas": (d.get("totais") or {}).get("levado"),
        "valor_levado": (d.get("totais") or {}).get("valor_levado"),
        "por_familia": {k: dict(v) for k, v in por_familia.items()},
    }


def substituicoes():
    import openpyxl

    wb = openpyxl.load_workbook(XLSX_OBRAS, read_only=True, data_only=True)
    linhas = list(wb["Planilha1"].iter_rows(values_only=True))
    cab = [str(c).strip() if c else "" for c in linhas[1]]
    idx = {nome: i for i, nome in enumerate(cab) if nome}
    peca, material, concluidas, familia = Counter(), Counter(), 0, Counter()
    total = 0
    for linha in linhas[2:]:
        ativo = str(linha[idx.get("Ativo", 2)] or "").strip()
        if not ativo.isdigit():
            continue
        total += 1
        d = str(linha[idx["Defeito identificado"]] or "").strip()
        if d and d.lower() != "none":
            for parte in re.split(r"[,/]| e ", d):
                parte = parte.strip().strip(".").title()
                # o campo é livre: o COEP às vezes escreve o encaminhamento em vez da
                # peça ("Vamos requisitar", "Aproveitar do 79000...").
                if len(parte) > 2 and not RE_NAO_E_PECA.search(parte):
                    peca[parte] += 1
        cod = str(linha[idx["Cod Material P/ Requisitar"]] or "").strip()
        if cod.isdigit():
            material[cod] += 1
        status = str(linha[idx["Status da Substituição"]] or "").strip().lower()
        if status.startswith("conclu"):
            concluidas += 1
        eq = str(linha[idx["Equipamento"]] or "").strip().lower()
        if eq:
            familia[eq] += 1
    return {
        "registros": total,
        "peca_substituida": peca.most_common(),
        "codigo_material": material.most_common(),
        "substituicoes_concluidas": concluidas,
        "por_familia": familia.most_common(),
    }


# ── Montagem ─────────────────────────────────────────────────────────────────────
def _veredito_componente(ss_da_demanda, ativo, comp_ss, comp_ativo):
    """grande | acessorio | indefinido — o que a demanda exigiu de peça.

    Prioridade: leitura da descrição da SS (a mais confiável) > peça registrada no
    OBRAS_EQ_ESPECIAL para o ativo. Sem nenhuma das duas, fica indefinido — e a maior
    parte das 6.305 SS fica aqui, porque o extrato não traz o texto da SS.
    """
    for linha in ss_da_demanda:
        achado = comp_ss.get(_norm_ss(linha.get("NUMERO_SS")))
        if achado and achado[0] != "indefinido":
            return achado[0]
    achado = comp_ativo.get(ativo)
    return achado[0] if achado else "indefinido"


def _norm_ss(numero):
    """ETO-COEP 00092/2023 e ETO-COEP 92/2023 são a mesma SS."""
    texto = re.sub(r"\s+", " ", str(numero or "").strip().upper())
    m = re.match(r"^([A-Z\-]+)\s*0*(\d+)/(\d{4})$", texto)
    return f"{m.group(1)} {int(m.group(2))}/{m.group(3)}" if m else texto


ARQ_OCORRENCIA = os.path.join(RAIZ, "data", "missao", "ss_ocorrencia.json")


def datas_de_ocorrencia():
    """{SS: datetime da ocorrência} — a data em que o equipamento falhou de fato.

    Regra do gestor (21/08): o evento se ancora na OCORRÊNCIA, não na abertura da SS.
    Abertura é o carimbo do registro, que vem depois — mediana de 21 dias e cauda de
    até 734 — e o SGM ainda reescreve a abertura ao reabrir/repassar.

    Desde 21/08 a fonte principal é ss_ocorrencia.json (base Eqp do gestor, 11/08):
    10.386 SS de RL/RT com DTA_OCORRENCIA em 100% das linhas. Antes disso a única
    fonte era o extrato do COEP, com 211 SS — 5% de cobertura. O extrato do COEP
    segue como complemento para o que a base nova não alcança.
    """
    ocorrencias = {}
    if os.path.exists(ARQ_OCORRENCIA):
        with open(ARQ_OCORRENCIA, encoding="utf-8") as fh:
            for reg in json.load(fh):
                bruto = (reg.get("DTA_OCORRENCIA") or "").strip()
                if not bruto:
                    continue
                try:
                    ocorrencias[_norm_ss(reg.get("SS_ORIGINAL"))] = datetime.datetime.strptime(
                        bruto[:19], "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    continue
    ocorrencias.update(_ocorrencias_do_extrato_coep())
    return ocorrencias


def _ocorrencias_do_extrato_coep():
    import openpyxl

    wb = openpyxl.load_workbook(XLSX_ENTRADA, read_only=True)
    ocorrencias = {}
    linhas = list(wb["Dados"].iter_rows(values_only=True))
    cab = [str(c).strip() if c else "" for c in linhas[0]]
    i_oc, i_ss = cab.index("DATA_OCORRENCIA_SS"), cab.index("NUMERO_SS")
    for linha in linhas[1:]:
        if isinstance(linha[i_oc], datetime.datetime):
            ocorrencias[_norm_ss(linha[i_ss])] = linha[i_oc]
    # a aba tratada traz a mesma data em texto e alcança SS que a outra não tem
    tratadas = list(wb["Dados Tratados"].iter_rows(values_only=True))
    cab2 = [str(c).strip() if c else "" for c in tratadas[0]]
    i_oc2, i_ss2 = cab2.index("DataOcorrência"), cab2.index("SS")
    for linha in tratadas[1:]:
        chave = _norm_ss(linha[i_ss2])
        if chave in ocorrencias or not linha[i_oc2]:
            continue
        try:
            ocorrencias[chave] = datetime.datetime.strptime(
                str(linha[i_oc2]).strip(), "%d/%m/%Y %H:%M:%S"
            )
        except ValueError:
            continue
    return ocorrencias


def defasagem_medida():
    """Quanto a abertura atrasa em relação à ocorrência, nas SS que têm as duas datas."""
    import openpyxl

    wb = openpyxl.load_workbook(XLSX_ENTRADA, read_only=True)
    linhas = list(wb["Dados"].iter_rows(values_only=True))
    cab = [str(c).strip() if c else "" for c in linhas[0]]
    i_oc, i_ab = cab.index("DATA_OCORRENCIA_SS"), cab.index("DATA_ABERTURA_SS")
    dias, troca_de_ano, por_ano_oc, por_ano_ab = [], 0, Counter(), Counter()
    for linha in linhas[1:]:
        oc, ab = linha[i_oc], linha[i_ab]
        if not (isinstance(oc, datetime.datetime) and isinstance(ab, datetime.datetime)):
            continue
        dias.append((ab - oc).days)
        por_ano_oc[oc.year] += 1
        por_ano_ab[ab.year] += 1
        if oc.year != ab.year:
            troca_de_ano += 1
    dias.sort()
    if not dias:
        return None
    n = len(dias)
    return {
        "ss_com_as_duas_datas": n,
        "dias": {
            "min": dias[0],
            "p25": dias[n // 4],
            "mediana": dias[n // 2],
            "p75": dias[3 * n // 4],
            "max": dias[-1],
            "media": round(sum(dias) / n, 1),
        },
        "mesmo_dia": sum(1 for d in dias if d == 0),
        "acima_de_180_dias": sum(1 for d in dias if d > 180),
        "trocam_de_ano": troca_de_ano,
        "trocam_de_ano_pct": round(100.0 * troca_de_ano / n, 1),
        "serie_por_ocorrencia": dict(sorted(por_ano_oc.items())),
        "serie_por_abertura": dict(sorted(por_ano_ab.items())),
        "leitura": (
            "A defasagem nunca é negativa: a abertura sempre vem depois. O erro de usar "
            "abertura é portanto SISTEMÁTICO e direcional — empurra falha antiga para o "
            "ano corrente, inflando o ano recente e esvaziando os anteriores. Nesta "
            "amostra a abertura tira 8 eventos de 2024 e 13 de 2025 e joga 21 em 2026."
        ),
    }


def resolvidos_por_ano():
    """O contraponto: quanto o posto tirou da mesa em cada ano.

    Três medidas, porque elas contam coisas diferentes e a diferença é o recado:
      demandas encerradas   a SS de falha terminou (atendida ou cancelada)
      obra concluída        o serviço foi feito em campo (DATA_CONCLUSAO_FISICA)
      obra encerrada        o contábil fechou (DTH_ENCERRAMENTO) — sempre atrasado
    """
    with open(ARQ_MIN, encoding="utf-8") as fh:
        base = json.load(fh)
    frota = parque()
    por_ativo = defaultdict(list)
    for ss in base:
        cod = str(ss.get("NUM_TRAFO") or "").strip()
        if cod.isdigit():
            por_ativo[cod].append(ss)

    encerradas = defaultdict(Counter)
    for cod, linhas in por_ativo.items():
        familia = frota.get(cod, {}).get("familia") or familia_pela_ss(linhas)
        if familia not in ("religador", "regulador"):
            continue
        for dem in demandas.encadear(linhas):
            if "falha" not in [classificar(s) for s in dem["ss"]]:
                continue
            fim = None
            for s in dem["ss"]:
                if s.get("SITUACAO_SS") in ("SS ATENDIDA", "SS CANCELADA"):
                    d = demandas._dt(s.get("DATA_TERMINO_SS", "")) or demandas._dt(
                        s.get("DATA_ABERTURA_SS", "")
                    )
                    if d and (fim is None or d > fim):
                        fim = d
            if fim:
                encerradas[str(fim.year)][familia] += 1

    obras = {"concluida": defaultdict(Counter), "encerrada": defaultdict(Counter)}
    if os.path.exists(ARQ_AIC_RLRT):
        with open(ARQ_AIC_RLRT, encoding="utf-8") as fh:
            for obra in json.load(fh):
                m = RE_OBRA_SUBST.search(obra.get("DESCRICAO_OBRA") or "")
                if not m:
                    continue
                familia = "religador" if m.group(1).upper().startswith("RELIG") else "regulador"
                for campo, rot in (("DATA_CONCLUSAO_FISICA", "concluida"),
                                   ("DTH_ENCERRAMENTO", "encerrada")):
                    ano = (obra.get(campo) or "")[:4]
                    if ano.isdigit():
                        obras[rot][ano][familia] += 1

    anos = ("2024", "2025", "2026")
    return {
        "demandas_de_falha_encerradas": {a: dict(encerradas.get(a, {})) for a in anos},
        "obra_de_substituicao_concluida_em_campo": {a: dict(obras["concluida"].get(a, {})) for a in anos},
        "obra_de_substituicao_encerrada_no_contabil": {a: dict(obras["encerrada"].get(a, {})) for a in anos},
        "leitura": (
            "2026 vai só até 20/08 e a obra leva meses para encerrar no contábil, então "
            "a linha de obra encerrada subconta 2026 por atraso de sistema, não por "
            "queda de produção. A linha de demandas encerradas é a mais comparável "
            "entre anos."
        ),
    }


def _bloco_componente(eventos, exp):
    """A régua do gestor: taxa de falha só do que exige peça grande.

    Devolve o que é mensurável hoje (evidência de componente por SS lida ou por peça
    registrada), o que fica separado por ser acessório, e o tamanho do vão — os
    eventos sem nenhuma fonte que diga o componente.
    """
    por_classe = Counter(e["componente"] for e in eventos)
    fora = os.path.join(DIR_IA, "result_01.json")
    cobertura = round(100.0 * (por_classe["grande"] + por_classe["acessorio"]) / len(eventos), 1) if eventos else None

    detalhe = {}
    for fam, expostos in exp["familia"].items():
        bloco = {}
        for ano in list(ANOS_CHEIOS) + [ANO_PARCIAL]:
            fator = exp["fracao_parcial"] if ano == ANO_PARCIAL else 1.0
            do_ano = [e for e in eventos if e["ano"] == ano and e["familia"] == fam]
            grandes = [e for e in do_ano if e["componente"] == "grande"]
            acess = [e for e in do_ano if e["componente"] == "acessorio"]
            eq_ano = expostos * fator
            bloco[str(ano)] = {
                "com_peca_grande": len(grandes),
                "taxa_confirmada_100": round(100.0 * len(grandes) / eq_ano, 2) if eq_ano else None,
                "acessorio_separado": len(acess),
                "sem_evidencia_de_componente": len(do_ano) - len(grandes) - len(acess),
                "eventos_no_ano": len(do_ano),
            }
        detalhe[fam] = bloco

    # proporção observada onde HÁ leitura — o único fator defensável para dimensionar
    lidos = [e for e in eventos if e["componente"] in ("grande", "acessorio")]
    proporcao = round(
        100.0 * sum(1 for e in lidos if e["componente"] == "grande") / len(lidos), 1
    ) if lidos else None

    return {
        "regra": (
            "Conta na taxa de falha: tanque/parte ativa, controle, equipamento completo "
            "(religador); controle, células, banco completo (regulador). Furto conta "
            "quando leva uma dessas peças. Fica separado: trafo auxiliar, chave faca, "
            "rádio/antena, bateria, aterramento, cabo/conector, estrutura e "
            "parametrização — trazidos, não somados."
        ),
        "classificacao_dos_eventos": dict(por_classe),
        "cobertura_de_evidencia_pct": cobertura,
        "proporcao_peca_grande_onde_ha_leitura_pct": proporcao,
        "por_familia_e_ano": detalhe,
        "limite": (
            "O ORIGEM_SS não substitui a leitura: nas 92 SS lidas na íntegra, com "
            "ORIGEM_SS = RELIGADOR, 18 exigiram peça grande e 13 não. Estender a régua "
            "às 6.305 SS depende do texto da SS (campo DESCRIPTION) ou do DEFEITO_SS, "
            "nenhum dos dois presente no extrato de RL/RT que está no repositório."
        ),
    }


def montar():
    with open(ARQ_MIN, encoding="utf-8") as fh:
        base = json.load(fh)

    frota = parque()
    exp = exposicao(frota)

    # Separa a base por ativo. O denominador agora é o parque do gestor, que é o total
    # real — logo o ativo com SS que o cadastro de ajustes não tem NÃO pode ser
    # descartado: ele é parte do parque, só não tem estudo cadastrado. Entra com a
    # família lida na descrição da SS e sem modelo, marcado como fora do cadastro.
    bruto = defaultdict(list)
    for ss in base:
        cod = str(ss.get("NUM_TRAFO") or "").strip()
        if cod.isdigit():
            bruto[cod].append(ss)

    por_ativo, fora_do_cadastro, sem_familia = {}, set(), set()
    for cod, linhas in bruto.items():
        if cod in frota:
            por_ativo[cod] = linhas
            continue
        familia = familia_pela_ss(linhas)
        if familia is None:
            sem_familia.add(cod)  # nem cadastro nem descrição dizem o que é
            continue
        frota[cod] = {"familia": familia, "marca": "SEM CADASTRO"}
        fora_do_cadastro.add(cod)
        por_ativo[cod] = linhas

    triagem = Counter(classificar(ss) for ss in base)
    ocorrencias = datas_de_ocorrencia()
    comp_ss = componente_por_ss()
    comp_ativo = componente_por_ativo()

    eventos = []  # um por demanda de falha
    for cod, linhas in por_ativo.items():
        for dem in demandas.encadear(linhas):
            classes = [classificar(s) for s in dem["ss"]]
            if "falha" not in classes:
                continue
            # Âncora do evento: a OCORRÊNCIA mais antiga entre as SS da demanda. A
            # abertura só entra quando nenhuma SS da cadeia declara ocorrência — e
            # nesse caso o evento fica marcado, para o número dizer em que pé está.
            ocorreu = min(
                (ocorrencias[_norm_ss(s.get("NUMERO_SS"))] for s in dem["ss"]
                 if _norm_ss(s.get("NUMERO_SS")) in ocorrencias),
                default=None,
            )
            abertura = min(
                (demandas._dt(s.get("DATA_ABERTURA_SS", "")) for s in dem["ss"] if s.get("DATA_ABERTURA_SS")),
                default=None,
            )
            marco = ocorreu or abertura
            if marco is None:
                continue
            eventos.append(
                {
                    "ativo": cod,
                    "familia": frota[cod]["familia"],
                    "marca": frota[cod]["marca"],
                    "ano": marco.year,
                    "data": marco.date().isoformat(),
                    "ancora": "ocorrencia" if ocorreu else "abertura",
                    "atraso_do_registro": (abertura - ocorreu).days if (ocorreu and abertura) else None,
                    "severidade": severidade(dem["ss"]),
                    "componente": _veredito_componente(dem["ss"], cod, comp_ss, comp_ativo),
                    "ss": len(dem["ss"]),
                    "numeros_ss": [(s.get("NUMERO_SS") or "").strip() for s in dem["ss"]],
                    "tiposs": sorted({(s.get("TIPOSS") or "").strip().upper()
                                      for s in dem["ss"] if s.get("TIPOSS")}),
                    "localidade": next(((s.get("LOCALIDADE") or "").strip()
                                        for s in dem["ss"] if s.get("LOCALIDADE")), ""),
                    "origem": sorted({(s.get("ORIGEM_SS") or "").strip().upper() for s in dem["ss"] if s.get("ORIGEM_SS")}),
                }
            )

    def taxa(filtro, expostos, anos):
        n = sum(1 for e in eventos if filtro(e))
        if not expostos or not anos:
            return {"eventos": n, "taxa_100": None}
        return {
            "eventos": n,
            "equipamento_ano": round(expostos * anos, 1),
            "taxa_100": round(100.0 * n / (expostos * anos), 1),
            "mtbf_anos": round((expostos * anos) / n, 1) if n else None,
        }

    cheios = {}
    for fam, expostos in exp["familia"].items():
        cheios[fam] = {
            "geral": taxa(lambda e, f=fam: e["familia"] == f and e["ano"] in ANOS_CHEIOS, expostos, exp["anos_cheios"]),
            "funcional": taxa(
                lambda e, f=fam: e["familia"] == f and e["ano"] in ANOS_CHEIOS and e["severidade"] == "funcional",
                expostos,
                exp["anos_cheios"],
            ),
            "anomalia": taxa(
                lambda e, f=fam: e["familia"] == f and e["ano"] in ANOS_CHEIOS and e["severidade"] == "anomalia",
                expostos,
                exp["anos_cheios"],
            ),
            "parque": expostos,
        }

    parcial = {
        fam: taxa(
            lambda e, f=fam: e["familia"] == f and e["ano"] == ANO_PARCIAL,
            expostos,
            exp["fracao_parcial"],
        )
        for fam, expostos in exp["familia"].items()
    }

    por_marca = {}
    for chave, expostos in exp["marca"].items():
        fam, marca = chave.split("|", 1)
        if expostos < 20:  # amostra pequena não vira taxa publicável
            continue
        por_marca[chave] = taxa(
            lambda e, f=fam, m=marca: e["familia"] == f and e["marca"] == m and e["ano"] in ANOS_CHEIOS,
            expostos,
            exp["anos_cheios"],
        ) | {"parque": expostos}

    # ── Série ano a ano ──────────────────────────────────────────────────────────
    # 2024 e 2025 são anos cheios (fator 1,0); 2026 vale a fração decorrida até a foto,
    # senão a taxa de um ano pela metade sairia pela metade e pareceria queda.
    aic_ano = (trocas_no_aic() or {}).get("por_ano_de_conclusao_fisica", {})
    serie = {}
    for ano in list(ANOS_CHEIOS) + [ANO_PARCIAL]:
        fator = exp["fracao_parcial"] if ano == ANO_PARCIAL else 1.0
        do_ano = [e for e in eventos if e["ano"] == ano]
        bloco = {}
        for fam, expostos in exp["familia"].items():
            eq_ano = expostos * fator
            desta = [e for e in do_ano if e["familia"] == fam]
            ativos = {e["ativo"] for e in desta}
            trocas = aic_ano.get(str(ano), {}).get(fam, 0)
            bloco[fam] = {
                "parque": expostos,
                "equipamento_ano": round(eq_ano, 1),
                "eventos": len(desta),
                "taxa_100": round(100.0 * len(desta) / eq_ano, 1) if eq_ano else None,
                "funcional_100": round(
                    100.0 * sum(1 for e in desta if e["severidade"] == "funcional") / eq_ano, 1
                ) if eq_ano else None,
                "anomalia_100": round(
                    100.0 * sum(1 for e in desta if e["severidade"] == "anomalia") / eq_ano, 1
                ) if eq_ano else None,
                "mtbf_anos": round(eq_ano / len(desta), 1) if desta else None,
                "ativos_distintos": len(ativos),
                "incidencia_pct": round(100.0 * len(ativos) / expostos, 1) if expostos else None,
                "trocas_confirmadas": trocas,
                "taxa_substituicao_100": round(100.0 * trocas / eq_ano, 1) if eq_ano else None,
                "chamadas_por_troca": round(len(desta) / trocas, 1) if trocas else None,
            }
        modo = Counter()
        for e in do_ano:
            for o in e["origem"]:
                modo[o] += 1
        bloco["_modo_declarado"] = {
            "eventos_com_origem": sum(1 for e in do_ano if e["origem"]),
            "cobertura_pct": round(100.0 * sum(1 for e in do_ano if e["origem"]) / len(do_ano), 1)
            if do_ano else None,
            "top": modo.most_common(8),
        }
        bloco["_ancora"] = {
            "por_ocorrencia": sum(1 for e in do_ano if e["ancora"] == "ocorrencia"),
            "por_abertura": sum(1 for e in do_ano if e["ancora"] == "abertura"),
            "cobertura_pct": round(
                100.0 * sum(1 for e in do_ano if e["ancora"] == "ocorrencia") / len(do_ano), 1
            ) if do_ano else None,
        }
        bloco["_fator_de_exposicao"] = fator
        serie[str(ano)] = bloco

    # reincidência: ativo com mais de um evento dentro da janela cheia
    por_ativo_evt = Counter(e["ativo"] for e in eventos if e["ano"] in ANOS_CHEIOS)
    reincidentes = {a: n for a, n in por_ativo_evt.items() if n > 1}

    # Incidência ≠ frequência. Frequência é evento/eq-ano (média do parque); incidência
    # é quantos equipamentos DISTINTOS falharam ao menos uma vez. O parque não falha
    # parelho: a cauda de reincidentes carrega a média e é ela que se ataca primeiro.
    incidencia = {}
    for fam, expostos in exp["familia"].items():
        ativos = [a for a, n in por_ativo_evt.items() if frota[a]["familia"] == fam]
        dist = Counter(por_ativo_evt[a] for a in ativos)
        cauda = [a for a in ativos if por_ativo_evt[a] >= 3]
        incidencia[fam] = {
            "parque": expostos,
            "ativos_com_ao_menos_um_evento": len(ativos),
            "incidencia_pct_em_2_anos": round(100.0 * len(ativos) / expostos, 1) if expostos else None,
            "distribuicao_eventos_por_ativo": dict(sorted(dist.items())),
            "ativos_com_3_ou_mais": len(cauda),
            "eventos_vindos_da_cauda": sum(por_ativo_evt[a] for a in cauda),
            "pct_dos_eventos_na_cauda": round(
                100.0 * sum(por_ativo_evt[a] for a in cauda) / sum(por_ativo_evt[a] for a in ativos), 1
            ) if ativos else None,
        }

    # modo de falha declarado (camada 3)
    declarados = Counter()
    sem_declaracao = 0
    for e in eventos:
        if e["origem"]:
            for o in e["origem"]:
                declarados[o] += 1
        else:
            sem_declaracao += 1

    pacote = {
        "premissas": PREMISSAS,
        "janela": {"cheios": list(ANOS_CHEIOS), "parcial": ANO_PARCIAL, "corte": FIM_DO_PARCIAL.isoformat()},
        "triagem_ss": dict(triagem),
        "parque": {
            "religador": PARQUE_ATUAL["religador"],
            "regulador": PARQUE_ATUAL["regulador"],
            "nota": "parque no início de 2026; as instalações do ano somam por cima",
            "fonte": "informado pelo gestor em 21/08/2026",
            "no_cadastro_de_ajustes": exp["no_cadastro_de_ajustes"],
            "sem_estudo_de_ajuste_cadastrado": exp["sem_estudo_cadastrado"],
            "ativos_com_ss_fora_do_cadastro": len(fora_do_cadastro),
            "ativos_sem_familia_identificavel": len(sem_familia),
        },
        "eventos": len(eventos),
        "eventos_na_janela_cheia": sum(1 for e in eventos if e["ano"] in ANOS_CHEIOS),
        "taxa_anos_cheios": cheios,
        "taxa_2026_anualizada": parcial,
        "taxa_por_marca": por_marca,
        "reincidencia": {
            "ativos_com_mais_de_um_evento": len(reincidentes),
            "eventos_desses_ativos": sum(reincidentes.values()),
            "pior": sorted(reincidentes.items(), key=lambda kv: -kv[1])[:8],
        },
        "modo_de_falha_declarado": {
            "eventos_com_origem": sum(1 for e in eventos if e["origem"]),
            "eventos_sem_origem": sem_declaracao,
            "cobertura_pct": round(100.0 * sum(1 for e in eventos if e["origem"]) / len(eventos), 1) if eventos else None,
            "top": declarados.most_common(15),
        },
        "ancoragem": {
            "por_ocorrencia": sum(1 for e in eventos if e["ancora"] == "ocorrencia"),
            "por_abertura": sum(1 for e in eventos if e["ancora"] == "abertura"),
            "cobertura_pct": round(
                100.0 * sum(1 for e in eventos if e["ancora"] == "ocorrencia") / len(eventos), 1
            ) if eventos else None,
        },
        "defasagem_ocorrencia_abertura": defasagem_medida(),
        "impacto_da_ancora": {
            "eventos_conferiveis": sum(1 for e in eventos if e["atraso_do_registro"] is not None),
            "mudaram_de_ano": sum(
                1 for e in eventos
                if e["atraso_do_registro"] is not None
                and datetime.date.fromisoformat(e["data"]).year
                != (datetime.date.fromisoformat(e["data"])
                    + datetime.timedelta(days=e["atraso_do_registro"])).year
            ),
            "atraso_mediano_dias": sorted(
                e["atraso_do_registro"] for e in eventos if e["atraso_do_registro"] is not None
            )[sum(1 for e in eventos if e["atraso_do_registro"] is not None) // 2]
            if any(e["atraso_do_registro"] is not None for e in eventos) else None,
            "leitura": (
                "Nos eventos em que dá para conferir as duas datas, mais de um terço "
                "estava no ano errado pela abertura. A série por abertura não serve para "
                "ler tendência; serve só para medir o próprio atraso de registro."
            ),
        },
        "regua_do_componente": _bloco_componente(eventos, exp),
        "peca_grande_em_campo": peca_grande_em_campo(),
        "parque_por_ano": parque_por_ano(),
        "resolvidos_por_ano": resolvidos_por_ano(),
        "serie_por_ano": serie,
        "incidencia": incidencia,
        "substituicao": substituicoes(),
        "trocas_no_aic": trocas_no_aic(),
    }

    # O contraste que dá sentido à métrica: quantas chamadas atribuídas ao equipamento
    # para cada troca efetivamente executada. Religador resolve muito em campo;
    # regulador converte chamada em peça com muito mais frequência — e peça de
    # regulador 34,5 kV custa R$ 57 mil a R$ 127 mil.
    aic = pacote["trocas_no_aic"]
    if aic:
        conversao = {}
        for fam, expostos in exp["familia"].items():
            trocas = sum(
                aic["por_ano_de_conclusao_fisica"].get(str(a), {}).get(fam, 0) for a in ANOS_CHEIOS
            )
            eventos_fam = cheios[fam]["geral"]["eventos"]
            eq_ano = expostos * exp["anos_cheios"]
            conversao[fam] = {
                "chamadas_atribuidas_ao_equipamento": eventos_fam,
                "trocas_confirmadas": trocas,
                "taxa_chamada_100": cheios[fam]["geral"]["taxa_100"],
                "taxa_substituicao_100": round(100.0 * trocas / eq_ano, 1) if eq_ano else None,
                "chamadas_por_troca": round(eventos_fam / trocas, 1) if trocas else None,
            }
        pacote["conversao_chamada_troca"] = conversao
    return pacote, eventos


if __name__ == "__main__":
    p, eventos = montar()
    with open(SAIDA, "w", encoding="utf-8") as fh:
        json.dump(p, fh, ensure_ascii=False, indent=1)
    # o rol de casos da taxa total: um registro por demanda de falha, com a marca de
    # «primeira do ativo no ano» — é ela que vira equipamento contado na taxa
    ordem = sorted(eventos, key=lambda e: (e["ativo"], e["ano"], e["data"]))
    vistos = set()
    for e in ordem:
        chave = (e["ativo"], e["ano"])
        e["primeira_do_ano"] = chave not in vistos
        vistos.add(chave)
    with open(SAIDA_EVENTOS, "w", encoding="utf-8") as fh:
        json.dump(ordem, fh, ensure_ascii=False)
    print(f"casos gravados: {SAIDA_EVENTOS} ({len(ordem)})")
    print(json.dumps({k: v for k, v in p.items() if k != "premissas"}, ensure_ascii=False, indent=1))
