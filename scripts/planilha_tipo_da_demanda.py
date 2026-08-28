"""
A planilha do TIPOSS dos que passaram pelo posto — dist/TIPO_DA_DEMANDA.xlsx.

Sai de data/missao/tipo_da_demanda.json. Cinco abas: o cruzamento tipo x situação,
o agrupamento por família, o desfecho dos encerrados, a base com um ativo por linha
para pivotar, e a régua escrita.

Rodar: python3 scripts/planilha_tipo_da_demanda.py
"""

import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAIZ = "/home/user/vera-pc"
p = json.load(open(f"{RAIZ}/data/missao/tipo_da_demanda.json", encoding="utf-8"))
wb = Workbook(); wb.remove(wb.active)

TINTA = "FF211D15"; PAPEL = "FFF2EFE6"; SINAL = "FFBC4B0E"
cab = Font(name="Calibri", bold=True, color=PAPEL, size=10)
fill_cab = PatternFill("solid", fgColor=TINTA)
neg = Font(bold=True)
fina = Side(style="thin", color="FFC8C2AF")
borda = Border(bottom=fina)

def escreve(ws, linhas, larguras, congela="A2"):
    for r in linhas:
        ws.append(r)
    for c in range(1, len(linhas[0]) + 1):
        cel = ws.cell(row=1, column=c)
        cel.font = cab; cel.fill = fill_cab
        cel.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = congela
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30

# 1 — o cruzamento
ws = wb.create_sheet("Tipo x situação")
baldes = p["baldes"]; rot = p["rotulo_balde"]
linhas = [["Tipo da SS (TIPOSS)"] + [rot[b] for b in baldes] + ["Total"]]
for t, d in p["cruzado"].items():
    linhas.append([t] + [d[b] for b in baldes] + [d["total"]])
linhas.append(["TOTAL"] + [sum(d[b] for d in p["cruzado"].values()) for b in baldes]
              + [sum(d["total"] for d in p["cruzado"].values())])
linhas.append([])
linhas.append(["Entram na conta do posto", p["resumo"]["passaram"]])
linhas.append(["Fora da conta (não é manutenção)", p["resumo"]["fora_da_conta"]])
linhas.append(["Passaram pelo posto, bruto", p["resumo"]["passaram_bruto"]])
escreve(ws, linhas, [40, 18, 24, 16, 20, 16, 9])
linha_total = len(p["cruzado"]) + 2
for c in range(1, 8):
    ws.cell(row=linha_total, column=c).font = neg
    ws.cell(row=linha_total, column=c).border = Border(top=Side(style="medium", color=TINTA))
for r in range(linha_total + 2, ws.max_row + 1):
    ws.cell(row=r, column=1).font = Font(italic=True)
ws.auto_filter.ref = f"A1:G{linha_total}"

# 2 — por família
ws = wb.create_sheet("Por família")
linhas = [["Família"] + [rot[b] for b in baldes] + ["Total"]]
for f, d in p["por_familia"].items():
    linhas.append([f] + [d[b] for b in baldes] + [d["total"]])
escreve(ws, linhas, [30, 18, 24, 16, 20, 16, 9])

# 3 — desfecho dos encerrados
ws = wb.create_sheet("Encerrados por desfecho")
linhas = [["Tipo da SS (TIPOSS)", "SS atendida", "SS cancelada", "Total encerrado"]]
for t, d in p["desfecho_por_tipo"].items():
    if d["total"]:
        linhas.append([t, d["atendida"], d["cancelada"], d["total"]])
tot = [sum(d["atendida"] for d in p["desfecho_por_tipo"].values()),
       sum(d["cancelada"] for d in p["desfecho_por_tipo"].values()),
       sum(d["total"] for d in p["desfecho_por_tipo"].values())]
linhas.append(["TOTAL"] + tot)
escreve(ws, linhas, [36, 14, 14, 16])
for c in range(1, 5):
    ws.cell(row=ws.max_row, column=c).font = neg

# 4 — a base, um ativo por linha
ws = wb.create_sheet("Base dos 143")
cols = [("ativo", "Ativo", 13), ("sigla", "RL/RT", 8), ("localidade", "Praça", 24),
        ("criticidade", "Criticidade", 14), ("balde_rotulo", "Situação", 26),
        ("tipo_da_demanda", "Tipo da SS (TIPOSS)", 34), ("familia", "Família", 26),
        ("como_terminou", "Como terminou", 15), ("qtd_ss_no_coep", "SS no COEP", 11),
        ("todos_os_tipos", "Todos os TIPOSS do ativo", 46),
        ("dias_no_posto", "Dias no posto", 13), ("ss", "SS do COEP", 46)]
linhas = [[c[1] for c in cols]]
for x in p["por_ativo"]:
    linhas.append([x[c[0]] for c in cols])
escreve(ws, linhas, [c[2] for c in cols])

# 5 — como foi feito
ws = wb.create_sheet("Como foi feito")
texto = [
    ["O que esta planilha responde"],
    [""],
    ["Quantos dos 143 equipamentos que passaram pelo posto do COEP em 2026 são de"],
    ["INDISPONIBILIDADE PARA OPERAÇÃO (o equipamento saiu de operação) e quantos são"],
    ["de EM OPERAÇÃO COM ANOMALIA (segue rodando, com defeito)."],
    [""],
    ["A régua"],
    [""],
    ["O tipo vem do campo TIPOSS da SS do COEP, na base de SS/OS " + p["fonte"] + "."],
    ["Quando o ativo teve mais de uma SS no posto, vale a MAIS PESADA: indisponibilidade"],
    ["ganha de em operação com anomalia, que ganha de aviso, que ganha do resto."],
    ["São 9 ativos com tipos misturados; a coluna 'Todos os TIPOSS do ativo' mostra tudo."],
    [""],
    ["Duas SS de janeiro de 2023 (7915029003 Gurupi e 7923674004 Araguaína) não estão"],
    ["na base de SS/OS — o export só alcança 24 SS do COEP daquele ano. Ficam sem tipo,"],
    ["sem chute."],
    [""],
    ["O que ficou FORA da conta"],
    [""],
    ["Por decisão do gestor (29/08) saem da conta do posto os tipos que não são"],
    ["manutenção de equipamento — instalar, energizar e ajustar não é consertar:"],
    [""],
    ["   OBRAS (NOVOS EQUIPAMENTOS)   13    instalação de equipamento novo"],
    ["   COMISSIONAMENTO               9    energização"],
    ["   AJUSTES DE PROTEÇÃO           2    parametrização"],
    ["                                24"],
    [""],
    ["Eles continuam na aba 'Base dos 143', marcados como 'Fora da conta', para"],
    ["conferência. Passaram pelo posto 143; a conta de manutenção é sobre 119."],
    [""],
    ["O que continua na conta, mas não é falha"],
    [""],
    ["SOLICITAÇÃO DE SERVIÇO (4) e AVISOS DE ANOMALIA (4, somando os três tipos de"],
    ["aviso) também não são falha de equipamento. Somam 8 e seguem na conta até o"],
    ["gestor dizer o contrário — basta mandar que saem junto."],
    [""],
    ["Cada ativo aparece UMA VEZ"],
    [""],
    ["São 143 ativos distintos em 143 linhas. Equipamento que saiu de operação duas"],
    ["vezes conta uma vez só; a coluna 'SS no COEP' mostra quantas SS ele teve."],
]
for r in texto:
    ws.append(r)
ws.column_dimensions["A"].width = 96
ws["A1"].font = Font(bold=True, size=12, color=SINAL)
for r, txt in enumerate(texto, start=1):
    if txt and txt[0] and not txt[0].startswith(" ") and len(txt[0]) < 46 \
            and not txt[0].endswith("."):
        ws.cell(row=r, column=1).font = Font(bold=True, size=11)
ws["A1"].font = Font(bold=True, size=12, color=SINAL)

os.makedirs(f"{RAIZ}/dist", exist_ok=True)
wb.save(f"{RAIZ}/dist/TIPO_DA_DEMANDA.xlsx")
print("gravado dist/TIPO_DA_DEMANDA.xlsx")
