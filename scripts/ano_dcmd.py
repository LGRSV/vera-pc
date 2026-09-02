"""
O ano de 2026 na visão DCMD, de ponta a ponta — dist/ANO_DCMD_2026.xlsx.

O gestor corrigiu em 02/09: o quadro de set–dez é VISÃO DCMD, e nela «resolvemos 71
até agosto». E pediu que a série de jan a ago entregue o bastão para setembro — a
história e a premissa têm de ser a mesma conta.

O QUE É A VISÃO DCMD AQUI: os 143 equipamentos que passaram pelo posto do COEP em
2026, na partição fechada em 28 e 29/08 — 71 resolvidos · 54 na fila · 15 despachados
para outra mesa · 3 em execução no campo. O 71 é o número que o gestor reconhece, e
esta planilha o mensaliza: 4 · 1 · 4 · 2 · 11 · 29 · 14 · 6.

O BACKLOG DE 100 DO QUADRO É NÚMERO REDONDO DE REFERÊNCIA (gestor, 02/09) — não é o
estoque apurado. O estoque real no fim de agosto é 72, que é 54 + 15 + 3, os três
baldes que sobraram. Reancorando a premissa nesse 72, DEZEMBRO FECHA EM 18 PENDENTES,
não em 29. É a mesma premissa dele, só que partindo do número apurado.

  fim de agosto  72
  setembro       72 + 8 − 6  = 74
  outubro        74 + 5 − 29 = 50
  novembro       50 + 4 − 22 = 32
  dezembro       32 + 6 − 20 = 18

O RITMO DA PREMISSA É OUTRO. De janeiro a agosto o posto resolveu 8,9 por mês e
recebeu 11,6. A premissa de set–dez prevê resolver 19,2 por mês e receber 5,8 — mais
que o dobro de saída com metade de entrada. A aba «Ritmo e cenários» põe isso lado a
lado e mostra dezembro em quatro hipóteses, de 18 a 82.

DUAS RESSALVAS DA BASE:
  · a posição é 18/08/2026 (coep_2026.json), então agosto é mês parcial;
  · dois ativos têm data de fechamento anterior à chegada ao posto (7921040031 e
    7955430075). A saída é travada na data de entrada, senão o saldo não fecha.

Rodar: python3 scripts/ano_dcmd.py
"""

import datetime as dt
import json
import os
import sys
from collections import Counter

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Font, PatternFill

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
import backlog_mensal as bm  # noqa: E402  — reaproveita o estilo de gráfico

COEP = os.path.join(RAIZ, "data", "missao", "coep_2026.json")
SAIDA = os.path.join(RAIZ, "dist", "ANO_DCMD_2026.xlsx")
JSON_SAIDA = os.path.join(RAIZ, "data", "missao", "ano_dcmd.json")

TINTA, PAPEL, SOMBRA, SINAL = bm.TINTA, bm.PAPEL, bm.SOMBRA, bm.SINAL
VERDE, LARANJA, NEUTRO = bm.VERDE, bm.LARANJA, bm.NEUTRO
MESES = ["janeiro", "fevereiro", "março", "abril", "maio", "junho", "julho", "agosto",
         "setembro", "outubro", "novembro", "dezembro"]
POSICAO = dt.date(2026, 8, 18)

# a premissa do gestor, set–dez (resolvidos acumulados 6·35·57·77 → por mês)
PREM_ENTRA = [8, 5, 4, 6]
PREM_RESOLVE = [6, 29, 22, 20]
QUADRO_PENDENTES = [102, 70, 47, 29]      # como está no quadro dele, com backlog 100


def data(s):
    try:
        return dt.datetime.strptime((s or "").strip(), "%d/%m/%Y").date()
    except (ValueError, AttributeError):
        return None


def apurar():
    with open(COEP, encoding="utf-8") as fh:
        cp = json.load(fh)
    at = {a["ativo"]: a for a in cp["ativos"]}
    res = [r for r in cp["resolvidos_do_coep"] if r["conta_como_resolvido_pelo_coep"]]
    fila = {a for a, v in at.items() if v["segue_no_posto"]}
    voltaram = {r["ativo"] for r in res} & fila
    fech = {r["ativo"]: data(r["data_do_fechamento"]) for r in res if r["ativo"] not in voltaram}
    despachados = {p["ativo"] for p in cp["pendentes_em_outra_mesa"]}

    itens, invertidos = [], []
    for ativo, v in at.items():
        chegada = data(v.get("primeira_chegada"))
        saida = fech.get(ativo)
        if saida and chegada and saida < chegada:      # trava: não sai antes de entrar
            invertidos.append((ativo, chegada, saida))
            saida = chegada
        itens.append({"ativo": ativo, "tipo": "RT" if ativo[:2] == "58" else "RL",
                      "localidade": v.get("localidade", ""), "chegada": chegada, "saida": saida,
                      "ss": v.get("ss", ""), "criticidade": v.get("criticidade", ""),
                      # os 18 de outra mesa se repartem em 15 despachados e 3 em execução no
                      # campo (partição de 29/08); aqui vão num balde só, que é o que o saldo usa
                      "balde": ("resolvido" if saida else
                                "em outra mesa" if ativo in despachados else
                                "na fila do posto" if ativo in fila else "em outro estado")})
    return itens, invertidos


def mensal(itens):
    herdados = sum(1 for i in itens if i["chegada"] and i["chegada"].year < 2026)
    linhas, saldo = [], herdados
    for m in range(1, 9):
        ent = [i for i in itens if i["chegada"] and i["chegada"].year == 2026
               and i["chegada"].month == m]
        sai = [i for i in itens if i["saida"] and i["saida"].year == 2026 and i["saida"].month == m]
        ini = saldo
        saldo += len(ent) - len(sai)
        linhas.append({"mes": m, "rotulo": MESES[m - 1], "origem": "apurado", "inicio": ini,
                       "entraram": len(ent), "resolvidos": len(sai), "fim": saldo,
                       "rl_entraram": sum(1 for i in ent if i["tipo"] == "RL"),
                       "rt_entraram": sum(1 for i in ent if i["tipo"] == "RT"),
                       "rl_resolvidos": sum(1 for i in sai if i["tipo"] == "RL"),
                       "rt_resolvidos": sum(1 for i in sai if i["tipo"] == "RT")})
    for j in range(4):
        ini = saldo
        saldo += PREM_ENTRA[j] - PREM_RESOLVE[j]
        linhas.append({"mes": 9 + j, "rotulo": MESES[8 + j], "origem": "premissa", "inicio": ini,
                       "entraram": PREM_ENTRA[j], "resolvidos": PREM_RESOLVE[j], "fim": saldo,
                       "rl_entraram": None, "rt_entraram": None,
                       "rl_resolvidos": None, "rt_resolvidos": None})
    return herdados, linhas


# ----------------------------------------------------------------------------- abas
def aba_ano(wb, linhas, herdados, bc):
    ws = wb.create_sheet("O ano inteiro")
    bm.titulo(ws, "2026 NA VISÃO DCMD — janeiro a agosto apurado, setembro a dezembro pela premissa",
              "Os 143 equipamentos que passaram pelo posto do COEP. O ano abre com %d herdados de "
              "2025 ou antes. De janeiro a agosto os números vêm da base; de setembro em diante é a "
              "premissa do gestor, reancorada no estoque apurado de %d — por isso dezembro fecha em "
              "%d, e não nos 29 do quadro, que partia de um backlog redondo de 100. O banco de "
              "capacitor vai em coluna própria; de setembro em diante ele fica parado em %d, "
              "porque a premissa não projeta BC."
              % (herdados, linhas[7]["fim"], linhas[-1]["fim"], bc[-1]["fim"]))
    cab = ["Mês", "Origem", "Backlog no início", "Entraram", "Resolvidos", "Backlog no fim",
           "Variação", "BC no fim", "TOTAL com BC", "RL entraram", "RT entraram",
           "RL resolvidos", "RT resolvidos"]
    bm.cabecalho(ws, 4, cab, [13, 11, 17, 12, 12, 16, 11, 12, 14, 13, 13, 14, 14])
    r = 5
    bc_fim = [b["fim"] for b in bc] + [bc[-1]["fim"]] * 4      # a premissa não projeta BC
    for L, b in zip(linhas, bc_fim):
        ws.append([L["rotulo"], L["origem"], L["inicio"], L["entraram"], L["resolvidos"], L["fim"],
                   L["fim"] - L["inicio"], b, L["fim"] + b, L["rl_entraram"], L["rt_entraram"],
                   L["rl_resolvidos"], L["rt_resolvidos"]])
        for c in range(2, 14):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        if L["origem"] == "premissa":
            for c in range(1, 14):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=SOMBRA)
        r += 1
    fim = r - 1
    ws.cell(row=r, column=1, value="no ano").font = Font(bold=True)
    ws.cell(row=r, column=3, value=herdados).font = Font(bold=True)
    ws.cell(row=r, column=4, value=f"=SUM(D5:D{fim})").font = Font(bold=True)
    ws.cell(row=r, column=5, value=f"=SUM(E5:E{fim})").font = Font(bold=True)
    ws.cell(row=r, column=6, value=linhas[-1]["fim"]).font = Font(bold=True)
    for c in range(3, 7):
        ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth, ch.overlap = "col", "clustered", 80, -12
    ch.add_data(Reference(ws, min_col=4, min_row=4, max_col=5, max_row=fim), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=fim))
    ch.title = "O ano inteiro: entradas, resolvidos e o saldo (set–dez sombreado é premissa)"
    ch.y_axis.title = "equipamentos"
    bm.cor_barra(ch.series[0], LARANJA)
    bm.cor_barra(ch.series[1], VERDE)
    for s in ch.series:
        bm.rotulos(s)
    lin = LineChart()
    lin.add_data(Reference(ws, min_col=6, min_row=4, max_row=fim), titles_from_data=True)
    ls = lin.series[0]
    ls.graphicalProperties = GraphicalProperties()
    ls.graphicalProperties.line = LineProperties(solidFill=TINTA[2:], w=28000)
    ls.marker = Marker(symbol="circle", size=8)
    ls.marker.graphicalProperties = GraphicalProperties(solidFill=TINTA[2:])
    ls.smooth = False
    bm.rotulos(ls)
    ch += lin
    ch.legend.position, ch.legend.overlay = "b", False
    bm.categorias(ch, ws, "$A$5:$A$%d" % fim)
    ws.add_chart(bm.estilo(ch, 12, 28), "A20")


def aba_quadro(wb, linhas):
    ws = wb.create_sheet("Como o quadro muda")
    bm.titulo(ws, "O mesmo quadro, partindo do número apurado",
              "Mesma premissa de entrada e de resolução, só que ancorada no estoque real de %d no fim "
              "de agosto em vez do backlog redondo de 100. O quadro original não estava errado — o 100 "
              "é referência, e a diferença de %d equipamentos anda junto o ano inteiro."
              % (linhas[7]["fim"], 100 - linhas[7]["fim"]))
    bm.cabecalho(ws, 4, ["Pendentes no fim do mês", "setembro", "outubro", "novembro", "dezembro"],
                 [36, 15, 15, 15, 15])
    prem = [L["fim"] for L in linhas[8:]]
    for i, (rot, vals) in enumerate([("Como no quadro (backlog 100)", QUADRO_PENDENTES),
                                     ("Reancorado no apurado (72)", prem),
                                     ("Diferença", [b - a for a, b in zip(QUADRO_PENDENTES, prem)])]):
        ws.append([rot] + list(vals))
        for c in range(2, 6):
            ws.cell(row=5 + i, column=c).alignment = Alignment(horizontal="center")
        if i == 2:
            for c in range(1, 6):
                ws.cell(row=5 + i, column=c).font = Font(italic=True, color=SINAL)
    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth, ch.overlap = "col", "clustered", 80, -12
    ch.add_data(Reference(ws, min_col=1, min_row=5, max_col=5, max_row=6),
                from_rows=True, titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=2, min_row=4, max_col=5, max_row=4))
    ch.title = "Pendentes no fim do mês: o quadro e a versão ancorada no apurado"
    ch.y_axis.title = "equipamentos"
    bm.cor_barra(ch.series[0], NEUTRO)
    bm.cor_barra(ch.series[1], LARANJA)
    for s in ch.series:
        bm.rotulos(s)
    ch.legend.position, ch.legend.overlay = "b", False
    bm.categorias(ch, ws, "$B$4:$E$4")
    ws.add_chart(bm.estilo(ch, 12, 24), "A10")


def aba_ritmo(wb, linhas):
    ws = wb.create_sheet("Ritmo e cenários")
    real_res = sum(L["resolvidos"] for L in linhas[:8])
    real_ent = sum(L["entraram"] for L in linhas[:8])
    base = linhas[7]["fim"]
    prem_res, prem_ent = sum(PREM_RESOLVE), sum(PREM_ENTRA)
    r_res, r_ent = real_res / 8, real_ent / 8
    bm.titulo(ws, "O ritmo da premissa é o dobro do realizado",
              "De janeiro a agosto o posto resolveu %.1f por mês e recebeu %.1f. A premissa de set–dez "
              "prevê resolver %.1f por mês e receber %.1f — mais que o dobro de saída com metade de "
              "entrada. Os quatro cenários abaixo partem todos do estoque apurado de %d."
              % (r_res, r_ent, prem_res / 4, prem_ent / 4, base))
    bm.cabecalho(ws, 4, ["Ritmo por mês", "Janeiro a agosto (apurado)", "Setembro a dezembro (premissa)",
                         "Quantas vezes"], [26, 24, 28, 14])
    for rot, a, b in (("Resolvidos", r_res, prem_res / 4), ("Entradas", r_ent, prem_ent / 4)):
        ws.append([rot, round(a, 1), round(b, 1), round(b / a, 2)])
    for r in (5, 6):
        for c in range(2, 5):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=c).number_format = "0.0" if c < 4 else "0.00"

    bm.cabecalho(ws, 8, ["Dezembro em quatro cenários", "Entram set–dez", "Resolvem set–dez",
                         "Pendentes em dezembro"], [40, 17, 18, 21])
    cen = [("Premissa do gestor, como está", prem_ent, prem_res),
           ("Entradas no ritmo realizado, resolução da premissa", round(r_ent * 4), prem_res),
           ("Entradas da premissa, resolução no ritmo realizado", prem_ent, round(r_res * 4)),
           ("Tudo no ritmo realizado de jan–ago", round(r_ent * 4), round(r_res * 4))]
    r = 9
    for rot, e, s in cen:
        ws.append([rot, e, s, base + e - s])
        for c in range(2, 5):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1
    fim = r - 1
    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth = "col", "clustered", 70
    ch.add_data(Reference(ws, min_col=4, min_row=8, max_row=fim), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=9, max_row=fim))
    ch.title = "Onde dezembro fecha, conforme o ritmo que se mantiver"
    ch.y_axis.title = "pendentes em dezembro"
    bm.cor_barra(ch.series[0], LARANJA)
    bm.rotulos(ch.series[0])
    ch.legend = None
    bm.categorias(ch, ws, "$A$9:$A$%d" % fim)
    ws.add_chart(bm.estilo(ch, 12, 26), "A15")


def aba_71(wb, linhas, itens):
    ws = wb.create_sheet("Os 71 resolvidos")
    bm.titulo(ws, "Os 71 resolvidos até agosto, mês a mês",
              "É o número que o gestor reconhece: 71 demandas encerradas de ponta a ponta, 45 "
              "religadores e 26 reguladores. Junho responde por 29 deles — o mês do mutirão.")
    bm.cabecalho(ws, 4, ["Mês", "Resolvidos", "RL", "RT", "Acumulado"], [14, 13, 10, 10, 13])
    acum = 0
    r = 5
    for L in linhas[:8]:
        acum += L["resolvidos"]
        ws.append([L["rotulo"], L["resolvidos"], L["rl_resolvidos"], L["rt_resolvidos"], acum])
        for c in range(2, 6):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1
    fim = r - 1
    ws.cell(row=r, column=1, value="total").font = Font(bold=True)
    for c, v in ((2, acum), (3, sum(L["rl_resolvidos"] for L in linhas[:8])),
                 (4, sum(L["rt_resolvidos"] for L in linhas[:8]))):
        ws.cell(row=r, column=c, value=v).font = Font(bold=True)
        ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth, ch.overlap = "col", "clustered", 80, -12
    ch.add_data(Reference(ws, min_col=3, min_row=4, max_col=4, max_row=fim), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=fim))
    ch.title = "Os 71 por mês, separados em religador e regulador"
    ch.y_axis.title = "equipamentos resolvidos"
    bm.cor_barra(ch.series[0], LARANJA)
    bm.cor_barra(ch.series[1], VERDE)
    for s in ch.series:
        bm.rotulos(s)
    ch.legend.position, ch.legend.overlay = "b", False
    bm.categorias(ch, ws, "$A$5:$A$%d" % fim)
    ws.add_chart(bm.estilo(ch, 12, 26), "A16")


def aba_base(wb, itens):
    ws = wb.create_sheet("Base dos 143")
    bm.titulo(ws, "Os 143 que passaram pelo posto em 2026",
              "Um por linha, com o balde em que terminou. Os 71 resolvidos têm data de saída; "
              "os outros 72 são o estoque do fim de agosto: 54 na fila do posto e 18 em outra mesa.")
    bm.cabecalho(ws, 4, ["Ativo", "Tipo", "Localidade", "Criticidade", "SS no COEP",
                         "Chegou ao posto", "Saiu em", "Dias", "Balde"],
                 [13, 6, 24, 14, 22, 15, 13, 8, 26])
    r = 5
    for i in sorted(itens, key=lambda x: (x["balde"], x["chegada"] or dt.date(2000, 1, 1))):
        dias = ((i["saida"] or POSICAO) - i["chegada"]).days if i["chegada"] else None
        ws.append([i["ativo"], i["tipo"], i["localidade"], i["criticidade"], i["ss"],
                   i["chegada"], i["saida"], dias, i["balde"]])
        ws.cell(row=r, column=1).number_format = "@"
        for c in (6, 7):
            ws.cell(row=r, column=c).number_format = "dd/mm/yyyy"
        r += 1
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:I%d" % (r - 1)
    return Counter(i["balde"] for i in itens)


def aba_como(wb, linhas, herdados, baldes, invertidos, bc):
    ws = wb.create_sheet("Como foi feito")
    ws.column_dimensions["A"].width = 112
    base = linhas[7]["fim"]
    texto = [
        ("A REGRA QUE MANDA AQUI", True),
        ("A série de janeiro a agosto TEM DE BATER com os números de setembro em diante. Não são "
         "duas contas: é uma só, cortada no meio. O que fecha agosto é o que abre setembro — por "
         "isso a premissa foi reancorada no estoque apurado, e não no backlog redondo de 100. "
         "Qualquer número novo que chegue para set–dez tem de fechar contra este saldo de agosto; "
         "se não fechar, é sinal de que as duas metades estão em réguas diferentes.", False),
        ("", False),
        ("A CONTA, EM UMA LINHA", True),
        ("%d herdados de 2025 ou antes + %d que chegaram em 2026 = 143 que passaram pelo posto. "
         "Desses, 71 resolvidos até agosto e %d ainda abertos — que é %d na fila do posto e %d em "
         "outra mesa; esses 18 se repartem, pela partição de 29/08, em 15 despachados com a peça já "
         "trocada e 3 ainda em execução no campo."
         % (herdados, sum(L["entraram"] for L in linhas[:8]), base,
            baldes["na fila do posto"], baldes["em outra mesa"]), False),
        ("", False),
        ("POR QUE O QUADRO PARTIA DE 100 E ESTA SÉRIE PARTE DE %d" % base, True),
        ("O gestor confirmou em 02/09 que o backlog de 100 do quadro é NÚMERO REDONDO DE REFERÊNCIA, "
         "não o estoque apurado. O apurado no fim de agosto é %d. Reancorando a premissa nesse número, "
         "e mantendo as mesmas entradas e resoluções que ele previu, dezembro fecha em %d em vez de 29." % (base, linhas[-1]["fim"]), False),
        ("A diferença de %d equipamentos anda junto o ano inteiro: setembro %d contra 102, outubro %d "
         "contra 70, novembro %d contra 47, dezembro %d contra 29."
         % (100 - base, linhas[8]["fim"], linhas[9]["fim"], linhas[10]["fim"], linhas[11]["fim"]), False),
        ("", False),
        ("DE ONDE VEM CADA METADE", True),
        ("Janeiro a agosto: apurado sobre coep_2026.json, a partição dos 143 fechada em 28 e 29/08 — "
         "71 resolvidos · 54 na fila · 15 despachados · 3 em execução. Resolvido é a demanda que "
         "acabou; quem voltou para a fila não conta, e quem saiu para ajuste ou comissionamento fica "
         "em balde próprio.", False),
        ("Setembro a dezembro: a premissa do gestor, como ele mandou — entram 8·5·4·6 e resolvem "
         "6·29·22·20 (que são os acumulados 6·35·57·77 do quadro, mês a mês).", False),
        ("", False),
        ("O RITMO", True),
        ("Jan–ago: %.1f resolvidos por mês e %.1f entradas. Premissa set–dez: %.1f resolvidos e %.1f "
         "entradas. Ou seja, a premissa promete mais que o dobro de saída com metade de entrada. "
         "A aba «Ritmo e cenários» mostra dezembro em quatro hipóteses, de %d a %d."
         % (sum(L["resolvidos"] for L in linhas[:8]) / 8, sum(L["entraram"] for L in linhas[:8]) / 8,
            sum(PREM_RESOLVE) / 4, sum(PREM_ENTRA) / 4, linhas[-1]["fim"],
            base + round(sum(L["entraram"] for L in linhas[:8]) / 8 * 4)
            - round(sum(L["resolvidos"] for L in linhas[:8]) / 8 * 4)), False),
        ("", False),
        ("RESSALVAS", True),
        ("A posição do apurado é 18/08/2026 — agosto é mês parcial, e por isso ele mostra só 1 entrada.", False),
        ("%d ativos têm data de fechamento anterior à data de chegada ao posto (%s). A saída foi "
         "travada na data de entrada; sem isso o saldo não fecha."
         % (len(invertidos), ", ".join(a for a, _, _ in invertidos)), False),
        ("", False),
        ("O BANCO DE CAPACITOR", True),
        ("Entra a pedido do gestor (02/09): «tem que ter a visão de BC, que começa com código "
         "operativo 59». Ele nunca aparecia nas contas porque a consulta da base de repasse o "
         "descarta na cláusula «COD_ELE NOT IN ('59','BR')» — só indo direto na base de SS/OS. "
         "São 242 SS em 102 ativos, e o backlog cai de %d no começo do ano para %d. De setembro "
         "em diante ele fica parado nesse número na tabela, porque a premissa do gestor não "
         "projeta BC — quando houver premissa de BC, é só substituir." % (bc[0]["inicio"], bc[-1]["fim"]), False),
        ("", False),
        ("Esta é a visão DCMD, do posto. Não confundir com o backlog do parque em "
         "dist/BACKLOG_MENSAL_2026.xlsx, que conta todo RL/RT com indisponibilidade aberta (93 no "
         "fecho da base) — é outro recorte, mais largo, e serve para outra pergunta.", False),
    ]
    for i, (t, negrito) in enumerate(texto, 1):
        c = ws.cell(row=i, column=1, value=t)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if negrito:
            c.font = Font(bold=True, size=11, color=SINAL if i == 1 else TINTA)


def bc_mensal():
    """O banco de capacitor (59), que a consulta da base de repasse descarta. Vem da base
    de SS/OS pela mesma régua de indisponibilidade do backlog do parque."""
    ss, posicao = bm.ler()
    dem, _ = bm.demandas(ss, {bm.IND})
    return bm.serie(dem, posicao, ("BC",)), posicao


def aba_bc(wb, bc, posicao):
    ws = wb.create_sheet("Banco de capacitor")
    bm.titulo(ws, "O banco de capacitor — a visão que faltava",
              "Código operativo 59. Ele nunca aparecia nas contas porque a consulta da base de "
              "repasse o descarta na cláusula «COD_ELE NOT IN ('59','BR')». Indo direto na base de "
              "SS/OS são 242 SS em 102 ativos, e o backlog cai de %d no começo do ano para %d na "
              "posição de %s. Mesma régua do parque: ativo com SS de indisponibilidade em aberto."
              % (bc[0]["inicio"], bc[-1]["fim"], posicao.strftime("%d/%m/%Y")))
    bm.cabecalho(ws, 4, ["Mês", "BC no início", "Entraram", "Saíram", "BC no fim"],
                 [14, 15, 12, 12, 14])
    r = 5
    for L in bc:
        ws.append([L["rotulo"], L["inicio"], L["entraram"], L["sairam"], L["fim"]])
        for c in range(2, 6):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1
    fim = r - 1
    ch = BarChart()
    ch.type, ch.grouping, ch.gapWidth = "col", "clustered", 70
    ch.add_data(Reference(ws, min_col=5, min_row=4, max_row=fim), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=fim))
    ch.title = "Bancos de capacitor com demanda aberta, no fim de cada mês"
    ch.y_axis.title = "bancos"
    bm.cor_barra(ch.series[0], NEUTRO)
    bm.rotulos(ch.series[0])
    ch.legend = None
    bm.categorias(ch, ws, "$A$5:$A$%d" % fim)
    ws.add_chart(bm.estilo(ch, 12, 24), "A16")


def montar(saida=SAIDA):
    itens, invertidos = apurar()
    herdados, linhas = mensal(itens)
    bc, pos_bc = bc_mensal()
    assert sum(L["resolvidos"] for L in linhas[:8]) == 71, "os resolvidos de jan–ago têm de dar 71"
    assert linhas[7]["fim"] == 72, "o estoque de agosto tem de dar 72"
    wb = Workbook()
    wb.remove(wb.active)
    aba_ano(wb, linhas, herdados, bc)
    aba_quadro(wb, linhas)
    aba_ritmo(wb, linhas)
    aba_71(wb, linhas, itens)
    aba_bc(wb, bc, pos_bc)
    baldes = aba_base(wb, itens)
    aba_como(wb, linhas, herdados, baldes, invertidos, bc)
    os.makedirs(os.path.dirname(saida), exist_ok=True)
    wb.save(saida)
    with open(JSON_SAIDA, "w", encoding="utf-8") as fh:
        json.dump({"posicao": POSICAO.isoformat(), "herdados": herdados, "baldes": dict(baldes),
                   "mensal": linhas, "mensal_bc": bc, "quadro_do_gestor": QUADRO_PENDENTES,
                   "invertidos": [[a, c.isoformat(), s.isoformat()] for a, c, s in invertidos]},
                  fh, ensure_ascii=False, indent=1)
    print("%s | herdados %d | baldes %s" % (saida, herdados, dict(baldes)))
    print("  saldo: %s" % " · ".join("%s %d%s" % (L["rotulo"][:3], L["fim"],
                                                  "*" if L["origem"] == "premissa" else "")
                                     for L in linhas))
    from planilha_automatica import grava_cache
    print("  cache: %d células" % grava_cache(saida))
    return saida


if __name__ == "__main__":
    montar(sys.argv[1] if len(sys.argv) > 1 else SAIDA)
