"""
Carteira de entrada, mês a mês — os 117 ativos da foto de junho distribuídos
pela data de abertura da SS.

Regra do gestor (13/08): SS aberta antes de 2026 entra em janeiro de 2026. O
mês de janeiro passa a ser, na prática, «o que eu já herdei velho», e os meses
seguintes mostram o que chegou depois.

A data sai da coluna DATA_ABERTURA_SS da aba «Dados» da planilha de entrada.
Quando a SS só existe na aba «Dados Tratados», que não traz abertura, a data
vem do cruzamento pelo número da SS com a base de SS/OS cheia — o caminho que
o gestor mandou usar.

Ativo com mais de uma SS na foto entra pela mais antiga: é quando a demanda
chegou ao posto.
"""

import datetime
import json
import os
import re
from collections import Counter, defaultdict

import openpyxl

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(RAIZ, "data", "raw", "BASE_SS_OS_EQ_ESPECIAIS_ENTRADA.xlsx")
ARQ_MIN = os.path.join(RAIZ, "data", "missao", "ssos_min.json")

CORTE = 2026
MES_CORTE = "2026-01"
MESES_PT = ["jan", "fev", "mar", "abr", "mai", "jun",
            "jul", "ago", "set", "out", "nov", "dez"]

PLANILHA = "planilha de entrada"
CRUZAMENTO = "cruzamento pela SS na base de SS/OS"

FORMATOS = ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y")


def _data(texto):
    t = str(texto or "").strip()
    if not t:
        return None
    for f in FORMATOS:
        try:
            return datetime.datetime.strptime(t[:19] if " " in t else t[:10], f).date()
        except ValueError:
            continue
    return None


def _aberturas_da_planilha():
    """NUMERO_SS → data de abertura, da aba «Dados» (extrato cru do SGM)."""
    if not os.path.exists(XLSX):
        return {}
    ws = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)["Dados"]
    linhas = ws.iter_rows(values_only=True)
    cabecalho = list(next(linhas))
    try:
        i_ss = cabecalho.index("NUMERO_SS")
        i_ab = cabecalho.index("DATA_ABERTURA_SS")
    except ValueError:
        return {}
    fora = {}
    for r in linhas:
        ss = str(r[i_ss] or "").strip()
        if ss and ss not in fora:
            fora[ss] = _data(r[i_ab])
    return {k: v for k, v in fora.items() if v}


def _aberturas_da_base():
    """NUMERO_SS → data de abertura, da base de SS/OS cheia."""
    if not os.path.exists(ARQ_MIN):
        return {}
    with open(ARQ_MIN, encoding="utf-8") as fh:
        linhas = json.load(fh)
    fora = {}
    for r in linhas:
        ss = str(r.get("NUMERO_SS") or "").strip()
        if ss and ss not in fora:
            d = _data(r.get("DATA_ABERTURA_SS"))
            if d:
                fora[ss] = d
    return fora


def rotulo(mes):
    ano, m = mes.split("-")
    return f"{MESES_PT[int(m) - 1]}/{ano}"


RECORTE = "INDISPONIBILIDADE"


def _tiposs_por_ss(base):
    """NUMERO_SS → TIPOSS, da aba «Dados» com fallback na base de SS/OS."""
    tp = {}
    if os.path.exists(XLSX):
        ws = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)["Dados"]
        linhas = ws.iter_rows(values_only=True)
        cabecalho = list(next(linhas))
        try:
            i_ss, i_tp = cabecalho.index("NUMERO_SS"), cabecalho.index("TIPOSS")
        except ValueError:
            i_ss = i_tp = None
        if i_ss is not None:
            for r in linhas:
                n = str(r[i_ss] or "").strip()
                if n and n not in tp and str(r[i_tp] or "").strip():
                    tp[n] = str(r[i_tp] or "").strip()
    for r in base:
        n = (r.get("NUMERO_SS") or "").strip()
        if n and n not in tp and (r.get("TIPOSS") or "").strip():
            tp[n] = (r.get("TIPOSS") or "").strip()
    return tp


def _base_ssos():
    if not os.path.exists(ARQ_MIN):
        return []
    with open(ARQ_MIN, encoding="utf-8") as fh:
        return json.load(fh)


ANO_DA_SS = re.compile(r"/(\d{4})\s*$")


def entrantes_no_coep(base, foto=frozenset()):
    """Ativos que passaram pelo posto ETO-COEP, mês a mês, pela abertura da SS.

    Entrante = a primeira SS do ETO-COEP daquele ativo em toda a base. Ativo que
    volta ao posto conta como revisita, senão o mesmo equipamento inflaria a
    entrada duas vezes.
    """
    coep = []
    for r in base:
        if (r.get("COD_EQUIPE") or "").strip() != "ETO-COEP":
            continue
        d = _data(r.get("DATA_ABERTURA_SS"))
        a = (r.get("NUM_TRAFO") or "").strip()
        if d and a:
            coep.append({**r, "_d": d, "_a": a})
    coep.sort(key=lambda r: r["_d"])

    primeira = {}
    for r in coep:
        primeira.setdefault(r["_a"], r)

    def recarimbada(r):
        """SS cujo ano do número não bate com o ano da abertura.

        O SGM re-carimba a DATA_ABERTURA_SS quando a SS é reaberta ou repassada:
        uma SS de 2023 pode aparecer com abertura em 2026. Onde isso acontece, a
        «entrada» do mês é demanda velha voltando, não demanda nova.
        """
        m = ANO_DA_SS.search(str(r.get("NUMERO_SS") or ""))
        return bool(m) and int(m.group(1)) != r["_d"].year

    por_mes = defaultdict(lambda: {"ss": 0, "ativos": set(), "novos": set()})
    for r in coep:
        e = por_mes[f"{r['_d'].year}-{r['_d'].month:02d}"]
        e["ss"] += 1
        e["ativos"].add(r["_a"])
        if primeira[r["_a"]] is r:
            e["novos"].add(r["_a"])

    serie = []
    for k, e in sorted(por_mes.items()):
        novos = [primeira[a] for a in e["novos"]]
        serie.append({
            "mes": k, "rotulo": rotulo(k), "ss": e["ss"],
            "ativos": len(e["ativos"]), "novos": len(novos),
            "revisita": len(e["ativos"]) - len(novos),
            "na_foto": sum(1 for r in novos if r["_a"] in foto),
            "fora_da_foto": sum(1 for r in novos if r["_a"] not in foto),
            "ss_do_ano": sum(1 for r in novos if not recarimbada(r)),
            "ss_recarimbada": sum(1 for r in novos if recarimbada(r)),
        })

    detalhe = [{
        "ativo": a, "primeira_ss": r["NUMERO_SS"], "abertura": r["_d"],
        "mes": f"{r['_d'].year}-{r['_d'].month:02d}",
        "localidade": r.get("LOCALIDADE", ""),
        "equipamento": r.get("DESCICAO_DO_ATIVO", ""),
        "criticidade_ss": r.get("CRITICIDADE_SS", ""),
        "tiposs": r.get("TIPOSS", ""),
        "situacao_hoje": r.get("SITUACAO_SS", ""),
        "na_foto": "sim" if a in foto else "não",
        "recarimbada": "sim" if recarimbada(r) else "não",
    } for a, r in sorted(primeira.items(), key=lambda kv: kv[1]["_d"])]
    return serie, detalhe


def _datas_de_tratativa(lista, entrada, base):
    """Quando cada resolvido saiu da carteira — término da SS ou repasse.

    Ordem das fontes, da mais forte para a mais fraca: término da SS de entrada
    (separando cancelamento de encerramento normal), data de repasse (a abertura
    da SS seguinte da mesma demanda), obra encerrada no AIC, reporte de campo,
    decisão do gestor e, em último caso, a SS mais recente atendida no ativo.
    """
    por_ss, por_ativo_ss = defaultdict(list), defaultdict(list)
    for r in base:
        por_ss[(r.get("NUMERO_SS") or "").strip()].append(r)
        a = (r.get("NUM_TRAFO") or "").strip()
        if a:
            por_ativo_ss[a].append(r)

    aic = {}
    arq = os.path.join(RAIZ, "data", "raw", "aic_obras.csv")
    if os.path.exists(arq):
        import csv
        with open(arq, encoding="utf-8") as fh:
            for row in csv.DictReader(fh, delimiter=";"):
                if row.get("data_encerramento"):
                    aic.setdefault(row["obra"], row["data_encerramento"])

    reportes = defaultdict(list)
    arq = os.path.join(RAIZ, "data", "raw", "reportes_campo.json")
    if os.path.exists(arq):
        with open(arq, encoding="utf-8") as fh:
            for r in json.load(fh):
                reportes[r["ativo"]].append(r)

    ficha = {}
    for balde in ("resolvidos", "verificar", "em_andamento"):
        for x in (entrada.get(balde) or {}).get("lista", []):
            ficha.setdefault(x["ativo"], x)

    fora = []
    for x in lista:
        e = ficha.get(x["ativo"], {})
        d = via = None
        if x["resolvido"]:
            for r in por_ss.get(x["numero_ss"], []):
                t = _data(r.get("DATA_TERMINO_SS"))
                if t and (d is None or t < d):
                    d, via = t, ("cancelamento da SS de entrada"
                                 if "CANCELADA" in (r.get("SITUACAO_SS") or "").upper()
                                 else "término da SS de entrada")
            if not d and e.get("cauda_mesma_demanda"):
                ds = [y for y in (_data(c.get("abertura")) for c in e["cauda_mesma_demanda"]) if y]
                if ds:
                    d, via = min(ds), "repasse para a etapa seguinte"
            if not d and e.get("obras_encerradas"):
                ds = [y for y in (_data(aic.get(o)) for o in e["obras_encerradas"]) if y]
                if ds:
                    d, via = max(ds), "obra encerrada no AIC"
            if not d and reportes.get(x["ativo"]):
                ds = [y for y in (_data(r["data"]) for r in reportes[x["ativo"]]) if y]
                if ds:
                    d, via = max(ds), "reporte de campo"
            if not d and e.get("decisao_gestor"):
                d = _data(e["decisao_gestor"].get("data"))
                via = "decisão do gestor" if d else None
            if not d:
                abertura = _data(x["abertura"])
                ds = [t for t in (_data(r.get("DATA_TERMINO_SS"))
                                  for r in por_ativo_ss.get(x["ativo"], []))
                      if t and (not abertura or t >= abertura)]
                if ds:
                    d, via = max(ds), "última SS atendida do ativo"
        fora.append({
            "ativo": x["ativo"], "localidade": x["localidade"],
            "numero_ss": x["numero_ss"], "resolvido": x["resolvido"],
            "parecer_coep": e.get("parecer_coep", ""),
            "resolucao": d.isoformat() if d else "",
            "mes_resolucao": f"{d.year}-{d.month:02d}" if d else "",
            "via": via or ("" if x["resolvido"] else "ainda no fluxo"),
        })
    return fora


def montar(entrada):
    if not entrada:
        return None

    itens = []
    for balde in ("resolvidos", "verificar", "em_andamento"):
        for x in (entrada.get(balde) or {}).get("lista", []):
            itens.append({**x, "balde": balde})
    if not itens:
        return None

    base = _base_ssos()
    tiposs = _tiposs_por_ss(base)
    da_planilha = _aberturas_da_planilha()
    da_base = _aberturas_da_base()

    for i in itens:
        ss = i["numero_ss"]
        d, fonte = da_planilha.get(ss), PLANILHA
        if not d:
            d, fonte = da_base.get(ss), CRUZAMENTO
        i["abertura"] = d
        i["fonte_data"] = fonte if d else "não encontrada"
        i["tiposs"] = tiposs.get(ss, "")

    # Decisão final do gestor (13/08): TODOS os tipos de SS contam. O tiposs
    # fica anotado em cada item, mas nada é filtrado por ele.
    dentro = list(itens)
    ativos_dentro = {i["ativo"] for i in dentro}
    fora_por_ativo = {}
    for i in sorted((x for x in itens if x["ativo"] not in ativos_dentro),
                    key=lambda x: x["abertura"] or datetime.date.max):
        fora_por_ativo.setdefault(i["ativo"], i)
    fora_do_recorte = {
        "qtd": len(fora_por_ativo),
        "por_tipo": Counter(i["tiposs"] or "sem tipo"
                            for i in fora_por_ativo.values()).most_common(),
        "lista": [{
            "ativo": i["ativo"], "numero_ss": i["numero_ss"],
            "tiposs": i["tiposs"], "localidade": i.get("localidade", ""),
            "resolvido": i["balde"] == "resolvidos",
        } for i in fora_por_ativo.values()],
    }
    itens = dentro

    sem_data = [i for i in itens if not i["abertura"]]

    # Um ativo pode ter mais de uma SS na foto; vale a mais antiga.
    por_ativo = {}
    ss_por_ativo = defaultdict(list)
    for i in itens:
        if not i["abertura"]:
            continue
        ss_por_ativo[i["ativo"]].append(i)
        atual = por_ativo.get(i["ativo"])
        if not atual or i["abertura"] < atual["abertura"]:
            por_ativo[i["ativo"]] = i

    lista, meses = [], defaultdict(list)
    por_ano_legado = Counter()
    fonte = Counter()
    for ativo, i in sorted(por_ativo.items()):
        d = i["abertura"]
        legado = d.year < CORTE
        mes = MES_CORTE if legado else f"{d.year}-{d.month:02d}"
        if legado:
            por_ano_legado[d.year] += 1
        fonte[i["fonte_data"]] += 1
        outras = [o["numero_ss"] for o in ss_por_ativo[ativo] if o is not i]
        registro = {
            "ativo": ativo,
            "numero_ss": i["numero_ss"],
            "tipo": i.get("tipo", ""),
            "localidade": i.get("localidade", ""),
            "abertura": d.isoformat(),
            "mes": mes,
            "legado": legado,
            "resolvido": i["balde"] == "resolvidos",
            "balde": i["balde"],
            # ativo que saiu da lista de hoje não tem ficha no painel — a carta dele
            # não pode virar botão, senão o clique não faz nada.
            "na_carteira": bool(i.get("na_carteira")),
            "motivo": i.get("motivo", ""),
            "parecer_coep": i.get("parecer_coep", ""),
            "situacao_hoje": i.get("situacao_hoje", ""),
            "fonte_data": i["fonte_data"],
            "outras_ss": outras,
        }
        lista.append(registro)
        meses[mes].append(registro)

    blocos = []
    for mes in sorted(meses):
        g = meses[mes]
        resolvidos = sum(1 for x in g if x["resolvido"])
        blocos.append({
            "mes": mes,
            "rotulo": rotulo(mes),
            "qtd": len(g),
            "legado": sum(1 for x in g if x["legado"]),
            "no_mes": sum(1 for x in g if not x["legado"]),
            "resolvidos": resolvidos,
            "em_andamento": len(g) - resolvidos,
            "percentual": round(100 * resolvidos / len(g), 1) if g else 0.0,
            "ativos": [x["ativo"] for x in g],
        })

    mais_antiga = min(lista, key=lambda x: x["abertura"]) if lista else None
    total = len(lista)
    resolvidos = sum(1 for x in lista if x["resolvido"])

    # As três colunas do gestor no mesmo eixo de tempo: o estoque herdado pela
    # abertura da SS, o que entrou novo no posto e o que foi tratado de fato.
    serie_coep, detalhe_coep = entrantes_no_coep(base, {x["ativo"] for x in lista})
    tratativas = _datas_de_tratativa(lista, entrada, base)
    entrantes = {x["mes"]: x["novos"] for x in serie_coep}
    saidas = Counter(t["mes_resolucao"] for t in tratativas if t["mes_resolucao"])
    # Janela de exibição: janeiro a agosto, até 20/08, que é a posição da base de SS/OS.
    # Agosto é mês parcial e vai dito assim no rótulo — antes ele ficava de fora para
    # não virar um toco no fim da curva, mas isso desencontrava esta curva da do posto,
    # que já mostra agosto. Melhor mostrar e avisar do que esconder.
    JANELA_FIM = "2026-08"
    do_ano = sorted(m for m in ({b["mes"] for b in blocos}
                    | {m for m in entrantes if m.startswith("2026")}
                    | {m for m in saidas if m.startswith("2026")})
                    if m <= JANELA_FIM)
    apos_janela = {
        "resolvidos": sum(q for m, q in saidas.items()
                          if m.startswith("2026") and m > JANELA_FIM),
        "entrantes": sum(q for m, q in entrantes.items()
                         if m.startswith("2026") and m > JANELA_FIM),
        "lista": [{"ativo": t["ativo"], "localidade": t["localidade"]}
                  for t in tratativas
                  if t["mes_resolucao"].startswith("2026") and t["mes_resolucao"] > JANELA_FIM],
    }
    # A conta por SS difere da conta por ativo: ativos com duas SS resolvidas na
    # foto contam uma vez aqui e duas lá. Guardo as duas para a conciliação.
    ss_resolvidas = sum(1 for i in itens if i["balde"] == "resolvidos")
    duplicados = sorted(a for a, g in ss_por_ativo.items()
                        if sum(1 for i in g if i["balde"] == "resolvidos") > 1)
    # Duas séries, por decisão do gestor: a carteira herdada JÁ É entrada.
    # Entrantes = os ativos da foto pelo mês de entrada, janeiro com o acervo;
    # Resolvidos = pelo mês da tratativa. Cada ativo conta uma vez em cada série.
    curva = [{
        "mes": m, "rotulo": rotulo(m),
        "entrantes": next((b["qtd"] for b in blocos if b["mes"] == m), 0),
        "resolvidos": saidas.get(m, 0),
    } for m in do_ano]

    # Livro-caixa da carteira herdada: o acervo de antes de 2026 é o saldo de
    # abertura; cada mês soma as SS abertas no próprio mês e desconta as tratadas.
    # Universo = os 117 da foto, que têm entrada E saída rastreadas. Quem passou
    # pelo posto por fora da foto fica anotado à parte: sem SS na foto de entrada
    # não há data de tratativa para dar baixa, e o ativo inflaria o saldo para
    # sempre.
    abertura_acervo = sum(1 for x in lista if x["legado"])
    abriu_no_mes = Counter(x["mes"] for x in lista if not x["legado"])
    baixas = Counter(t["mes_resolucao"] for t in tratativas if t["mes_resolucao"])
    primeiro_coep = {x["ativo"]: x["mes"] for x in detalhe_coep}
    na_foto = {x["ativo"] for x in lista}
    fora_do_livro = sum(1 for a, m in primeiro_coep.items()
                        if a not in na_foto and m in do_ano)

    saldo, corrente = [], abertura_acervo
    for m in do_ano:
        inicial = corrente
        e, s_ = abriu_no_mes.get(m, 0), baixas.get(m, 0)
        corrente = inicial + e - s_
        saldo.append({
            "mes": m, "rotulo": rotulo(m),
            "inicial": inicial, "entram": e, "saem": s_, "final": corrente,
        })

    return {
        "recorte": "todos os tipos de SS",
        "janela": "janeiro a agosto de 2026, com agosto até o dia 20",
        "apos_janela": apos_janela,
        "ss_resolvidas": ss_resolvidas,
        "resolvidos_duplicados": duplicados,
        "fora_do_recorte": fora_do_recorte,
        "curva": curva,
        "saldo": saldo,
        "abertura": abertura_acervo,
        "fora_do_livro": fora_do_livro,
        "serie_coep": serie_coep,
        "tratativas": tratativas,
        "total": total,
        "total_ss": len(itens),
        "resolvidos": resolvidos,
        "em_andamento": total - resolvidos,
        "fora_da_carteira": sum(1 for x in lista if not x["na_carteira"]),
        "meses": blocos,
        "lista": sorted(lista, key=lambda x: (x["mes"], x["abertura"], x["ativo"])),
        "legado": {
            "qtd": sum(1 for x in lista if x["legado"]),
            "por_ano": [{"ano": a, "qtd": q} for a, q in sorted(por_ano_legado.items())],
            "mais_antiga": mais_antiga if mais_antiga and mais_antiga["legado"] else None,
        },
        "fonte_data": [{"fonte": f, "qtd": q} for f, q in fonte.most_common()],
        "multiplas_ss": [
            {"ativo": a, "ss": sorted(o["numero_ss"] for o in g),
             "usada": por_ativo[a]["numero_ss"]}
            for a, g in sorted(ss_por_ativo.items()) if len(g) > 1
        ],
        "sem_data": [x["numero_ss"] for x in sem_data],
        "regra": (
            "Todos os tipos de SS contam. "
            "Mês da carteira de entrada = mês em que a SS foi aberta. SS aberta antes de "
            "2026 cai em janeiro de 2026, por decisão do gestor — janeiro concentra o que "
            "já era antigo quando o posto assumiu. Ativo com mais de uma SS na foto entra "
            "pela mais antiga."
        ),
    }


if __name__ == "__main__":
    with open(os.path.join(RAIZ, "data", "meta.json"), encoding="utf-8") as fh:
        meta = json.load(fh)
    d = montar(meta.get("entrada"))
    print(f"{d['total']} ativos em {len(d['meses'])} meses "
          f"({d['resolvidos']} resolvidos, {d['em_andamento']} em andamento)")
    for b in d["meses"]:
        extra = f"  ({b['legado']} vieram de antes de 2026)" if b["legado"] else ""
        print(f"  {b['rotulo']}  {b['qtd']:>3}  resolvidos {b['resolvidos']:>3} "
              f"({b['percentual']:>5.1f}%){extra}")
    print("  legado por ano:", {x["ano"]: x["qtd"] for x in d["legado"]["por_ano"]})
    print("  origem da data:", {x["fonte"]: x["qtd"] for x in d["fonte_data"]})
    print("  ativos com mais de uma SS:", len(d["multiplas_ss"]))
