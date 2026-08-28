"""
O painel RL × RT do posto — dist/PAINEL_RL_RT.xlsx.

Seis visões pedidas pelo gestor em 29/08, todas na régua de 28/08 (71 resolvidos ·
54 na fila · 125 na conta do posto), separadas por religador e regulador:

  1. Os 125: resolvidos e pendentes, por tipo
  2. O passivo — quantos vinham de anos anteriores, por tipo e por grupo
  3. Os 54 pendentes, por tipo, abertos pelo ano da ocorrência
  4. Os 125, por tipo e por ano da ocorrência
  5. Falha em 2026 e em 2025, mês a mês, da taxa de falha da planilha base, mais o
     confronto do mesmo mês nos dois anos
  6. O passivo mês a mês, pela data da ocorrência inicial
  7. A premissa de setembro a dezembro: os 54 pendentes mais o que 2025 queimou
     no mesmo quadrimestre
  8. A base dos 125, um por linha, com como cada um foi resolvido

REGRA DO ANO: vale sempre a DATA DE OCORRÊNCIA, nunca a abertura da SS nem o número
dela. Nos 71 é a ocorrência da demanda que o posto fechou; nos 54, a da demanda que
segue aberta — não a mais antiga do ativo, que pode ser de um ciclo já encerrado.

Rodar: python3 scripts/painel_rl_rt.py
"""

import json
import os
import sys
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.drawing.colors import ColorChoice
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(RAIZ, "scripts"))
import sla_manutencao as base  # noqa: E402
import sla_por_equipe as eq  # noqa: E402

COEP = os.path.join(RAIZ, "data", "missao", "coep_2026.json")
OCORR = os.path.join(RAIZ, "data", "missao", "ss_ocorrencia.json")
BASE = os.path.join(RAIZ, "data", "missao", "base_coep.json")
SAIDA = os.path.join(RAIZ, "dist", "PAINEL_RL_RT.xlsx")

TINTA, PAPEL, SINAL = "FF211D15", "FFF2EFE6", "FFBC4B0E"
# as duas séries, conferidas com o validador do dataviz sobre o papel claro
COR_RL, COR_RT = "1F7C50", "B8480C"
MES = ["jan", "fev", "mar", "abr", "mai", "jun",
       "jul", "ago", "set", "out", "nov", "dez"]
ANOS = ["2023", "2024", "2025", "2026"]


def apurar():
    reg = base.base_de_repasse(None)
    anterior = eq.raizes(reg)
    oc_ss = {}
    with open(OCORR, encoding="utf-8") as fh:
        for r in json.load(fh):
            k = base.norm(r["SS_ORIGINAL"])
            d = base.data(r.get("DTA_OCORRENCIA"))
            if k and d:
                oc_ss.setdefault(k, d)
    with open(COEP, encoding="utf-8") as fh:
        cp = json.load(fh)

    at = {a["ativo"]: a for a in cp["ativos"]}
    ss_info = {s["ss"]: s for s in cp["ss"]}
    res82 = {r["ativo"]: r for r in cp["resolvidos_do_coep"]
             if r["conta_como_resolvido_pelo_coep"]}
    fila = {a for a, v in at.items() if v["segue_no_posto"]}
    r71 = sorted(set(res82) - fila)
    assert len(res82) == 82 and len(r71) == 71 and len(fila) == 54, \
        f"esperado 82-11=71 e 54, veio {len(res82)}/{len(r71)}/{len(fila)}"

    def mes_da_demanda(ativo, pendente):
        if not pendente:
            d = res82[ativo].get("ocorrencia_da_demanda")   # dd/mm/aaaa
            if d:
                return f"{d[6:10]}-{d[3:5]}"
        melhor = None
        for p in str(at[ativo].get("ss") or "").split("|"):
            k = base.norm(p)
            if not k:
                continue
            s = ss_info.get(p.strip())
            if pendente and s and not s.get("segue_no_posto"):
                continue      # na fila só vale a cadeia da SS que segue aberta
            for c in (eq.do_comeco(k, anterior), k):
                d = oc_ss.get(c)
                if d and (melhor is None or d < melhor):
                    melhor = d
        return melhor.strftime("%Y-%m") if melhor else None

    mes = {a: mes_da_demanda(a, False) for a in r71}
    mes.update({a: mes_da_demanda(a, True) for a in fila})
    assert all(mes.values()), "ativo sem data de ocorrência"

    sig = {a: ("RL" if at[a]["tipo"] == "religador" else "RT") for a in mes}
    return {"resolvidos": r71, "fila": sorted(fila), "mes": mes, "sig": sig,
            "res": res82, "at": {a: at[a] for a in mes},
            "localidade": {a: at[a]["localidade"] for a in mes}}


def quadro(conj, sig, mes):
    """{tipo: {ano: n, passivo, total}} — o ano é o da ocorrência."""
    q = {}
    for t in ("RL", "RT", "Total"):
        d = {y: sum(1 for a in conj
                    if (sig[a] == t or t == "Total") and mes[a][:4] == y)
             for y in ANOS}
        d["passivo"] = sum(d[y] for y in ANOS if y != "2026")
        d["total"] = sum(d[y] for y in ANOS)
        q[t] = d
    return q


# ------------------------------------------------------------------ a planilha
def cabeca(ws, linha, titulo, cols):
    c = ws.cell(row=linha, column=1, value=titulo)
    c.font = Font(bold=True, size=11, color=SINAL)
    ws.append(cols)
    for i in range(1, len(cols) + 1):
        cel = ws.cell(row=linha + 1, column=i)
        cel.font = Font(bold=True, color=PAPEL, size=10)
        cel.fill = PatternFill("solid", fgColor=TINTA)
        cel.alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True)
    return linha + 2


def negrito(ws, linha, n):
    for i in range(1, n + 1):
        ws.cell(row=linha, column=i).font = Font(bold=True)
        ws.cell(row=linha, column=i).border = Border(
            top=Side(style="medium", color=TINTA))


def montar():
    d = apurar()
    sig, mes = d["sig"], d["mes"]
    res, fila = d["resolvidos"], d["fila"]
    todos = res + fila
    q_res, q_fila, q_125 = (quadro(res, sig, mes), quadro(fila, sig, mes),
                            quadro(todos, sig, mes))

    wb = Workbook()
    ws = wb.active
    ws.title = "Visão do posto"
    ws.column_dimensions["A"].width = 30
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 13

    # 1 — os 125: resolvidos e pendentes
    r = cabeca(ws, 1, "1 · Os 125 do posto em 2026",
               ["", "Resolvidos", "Na fila", "Total", "% resolvido"])
    for t in ("RL", "RT", "Total"):
        a, b = q_res[t]["total"], q_fila[t]["total"]
        ws.append(["Religador" if t == "RL" else
                   "Regulador" if t == "RT" else "Total", a, b, a + b,
                   round(a / (a + b), 4)])
        ws.cell(row=ws.max_row, column=5).number_format = "0.0%"
    negrito(ws, ws.max_row, 5)

    # 2 — o passivo
    r = cabeca(ws, ws.max_row + 2, "2 · Quantos vinham de anos anteriores",
               ["", "Passivo 23+24+25", "De 2026", "Total", "% passivo"])
    for rot, q in (("Resolvidos", q_res), ("Na fila", q_fila), ("Os 125", q_125)):
        for t in ("RL", "RT", "Total"):
            nome = ("Religador" if t == "RL" else
                    "Regulador" if t == "RT" else "Total")
            ws.append([f"{rot} · {nome}", q[t]["passivo"], q[t]["2026"],
                       q[t]["total"],
                       round(q[t]["passivo"] / q[t]["total"], 4) if q[t]["total"] else 0])
            ws.cell(row=ws.max_row, column=5).number_format = "0.0%"
            if t == "Total":
                negrito(ws, ws.max_row, 5)

    # 3 — os 54 pendentes
    r = cabeca(ws, ws.max_row + 2, "3 · Os 54 que estão na fila do posto",
               [""] + ANOS + ["Passivo", "Total"])
    for t in ("RL", "RT", "Total"):
        q = q_fila[t]
        ws.append(["Religador" if t == "RL" else
                   "Regulador" if t == "RT" else "Total"]
                  + [q[y] or "—" for y in ANOS] + [q["passivo"], q["total"]])
    negrito(ws, ws.max_row, 7)

    # 4 — os 125 por ano
    r = cabeca(ws, ws.max_row + 2, "4 · Os 125 pelo ano da ocorrência",
               [""] + ANOS + ["Passivo", "Total"])
    for t in ("RL", "RT", "Total"):
        q = q_125[t]
        ws.append(["Religador" if t == "RL" else
                   "Regulador" if t == "RT" else "Total"]
                  + [q[y] or "—" for y in ANOS] + [q["passivo"], q["total"]])
    negrito(ws, ws.max_row, 7)

    # 5 — falha mês a mês, da taxa de falha da planilha base
    with open(BASE, encoding="utf-8") as fh:
        bc = json.load(fh)
    aba_falha(wb, bc, 2026, 8,
              "5 · Quantos RL e RT falharam em 2026 (taxa de falha)")
    aba_falha(wb, bc, 2025, 12,
              "5b · Quantos RL e RT falharam em 2025 (taxa de falha)")
    aba_confronto(wb, bc)

    # 6 — o passivo mês a mês, pela ocorrência inicial
    ws3 = wb.create_sheet("Passivo mensal")
    ws3.column_dimensions["A"].width = 12
    for c in "BCDE":
        ws3.column_dimensions[c].width = 13
    cabeca(ws3, 1, "6 · O passivo pela data da ocorrência inicial",
           ["Mês", "Religador", "Regulador", "Total", "Acumulado"])
    passivo = [a for a in todos if mes[a][:4] != "2026"]
    c = Counter((mes[a], sig[a]) for a in passivo)
    ac = 0
    for m in sorted({k[0] for k in c}):
        a, b = c.get((m, "RL"), 0), c.get((m, "RT"), 0)
        ac += a + b
        ws3.append([f"{MES[int(m[5:7]) - 1]}/{m[2:4]}", a, b, a + b, ac])
    ws3.append(["Total", sum(1 for a in passivo if sig[a] == "RL"),
                sum(1 for a in passivo if sig[a] == "RT"), len(passivo), ""])
    negrito(ws3, ws3.max_row, 5)
    grafico(ws3, "Passivo por mês da ocorrência", 2, 3, ws3.max_row - 1, "G2")

    # 7 — a premissa de setembro a dezembro
    aba_premissa(wb, bc, q_fila)

    # 8 — a base dos 125
    aba_base(wb, d)

    # a régua
    ws4 = wb.create_sheet("Como foi feito")
    ws4.column_dimensions["A"].width = 98
    for t in TEXTO:
        ws4.append([t])
    ws4["A1"].font = Font(bold=True, size=12, color=SINAL)
    for i, t in enumerate(TEXTO, start=1):
        if t and not t.startswith(" ") and t.endswith(":"):
            ws4.cell(row=i, column=1).font = Font(bold=True, size=11)

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    return SAIDA, q_res, q_fila, q_125, len(passivo)



def aba_falha(wb, bc, ano, ate, titulo):
    """Uma aba de falha por ano — a série mensal da planilha base do gestor."""
    ws = wb.create_sheet(f"Falha {ano} mensal")
    ws.column_dimensions["A"].width = 12
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 13
    cabeca(ws, 1, titulo,
           ["Mês", "Religador", "Regulador", "Total", "Acumulado",
            "Parque RL", "Parque RT"])
    rl = bc["mensal"][f"RL|{ano}"][:ate]
    rt = bc["mensal"][f"RT|{ano}"][:ate]
    ac = 0
    for i in range(ate):
        f_rl, f_rt = rl[i]["falhas"], rt[i]["falhas"]
        ac += f_rl + f_rt
        ws.append([f"{MES[i]}/{str(ano)[2:]}", f_rl, f_rt, f_rl + f_rt, ac,
                   rl[i]["parque"], rt[i]["parque"]])
    s_rl = sum(x["falhas"] for x in rl)
    s_rt = sum(x["falhas"] for x in rt)
    ws.append(["Total", s_rl, s_rt, s_rl + s_rt, "",
               rl[-1]["parque"], rt[-1]["parque"]])
    negrito(ws, ws.max_row, 7)
    ws.append([])
    ws.append(["Taxa do ano", round(s_rl / rl[-1]["parque"], 4),
               round(s_rt / rt[-1]["parque"], 4),
               round((s_rl + s_rt) / (rl[-1]["parque"] + rt[-1]["parque"]), 4)])
    for c in (2, 3, 4):
        ws.cell(row=ws.max_row, column=c).number_format = "0.00%"
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    grafico(ws, f"Falhas de {ano} por mês", 2, 3, 2 + ate, "I2")
    return ws


def aba_confronto(wb, bc):
    """2025 contra 2026, mês a mês — o mesmo mês nos dois anos, lado a lado."""
    ws = wb.create_sheet("Falha 2025 x 2026")
    ws.column_dimensions["A"].width = 12
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 14
    cabeca(ws, 1, "5c · O mesmo mês nos dois anos",
           ["Mês", "RL 2025", "RL 2026", "RT 2025", "RT 2026",
            "Total 2025", "Total 2026"])
    d = {f"{t}|{a}": [x["falhas"] for x in bc["mensal"][f"{t}|{a}"]]
         for t in ("RL", "RT") for a in (2025, 2026)}
    for i in range(12):
        ate26 = i < 8      # 2026 só tem dado até agosto, a posição do relatório
        ws.append([MES[i],
                   d["RL|2025"][i], d["RL|2026"][i] if ate26 else "",
                   d["RT|2025"][i], d["RT|2026"][i] if ate26 else "",
                   d["RL|2025"][i] + d["RT|2025"][i],
                   (d["RL|2026"][i] + d["RT|2026"][i]) if ate26 else ""])
    p8 = lambda k: sum(d[k][:8])
    ws.append(["Jan–ago", p8("RL|2025"), p8("RL|2026"), p8("RT|2025"),
               p8("RT|2026"), p8("RL|2025") + p8("RT|2025"),
               p8("RL|2026") + p8("RT|2026")])
    negrito(ws, ws.max_row, 7)
    ws.append(["Ano fechado", sum(d["RL|2025"]), "", sum(d["RT|2025"]), "",
               sum(d["RL|2025"]) + sum(d["RT|2025"]), ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, italic=True)
    return ws


def aba_premissa(wb, bc, q_fila):
    """E se, de setembro a dezembro, queimar o mesmo que 2025 queimou?

    A premissa do gestor: os 54 que já estão parados em agosto, mais o que falhou no
    último quadrimestre de 2025. É cenário, não previsão — 2026 vinha com forma
    diferente de 2025, concentrando falha no começo do ano.
    """
    ws = wb.create_sheet("Premissa set–dez")
    ws.column_dimensions["A"].width = 34
    for c in "BCD":
        ws.column_dimensions[c].width = 14
    rl = [x["falhas"] for x in bc["mensal"]["RL|2025"]][8:]
    rt = [x["falhas"] for x in bc["mensal"]["RT|2025"]][8:]

    cabeca(ws, 1, "7 · Se de setembro a dezembro queimar o mesmo que 2025",
           ["", "Religador", "Regulador", "Total"])
    ws.append(["Parados no posto em 18/08", q_fila["RL"]["total"],
               q_fila["RT"]["total"], q_fila["Total"]["total"]])
    for i, m in enumerate(("setembro", "outubro", "novembro", "dezembro")):
        ws.append([f"+ falhas de {m} (como em 2025)", rl[i], rt[i], rl[i] + rt[i]])
    ws.append(["Somam no quadrimestre", sum(rl), sum(rt), sum(rl) + sum(rt)])
    negrito(ws, ws.max_row, 4)
    ws.append(["FORA NO FIM DO ANO, sem resolver nada",
               q_fila["RL"]["total"] + sum(rl), q_fila["RT"]["total"] + sum(rt),
               q_fila["Total"]["total"] + sum(rl) + sum(rt)])
    for c in range(1, 5):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True, size=11, color=SINAL)

    # como o estoque anda, mês a mês
    r = cabeca(ws, ws.max_row + 2, "O acúmulo, mês a mês",
               ["Mês", "Entram", "Estoque RL", "Estoque RT", "Estoque total"])
    erl, ert = q_fila["RL"]["total"], q_fila["RT"]["total"]
    ws.append(["ago/26 (hoje)", "—", erl, ert, erl + ert])
    for i, m in enumerate(("set", "out", "nov", "dez")):
        erl += rl[i]
        ert += rt[i]
        ws.append([f"{m}/26", rl[i] + rt[i], erl, ert, erl + ert])
    negrito(ws, ws.max_row, 5)
    grafico(ws, "Estoque de equipamentos fora, set a dez", r - 1, r, ws.max_row,
            "G2", cols=(3, 4))

    # e se o posto continuar resolvendo no ritmo de 2026?
    ritmo = round(62 / 8, 1)          # os 62 concluídos em oito meses
    ritmo_enc = round(48 / 8, 1)      # só as encerradas de ponta a ponta
    fim = q_fila["Total"]["total"] + sum(rl) + sum(rt)
    r = cabeca(ws, ws.max_row + 2, "E se o posto continuar resolvendo?",
               ["Cenário", "Por mês", "Resolve em 4 meses", "Fora no fim do ano"])
    ws.append(["Sem resolver nada", 0, 0, fim])
    ws.append(["No ritmo dos encerrados (48 em 8 meses)", ritmo_enc,
               round(ritmo_enc * 4), max(fim - round(ritmo_enc * 4), 0)])
    ws.append(["No ritmo do trabalho concluído (62 em 8 meses)", ritmo,
               round(ritmo * 4), max(fim - round(ritmo * 4), 0)])
    ws.append(["No ritmo da peça comprovada (25 em 8 meses)", round(25 / 8, 1),
               round(25 / 8 * 4), max(fim - round(25 / 8 * 4), 0)])
    for r2 in range(r, ws.max_row + 1):
        ws.cell(row=r2, column=4).font = Font(bold=True)
    return ws


def aba_base(wb, d):
    """Os 125, um por linha, com como cada um foi resolvido."""
    ws = wb.create_sheet("Base dos 125")
    cols = [("Ativo", 13), ("RL/RT", 7), ("Praça", 24), ("Criticidade", 13),
            ("Situação", 14), ("Ocorrência", 12), ("Ano", 7), ("Mês", 9),
            ("Passivo?", 10), ("SS no COEP", 40), ("SS que fechou", 20),
            ("Posto que fechou", 16), ("Como terminou", 14),
            ("Data do fechamento", 15), ("Dias da demanda", 13),
            ("Dias no posto", 12), ("Como foi resolvido", 62)]
    ws.append([c[0] for c in cols])
    at, res, sig, mes = d["at"], d["res"], d["sig"], d["mes"]
    for grupo, conj in (("Resolvido", d["resolvidos"]), ("Na fila", d["fila"])):
        for a in sorted(conj, key=lambda x: (sig[x], mes[x], x)):
            r = res.get(a, {})
            oc = r.get("ocorrencia_da_demanda") or ""
            if not oc and mes[a]:
                oc = f"{mes[a][5:7]}/{mes[a][:4]}"
            ws.append([
                a, sig[a], at[a]["localidade"], at[a].get("criticidade") or "—",
                grupo, oc, mes[a][:4], f"{MES[int(mes[a][5:7]) - 1]}/{mes[a][2:4]}",
                "passivo" if mes[a][:4] != "2026" else "de 2026",
                at[a].get("ss") or "", r.get("ss_que_fechou") or "",
                r.get("posto_que_fechou") or "", r.get("como_terminou") or "",
                r.get("data_do_fechamento") or "", r.get("dias_da_demanda") or "",
                at[a].get("dias_no_posto") or "",
                r.get("prova") or ("segue pendente no posto" if grupo == "Na fila"
                                   else ""),
            ])
    for i, (_, larg) in enumerate(cols, start=1):
        cel = ws.cell(row=1, column=i)
        cel.font = Font(bold=True, color=PAPEL, size=10)
        cel.fill = PatternFill("solid", fgColor=TINTA)
        cel.alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30

    # o resumo de como os 71 foram resolvidos
    ws2 = wb.create_sheet("Como os 71 foram resolvidos")
    ws2.column_dimensions["A"].width = 56
    for c in "BCD":
        ws2.column_dimensions[c].width = 13
    r = cabeca(ws2, 1, "8 · Por que cada um saiu da fila",
               ["Prova registrada", "Religador", "Regulador", "Total"])
    provas = Counter((res[a]["prova"], sig[a]) for a in d["resolvidos"])
    for pr in sorted({k[0] for k in provas}):
        ws2.append([pr, provas.get((pr, "RL"), 0), provas.get((pr, "RT"), 0),
                    provas.get((pr, "RL"), 0) + provas.get((pr, "RT"), 0)])
    ws2.append(["Total", sum(1 for a in d["resolvidos"] if sig[a] == "RL"),
                sum(1 for a in d["resolvidos"] if sig[a] == "RT"),
                len(d["resolvidos"])])
    negrito(ws2, ws2.max_row, 4)

    r = cabeca(ws2, ws2.max_row + 2, "Pelo desfecho da SS",
               ["Como terminou", "Religador", "Regulador", "Total"])
    c = Counter((res[a]["como_terminou"], sig[a]) for a in d["resolvidos"])
    for k in sorted({x[0] for x in c}):
        ws2.append([k, c.get((k, "RL"), 0), c.get((k, "RT"), 0),
                    c.get((k, "RL"), 0) + c.get((k, "RT"), 0)])
    negrito(ws2, ws2.max_row, 4)

    r = cabeca(ws2, ws2.max_row + 2, "Pelo posto que fechou a cadeia",
               ["Posto", "Religador", "Regulador", "Total"])
    c = Counter((res[a]["posto_que_fechou"], sig[a]) for a in d["resolvidos"])
    for k, _ in Counter(res[a]["posto_que_fechou"]
                        for a in d["resolvidos"]).most_common():
        ws2.append([k, c.get((k, "RL"), 0), c.get((k, "RT"), 0),
                    c.get((k, "RL"), 0) + c.get((k, "RT"), 0)])
    negrito(ws2, ws2.max_row, 4)
    return ws


def grafico(ws, titulo, cab, r0, r1, onde, cols=(2, 3)):
    """cab = linha do cabeçalho (de onde sai o nome da série); r0..r1 = os dados.

    `cols` são as duas colunas das séries — nem sempre B e C: na aba da premissa o
    que empilha é o estoque de RL e RT, que estão em C e D.
    """
    g = BarChart()
    g.type, g.grouping, g.overlap = "col", "stacked", 100
    g.title, g.height, g.width, g.gapWidth = titulo, 8.5, 22, 70
    for col, cor in zip(cols, (COR_RL, COR_RT)):
        s = Series(Reference(ws, min_col=col, min_row=cab, max_row=r1),
                   title_from_data=True)
        s.graphicalProperties.solidFill = ColorChoice(srgbClr=cor)
        s.graphicalProperties.line.noFill = True
        g.series.append(s)
    # mês é texto: sem StrRef o Excel numera as categorias
    cats = AxDataSource(strRef=StrRef(f=f"'{ws.title}'!$A${r0}:$A${r1}"))
    for s in g.series:
        s.cat = cats
    ws.add_chart(g, onde)


TEXTO = [
    "O painel RL × RT do posto do COEP",
    "",
    "A régua:",
    "Tudo aqui está na conta de 28/08: 71 resolvidos, 54 na fila, 125 na conta do",
    "posto. São 82 ativos que fecharam a demanda no ano, menos os 11 que resolveram e",
    "voltaram para a fila — pela decisão do gestor, quem voltou não conta como",
    "resolvido.",
    "",
    "O ano é sempre o da OCORRÊNCIA:",
    "Nunca a abertura da SS, nunca o número dela. A abertura vem em média 39 dias",
    "depois do fato e em 9,8% dos casos cai em outro ano; e a ETO-COEP 00149/2025 foi",
    "aberta em 29/06/2026. Numerar não é abrir, e abrir não é o fato acontecer.",
    "",
    "Qual ocorrência, quando o ativo tem várias:",
    "Nos 71 resolvidos vale a ocorrência da demanda que o posto FECHOU. Nos 54 da fila",
    "vale a da demanda que SEGUE ABERTA — não a mais antiga do ativo, que pode ser de",
    "um ciclo já encerrado. Equipamento reincidente tem mais de uma cadeia, e misturar",
    "as duas envelhece o passivo artificialmente.",
    "",
    "As abas 5 e 6 respondem a perguntas diferentes:",
    "A aba 5 é a TAXA DE FALHA: quantos equipamentos falharam em 2026, da planilha base",
    "de equipamentos especiais. Conta falha do parque inteiro, tenha ela passado pelo",
    "COEP ou não.",
    "",
    "A aba 6 é o PASSIVO DO POSTO: dos 125 que passaram pelo COEP, quantos vinham de",
    "ocorrência de 2023, 2024 ou 2025, e em que mês esse fato aconteceu. É a idade da",
    "fila, não a taxa.",
    "",
    "Os dois números não se somam nem se comparam direto: o primeiro é do parque, o",
    "segundo é da mesa.",
    "",
    "A premissa de setembro a dezembro é CENÁRIO, não previsão:",
    "Ela repete em 2026 o que 2025 queimou de setembro a dezembro — 11 religadores e",
    "6 reguladores — em cima dos 54 já parados em 18/08. Dá 71 fora no fim do ano se o",
    "posto não resolver mais nada.",
    "",
    "Mas 2026 não vinha com a forma de 2025. Este ano concentrou falha no começo: 24",
    "das 37 caíram de janeiro a março, e de maio a agosto foram só 8. Repetir o",
    "quadrimestre de 2025 é, por isso, uma premissa PESSIMISTA para o religador — em",
    "2025 setembro a dezembro trouxe 11 RL, e o ritmo de 2026 desde maio é bem menor.",
    "",
    "Por isso a aba traz também o outro lado: quanto o posto derruba se continuar no",
    "ritmo que teve. No ritmo do trabalho concluído (62 em oito meses) sobrariam 40",
    "fora no fim do ano, não 71.",
    "",
    "Cuidado com o parque das abas de falha:",
    "A coluna Parque é a da série mensal da planilha base do gestor — 1.281 RL e 180 RT",
    "em janeiro de 2026, mais a expansão somada no próprio mês. Para 2025 essa base foi",
    "carregada para trás, então ela é o parque do fim de 2025, não a média do ano.",
    "",
    "A taxa oficial dos três anos usa outro parque: 1.307 RL e 207 RT. Nessa base as",
    "taxas de 2025 ficam RL 2,68% e RT 8,70%, contra os 2,73% e 10,00% da aba. A",
    "diferença é só o denominador; o número de falhas é o mesmo.",
    "",
    "Cada ativo conta uma vez:",
    "125 ativos distintos em 125 linhas. Equipamento que passou pelo posto duas vezes",
    "no ano ocupa uma linha só.",
]


if __name__ == "__main__":
    caminho, q_res, q_fila, q_125, npas = montar()
    print("gravado:", caminho)
    print(f"  resolvidos {q_res['Total']['total']} "
          f"(RL {q_res['RL']['total']} · RT {q_res['RT']['total']})")
    print(f"  na fila    {q_fila['Total']['total']} "
          f"(RL {q_fila['RL']['total']} · RT {q_fila['RT']['total']})")
    print(f"  os 125     RL {q_125['RL']['total']} · RT {q_125['RT']['total']}")
    print(f"  passivo    {npas} de {q_125['Total']['total']} "
          f"(RL {q_125['RL']['passivo']} · RT {q_125['RT']['passivo']})")
